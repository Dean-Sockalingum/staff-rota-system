"""
Head of Service Dashboard View
===============================

Aggregated multi-home dashboard for Head of Service team oversight.
Provides governance and strategic oversight across all 5 care homes for:
- Quality assurance
- Fiscal monitoring
- Staff allocation
- Compliance tracking

Access: SM (Service Manager), OM (Operations Manager), HOS (Head of Service), IDI (IDI Team)
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q, F, Prefetch
from django.db import models
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import json

from .models import Shift, Unit, User, LeaveRequest, StaffReallocation, Resident, CarePlanReview
from .models_multi_home import CareHome
from .decorators_api import api_login_required
# Automated workflow models - to be integrated in Phase 2
# from .models_automated_workflow import (
#     StaffingCoverRequest, ReallocationRequest, AgencyRequest
# )


@login_required
def senior_management_dashboard(request):
    """
    Head of Service Dashboard - Multi-home governance oversight
    
    Provides aggregated strategic view across all 5 care homes:
    - Orchard Grove, Meadowburn, Hawthorn House, Riverside, Victoria Gardens
    
    Restricted to: SM, OM, HOS, IDI roles
    
    Supports date range filtering via GET parameters:
    - start_date: YYYY-MM-DD format (default: today)
    - end_date: YYYY-MM-DD format (default: today)
    """
    
    # Check if user has Head of Service team permissions
    # HOS team (SM, OM, HOS, IDI) sits above home-level staff and provides governance oversight across all 5 homes
    # Superusers also have full access regardless of role
    if not (request.user.is_superuser or (request.user.role and request.user.role.is_senior_management_team)):
        return render(request, 'scheduling/access_denied.html', {
            'message': 'This dashboard is restricted to Head of Service team members only (SM, OM, HOS, IDI). Access provides governance oversight across all care homes.'
        })
    
    today = timezone.now().date()
    
    # Handle date range parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_home = request.GET.get('care_home', '')  # Filter by care home
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today
    else:
        start_date = today
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = today
    else:
        end_date = today
    
    # Ensure start_date is before end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    
    # Check if export is requested
    export_format = request.GET.get('export')
    if export_format == 'csv':
        return senior_dashboard_export(request)
    
    days_in_range = (end_date - start_date).days + 1
    
    current_month_start = today.replace(day=1)
    next_month_start = (current_month_start + timedelta(days=32)).replace(day=1)
    
    # Get all care homes with optimized prefetch for units and today's shifts
    care_homes = CareHome.objects.prefetch_related(
        'units',  # Load all units in one query
        Prefetch(
            'units__shift_set',
            queryset=Shift.objects.filter(
                date=today,
                status__in=['SCHEDULED', 'CONFIRMED']
            ).select_related('user', 'shift_type'),
            to_attr='today_shifts'
        )
    ).order_by('name')
    
    # Filter by selected home if provided
    if selected_home:
        care_homes = care_homes.filter(name=selected_home)
    
    # =================================================================
    # SECTION 1: HOME OVERVIEW - Occupancy & Capacity
    # =================================================================
    home_overview = []
    total_capacity = 0
    total_occupancy = 0
    
    for home in care_homes:
        # Use prefetched units instead of querying database
        units = [u for u in home.units.all() if u.is_active]
        unit_count = len(units)
        
        occupancy_rate = (home.current_occupancy / home.bed_capacity * 100) if home.bed_capacity > 0 else 0
        
        home_overview.append({
            'home': home,
            'display_name': home.get_name_display(),
            'occupancy': home.current_occupancy,
            'capacity': home.bed_capacity,
            'occupancy_rate': occupancy_rate,
            'units': unit_count,
            'location': home.location_address,
        })
        
        total_capacity += home.bed_capacity
        total_occupancy += home.current_occupancy
    
    overall_occupancy_rate = (total_occupancy / total_capacity * 100) if total_capacity > 0 else 0
    
    # =================================================================
    # SECTION 2: STAFFING LEVELS - Date Range Coverage
    # =================================================================
    staffing_today = []
    
    # Home-specific staffing requirements
    # Minimum = Regulatory/safety minimum (after factoring leave/sickness)
    # Ideal = Full rota when everyone scheduled (25-41 range from base model)
    home_staffing = {
        'HAWTHORN_HOUSE': {'day_min': 18, 'day_ideal': 41, 'night_min': 18, 'night_ideal': 41},
        'MEADOWBURN': {'day_min': 17, 'day_ideal': 41, 'night_min': 17, 'night_ideal': 41},
        'ORCHARD_GROVE': {'day_min': 17, 'day_ideal': 41, 'night_min': 17, 'night_ideal': 41},
        'RIVERSIDE': {'day_min': 17, 'day_ideal': 41, 'night_min': 17, 'night_ideal': 41},
        'VICTORIA_GARDENS': {'day_min': 10, 'day_ideal': 18, 'night_min': 10, 'night_ideal': 14},
    }
    
    for home in care_homes:
        # Use prefetched units and shifts
        units = [u for u in home.units.all() if u.is_active]
        
        # Collect today's shifts from prefetched data
        today_shifts = []
        for unit in units:
            if hasattr(unit, 'today_shifts'):
                today_shifts.extend(unit.today_shifts)
        
        # Count by shift type - use set to ensure unique users
        day_users = set()
        night_users = set()
        for shift in today_shifts:
            if shift.user:
                if 'DAY' in shift.shift_type.name.upper():
                    day_users.add(shift.user.pk)
                elif 'NIGHT' in shift.shift_type.name.upper():
                    night_users.add(shift.user.pk)
        
        day_shifts = len(day_users)
        night_shifts = len(night_users)
        
        # Get home-specific requirements
        requirements = home_staffing.get(home.name, {'day_min': 17, 'day_ideal': 24, 'night_min': 17, 'night_ideal': 21})
        
        # Calculate required vs actual (for TODAY only, not per day in range)
        day_required = requirements['day_min']
        night_required = requirements['night_min']
        
        day_coverage = (day_shifts / day_required * 100) if day_required > 0 else 100
        night_coverage = (night_shifts / night_required * 100) if night_required > 0 else 100
        
        staffing_today.append({
            'home': home.get_name_display(),
            'day_actual': day_shifts,
            'day_required': day_required,
            'day_coverage': day_coverage,
            'night_actual': night_shifts,
            'night_required': night_required,
            'night_coverage': night_coverage,
            'status': 'good' if day_coverage >= 100 and night_coverage >= 100 else 'critical',
        })
    
    # =================================================================
    # SECTION 3: FISCAL MONITORING - Monthly Budget Tracking
    # =================================================================
    fiscal_summary = []
    total_agency_budget = 0
    total_agency_spend = 0
    total_ot_budget = 0
    total_ot_spend = 0
    
    # Optimize: Single aggregated query for all homes at once
    all_units = []
    for home in care_homes:
        all_units.extend([u for u in home.units.all()])
    
    # Get shifts grouped by care_home in single query
    fiscal_stats = Shift.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        status__in=['SCHEDULED', 'CONFIRMED']
    ).values('unit__care_home').annotate(
        agency_count=Count('pk', filter=Q(user__isnull=True)),
        ot_count=Count('pk', filter=Q(shift_classification='OVERTIME'))
    )
    
    # Create lookup dict for quick access
    fiscal_lookup = {stat['unit__care_home']: stat for stat in fiscal_stats}
    
    for home in care_homes:
        # Get stats from aggregated query
        stats = fiscal_lookup.get(home.pk, {'agency_count': 0, 'ot_count': 0})
        agency_shifts_count = stats['agency_count']
        ot_shifts = stats['ot_count']
        
        # Estimate: £300 per agency shift (rough average)
        agency_spend = Decimal(str(agency_shifts_count * 300))
        # OT spend for date range (estimated at £25/hour, 12 hours per shift)
        ot_spend = Decimal(str(ot_shifts * 25 * 12))  # Rough estimate
        
        agency_budget = home.budget_agency_monthly
        ot_budget = home.budget_overtime_monthly
        
        agency_utilization = (float(agency_spend) / float(agency_budget) * 100) if agency_budget > 0 else 0
        ot_utilization = (float(ot_spend) / float(ot_budget) * 100) if ot_budget > 0 else 0
        
        fiscal_summary.append({
            'home': home.get_name_display(),
            'agency_budget': agency_budget,
            'agency_spend': agency_spend,
            'agency_utilization': agency_utilization,
            'ot_budget': ot_budget,
            'ot_spend': ot_spend,
            'ot_utilization': ot_utilization,
            'total_budget': agency_budget + ot_budget,
            'total_spend': agency_spend + ot_spend,
            'fiscal_status': 'good' if agency_utilization < 80 and ot_utilization < 80 else 'warning' if agency_utilization < 100 and ot_utilization < 100 else 'over_budget',
        })
        
        total_agency_budget += float(agency_budget)
        total_agency_spend += float(agency_spend)
        total_ot_budget += float(ot_budget)
        total_ot_spend += float(ot_spend)
    
    # =================================================================
    # SECTION 4: STAFFING ALERTS & CRITICAL ISSUES
    # =================================================================
    critical_alerts = []
    
    # TODO: Integrate with StaffingAlert model
    # For now, identify unfilled shifts as "alerts"
    next_week = today + timedelta(days=7)
    
    # Optimize: Get all unfilled shifts in single query, then group by home
    all_home_units = [u for home in care_homes for u in home.units.all()]
    all_unfilled = Shift.objects.filter(
        unit__in=all_home_units,
        date__gte=today,
        date__lte=next_week,
        user__isnull=True,
        status='SCHEDULED'
    ).select_related('unit', 'unit__care_home')[:25]  # Top 25 total
    
    # Group by home
    unfilled_by_home = {}
    for shift in all_unfilled:
        home_name = shift.unit.care_home.name
        if home_name not in unfilled_by_home:
            unfilled_by_home[home_name] = []
        if len(unfilled_by_home[home_name]) < 5:  # Top 5 per home
            unfilled_by_home[home_name].append(shift)
    
    for home in care_homes:
        unfilled = unfilled_by_home.get(home.name, [])
        
        for shift in unfilled:
            critical_alerts.append({
                'home': home.get_name_display(),
                'date': shift.date,
                'unit': shift.unit.name,
                'severity': 'HIGH' if shift.date <= today + timedelta(days=2) else 'MEDIUM',
                'created': shift.created_at if hasattr(shift, 'created_at') else today,
                'age_hours': ((timezone.now() - shift.created_at).total_seconds() / 3600) if hasattr(shift, 'created_at') else 0,
            })
    
    # Sort by age (oldest first)
    critical_alerts = sorted(critical_alerts, key=lambda x: x['age_hours'], reverse=True)[:20]
    
    # =================================================================
    # SECTION 5: PENDING ACTIONS - Require Management Attention
    # =================================================================
    
    # Manual review leave requests (all homes)
    manual_reviews = LeaveRequest.objects.filter(
        status='MANUAL_REVIEW'
    ).select_related('user', 'user__unit').order_by('created_at')
    
    pending_by_home = {}
    for request in manual_reviews:
        if request.user and request.user.unit:
            home = request.user.unit.care_home
            home_name = home.get_name_display() if home else 'Unknown'
            if home_name not in pending_by_home:
                pending_by_home[home_name] = []
            pending_by_home[home_name].append(request)
    
    # Pending reallocations
    pending_reallocations = StaffReallocation.objects.filter(
        status='NEEDED'
    ).select_related('target_unit', 'target_unit__care_home').count()
    
    # Unfilled cover requests (count unfilled shifts)
    unfilled_covers = Shift.objects.filter(
        date__gte=today,
        user__isnull=True,
        status='SCHEDULED'
    ).count()
    
    # =================================================================
    # SECTION 6: QUALITY METRICS
    # =================================================================
    quality_metrics = []
    
    # Last 30 days metrics
    last_30_days = today - timedelta(days=30)
    
    # Optimize: Single aggregated query for quality metrics
    quality_stats = Shift.objects.filter(
        date__gte=last_30_days,
        date__lte=today
    ).values('unit__care_home').annotate(
        total_shifts=Count('pk'),
        unfilled_shifts=Count('pk', filter=Q(user__isnull=True))
    )
    
    quality_lookup = {stat['unit__care_home']: stat for stat in quality_stats}
    
    # Get staff counts by home in single query
    staff_counts = User.objects.filter(
        is_active=True
    ).values('unit__care_home').annotate(
        staff_count=Count('pk')
    )
    staff_lookup = {stat['unit__care_home']: stat['staff_count'] for stat in staff_counts}
    
    for home in care_homes:
        # Get stats from aggregated queries
        stats = quality_lookup.get(home.pk, {'total_shifts': 0, 'unfilled_shifts': 0})
        total_shifts = stats['total_shifts']
        unfilled_shifts = stats['unfilled_shifts']
        
        agency_rate = (unfilled_shifts / total_shifts * 100) if total_shifts > 0 else 0
        
        # Staff turnover (placeholder - would need actual termination data)
        staff_count = staff_lookup.get(home.pk, 0)
        
        quality_metrics.append({
            'home': home.get_name_display(),
            'agency_rate': agency_rate,
            'staff_count': staff_count,
            'shifts_30d': total_shifts,
            'quality_score': 100 - agency_rate,  # Simple inverse for now
        })
    
    # =================================================================
    # SECTION 7: CARE PLAN COMPLIANCE - By Home
    # =================================================================
    # Always show all homes for governance reporting regardless of filter
    care_plan_compliance = []
    all_homes_for_compliance = CareHome.objects.all().order_by('name')
    
    for home in all_homes_for_compliance:
        # Get all active residents for this home
        residents = Resident.objects.filter(
            unit__care_home=home,
            is_active=True
        )
        
        total_residents = residents.count()
        
        # Get latest review for each resident
        reviews = []
        for resident in residents:
            latest_review = resident.care_plan_reviews.order_by('-due_date').first()
            if latest_review:
                reviews.append(latest_review)
        
        # Count by status
        completed = sum(1 for r in reviews if r.status == 'COMPLETED')
        overdue = sum(1 for r in reviews if r.status == 'OVERDUE')
        due_soon = sum(1 for r in reviews if r.status == 'DUE')
        upcoming = sum(1 for r in reviews if r.status == 'UPCOMING')
        in_progress = sum(1 for r in reviews if r.status == 'IN_PROGRESS')
        
        # Calculate compliance rate (completed / total reviews)
        compliance_rate = (completed / len(reviews) * 100) if reviews else 0
        overdue_rate = (overdue / len(reviews) * 100) if reviews else 0
        
        # Always add home to the list for governance visibility
        care_plan_compliance.append({
            'home': home.get_name_display(),
            'home_obj': home,
            'total_residents': total_residents,
            'total_reviews': len(reviews),
            'completed': completed,
            'overdue': overdue,
            'overdue_rate': overdue_rate,
            'due_soon': due_soon,
            'upcoming': upcoming,
            'in_progress': in_progress,
            'compliance_rate': compliance_rate,
        })
    
    # =================================================================
    # SECTION 8: STAFF VACANCIES TRACKING
    # =================================================================
    from staff_records.models import StaffProfile
    
    vacancies = []
    
    # Get all leavers in the last 90 days and upcoming leavers
    ninety_days_ago = today - timedelta(days=90)
    thirty_days_forward = today + timedelta(days=30)
    
    # Get staff profiles for leavers
    leaver_profiles = StaffProfile.objects.filter(
        employment_status='LEAVER',
        end_date__isnull=False
    ).select_related(
        'user',
        'user__role',
        'user__unit',
        'user__unit__care_home'
    ).filter(
        Q(end_date__gte=ninety_days_ago) | Q(end_date__lte=thirty_days_forward)
    ).order_by('-end_date')
    
    for profile in leaver_profiles:
        if not profile.user or not profile.user.unit or not profile.user.unit.care_home:
            continue
            
        # Calculate days vacant
        if profile.end_date <= today:
            days_vacant = (today - profile.end_date).days
            status = 'VACANT'
        else:
            days_vacant = 0
            status = 'UPCOMING'
        
        # Get contracted hours from shifts_per_week (assume 12-hour shifts)
        hours_per_week = profile.user.shifts_per_week_override or 40  # Default 40 hours
        
        vacancies.append({
            'home': profile.user.unit.care_home.get_name_display(),
            'home_obj': profile.user.unit.care_home,
            'role': profile.user.role.get_name_display() if profile.user.role else 'Unknown',
            'role_code': profile.user.role.name if profile.user.role else 'UNKNOWN',
            'staff_name': profile.user.full_name,
            'sap': profile.user.sap,
            'unit': profile.user.unit.name,
            'end_date': profile.end_date,
            'days_vacant': days_vacant,
            'hours_per_week': hours_per_week,
            'status': status,
            'severity': 'HIGH' if days_vacant > 30 else 'MEDIUM' if days_vacant > 14 else 'LOW' if status == 'VACANT' else 'INFO',
        })
    
    # Group by home for summary
    vacancies_by_home = {}
    for v in vacancies:
        home_name = v['home']
        if home_name not in vacancies_by_home:
            vacancies_by_home[home_name] = {
                'home': home_name,
                'total_vacancies': 0,
                'vacant_now': 0,
                'upcoming_leavers': 0,
                'total_hours_vacant': 0,
                'roles': {}
            }
        
        vacancies_by_home[home_name]['total_vacancies'] += 1
        
        if v['status'] == 'VACANT':
            vacancies_by_home[home_name]['vacant_now'] += 1
            vacancies_by_home[home_name]['total_hours_vacant'] += v['hours_per_week']
        else:
            vacancies_by_home[home_name]['upcoming_leavers'] += 1
        
        role = v['role_code']
        if role not in vacancies_by_home[home_name]['roles']:
            vacancies_by_home[home_name]['roles'][role] = 0
        vacancies_by_home[home_name]['roles'][role] += 1
    
    # =================================================================
    # SECTION 9: CITYWIDE SUMMARY - Weekly Overview
    # =================================================================
    # Generate 7-day view starting from start_date
    citywide_day_summary = []
    citywide_night_summary = []
    
    # Create list of dates for the week
    week_dates = [start_date + timedelta(days=i) for i in range(min(7, days_in_range))]
    
    for home in care_homes:
        units = Unit.objects.filter(care_home=home, is_active=True)
        
        # Specific staffing requirements per home (from Head of Service specification)
        home_staffing = {
            'HAWTHORN_HOUSE': {'day_sscw_min': 2, 'day_sscw_ideal': 2, 'day_min': 18, 'day_ideal': 24, 'night_sscwn_min': 2, 'night_sscwn_ideal': 2, 'night_min': 18, 'night_ideal': 21},
            'MEADOWBURN': {'day_sscw_min': 2, 'day_sscw_ideal': 2, 'day_min': 17, 'day_ideal': 24, 'night_sscwn_min': 2, 'night_sscwn_ideal': 2, 'night_min': 17, 'night_ideal': 21},
            'ORCHARD_GROVE': {'day_sscw_min': 2, 'day_sscw_ideal': 2, 'day_min': 17, 'day_ideal': 24, 'night_sscwn_min': 2, 'night_sscwn_ideal': 2, 'night_min': 17, 'night_ideal': 21},
            'RIVERSIDE': {'day_sscw_min': 2, 'day_sscw_ideal': 2, 'day_min': 17, 'day_ideal': 24, 'night_sscwn_min': 2, 'night_sscwn_ideal': 2, 'night_min': 17, 'night_ideal': 21},
            'VICTORIA_GARDENS': {'day_sscw_min': 1, 'day_sscw_ideal': 2, 'day_min': 10, 'day_ideal': 15, 'night_sscwn_min': 1, 'night_sscwn_ideal': 2, 'night_min': 10, 'night_ideal': 12},
        }
        
        requirements = home_staffing.get(home.name, {'day_sscw_min': 2, 'day_sscw_ideal': 2, 'day_min': 17, 'day_ideal': 24, 'night_sscwn_min': 2, 'night_sscwn_ideal': 2, 'night_min': 17, 'night_ideal': 21})
        
        # Day shift summary for this home
        day_data = {
            'home': home.get_name_display(),
            'home_obj': home,
            'sscw_min': requirements['day_sscw_min'],
            'sscw_ideal': requirements['day_sscw_ideal'],
            'min_required': requirements['day_min'],
            'ideal_required': requirements['day_ideal'],
            'days': []
        }
        
        # Night shift summary for this home
        night_data = {
            'home': home.get_name_display(),
            'home_obj': home,
            'sscwn_min': requirements['night_sscwn_min'],
            'sscwn_ideal': requirements['night_sscwn_ideal'],
            'min_required': requirements['night_min'],
            'ideal_required': requirements['night_ideal'],
            'days': []
        }
        
        # For each day in the week
        for date in week_dates:
            # Get shifts for this day
            day_shifts_for_date = Shift.objects.filter(
                unit__in=units,
                date=date,
                shift_type__name__icontains='DAY',
                status__in=['SCHEDULED', 'CONFIRMED']
            )
            
            night_shifts_for_date = Shift.objects.filter(
                unit__in=units,
                date=date,
                shift_type__name__icontains='NIGHT',
                status__in=['SCHEDULED', 'CONFIRMED']
            )
            
            # Count by role and type
            day_seniors = day_shifts_for_date.filter(
                user__role__name='SSCW'
            ).count()
            day_staff = day_shifts_for_date.filter(
                user__role__name__in=['SCW', 'SCA']
            ).count()
            day_leave = LeaveRequest.objects.filter(
                user__unit__in=units,
                start_date__lte=date,
                end_date__gte=date,
                status='APPROVED',
                user__shift_preference__icontains='DAY'
            ).count()
            day_agency = day_shifts_for_date.filter(
                user__sap__startswith='AGENCY'
            ).count() if day_shifts_for_date.filter(user__sap__isnull=False).exists() else 0
            day_overtime = day_shifts_for_date.filter(
                is_overtime=True
            ).count() if hasattr(day_shifts_for_date.first(), 'is_overtime') else 0
            day_absent = day_shifts_for_date.filter(
                status='ABSENT'
            ).count()
            
            night_seniors = night_shifts_for_date.filter(
                user__role__name='SSCWN'
            ).count()
            night_staff = night_shifts_for_date.filter(
                user__role__name__in=['SCWN', 'SCAN']
            ).count()
            night_leave = LeaveRequest.objects.filter(
                user__unit__in=units,
                start_date__lte=date,
                end_date__gte=date,
                status='APPROVED',
                user__shift_preference__icontains='NIGHT'
            ).count()
            night_agency = night_shifts_for_date.filter(
                user__sap__startswith='AGENCY'
            ).count() if night_shifts_for_date.filter(user__sap__isnull=False).exists() else 0
            night_overtime = night_shifts_for_date.filter(
                is_overtime=True
            ).count() if hasattr(night_shifts_for_date.first(), 'is_overtime') else 0
            night_absent = night_shifts_for_date.filter(
                status='ABSENT'
            ).count()
            
            # Add day data
            day_data['days'].append({
                'date': date,
                'day_name': date.strftime('%a'),
                'day_num': date.day,
                'seniors': day_seniors,
                'staff': day_staff,
                'leave': day_leave,
                'agency': day_agency,
                'overtime': day_overtime,
                'absent': day_absent,
                'total': day_seniors + day_staff,
            })
            
            # Add night data
            night_data['days'].append({
                'date': date,
                'day_name': date.strftime('%a'),
                'day_num': date.day,
                'seniors': night_seniors,
                'staff': night_staff,
                'leave': night_leave,
                'agency': night_agency,
                'overtime': night_overtime,
                'absent': night_absent,
                'total': night_seniors + night_staff,
            })
        
        citywide_day_summary.append(day_data)
        citywide_night_summary.append(night_data)
    
    # =================================================================
    # SECTION 8: WEEKLY SNAPSHOT - Simplified Daily Summary per Home
    # =================================================================
    # Generate simplified weekly summary showing SSCW/SSCWN and total Care staff
    weekly_snapshot = []
    
    # Get all care homes for snapshot (respect filtering)
    all_care_homes = CareHome.objects.all().order_by('name')
    if selected_home:
        all_care_homes = all_care_homes.filter(name=selected_home)
    
    for home in all_care_homes:
        units = Unit.objects.filter(care_home=home, is_active=True)
        
        # Home-specific requirements (care staff only, not including SSCW/SSCWN)
        home_staffing_config = {
            'HAWTHORN_HOUSE': {'day_min': 18, 'day_ideal': 18, 'night_min': 18, 'night_ideal': 18, 'day_sscw': 2, 'night_sscw': 2},
            'MEADOWBURN': {'day_min': 17, 'day_ideal': 17, 'night_min': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
            'ORCHARD_GROVE': {'day_min': 17, 'day_ideal': 17, 'night_min': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
            'RIVERSIDE': {'day_min': 17, 'day_ideal': 17, 'night_min': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
            'VICTORIA_GARDENS': {'day_min': 10, 'day_ideal': 10, 'night_min': 10, 'night_ideal': 10, 'day_sscw': 1, 'night_sscw': 1},
        }
        
        requirements = home_staffing_config.get(home.name, {'day_min': 17, 'day_ideal': 17, 'night_min': 17, 'night_ideal': 17})
        
        home_snapshot = {
            'home': home.get_name_display(),
            'home_obj': home,
            'day_ideal': requirements['day_ideal'],
            'night_ideal': requirements['night_ideal'],
            'days': []
        }
        
        # For each day in the week
        for date in week_dates:
            # Day shifts
            day_shifts = Shift.objects.filter(
                unit__in=units,
                date=date,
                shift_type__name__icontains='DAY',
                status__in=['SCHEDULED', 'CONFIRMED']
            )
            
            # Count SSCW only (not SM/OM)
            day_sscw = day_shifts.filter(user__role__name='SSCW').count()
            
            # Count all care staff (SCW + SCA)
            day_care = day_shifts.filter(user__role__name__in=['SCW', 'SCA']).count()
            
            # Night shifts
            night_shifts = Shift.objects.filter(
                unit__in=units,
                date=date,
                shift_type__name__icontains='NIGHT',
                status__in=['SCHEDULED', 'CONFIRMED']
            )
            
            # Count SSCWN only
            night_sscwn = night_shifts.filter(user__role__name='SSCWN').count()
            
            # Count all night care staff (SCWN + SCAN)
            night_care = night_shifts.filter(user__role__name__in=['SCWN', 'SCAN']).count()
            
            home_snapshot['days'].append({
                'date': date,
                'day_sscw': day_sscw,
                'day_care': day_care,
                'night_sscwn': night_sscwn,
                'night_care': night_care,
            })
        
        weekly_snapshot.append(home_snapshot)
    
    # =================================================================
    # Training Compliance Summary (All Homes)
    # =================================================================
    from scheduling.models import TrainingCourse, TrainingRecord
    
    training_compliance_by_home = []
    mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True)
    
    if mandatory_courses.exists():
        for home in care_homes:
            home_staff = User.objects.filter(
                unit__care_home=home,
                is_active=True
            ).distinct()
            
            total_staff = home_staff.count()
            total_required = total_staff * mandatory_courses.count()
            total_compliant = 0
            
            if total_staff > 0:
                for staff in home_staff:
                    for course in mandatory_courses:
                        latest_record = TrainingRecord.objects.filter(
                            staff_member=staff,
                            course=course
                        ).order_by('-completion_date').first()
                        
                        if latest_record and latest_record.get_status() == 'CURRENT':
                            total_compliant += 1
                
                compliance_percentage = (total_compliant / total_required * 100) if total_required > 0 else 0
                
                training_compliance_by_home.append({
                    'home': home,
                    'total_staff': total_staff,
                    'total_required': total_required,
                    'compliant': total_compliant,
                    'percentage': round(compliance_percentage, 1),
                })
    
    # =================================================================
    # Context for Template
    # =================================================================
    context = {
        'today': today,
        'current_time': timezone.now(),
        'current_month': current_month_start.strftime('%B %Y'),
        
        # Date range parameters
        'start_date': start_date,
        'end_date': end_date,
        'days_in_range': days_in_range,
        
        # Filtering
        'all_care_homes': CareHome.objects.all().order_by('name'),
        'selected_home': selected_home,
        
        # Overview
        'home_overview': home_overview,
        'total_capacity': total_capacity,
        'total_occupancy': total_occupancy,
        'overall_occupancy_rate': overall_occupancy_rate,
        'total_homes': care_homes.count(),
        
        # Staffing
        'staffing_today': staffing_today,
        
        # Fiscal
        'fiscal_summary': fiscal_summary,
        'total_agency_budget': total_agency_budget,
        'total_agency_spend': total_agency_spend,
        'total_ot_budget': total_ot_budget,
        'total_ot_spend': total_ot_spend,
        'total_budget': total_agency_budget + total_ot_budget,
        'total_spend': total_agency_spend + total_ot_spend,
        'overall_utilization': ((total_agency_spend + total_ot_spend) / (total_agency_budget + total_ot_budget) * 100) if (total_agency_budget + total_ot_budget) > 0 else 0,
        
        # Alerts & Actions
        'critical_alerts': critical_alerts,
        'pending_by_home': pending_by_home,
        'pending_reallocations': pending_reallocations,
        'unfilled_covers': unfilled_covers,
        
        # Quality
        'quality_metrics': quality_metrics,
        
        # Care Plan Compliance
        'care_plan_compliance': care_plan_compliance,
        
        # Staff Vacancies
        'vacancies': vacancies,
        'vacancies_by_home': vacancies_by_home,
        'total_vacancies': len([v for v in vacancies if v['status'] == 'VACANT']),
        'total_upcoming_leavers': len([v for v in vacancies if v['status'] == 'UPCOMING']),
        
        # Citywide Summary
        'citywide_day_summary': citywide_day_summary,
        'citywide_night_summary': citywide_night_summary,
        'week_dates': week_dates,
        
        # Weekly Snapshot - Simplified view
        'weekly_snapshot': weekly_snapshot,
        
        # Training Compliance
        'training_compliance_by_home': training_compliance_by_home,
    }
    
    return render(request, 'scheduling/senior_management_dashboard.html', context)


@login_required
def senior_dashboard_export(request):
    """
    Export senior dashboard data as CSV
    
    Supports date range filtering via GET parameters:
    - start_date: YYYY-MM-DD format (default: today)
    - end_date: YYYY-MM-DD format (default: today)
    - format: 'csv' (default)
    """
    
    # Check permissions - only Head of Service team members
    if not (request.user.role and request.user.role.is_senior_management_team):
        return HttpResponse('Access Denied - Head of Service Team Only', status=403)
    
    today = timezone.now().date()
    
    # Handle date range parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today
    else:
        start_date = today
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = today
    else:
        end_date = today
    
    # Ensure start_date is before end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="senior_dashboard_{start_date}_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Header with metadata
    writer.writerow(['Senior Management Dashboard Report'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Date Range:', f'{start_date} to {end_date}'])
    writer.writerow([])
    
    # Get all care homes
    care_homes = CareHome.objects.all().order_by('name')
    
    # ========================================
    # SECTION 1: HOME OVERVIEW
    # ========================================
    writer.writerow(['HOME OVERVIEW'])
    writer.writerow(['Home', 'Current Occupancy', 'Bed Capacity', 'Occupancy Rate %', 'Active Units', 'Location'])
    
    for home in care_homes:
        units = Unit.objects.filter(care_home=home, is_active=True)
        occupancy_rate = (home.current_occupancy / home.bed_capacity * 100) if home.bed_capacity > 0 else 0
        
        writer.writerow([
            home.get_name_display(),
            home.current_occupancy,
            home.bed_capacity,
            f'{occupancy_rate:.1f}',
            units.count(),
            home.location_address
        ])
    
    writer.writerow([])
    
    # ========================================
    # SECTION 2: STAFFING COVERAGE
    # ========================================
    writer.writerow(['STAFFING COVERAGE'])
    writer.writerow(['Home', 'Day Shifts Actual', 'Day Shifts Required', 'Day Coverage %', 
                     'Night Shifts Actual', 'Night Shifts Required', 'Night Coverage %', 'Status'])
    
    days_in_range = (end_date - start_date).days + 1
    
    for home in care_homes:
        units = Unit.objects.filter(care_home=home, is_active=True)
        
        range_shifts = Shift.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            unit__in=units,
            status__in=['SCHEDULED', 'CONFIRMED']
        )
        
        day_shifts = range_shifts.filter(shift_type__name__icontains='DAY').count()
        night_shifts = range_shifts.filter(shift_type__name__icontains='NIGHT').count()
        
        total_required_day = sum(u.min_day_staff for u in units) * days_in_range
        total_required_night = sum(u.min_night_staff for u in units) * days_in_range
        
        day_coverage = (day_shifts / total_required_day * 100) if total_required_day > 0 else 100
        night_coverage = (night_shifts / total_required_night * 100) if total_required_night > 0 else 100
        
        status = 'GOOD' if day_coverage >= 100 and night_coverage >= 100 else 'WARNING' if day_coverage >= 80 and night_coverage >= 80 else 'CRITICAL'
        
        writer.writerow([
            home.get_name_display(),
            day_shifts,
            total_required_day,
            f'{day_coverage:.1f}',
            night_shifts,
            total_required_night,
            f'{night_coverage:.1f}',
            status
        ])
    
    writer.writerow([])
    
    # ========================================
    # SECTION 3: FISCAL SUMMARY
    # ========================================
    writer.writerow(['FISCAL SUMMARY'])
    writer.writerow(['Home', 'Agency Budget', 'Agency Spend', 'Agency Utilization %',
                     'OT Budget', 'OT Spend', 'OT Utilization %', 'Total Budget', 'Total Spend'])
    
    for home in care_homes:
        units = Unit.objects.filter(care_home=home)
        
        # Agency spend for date range
        agency_shifts_count = Shift.objects.filter(
            unit__in=units,
            date__gte=start_date,
            date__lte=end_date,
            user__isnull=True,
            status__in=['SCHEDULED', 'CONFIRMED']
        ).count()
        agency_spend = Decimal(str(agency_shifts_count * 300))
        
        # OT spend for date range
        ot_shifts = Shift.objects.filter(
            unit__in=units,
            date__gte=start_date,
            date__lte=end_date,
            shift_classification='OVERTIME'
        ).count()
        ot_spend = Decimal(str(ot_shifts * 25 * 12))
        
        agency_budget = home.budget_agency_monthly
        ot_budget = home.budget_overtime_monthly
        
        agency_utilization = (float(agency_spend) / float(agency_budget) * 100) if agency_budget > 0 else 0
        ot_utilization = (float(ot_spend) / float(ot_budget) * 100) if ot_budget > 0 else 0
        
        writer.writerow([
            home.get_name_display(),
            f'£{agency_budget}',
            f'£{agency_spend}',
            f'{agency_utilization:.1f}',
            f'£{ot_budget}',
            f'£{ot_spend}',
            f'{ot_utilization:.1f}',
            f'£{agency_budget + ot_budget}',
            f'£{agency_spend + ot_spend}'
        ])
    
    writer.writerow([])
    
    # ========================================
    # SECTION 4: DETAILED SHIFT DATA
    # ========================================
    writer.writerow(['DETAILED SHIFT DATA'])
    writer.writerow(['Date', 'Home', 'Unit', 'Staff Name', 'Staff SAP', 'Shift Type', 
                     'Classification', 'Status', 'Start Time', 'End Time'])
    
    for home in care_homes:
        units = Unit.objects.filter(care_home=home, is_active=True)
        
        shifts = Shift.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            unit__in=units
        ).select_related('user', 'shift_type', 'unit').order_by('date', 'shift_type__name')
        
        for shift in shifts:
            writer.writerow([
                shift.date,
                home.get_name_display(),
                shift.unit.name,
                f'{shift.user.first_name} {shift.user.last_name}' if shift.user else 'UNFILLED',
                shift.user.sap if shift.user else 'N/A',
                shift.shift_type.name,
                shift.shift_classification,
                shift.status,
                shift.custom_start_time or shift.shift_type.start_time,
                shift.custom_end_time or shift.shift_type.end_time
            ])
    
    return response


@login_required
def custom_report_builder(request):
    """
    Custom Report Builder - Generate reports with user-selected fields and date ranges
    
    Supports multiple output formats:
    - PDF: Professional formatted report
    - Excel: Spreadsheet with data and charts
    - CSV: Raw data export
    
    Restricted to: Senior management team only
    """
    from django.db.models import Model
    from django.apps import apps
    import json
    
    # Check permissions
    if not (request.user.role and request.user.role.is_senior_management_team):
        return render(request, 'scheduling/access_denied.html', {
            'message': 'Report builder is restricted to Head of Service team members only.'
        })
    
    if request.method == 'GET':
        # Display report builder form
        today = timezone.now().date()
        
        # Define available report types and their fields
        report_types = {
            'staffing_coverage': {
                'name': 'Staffing Coverage Report',
                'description': 'Comprehensive staffing levels, shortages, and coverage analysis',
                'fields': [
                    {'id': 'date', 'label': 'Date', 'selected': True},
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'unit', 'label': 'Unit', 'selected': True},
                    {'id': 'shift_type', 'label': 'Shift Type (Day/Night)', 'selected': True},
                    {'id': 'scheduled_staff', 'label': 'Scheduled Staff Count', 'selected': True},
                    {'id': 'required_staff', 'label': 'Required Staff Count', 'selected': True},
                    {'id': 'shortage', 'label': 'Shortage/Surplus', 'selected': True},
                    {'id': 'sscw_count', 'label': 'SSCW/SSCWN Count', 'selected': False},
                    {'id': 'agency_count', 'label': 'Agency Staff Count', 'selected': False},
                    {'id': 'overtime_count', 'label': 'Overtime Shifts', 'selected': False},
                ]
            },
            'leave_usage': {
                'name': 'Leave Usage Report',
                'description': 'Annual leave, sickness, and other absences across homes',
                'fields': [
                    {'id': 'staff_name', 'label': 'Staff Name', 'selected': True},
                    {'id': 'staff_sap', 'label': 'Staff SAP', 'selected': True},
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'leave_type', 'label': 'Leave Type', 'selected': True},
                    {'id': 'start_date', 'label': 'Start Date', 'selected': True},
                    {'id': 'end_date', 'label': 'End Date', 'selected': True},
                    {'id': 'days_requested', 'label': 'Days Requested', 'selected': True},
                    {'id': 'status', 'label': 'Status (Approved/Denied/Pending)', 'selected': True},
                    {'id': 'unit', 'label': 'Unit', 'selected': False},
                    {'id': 'role', 'label': 'Role', 'selected': False},
                ]
            },
            'budget_variance': {
                'name': 'Budget Variance Report',
                'description': 'Planned vs actual staffing costs, agency usage, overtime',
                'fields': [
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'period', 'label': 'Period (Week/Month)', 'selected': True},
                    {'id': 'planned_cost', 'label': 'Planned Cost', 'selected': True},
                    {'id': 'actual_cost', 'label': 'Actual Cost', 'selected': True},
                    {'id': 'variance', 'label': 'Variance (£)', 'selected': True},
                    {'id': 'variance_percent', 'label': 'Variance (%)', 'selected': True},
                    {'id': 'agency_cost', 'label': 'Agency Cost', 'selected': False},
                    {'id': 'overtime_cost', 'label': 'Overtime Cost', 'selected': False},
                    {'id': 'regular_staff_cost', 'label': 'Regular Staff Cost', 'selected': False},
                ]
            },
            'compliance': {
                'name': 'Compliance Report',
                'description': 'Shift compliance, minimum staffing violations, care ratios',
                'fields': [
                    {'id': 'date', 'label': 'Date', 'selected': True},
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'unit', 'label': 'Unit', 'selected': True},
                    {'id': 'shift_type', 'label': 'Shift Type', 'selected': True},
                    {'id': 'staffing_met', 'label': 'Minimum Staffing Met (Yes/No)', 'selected': True},
                    {'id': 'sscw_present', 'label': 'SSCW Present (Yes/No)', 'selected': True},
                    {'id': 'violation_type', 'label': 'Violation Type', 'selected': False},
                    {'id': 'remediation', 'label': 'Remediation Action', 'selected': False},
                ]
            },
            'incidents': {
                'name': 'Incident Summary Report',
                'description': 'Incident tracking, types, and resolutions across homes',
                'fields': [
                    {'id': 'date', 'label': 'Date', 'selected': True},
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'unit', 'label': 'Unit', 'selected': True},
                    {'id': 'incident_type', 'label': 'Incident Type', 'selected': True},
                    {'id': 'severity', 'label': 'Severity', 'selected': True},
                    {'id': 'reported_by', 'label': 'Reported By', 'selected': False},
                    {'id': 'resolution_status', 'label': 'Resolution Status', 'selected': False},
                    {'id': 'follow_up_required', 'label': 'Follow-up Required', 'selected': False},
                ]
            },
            'comparative_analytics': {
                'name': 'Comparative Analytics Report',
                'description': 'Cross-home performance comparison and benchmarking',
                'fields': [
                    {'id': 'care_home', 'label': 'Care Home', 'selected': True},
                    {'id': 'avg_daily_staff', 'label': 'Average Daily Staff', 'selected': True},
                    {'id': 'fill_rate', 'label': 'Shift Fill Rate (%)', 'selected': True},
                    {'id': 'agency_usage_rate', 'label': 'Agency Usage Rate (%)', 'selected': True},
                    {'id': 'overtime_rate', 'label': 'Overtime Rate (%)', 'selected': True},
                    {'id': 'sickness_rate', 'label': 'Sickness Absence Rate (%)', 'selected': True},
                    {'id': 'leave_approval_rate', 'label': 'Leave Approval Rate (%)', 'selected': False},
                    {'id': 'budget_variance', 'label': 'Budget Variance (%)', 'selected': False},
                ]
            }
        }
        
        context = {
            'report_types': json.dumps(report_types),
            'report_types_dict': report_types,  # Keep dict for iteration
            'today': today,
            'all_care_homes': CareHome.objects.all().order_by('name'),
        }
        
        return render(request, 'scheduling/custom_report_builder.html', context)
    
    elif request.method == 'POST':
        # Generate and export report
        report_type = request.POST.get('report_type')
        export_format = request.POST.get('export_format', 'csv')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        selected_homes = request.POST.getlist('care_homes')
        selected_fields = request.POST.getlist('fields')
        
        # Parse dates
        from django.utils.dateparse import parse_date
        start_date = parse_date(start_date_str) if start_date_str else timezone.now().date()
        end_date = parse_date(end_date_str) if end_date_str else timezone.now().date()
        
        # Generate report data based on type
        report_data = _generate_report_data(report_type, start_date, end_date, selected_homes, selected_fields)
        
        # Export in requested format
        if export_format == 'csv':
            return _export_csv(report_type, report_data, start_date, end_date)
        elif export_format == 'excel':
            return _export_excel(report_type, report_data, start_date, end_date)
        elif export_format == 'pdf':
            return _export_pdf(report_type, report_data, start_date, end_date)
        else:
            return HttpResponse('Invalid export format', status=400)


def _generate_report_data(report_type, start_date, end_date, selected_homes, selected_fields):
    """Generate report data based on report type and parameters"""
    from django.db.models import Count, Avg, Sum, F, Q
    
    data = []
    
    # Filter homes
    if selected_homes:
        care_homes = CareHome.objects.filter(name__in=selected_homes)
    else:
        care_homes = CareHome.objects.all()
    
    if report_type == 'staffing_coverage':
        # Staffing coverage analysis - aggregated by home for high-level reporting
        for home in care_homes:
            units = Unit.objects.filter(care_home=home, is_active=True)
            
            # Get all shifts for this home across all units
            shifts = Shift.objects.filter(
                unit__care_home=home,
                unit__is_active=True,
                date__gte=start_date,
                date__lte=end_date
            ).select_related('shift_type', 'user', 'user__role')
            
            # Group by date and shift type
            current_date = start_date
            while current_date <= end_date:
                # Day shifts across all units in this home
                day_shifts = shifts.filter(
                    date=current_date,
                    shift_type__name__in=['DAY', 'DAY_SENIOR', 'DAY_ASSISTANT']
                )
                day_count = day_shifts.count()
                day_sscw = day_shifts.filter(user__role__name='SSCW').count()
                
                # Night shifts across all units in this home
                night_shifts = shifts.filter(
                    date=current_date,
                    shift_type__name__in=['NIGHT', 'NIGHT_SENIOR', 'NIGHT_ASSISTANT']
                )
                night_count = night_shifts.count()
                night_sscw = night_shifts.filter(user__role__name='SSCWN').count()
                
                # Calculate required staff as sum across all units
                day_required = sum(unit.min_day_staff for unit in units)
                night_required = sum(unit.min_night_staff for unit in units)
                
                if 'shift_type' in selected_fields:
                    # Separate rows for day and night
                    data.append({
                        'date': current_date,
                        'care_home': home.get_name_display(),
                        'shift_type': 'Day',
                        'scheduled_staff': day_count,
                        'required_staff': day_required,
                        'shortage': day_required - day_count,
                        'sscw_count': day_sscw,
                    })
                    
                    data.append({
                        'date': current_date,
                        'care_home': home.get_name_display(),
                        'shift_type': 'Night',
                        'scheduled_staff': night_count,
                        'required_staff': night_required,
                        'shortage': night_required - night_count,
                        'sscw_count': night_sscw,
                    })
                
                current_date += timedelta(days=1)
    
    elif report_type == 'leave_usage':
        # Leave usage analysis
        leave_requests = LeaveRequest.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date,
            user__unit__care_home__in=care_homes
        ).select_related('user', 'user__unit', 'user__unit__care_home', 'user__role')
        
        for leave in leave_requests:
            data.append({
                'staff_name': leave.user.full_name,
                'staff_sap': leave.user.sap,
                'care_home': leave.user.unit.care_home.get_name_display() if leave.user.unit and leave.user.unit.care_home else 'N/A',
                'leave_type': leave.get_leave_type_display(),
                'start_date': leave.start_date,
                'end_date': leave.end_date,
                'days_requested': leave.days_requested,
                'status': leave.get_status_display(),
                'unit': leave.user.unit.get_name_display() if leave.user.unit else 'N/A',
                'role': leave.user.role.get_name_display() if leave.user.role else 'N/A',
            })
    
    elif report_type == 'comparative_analytics':
        # Cross-home comparison
        for home in care_homes:
            shifts = Shift.objects.filter(
                unit__care_home=home,
                date__gte=start_date,
                date__lte=end_date
            )
            
            total_shifts = shifts.count()
            total_days = (end_date - start_date).days + 1
            avg_daily_staff = total_shifts / total_days if total_days > 0 else 0
            
            # Calculate fill rate
            unfilled_shifts = shifts.filter(user__isnull=True).count()
            fill_rate = ((total_shifts - unfilled_shifts) / total_shifts * 100) if total_shifts > 0 else 0
            
            data.append({
                'care_home': home.get_name_display(),
                'avg_daily_staff': round(avg_daily_staff, 1),
                'fill_rate': round(fill_rate, 1),
                'agency_usage_rate': 0,  # Placeholder
                'overtime_rate': 0,  # Placeholder
                'sickness_rate': 0,  # Placeholder
            })
    
    return data


def _export_csv(report_type, report_data, start_date, end_date):
    """Export report data as CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{start_date}_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([f'{report_type.replace("_", " ").title()} Report'])
    writer.writerow([f'Period: {start_date} to {end_date}'])
    writer.writerow([f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])
    
    # Write column headers
    if report_data:
        writer.writerow(report_data[0].keys())
        
        # Write data rows
        for row in report_data:
            writer.writerow(row.values())
    
    return response


def _export_excel(report_type, report_data, start_date, end_date):
    """Export report data as Excel with formatting"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_type[:31]  # Excel sheet name limit
        
        # Header
        ws['A1'] = f'{report_type.replace("_", " ").title()} Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Period: {start_date} to {end_date}'
        ws['A3'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Data header (row 5)
        if report_data:
            headers = list(report_data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_num, row_data in enumerate(report_data, 6):
                for col_num, value in enumerate(row_data.values(), 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{start_date}_{end_date}.xlsx"'
        
        return response
        
    except ImportError:
        # openpyxl not installed, fall back to CSV
        return _export_csv(report_type, report_data, start_date, end_date)


def _export_pdf(report_type, report_data, start_date, end_date):
    """Export report data as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
        )
        
        title = Paragraph(f'{report_type.replace("_", " ").title()} Report', title_style)
        elements.append(title)
        
        # Metadata
        meta_style = styles['Normal']
        elements.append(Paragraph(f'<b>Period:</b> {start_date} to {end_date}', meta_style))
        elements.append(Paragraph(f'<b>Generated:</b> {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', meta_style))
        elements.append(Spacer(1, 0.5*inch))
        
        # Data table
        if report_data:
            # Prepare table data
            headers = [[h.replace('_', ' ').title() for h in report_data[0].keys()]]
            data_rows = [[str(v) for v in row.values()] for row in report_data]
            table_data = headers + data_rows
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{start_date}_{end_date}.pdf"'
        response.write(pdf_data)
        
        return response
        
    except ImportError:
        # reportlab not installed, fall back to CSV
        return _export_csv(report_type, report_data, start_date, end_date)


@api_login_required
def api_staffing_gaps(request):
    """
    API endpoint for staffing gap details
    Returns detailed breakdown of staffing gaps by home, shift type, and date
    
    GET parameters:
    - date: YYYY-MM-DD (default: today)
    - days: number of days to include (default: 7)
    - home_id: filter by specific home (optional)
    
    Returns JSON with gap details for GAP breakdown modal
    """
    try:
        # Parse request parameters
        date_str = request.GET.get('date')
        if date_str:
            start_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date()
        
        days = int(request.GET.get('days', 7))
        end_date = start_date + timedelta(days=days)
        
        home_id = request.GET.get('home_id')
        
        # Get all care homes
        homes = CareHome.objects.all()
        if home_id:
            homes = homes.filter(id=home_id)
        
        gaps_data = []
        total_gap = 0
        total_filled = 0
        total_required = 0
        
        # Calculate gaps for each home
        for home in homes:
            # Get units for this home
            units = Unit.objects.filter(care_home=home)
            
            # Get shifts in date range
            shifts = Shift.objects.filter(
                unit__in=units,
                date__gte=start_date,
                date__lt=end_date
            ).select_related('unit', 'shift_type', 'user')
            
            # Group by date and shift type
            shift_groups = {}
            for shift in shifts:
                key = (shift.date, shift.shift_type.name if shift.shift_type else 'Unknown')
                if key not in shift_groups:
                    shift_groups[key] = {
                        'required': 0,
                        'filled': 0,
                        'shifts': []
                    }
                
                shift_groups[key]['required'] += 1
                if shift.user:
                    shift_groups[key]['filled'] += 1
                shift_groups[key]['shifts'].append(shift)
            
            # Calculate gaps for each group
            for (date, shift_type), data in shift_groups.items():
                gap = data['required'] - data['filled']
                
                if gap > 0:  # Only include if there's actually a gap
                    gaps_data.append({
                        'home': home.name,
                        'home_id': home.id,
                        'shift_type': shift_type,
                        'date': date.strftime('%Y-%m-%d'),
                        'date_display': date.strftime('%a, %b %d'),
                        'required': data['required'],
                        'filled': data['filled'],
                        'gap': gap,
                        'percentage_filled': round((data['filled'] / data['required'] * 100), 1) if data['required'] > 0 else 0
                    })
                    
                    total_gap += gap
                    total_required += data['required']
                    total_filled += data['filled']
        
        # Sort by date, then home, then shift type
        gaps_data.sort(key=lambda x: (x['date'], x['home'], x['shift_type']))
        
        # Calculate overall percentage
        overall_percentage = round((total_filled / total_required * 100), 1) if total_required > 0 else 0
        
        return JsonResponse({
            'success': True,
            'period': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'days': days
            },
            'summary': {
                'total_required': total_required,
                'total_filled': total_filled,
                'total_gap': total_gap,
                'percentage_filled': overall_percentage
            },
            'gaps': gaps_data,
            'homes_count': homes.count(),
            'generated_at': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_login_required
def api_multi_home_staffing(request, date_str):
    """
    API endpoint for multi-home staffing detail breakdown
    Returns detailed staffing levels across all homes for a specific date
    
    URL parameter:
    - date_str: YYYY-MM-DD format
    
    Returns JSON with staffing details by home, unit, and shift type
    Used when clicking data points on Multi-Home Staffing Chart
    """
    try:
        # Parse date
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Get all care homes
        homes = CareHome.objects.all().order_by('name')
        
        staffing_data = []
        total_shifts = 0
        total_filled = 0
        total_gap = 0
        
        for home in homes:
            # Get units for this home
            units = Unit.objects.filter(care_home=home, is_active=True)
            
            # Get shifts for this date
            shifts = Shift.objects.filter(
                unit__in=units,
                date=target_date
            ).select_related('unit', 'shift_type', 'user', 'user__role')
            
            # Group by shift type
            shift_type_groups = {}
            for shift in shifts:
                shift_type_name = shift.shift_type.name if shift.shift_type else 'Unknown'
                
                if shift_type_name not in shift_type_groups:
                    shift_type_groups[shift_type_name] = {
                        'required': 0,
                        'filled': 0,
                        'staff': []
                    }
                
                shift_type_groups[shift_type_name]['required'] += 1
                total_shifts += 1
                
                if shift.user:
                    shift_type_groups[shift_type_name]['filled'] += 1
                    total_filled += 1
                    shift_type_groups[shift_type_name]['staff'].append({
                        'name': f"{shift.user.first_name} {shift.user.last_name}",
                        'role': shift.user.role.name if shift.user.role else 'Unknown',
                        'unit': shift.unit.name
                    })
            
            # Calculate home-level stats
            home_required = sum(g['required'] for g in shift_type_groups.values())
            home_filled = sum(g['filled'] for g in shift_type_groups.values())
            home_gap = home_required - home_filled
            total_gap += home_gap
            
            # Build shift type breakdown
            shift_types_breakdown = []
            for shift_type, data in shift_type_groups.items():
                gap = data['required'] - data['filled']
                shift_types_breakdown.append({
                    'shift_type': shift_type,
                    'required': data['required'],
                    'filled': data['filled'],
                    'gap': gap,
                    'percentage_filled': round((data['filled'] / data['required'] * 100), 1) if data['required'] > 0 else 0,
                    'staff': data['staff']
                })
            
            staffing_data.append({
                'home': home.name,
                'home_id': home.id,
                'location': home.location_address,
                'total_required': home_required,
                'total_filled': home_filled,
                'total_gap': home_gap,
                'percentage_filled': round((home_filled / home_required * 100), 1) if home_required > 0 else 0,
                'shift_types': shift_types_breakdown
            })
        
        return JsonResponse({
            'success': True,
            'date': target_date.strftime('%Y-%m-%d'),
            'date_display': target_date.strftime('%A, %B %d, %Y'),
            'summary': {
                'total_shifts': total_shifts,
                'total_filled': total_filled,
                'total_gap': total_gap,
                'percentage_filled': round((total_filled / total_shifts * 100), 1) if total_shifts > 0 else 0,
                'homes_count': homes.count()
            },
            'homes': staffing_data,
            'generated_at': timezone.now().isoformat()
        })
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_login_required
def api_budget_breakdown(request, home_id):
    """
    API endpoint for budget breakdown by department/category
    Returns detailed cost breakdown for a specific care home
    
    URL parameter:
    - home_id: Care home ID
    
    GET parameters:
    - month: YYYY-MM (default: current month)
    
    Returns JSON with budget vs actual spending by category
    Used when clicking bars on Budget vs Actual Chart
    """
    try:
        # Get care home
        home = CareHome.objects.get(id=home_id)
        
        # Parse month parameter
        month_str = request.GET.get('month')
        if month_str:
            target_date = datetime.strptime(month_str, '%Y-%m').date()
        else:
            target_date = timezone.now().date()
        
        # Calculate month date range
        month_start = target_date.replace(day=1)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        
        # Get units for this home
        units = Unit.objects.filter(care_home=home, is_active=True)
        
        # Get all shifts for the month
        shifts = Shift.objects.filter(
            unit__in=units,
            date__gte=month_start,
            date__lt=month_end
        ).select_related('user', 'user__role', 'shift_type')
        
        # Calculate costs by category
        categories = {
            'Staff Wages': {'budget': 0, 'actual': 0, 'shifts': 0},
            'Agency Cover': {'budget': 0, 'actual': 0, 'shifts': 0},
            'Overtime': {'budget': 0, 'actual': 0, 'shifts': 0},
            'Training': {'budget': 0, 'actual': 0, 'shifts': 0},
        }
        
        # Hourly rates (simplified - should come from settings/database)
        base_rate = Decimal('12.50')
        agency_rate = Decimal('18.00')
        overtime_rate = Decimal('15.63')  # 1.25x base
        
        total_budget = Decimal('0')
        total_actual = Decimal('0')
        
        for shift in shifts:
            # Determine shift duration (simplified - assuming 8 hours)
            duration = 8
            
            if shift.user:
                # Regular staff wage
                cost = base_rate * duration
                categories['Staff Wages']['actual'] += float(cost)
                categories['Staff Wages']['shifts'] += 1
                total_actual += cost
            else:
                # Unfilled - budget for agency
                cost = agency_rate * duration
                categories['Agency Cover']['budget'] += float(cost)
                categories['Agency Cover']['shifts'] += 1
                total_budget += cost
        
        # Add sample overtime (20% of filled shifts)
        filled_shifts = shifts.filter(user__isnull=False).count()
        overtime_cost = float(overtime_rate * 8 * Decimal(filled_shifts) * Decimal('0.2'))
        categories['Overtime']['actual'] = overtime_cost
        categories['Overtime']['shifts'] = int(filled_shifts * 0.2)
        total_actual += Decimal(str(overtime_cost))
        
        # Training budget (fixed monthly)
        training_budget = 2000.0
        categories['Training']['budget'] = training_budget
        categories['Training']['actual'] = training_budget * 0.85  # 85% spent
        total_budget += Decimal(str(training_budget))
        total_actual += Decimal(str(training_budget * 0.85))
        
        # Build breakdown
        breakdown = []
        for category, data in categories.items():
            budget = data['budget']
            actual = data['actual']
            variance = actual - budget
            variance_pct = round((variance / budget * 100), 1) if budget > 0 else 0
            
            breakdown.append({
                'category': category,
                'budget': round(budget, 2),
                'actual': round(actual, 2),
                'variance': round(variance, 2),
                'variance_percentage': variance_pct,
                'shifts': data['shifts'],
                'status': 'over' if variance > 0 else 'under' if variance < 0 else 'on_budget'
            })
        
        return JsonResponse({
            'success': True,
            'home': {
                'id': home.id,
                'name': home.name,
                'location': home.location_address
            },
            'period': {
                'month': month_start.strftime('%B %Y'),
                'start_date': month_start.strftime('%Y-%m-%d'),
                'end_date': (month_end - timedelta(days=1)).strftime('%Y-%m-%d')
            },
            'summary': {
                'total_budget': round(float(total_budget), 2),
                'total_actual': round(float(total_actual), 2),
                'total_variance': round(float(total_actual - total_budget), 2),
                'variance_percentage': round(((total_actual - total_budget) / total_budget * 100), 1) if total_budget > 0 else 0
            },
            'breakdown': breakdown,
            'generated_at': timezone.now().isoformat()
        })
        
    except CareHome.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Care home not found'
        }, status=404)
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid month format. Use YYYY-MM'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_login_required
def api_overtime_detail(request, date_str):
    """
    API endpoint for overtime detail by staff member
    Returns detailed overtime breakdown for a specific date
    
    URL parameter:
    - date_str: YYYY-MM-DD format
    
    Returns JSON with overtime hours by staff member
    Used when clicking high-overtime days on Overtime Trend Chart
    """
    try:
        # Parse date
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Calculate week range (Monday to Sunday containing target date)
        days_since_monday = target_date.weekday()
        week_start = target_date - timedelta(days=days_since_monday)
        week_end = week_start + timedelta(days=6)
        
        # Get all care homes
        homes = CareHome.objects.all()
        all_units = Unit.objects.filter(care_home__in=homes, is_active=True)
        
        # Get shifts for the week
        shifts = Shift.objects.filter(
            unit__in=all_units,
            date__gte=week_start,
            date__lte=week_end,
            user__isnull=False
        ).select_related('user', 'user__role', 'unit', 'unit__care_home', 'shift_type')
        
        # Calculate hours per staff member
        staff_hours = {}
        for shift in shifts:
            staff_id = shift.user.id
            
            if staff_id not in staff_hours:
                staff_hours[staff_id] = {
                    'staff_name': f"{shift.user.first_name} {shift.user.last_name}",
                    'role': shift.user.role.name if shift.user.role else 'Unknown',
                    'home': shift.unit.care_home.name,
                    'regular_hours': 0,
                    'overtime_hours': 0,
                    'total_hours': 0,
                    'shifts': []
                }
            
            # Assume 8 hours per shift
            shift_hours = 8
            staff_hours[staff_id]['total_hours'] += shift_hours
            
            # Track shifts on target date
            if shift.date == target_date:
                staff_hours[staff_id]['shifts'].append({
                    'date': shift.date.strftime('%Y-%m-%d'),
                    'shift_type': shift.shift_type.name if shift.shift_type else 'Unknown',
                    'unit': shift.unit.name,
                    'hours': shift_hours
                })
        
        # Calculate overtime (over 37.5 hours per week)
        standard_hours = 37.5
        overtime_data = []
        total_overtime = 0
        total_staff_overtime = 0
        
        for staff_id, data in staff_hours.items():
            if data['total_hours'] > standard_hours:
                overtime = data['total_hours'] - standard_hours
                data['regular_hours'] = standard_hours
                data['overtime_hours'] = overtime
                total_overtime += overtime
                total_staff_overtime += 1
                
                overtime_data.append({
                    'staff_name': data['staff_name'],
                    'role': data['role'],
                    'home': data['home'],
                    'regular_hours': round(data['regular_hours'], 1),
                    'overtime_hours': round(overtime, 1),
                    'total_hours': round(data['total_hours'], 1),
                    'overtime_percentage': round((overtime / data['total_hours'] * 100), 1),
                    'shifts_on_date': data['shifts']
                })
            else:
                data['regular_hours'] = data['total_hours']
                data['overtime_hours'] = 0
        
        # Sort by overtime hours (descending)
        overtime_data.sort(key=lambda x: x['overtime_hours'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'date': target_date.strftime('%Y-%m-%d'),
            'date_display': target_date.strftime('%A, %B %d, %Y'),
            'week': {
                'start_date': week_start.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'week_display': f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}"
            },
            'summary': {
                'total_overtime_hours': round(total_overtime, 1),
                'staff_with_overtime': total_staff_overtime,
                'total_staff': len(staff_hours),
                'average_overtime': round(total_overtime / total_staff_overtime, 1) if total_staff_overtime > 0 else 0,
                'standard_weekly_hours': standard_hours
            },
            'overtime_staff': overtime_data,
            'generated_at': timezone.now().isoformat()
        })
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_login_required
def api_compliance_actions(request, home_id, metric):
    """
    API endpoint for compliance improvement actions
    Returns actionable recommendations for improving specific compliance metrics
    
    URL parameters:
    - home_id: Care home ID
    - metric: Compliance metric (training, documentation, care_planning, safeguarding, medication)
    
    Returns JSON with improvement actions and current compliance status
    Used when clicking low-scoring metrics on Compliance Score Chart
    """
    try:
        # Get care home
        home = CareHome.objects.get(id=home_id)
        
        # Validate metric
        valid_metrics = ['training', 'documentation', 'care_planning', 'safeguarding', 'medication']
        if metric not in valid_metrics:
            return JsonResponse({
                'success': False,
                'error': f'Invalid metric. Must be one of: {", ".join(valid_metrics)}'
            }, status=400)
        
        # Get units for this home
        units = Unit.objects.filter(care_home=home, is_active=True)
        staff = User.objects.filter(unit__in=units, is_active=True).distinct()
        
        # Calculate current compliance
        actions = []
        current_score = 0
        target_score = 100
        
        if metric == 'training':
            from scheduling.models import TrainingCourse, TrainingRecord
            
            mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True)
            total_required = staff.count() * mandatory_courses.count()
            completed = 0
            
            for staff_member in staff:
                for course in mandatory_courses:
                    latest = TrainingRecord.objects.filter(
                        staff_member=staff_member,
                        course=course,
                        completion_date__isnull=False
                    ).order_by('-completion_date').first()
                    
                    if latest and not latest.is_expired():
                        completed += 1
                    else:
                        # Add action for missing/expired training
                        actions.append({
                            'priority': 'high' if not latest else 'medium',
                            'action': f"Schedule {course.name} for {staff_member.first_name} {staff_member.last_name}",
                            'staff': f"{staff_member.first_name} {staff_member.last_name}",
                            'course': course.name,
                            'status': 'expired' if latest else 'not_started',
                            'deadline': 'Overdue' if latest and latest.is_expired() else '30 days'
                        })
            
            current_score = round((completed / total_required * 100), 1) if total_required > 0 else 0
            
        elif metric == 'documentation':
            # Sample documentation compliance check
            residents = Resident.objects.filter(unit__in=units, status='ACTIVE')
            total_residents = residents.count()
            compliant_count = 0
            
            for resident in residents:
                # Check if care plan review is up to date (within 28 days)
                latest_review = CarePlanReview.objects.filter(
                    resident=resident
                ).order_by('-review_date').first()
                
                if latest_review:
                    days_since = (timezone.now().date() - latest_review.review_date).days
                    if days_since <= 28:
                        compliant_count += 1
                    else:
                        actions.append({
                            'priority': 'high' if days_since > 35 else 'medium',
                            'action': f"Update care plan review for {resident.name}",
                            'resident': resident.name,
                            'last_review': latest_review.review_date.strftime('%Y-%m-%d'),
                            'days_overdue': days_since - 28,
                            'deadline': 'Immediate' if days_since > 35 else '7 days'
                        })
                else:
                    actions.append({
                        'priority': 'high',
                        'action': f"Create initial care plan review for {resident.name}",
                        'resident': resident.name,
                        'status': 'missing',
                        'deadline': 'Immediate'
                    })
            
            current_score = round((compliant_count / total_residents * 100), 1) if total_residents > 0 else 0
            
        elif metric == 'care_planning':
            # Care planning compliance
            residents = Resident.objects.filter(unit__in=units, status='ACTIVE')
            compliant_count = sum(1 for r in residents if hasattr(r, 'care_plan'))
            current_score = round((compliant_count / residents.count() * 100), 1) if residents.count() > 0 else 0
            
            for resident in residents:
                if not hasattr(resident, 'care_plan'):
                    actions.append({
                        'priority': 'high',
                        'action': f"Create care plan for {resident.name}",
                        'resident': resident.name,
                        'deadline': 'Immediate'
                    })
            
        elif metric == 'safeguarding':
            # Safeguarding training compliance
            safeguarding_trained = 0
            for staff_member in staff:
                # Check if they have safeguarding training
                from scheduling.models import TrainingRecord
                has_training = TrainingRecord.objects.filter(
                    staff_member=staff_member,
                    course__name__icontains='safeguard',
                    completion_date__isnull=False
                ).exists()
                
                if has_training:
                    safeguarding_trained += 1
                else:
                    actions.append({
                        'priority': 'high',
                        'action': f"Schedule Safeguarding training for {staff_member.first_name} {staff_member.last_name}",
                        'staff': f"{staff_member.first_name} {staff_member.last_name}",
                        'deadline': '30 days'
                    })
            
            current_score = round((safeguarding_trained / staff.count() * 100), 1) if staff.count() > 0 else 0
            
        elif metric == 'medication':
            # Medication management compliance
            current_score = 92.5  # Sample score
            actions.append({
                'priority': 'medium',
                'action': 'Review medication administration records for completeness',
                'deadline': '7 days'
            })
            actions.append({
                'priority': 'low',
                'action': 'Audit medication storage temperature logs',
                'deadline': '14 days'
            })
        
        # Sort actions by priority
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        actions.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 2))
        
        # Calculate gap
        gap = target_score - current_score
        
        return JsonResponse({
            'success': True,
            'home': {
                'id': home.id,
                'name': home.name,
                'location': home.location_address
            },
            'metric': {
                'name': metric.replace('_', ' ').title(),
                'current_score': round(current_score, 1),
                'target_score': target_score,
                'gap': round(gap, 1),
                'status': 'excellent' if current_score >= 95 else 'good' if current_score >= 85 else 'needs_improvement'
            },
            'actions': actions,
            'actions_count': {
                'high': sum(1 for a in actions if a.get('priority') == 'high'),
                'medium': sum(1 for a in actions if a.get('priority') == 'medium'),
                'low': sum(1 for a in actions if a.get('priority') == 'low'),
                'total': len(actions)
            },
            'generated_at': timezone.now().isoformat()
        })
        
    except CareHome.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Care home not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
