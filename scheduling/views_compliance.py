"""
Care Inspectorate Compliance Views
Web forms for staff to complete training, induction, supervision, and incident reports
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.http import JsonResponse
from .decorators_api import api_login_required
from django.views.decorators.http import require_http_methods
from datetime import date, timedelta
from decimal import Decimal
import json

from scheduling.models import (
    TrainingCourse, TrainingRecord, InductionProgress, 
    SupervisionRecord, IncidentReport, User, CareHome, Unit, ShiftType
)



# ============================================================================
# TRAINING MANAGEMENT
# ============================================================================

@login_required
def my_training_dashboard(request):
    """Staff member's personal training dashboard"""
    
    # Get all training records for current user
    training_records = TrainingRecord.objects.filter(
        staff_member=request.user
    ).select_related('course').order_by('-completion_date')
    
    # Calculate training statistics
    total_training = training_records.count()
    current_training = sum(1 for r in training_records if r.get_status() == 'CURRENT')
    expiring_soon = sum(1 for r in training_records if r.get_status() == 'EXPIRING_SOON')
    expired_training = sum(1 for r in training_records if r.get_status() == 'EXPIRED')
    
    # SSSC CPD hours
    current_year = timezone.now().year
    sssc_cpd_hours = training_records.filter(
        completion_date__year=current_year,
        sssc_cpd_hours_claimed__isnull=False
    ).aggregate(total=Sum('sssc_cpd_hours_claimed'))['total'] or Decimal('0.00')
    
    # Mandatory training compliance
    mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True)
    mandatory_completed = []
    mandatory_missing = []
    
    for course in mandatory_courses:
        latest_record = training_records.filter(course=course).first()
        if latest_record and latest_record.get_status() == 'CURRENT':
            mandatory_completed.append({
                'course': course,
                'record': latest_record,
                'days_until_expiry': latest_record.days_until_expiry()
            })
        else:
            mandatory_missing.append(course)
    
    context = {
        'training_records': training_records,
        'total_training': total_training,
        'current_training': current_training,
        'expiring_soon': expiring_soon,
        'expired_training': expired_training,
        'sssc_cpd_hours': sssc_cpd_hours,
        'sssc_target': Decimal('35.00'),
        'mandatory_completed': mandatory_completed,
        'mandatory_missing': mandatory_missing,
    }
    
    return render(request, 'compliance/my_training_dashboard.html', context)


@login_required
def training_compliance_dashboard(request):
    """Management view of training compliance across all homes"""
    
    # Get filters
    selected_home = request.GET.get('care_home', None)
    
    # Get mandatory courses
    mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True).order_by('name')
    
    # Get care homes
    care_homes = CareHome.objects.filter(is_active=True).order_by('name')
    
    # Build compliance data by home
    home_compliance_data = []
    
    for home in care_homes:
        # Filter if specific home selected
        if selected_home and home.name != selected_home:
            continue
            
        # Get active staff for this home
        home_staff = User.objects.filter(
            unit__care_home=home,
            is_active=True
        ).distinct()
        
        total_staff = home_staff.count()
        
        # Calculate compliance for each mandatory course
        course_compliance = []
        for course in mandatory_courses:
            compliant_count = 0
            expiring_count = 0
            expired_count = 0
            missing_count = 0
            
            compliant_staff = []
            expiring_staff = []
            expired_staff = []
            missing_staff = []
            
            for staff in home_staff:
                # Get latest training record for this course
                latest_record = TrainingRecord.objects.filter(
                    staff_member=staff,
                    course=course
                ).order_by('-completion_date').first()
                
                if latest_record:
                    status = latest_record.get_status()
                    if status == 'CURRENT':
                        compliant_count += 1
                        compliant_staff.append({'staff': staff, 'record': latest_record})
                    elif status == 'EXPIRING_SOON':
                        expiring_count += 1
                        expiring_staff.append({'staff': staff, 'record': latest_record})
                    elif status == 'EXPIRED':
                        expired_count += 1
                        expired_staff.append({
                            'staff': staff, 
                            'record': latest_record,
                            'days_overdue': abs(latest_record.days_until_expiry)
                        })
                else:
                    missing_count += 1
                    missing_staff.append(staff)
            
            compliance_percentage = (compliant_count / total_staff * 100) if total_staff > 0 else 0
            
            course_compliance.append({
                'course': course,
                'compliant': compliant_count,
                'expiring': expiring_count,
                'expired': expired_count,
                'missing': missing_count,
                'total': total_staff,
                'percentage': round(compliance_percentage, 1),
                'compliant_staff': compliant_staff,
                'expiring_staff': expiring_staff,
                'expired_staff': expired_staff,
                'missing_staff': missing_staff,
            })
        
        # Calculate overall home compliance
        total_required = total_staff * mandatory_courses.count()
        total_compliant = sum(c['compliant'] for c in course_compliance)
        overall_percentage = (total_compliant / total_required * 100) if total_required > 0 else 0
        
        home_compliance_data.append({
            'home': home,
            'total_staff': total_staff,
            'course_compliance': course_compliance,
            'overall_percentage': round(overall_percentage, 1),
            'total_compliant': total_compliant,
            'total_required': total_required
        })
    
    context = {
        'home_compliance_data': home_compliance_data,
        'mandatory_courses': mandatory_courses,
        'care_homes': care_homes,
        'selected_home': selected_home,
    }
    
    return render(request, 'compliance/training_compliance_dashboard.html', context)


@login_required
def add_staff_training_record(request):
    """Manager adds training record for a staff member"""
    
    if request.method == 'POST':
        staff_id = request.POST.get('staff_member')
        course_id = request.POST.get('course')
        completion_date = request.POST.get('completion_date')
        trainer_name = request.POST.get('trainer_name', '')
        training_provider = request.POST.get('training_provider', '')
        certificate_number = request.POST.get('certificate_number', '')
        sssc_cpd_hours = request.POST.get('sssc_cpd_hours', None)
        notes = request.POST.get('notes', '')
        
        try:
            staff_member = User.objects.get(id=staff_id)
            course = TrainingCourse.objects.get(id=course_id)
            completion_date_obj = date.fromisoformat(completion_date)
            
            # Calculate expiry date based on course validity
            expiry_date = completion_date_obj + timedelta(days=course.validity_months * 30)
            
            # Create training record
            record = TrainingRecord.objects.create(
                staff_member=staff_member,
                course=course,
                completion_date=completion_date_obj,
                expiry_date=expiry_date,
                trainer_name=trainer_name,
                training_provider=training_provider,
                certificate_number=certificate_number,
                sssc_cpd_hours_claimed=Decimal(sssc_cpd_hours) if sssc_cpd_hours else None,
                notes=notes,
                created_by=request.user
            )
            
            # Handle certificate file upload
            if 'certificate_file' in request.FILES:
                record.certificate_file = request.FILES['certificate_file']
                record.save()
            
            messages.success(request, f'Training record for {staff_member.full_name} - {course.name} has been added successfully.')
            return redirect('training_compliance_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error adding training record: {str(e)}')
    
    # GET request - show form
    selected_home = request.GET.get('care_home', None)
    selected_staff_id = request.GET.get('staff', None)
    
    # Get staff members
    if selected_home:
        staff_members = User.objects.filter(
            unit__care_home__name=selected_home,
            is_active=True
        ).distinct().order_by('first_name', 'last_name')
    else:
        staff_members = User.objects.filter(
            is_active=True,
            unit__isnull=False
        ).distinct().order_by('first_name', 'last_name')
    
    courses = TrainingCourse.objects.all().order_by('category', 'name')
    care_homes = CareHome.objects.filter(is_active=True).order_by('name')
    
    context = {
        'staff_members': staff_members,
        'courses': courses,
        'care_homes': care_homes,
        'selected_home': selected_home,
        'selected_staff_id': int(selected_staff_id) if selected_staff_id else None,
    }
    
    return render(request, 'compliance/add_staff_training_record.html', context)


@login_required
def submit_training_record(request):
    """Submit a new training record"""
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        completion_date = request.POST.get('completion_date')
        trainer_name = request.POST.get('trainer_name', '')
        training_provider = request.POST.get('training_provider', '')
        certificate_number = request.POST.get('certificate_number', '')
        sssc_cpd_hours = request.POST.get('sssc_cpd_hours', None)
        notes = request.POST.get('notes', '')
        
        try:
            course = TrainingCourse.objects.get(id=course_id)
            completion_date_obj = date.fromisoformat(completion_date)
            
            # Calculate expiry date based on course validity
            expiry_date = completion_date_obj + timedelta(days=course.validity_months * 30)
            
            # Create training record
            record = TrainingRecord.objects.create(
                staff_member=request.user,
                course=course,
                completion_date=completion_date_obj,
                expiry_date=expiry_date,
                trainer_name=trainer_name,
                training_provider=training_provider,
                certificate_number=certificate_number,
                sssc_cpd_hours_claimed=Decimal(sssc_cpd_hours) if sssc_cpd_hours else None,
                notes=notes,
                created_by=request.user
            )
            
            # Handle certificate file upload
            if 'certificate_file' in request.FILES:
                record.certificate_file = request.FILES['certificate_file']
                record.save()
            
            messages.success(request, f'Training record for {course.name} has been submitted successfully.')
            return redirect('my_training_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error submitting training record: {str(e)}')
    
    # GET request - show form
    courses = TrainingCourse.objects.all().order_by('category', 'name')
    
    context = {
        'courses': courses,
    }
    
    return render(request, 'compliance/submit_training_record.html', context)


@login_required
def training_breakdown_report(request):
    """Detailed training breakdown report - by person, course, or home"""
    import csv
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get filters
    selected_home = request.GET.get('care_home', None)
    selected_role = request.GET.get('role', None)
    selected_status = request.GET.get('status', None)
    view_type = request.GET.get('view_type', 'by_person')
    export_format = request.GET.get('export', None)
    show_all_courses = request.GET.get('show_all_courses', 'yes')  # Default to showing all
    
    # Get all courses or just mandatory based on filter
    if show_all_courses == 'yes':
        mandatory_courses = TrainingCourse.objects.all().order_by('is_mandatory', 'category', 'name')
    else:
        mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True).order_by('name')
    
    # Get care homes and roles
    care_homes = CareHome.objects.filter(is_active=True).order_by('name')
    from scheduling.models import Role
    staff_roles = Role.objects.all().order_by('name')
    
    # Build staff query
    staff_query = User.objects.filter(is_active=True)
    
    if selected_home:
        staff_query = staff_query.filter(unit__care_home__name=selected_home)
    
    if selected_role:
        staff_query = staff_query.filter(role__name=selected_role)
    
    staff_list = staff_query.distinct().select_related('role', 'unit__care_home').order_by('last_name', 'first_name')
    
    # ===== BY PERSON VIEW =====
    if view_type == 'by_person':
        staff_training_matrix = []
        
        for staff in staff_list:
            home_name = staff.unit.care_home.get_name_display() if staff.unit else "No Home"
            courses_status = []
            compliant_count = 0
            
            for course in mandatory_courses:
                latest_record = TrainingRecord.objects.filter(
                    staff_member=staff,
                    course=course
                ).order_by('-completion_date').first()
                
                if latest_record:
                    status = latest_record.get_status()
                    courses_status.append({
                        'course': course,
                        'status': status,
                        'completion_date': latest_record.completion_date,
                        'expiry_date': latest_record.expiry_date,
                        'days_until_expiry': latest_record.days_until_expiry()
                    })
                    if status == 'CURRENT':
                        compliant_count += 1
                else:
                    courses_status.append({
                        'course': course,
                        'status': 'MISSING',
                        'completion_date': None,
                        'expiry_date': None,
                        'days_until_expiry': None
                    })
            
            total_courses = mandatory_courses.count()
            compliance_pct = (compliant_count / total_courses * 100) if total_courses > 0 else 0
            
            # Filter by status if selected
            if selected_status:
                if selected_status == 'compliant' and compliance_pct < 100:
                    continue
                elif selected_status == 'expiring' and not any(c['status'] == 'EXPIRING_SOON' for c in courses_status):
                    continue
                elif selected_status == 'expired' and not any(c['status'] == 'EXPIRED' for c in courses_status):
                    continue
                elif selected_status == 'missing' and not any(c['status'] == 'MISSING' for c in courses_status):
                    continue
            
            staff_training_matrix.append({
                'staff': staff,
                'home_name': home_name,
                'courses': courses_status,
                'compliant_count': compliant_count,
                'total_courses': total_courses,
                'compliance_pct': compliance_pct
            })
    
    # ===== BY COURSE VIEW =====
    elif view_type == 'by_course':
        courses_breakdown = []
        
        for course in mandatory_courses:
            compliant = 0
            expiring = 0
            expired = 0
            missing = 0
            staff_records = []
            
            for staff in staff_list:
                home_name = staff.unit.care_home.get_name_display() if staff.unit else "No Home"
                latest_record = TrainingRecord.objects.filter(
                    staff_member=staff,
                    course=course
                ).order_by('-completion_date').first()
                
                if latest_record:
                    status = latest_record.get_status()
                    staff_records.append({
                        'staff': staff,
                        'home': home_name,
                        'completion_date': latest_record.completion_date,
                        'expiry_date': latest_record.expiry_date,
                        'status': status,
                        'days_until_expiry': latest_record.days_until_expiry()
                    })
                    
                    if status == 'CURRENT':
                        compliant += 1
                    elif status == 'EXPIRING_SOON':
                        expiring += 1
                    elif status == 'EXPIRED':
                        expired += 1
                else:
                    missing += 1
                    staff_records.append({
                        'staff': staff,
                        'home': home_name,
                        'completion_date': None,
                        'expiry_date': None,
                        'status': 'MISSING',
                        'days_until_expiry': None
                    })
            
            courses_breakdown.append({
                'course': course,
                'compliant': compliant,
                'expiring': expiring,
                'expired': expired,
                'missing': missing,
                'staff_records': staff_records
            })
    
    # ===== BY HOME SUMMARY VIEW =====
    else:
        home_summary = []
        
        for home in care_homes:
            if selected_home and home.name != selected_home:
                continue
            
            home_staff = staff_list.filter(unit__care_home=home)
            total_staff = home_staff.count()
            
            if total_staff == 0:
                continue
            
            # Calculate overall home compliance
            compliant_staff_count = 0
            expiring_staff_count = 0
            non_compliant_staff_count = 0
            
            course_breakdown = []
            
            for course in mandatory_courses:
                compliant = 0
                expiring = 0
                expired = 0
                missing = 0
                
                for staff in home_staff:
                    latest_record = TrainingRecord.objects.filter(
                        staff_member=staff,
                        course=course
                    ).order_by('-completion_date').first()
                    
                    if latest_record:
                        status = latest_record.get_status()
                        if status == 'CURRENT':
                            compliant += 1
                        elif status == 'EXPIRING_SOON':
                            expiring += 1
                        elif status == 'EXPIRED':
                            expired += 1
                    else:
                        missing += 1
                
                compliance_pct = (compliant / total_staff * 100) if total_staff > 0 else 0
                
                course_breakdown.append({
                    'course_name': course.name,
                    'compliant': compliant,
                    'expiring': expiring,
                    'expired': expired,
                    'missing': missing,
                    'percentage': compliance_pct
                })
            
            # Count staff by compliance status
            for staff in home_staff:
                staff_compliant_count = 0
                staff_has_expiring = False
                
                for course in mandatory_courses:
                    latest_record = TrainingRecord.objects.filter(
                        staff_member=staff,
                        course=course
                    ).order_by('-completion_date').first()
                    
                    if latest_record:
                        status = latest_record.get_status()
                        if status == 'CURRENT':
                            staff_compliant_count += 1
                        elif status == 'EXPIRING_SOON':
                            staff_has_expiring = True
                
                if staff_compliant_count == mandatory_courses.count():
                    compliant_staff_count += 1
                elif staff_has_expiring:
                    expiring_staff_count += 1
                else:
                    non_compliant_staff_count += 1
            
            compliance_percentage = (compliant_staff_count / total_staff * 100) if total_staff > 0 else 0
            
            home_summary.append({
                'home_name': home.get_name_display(),
                'total_staff': total_staff,
                'compliant_staff': compliant_staff_count,
                'expiring_staff': expiring_staff_count,
                'non_compliant_staff': non_compliant_staff_count,
                'compliance_percentage': compliance_percentage,
                'course_breakdown': course_breakdown
            })
    
    # Calculate overall statistics
    total_staff = staff_list.count()
    fully_compliant = 0
    has_expiring = 0
    has_expired_or_missing = 0
    
    for staff in staff_list:
        compliant_count = 0
        has_expiring_flag = False
        has_problem_flag = False
        
        for course in mandatory_courses:
            latest_record = TrainingRecord.objects.filter(
                staff_member=staff,
                course=course
            ).order_by('-completion_date').first()
            
            if latest_record:
                status = latest_record.get_status()
                if status == 'CURRENT':
                    compliant_count += 1
                elif status == 'EXPIRING_SOON':
                    has_expiring_flag = True
                elif status in ['EXPIRED', 'MISSING']:
                    has_problem_flag = True
            else:
                has_problem_flag = True
        
        if compliant_count == mandatory_courses.count():
            fully_compliant += 1
        elif has_expiring_flag:
            has_expiring += 1
        
        if has_problem_flag:
            has_expired_or_missing += 1
    
    compliance_percentage = (fully_compliant / total_staff * 100) if total_staff > 0 else 0
    
    # ===== EXPORT HANDLING =====
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="training_breakdown_{view_type}.csv"'
        
        writer = csv.writer(response)
        
        if view_type == 'by_person':
            # CSV Header
            header = ['Staff Name', 'Home', 'Role']
            for course in mandatory_courses:
                header.extend([f'{course.name} Status', f'{course.name} Expiry'])
            header.append('Compliance %')
            writer.writerow(header)
            
            # Data rows
            for item in staff_training_matrix:
                row = [item['staff'].full_name, item['home_name'], item['staff'].role.name]
                for course_status in item['courses']:
                    row.append(course_status['status'])
                    row.append(course_status['expiry_date'] if course_status['expiry_date'] else '—')
                row.append(f"{item['compliance_pct']:.1f}%")
                writer.writerow(row)
        
        return response
    
    elif export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Training Breakdown"
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        if view_type == 'by_person':
            # Excel Header
            headers = ['Staff Name', 'Home', 'Role']
            for course in mandatory_courses:
                headers.extend([f'{course.name} Status', f'{course.name} Expiry'])
            headers.append('Compliance %')
            
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for item in staff_training_matrix:
                row = [item['staff'].full_name, item['home_name'], item['staff'].role.name]
                for course_status in item['courses']:
                    row.append(course_status['status'])
                    row.append(str(course_status['expiry_date']) if course_status['expiry_date'] else '—')
                row.append(f"{item['compliance_pct']:.1f}%")
                ws.append(row)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="training_breakdown_{view_type}.xlsx"'
        wb.save(response)
        return response
    
    # Regular HTML view
    context = {
        'care_homes': care_homes,
        'staff_roles': staff_roles,
        'mandatory_courses': mandatory_courses,
        'selected_home': selected_home,
        'selected_role': selected_role,
        'selected_status': selected_status,
        'view_type': view_type,
        'show_all_courses': show_all_courses,
        'total_staff': total_staff,
        'fully_compliant': fully_compliant,
        'has_expiring': has_expiring,
        'has_expired_or_missing': has_expired_or_missing,
        'compliance_percentage': compliance_percentage,
    }
    
    if view_type == 'by_person':
        context['staff_training_matrix'] = staff_training_matrix
    elif view_type == 'by_course':
        context['courses_breakdown'] = courses_breakdown
    else:
        context['home_summary'] = home_summary
    
    return render(request, 'compliance/training_breakdown_report.html', context)


# ============================================================================
# INDUCTION TRACKING
# ============================================================================

@login_required
def my_induction_progress(request):
    """View personal induction progress"""
    
    try:
        induction = InductionProgress.objects.get(staff_member=request.user)
    except InductionProgress.DoesNotExist:
        messages.info(request, 'No induction record found. Contact your manager if you recently started.')
        return redirect('staff_dashboard')
    
    # Calculate progress
    completion_percentage = induction.get_completion_percentage()
    days_elapsed = (date.today() - induction.start_date).days
    weeks_elapsed = days_elapsed // 7
    
    context = {
        'induction': induction,
        'completion_percentage': completion_percentage,
        'weeks_elapsed': weeks_elapsed,
        'days_remaining': (induction.expected_completion_date - date.today()).days,
    }
    
    return render(request, 'compliance/my_induction_progress.html', context)


@login_required
def induction_management(request):
    """Manager view - list all staff inductions"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('staff_dashboard')
    
    # Get all induction records
    inductions = InductionProgress.objects.select_related(
        'staff_member', 'assessor'
    ).order_by('-start_date')
    
    # Calculate statistics
    total_inductions = inductions.count()
    in_progress = inductions.filter(final_assessment_complete=False).count()
    completed = inductions.filter(final_assessment_complete=True).count()
    
    # Add completion percentage to each induction
    for induction in inductions:
        induction.completion_pct = induction.get_completion_percentage()
    
    context = {
        'inductions': inductions,
        'total_inductions': total_inductions,
        'in_progress': in_progress,
        'completed': completed,
    }
    
    return render(request, 'compliance/induction_management.html', context)


@login_required
def update_induction_progress(request, induction_id):
    """Manager updates induction progress for a staff member"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('staff_dashboard')
    
    induction = get_object_or_404(InductionProgress, id=induction_id)
    
    if request.method == 'POST':
        # Update all checkbox fields
        induction.week1_orientation_complete = request.POST.get('week1_orientation_complete') == 'on'
        induction.week1_fire_safety_complete = request.POST.get('week1_fire_safety_complete') == 'on'
        induction.week1_infection_control_complete = request.POST.get('week1_infection_control_complete') == 'on'
        induction.week1_moving_handling_complete = request.POST.get('week1_moving_handling_complete') == 'on'
        induction.week1_health_safety_complete = request.POST.get('week1_health_safety_complete') == 'on'
        
        induction.week2_4_safeguarding_complete = request.POST.get('week2_4_safeguarding_complete') == 'on'
        induction.week2_4_person_centred_care_complete = request.POST.get('week2_4_person_centred_care_complete') == 'on'
        
        induction.week5_8_medication_complete = request.POST.get('week5_8_medication_complete') == 'on'
        induction.week5_8_clinical_skills_complete = request.POST.get('week5_8_clinical_skills_complete') == 'on'
        
        induction.week9_12_sssc_registration_complete = request.POST.get('week9_12_sssc_registration_complete') == 'on'
        induction.week9_12_quality_improvement_complete = request.POST.get('week9_12_quality_improvement_complete') == 'on'
        induction.week9_12_supervision_support_complete = request.POST.get('week9_12_supervision_support_complete') == 'on'
        
        # Update competency hours
        induction.personal_care_hours = Decimal(request.POST.get('personal_care_hours', 0))
        induction.meal_prep_hours = Decimal(request.POST.get('meal_prep_hours', 0))
        induction.documentation_hours = Decimal(request.POST.get('documentation_hours', 0))
        induction.medication_hours = Decimal(request.POST.get('medication_hours', 0))
        
        # Final assessment
        if request.POST.get('final_assessment_complete') == 'on':
            induction.final_assessment_complete = True
            if not induction.final_assessment_date:
                induction.final_assessment_date = date.today()
            induction.final_assessment_outcome = request.POST.get('final_assessment_outcome', '')
            induction.assessor = request.user
            
            if not induction.actual_completion_date:
                induction.actual_completion_date = date.today()
        
        induction.notes = request.POST.get('notes', '')
        induction.save()
        
        messages.success(request, f'Induction progress updated for {induction.staff_member.full_name}')
        return redirect('induction_management')
    
    context = {
        'induction': induction,
        'completion_percentage': induction.get_completion_percentage(),
    }
    
    return render(request, 'compliance/update_induction_progress.html', context)


# ============================================================================
# SUPERVISION RECORDS
# ============================================================================

@login_required
def my_supervision_records(request):
    """Staff member views their supervision history"""
    
    supervisions = SupervisionRecord.objects.filter(
        staff_member=request.user
    ).select_related('supervisor').order_by('-session_date')
    
    # Calculate statistics
    total_sessions = supervisions.count()
    this_year = supervisions.filter(session_date__year=timezone.now().year).count()
    
    # Find next scheduled supervision
    next_supervision = supervisions.filter(
        next_supervision_date__gte=date.today()
    ).order_by('next_supervision_date').first()
    
    # Check if supervision is overdue (more than 6 weeks since last)
    last_supervision = supervisions.first()
    supervision_overdue = False
    if last_supervision:
        days_since_last = (date.today() - last_supervision.session_date).days
        supervision_overdue = days_since_last > 42  # 6 weeks
    
    context = {
        'supervisions': supervisions,
        'total_sessions': total_sessions,
        'this_year': this_year,
        'next_supervision': next_supervision,
        'supervision_overdue': supervision_overdue,
        'last_supervision': last_supervision,
    }
    
    return render(request, 'compliance/my_supervision_records.html', context)


@login_required
def create_supervision_record(request):
    """Supervisor creates a supervision record"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to create supervision records.')
        return redirect('staff_dashboard')
    
    if request.method == 'POST':
        try:
            staff_member_id = request.POST.get('staff_member')
            staff_member = User.objects.get(sap=staff_member_id)
            
            supervision = SupervisionRecord.objects.create(
                staff_member=staff_member,
                supervisor=request.user,
                session_date=date.fromisoformat(request.POST.get('session_date')),
                session_type=request.POST.get('session_type'),
                duration_minutes=int(request.POST.get('duration_minutes', 60)),
                
                # Wellbeing
                wellbeing_score=int(request.POST.get('wellbeing_score', 5)),
                sickness_days_since_last=int(request.POST.get('sickness_days_since_last', 0)),
                wellbeing_concerns=request.POST.get('wellbeing_concerns', ''),
                support_offered=request.POST.get('support_offered', ''),
                
                # Performance
                performance_strengths=request.POST.get('performance_strengths', ''),
                performance_development=request.POST.get('performance_development', ''),
                
                # Training
                mandatory_training_current=request.POST.get('mandatory_training_current') == 'on',
                training_needs_identified=request.POST.get('training_needs_identified', ''),
                
                # SSSC
                sssc_registration_current=request.POST.get('sssc_registration_current') == 'on',
                sssc_cpd_hours_to_date=Decimal(request.POST.get('sssc_cpd_hours_to_date', 0)),
                
                # Safeguarding
                safeguarding_concerns_discussed=request.POST.get('safeguarding_concerns_discussed') == 'on',
                safeguarding_notes=request.POST.get('safeguarding_notes', ''),
                
                # Incidents
                incidents_since_last=int(request.POST.get('incidents_since_last', 0)),
                incident_learning=request.POST.get('incident_learning', ''),
                
                # Workload
                workload_manageable=request.POST.get('workload_manageable') == 'on',
                workload_notes=request.POST.get('workload_notes', ''),
                
                # Actions
                actions_from_previous=request.POST.get('actions_from_previous', ''),
                new_actions=request.POST.get('new_actions', ''),
                
                # Probationary
                is_probationary_review=request.POST.get('is_probationary_review') == 'on',
                probation_progress=request.POST.get('probation_progress', ''),
                probation_recommendation=request.POST.get('probation_recommendation', ''),
                
                # Next supervision
                next_supervision_date=date.fromisoformat(request.POST.get('next_supervision_date')) if request.POST.get('next_supervision_date') else None,
                
                supervisor_signature_date=date.today(),
            )
            
            messages.success(request, f'Supervision record created for {staff_member.full_name}')
            return redirect('supervision_management')
            
        except Exception as e:
            messages.error(request, f'Error creating supervision record: {str(e)}')
    
    # GET request - show form
    staff_members = User.objects.filter(
        is_active=True,
        is_staff=False
    ).order_by('last_name', 'first_name')
    
    context = {
        'staff_members': staff_members,
    }
    
    return render(request, 'compliance/create_supervision_record.html', context)


@login_required
def supervision_management(request):
    """Manager view - all supervision records"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('staff_dashboard')
    
    supervisions = SupervisionRecord.objects.select_related(
        'staff_member', 'supervisor'
    ).order_by('-session_date')
    
    # Calculate statistics
    total_sessions = supervisions.count()
    this_month = supervisions.filter(
        session_date__year=timezone.now().year,
        session_date__month=timezone.now().month
    ).count()
    
    # Staff needing supervision (no session in last 6 weeks)
    six_weeks_ago = date.today() - timedelta(weeks=6)
    staff_needing_supervision = []
    
    all_staff = User.objects.filter(is_active=True, is_staff=False)
    for staff in all_staff:
        last_supervision = SupervisionRecord.objects.filter(
            staff_member=staff
        ).order_by('-session_date').first()
        
        if not last_supervision or last_supervision.session_date < six_weeks_ago:
            staff_needing_supervision.append({
                'staff': staff,
                'last_session': last_supervision.session_date if last_supervision else None,
                'days_overdue': (date.today() - last_supervision.session_date).days if last_supervision else None
            })
    
    context = {
        'supervisions': supervisions,
        'total_sessions': total_sessions,
        'this_month': this_month,
        'staff_needing_supervision': staff_needing_supervision,
    }
    
    return render(request, 'compliance/supervision_management.html', context)


@login_required
def sign_supervision_record(request, record_id):
    """Allow staff member to sign their supervision record"""
    
    supervision = get_object_or_404(SupervisionRecord, id=record_id)
    
    # Check that the user is the staff member for this record
    if supervision.staff_member != request.user:
        messages.error(request, 'You can only sign your own supervision records.')
        return redirect('my_supervision_records')
    
    # Check if already signed
    if supervision.staff_signature_date:
        messages.warning(request, 'This supervision record has already been signed.')
        return redirect('my_supervision_records')
    
    if request.method == 'POST':
        # Get optional staff comments
        staff_comments = request.POST.get('staff_comments', '')
        
        # Sign the record
        supervision.staff_signature_date = date.today()
        supervision.staff_comments = staff_comments
        supervision.save()
        
        messages.success(request, 'Supervision record signed successfully.')
        return redirect('my_supervision_records')
    
    context = {
        'supervision': supervision,
    }
    
    return render(request, 'compliance/sign_supervision_record.html', context)


# ============================================================================
# INCIDENT REPORTING
# ============================================================================

@login_required
def report_incident(request):
    """Submit a new incident report"""
    
    if request.method == 'POST':
        try:
            # Generate reference number
            last_incident = IncidentReport.objects.order_by('-id').first()
            if last_incident:
                last_num = int(last_incident.reference_number.split('-')[-1])
                ref_number = f"IR-{timezone.now().year}-{last_num + 1:04d}"
            else:
                ref_number = f"IR-{timezone.now().year}-0001"
            
            incident = IncidentReport.objects.create(
                reference_number=ref_number,
                incident_date=date.fromisoformat(request.POST.get('incident_date')),
                incident_time=request.POST.get('incident_time'),
                reported_by=request.user,
                incident_type=request.POST.get('incident_type'),
                location=request.POST.get('location', ''),
                
                # Service user
                service_user_name=request.POST.get('service_user_name', ''),
                service_user_dob=date.fromisoformat(request.POST.get('service_user_dob')) if request.POST.get('service_user_dob') else None,
                
                # Description
                description=request.POST.get('description', ''),
                witnesses=request.POST.get('witnesses', ''),
                was_witnessed=request.POST.get('was_witnessed') == 'on',
                
                # Immediate actions
                immediate_actions=request.POST.get('immediate_actions', ''),
                injuries_sustained=request.POST.get('injuries_sustained', ''),
                body_map_completed=request.POST.get('body_map_completed') == 'on',
                photos_taken=request.POST.get('photos_taken') == 'on',
                
                # Medical
                gp_contacted=request.POST.get('gp_contacted') == 'on',
                ambulance_called=request.POST.get('ambulance_called') == 'on',
                hospital_attendance=request.POST.get('hospital_attendance') == 'on',
                hospital_admission=request.POST.get('hospital_admission') == 'on',
                medical_notes=request.POST.get('medical_notes', ''),
                
                # Severity
                severity=request.POST.get('severity'),
                risk_rating=request.POST.get('risk_rating'),
            )
            
            messages.success(request, f'Incident {ref_number} has been reported successfully. A manager will review it shortly.')
            
            # Check if Care Inspectorate notification required
            if incident.requires_care_inspectorate_notification():
                messages.warning(request, 'This incident requires Care Inspectorate notification. A manager has been notified.')
            
            return redirect('my_incident_reports')
            
        except Exception as e:
            messages.error(request, f'Error reporting incident: {str(e)}')
    
    # GET request - show form
    context = {
        'incident_types': IncidentReport._meta.get_field('incident_type').choices,
        'severity_choices': IncidentReport._meta.get_field('severity').choices,
        'risk_ratings': IncidentReport._meta.get_field('risk_rating').choices,
    }
    
    return render(request, 'compliance/report_incident.html', context)


@login_required
def my_incident_reports(request):
    """View incidents reported by current user"""
    
    incidents = IncidentReport.objects.filter(
        reported_by=request.user
    ).order_by('-incident_date', '-incident_time')
    
    context = {
        'incidents': incidents,
    }
    
    return render(request, 'compliance/my_incident_reports.html', context)


@login_required
def incident_management(request):
    """Manager view - all incident reports"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('staff_dashboard')
    
    # Filter options
    filter_type = request.GET.get('filter', 'all')
    
    incidents = IncidentReport.objects.select_related('reported_by', 'manager').order_by('-incident_date', '-incident_time')
    
    if filter_type == 'open':
        incidents = incidents.filter(incident_closed=False)
    elif filter_type == 'high_risk':
        incidents = incidents.filter(risk_rating__in=['HIGH', 'CRITICAL'])
    elif filter_type == 'care_inspectorate':
        incidents = incidents.filter(care_inspectorate_notified=True)
    elif filter_type == 'pending_review':
        incidents = incidents.filter(manager_reviewed=False)
    
    # Statistics
    total_incidents = IncidentReport.objects.count()
    this_month = IncidentReport.objects.filter(
        incident_date__year=timezone.now().year,
        incident_date__month=timezone.now().month
    ).count()
    open_incidents = IncidentReport.objects.filter(incident_closed=False).count()
    requiring_ci_notification = IncidentReport.objects.filter(
        care_inspectorate_notified=False
    ).filter(
        Q(severity='DEATH') | Q(severity='MAJOR_HARM') | 
        Q(incident_type__contains='SUSPECTED_') | Q(incident_type__contains='ALLEGATION')
    ).count()
    
    context = {
        'incidents': incidents,
        'total_incidents': total_incidents,
        'this_month': this_month,
        'open_incidents': open_incidents,
        'requiring_ci_notification': requiring_ci_notification,
        'current_filter': filter_type,
    }
    
    return render(request, 'compliance/incident_management.html', context)


@login_required
def view_incident(request, incident_id):
    """View detailed incident report"""
    
    incident = get_object_or_404(
        IncidentReport.objects.select_related(
            'reported_by', 
            'manager', 
            'investigation_assigned_to', 
            'closed_by'
        ), 
        id=incident_id
    )
    
    # Check permissions - staff can view their own, managers can view all
    if incident.reported_by != request.user and not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to view this incident.')
        return redirect('staff_dashboard')
    
    context = {
        'incident': incident,
        'can_edit': request.user.role and request.user.role.is_management,
    }
    
    return render(request, 'compliance/view_incident.html', context)


# ============================================================================
# REAL-TIME COMPLIANCE MONITOR - TASK 6 (Phase 2)
# ============================================================================

@login_required
def compliance_dashboard_api(request):
    """
    API endpoint for real-time compliance dashboard
    
    Returns JSON with:
    - Summary statistics (violations, compliance rate)
    - Active violations
    - Staff at risk of WTD violations
    - Upcoming risks
    - Weekly trends
    
    Usage:
        GET /compliance/dashboard/
        Response: {
            'summary': {
                'total_violations': int,
                'compliance_rate': float,
                ...
            },
            'active_violations': [...],
            'at_risk_staff': [...],
            ...
        }
    """
    from .compliance_monitor import get_compliance_dashboard
    
    # Get date range from query params (default 7 days)
    date_range_days = int(request.GET.get('days', 7))
    
    # Get dashboard data
    dashboard_data = get_compliance_dashboard(date_range_days)
    
    return JsonResponse(dashboard_data, status=200)


@login_required
def staff_compliance_status_api(request, user_id):
    """
    API endpoint for individual staff compliance status
    
    Returns WTD status for specific staff member:
    - Current weekly hours
    - 17-week rolling average
    - Upcoming shifts
    - Risk level
    
    Usage:
        GET /compliance/staff/123/status/
        Response: {
            'user_id': 123,
            'full_name': 'John Smith',
            'current_weekly_hours': 42.5,
            'rolling_average': 38.2,
            'risk_level': 'MEDIUM',
            ...
        }
    """
    from .compliance_monitor import ComplianceMonitor
    from .wdt_compliance import (
        calculate_weekly_hours,
        calculate_rolling_average_hours
    )
    
    # Get user
    try:
        staff_member = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    
    # Calculate compliance metrics
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    
    weekly_hours = calculate_weekly_hours(staff_member, week_start, weeks=1)
    rolling_avg = calculate_rolling_average_hours(staff_member, weeks=17)
    
    # Determine risk level
    monitor = ComplianceMonitor()
    if weekly_hours >= monitor.WTD_WARNING_THRESHOLD:
        risk_level = 'HIGH'
    elif weekly_hours >= Decimal('40.0'):
        risk_level = 'MEDIUM'
    else:
        risk_level = 'LOW'
    
    # Get upcoming shifts
    upcoming_shifts = list(staff_member.shift_set.filter(
        date__gte=today,
        date__lte=today + timedelta(days=7),
        status__in=['SCHEDULED', 'CONFIRMED']
    ).values('date', 'shift_type__name', 'unit__name'))
    
    return JsonResponse({
        'user_id': staff_member.id,
        'full_name': staff_member.full_name,
        'sap': staff_member.sap,
        'current_weekly_hours': float(weekly_hours),
        'rolling_average': float(rolling_avg),
        'risk_level': risk_level,
        'upcoming_shifts': upcoming_shifts,
        'wdt_limit': float(monitor.WTD_MAX_WEEKLY_HOURS),
        'hours_remaining': float(monitor.WTD_MAX_WEEKLY_HOURS - weekly_hours)
    }, status=200)


@login_required
def validate_assignment_api(request):
    """
    API endpoint to pre-validate shift assignment
    
    Checks if a proposed shift assignment is WTD/CI compliant
    before actually creating the shift
    
    Usage:
        POST /compliance/validate-assignment/
        Body: {
            'user_id': 123,
            'shift_date': '2025-12-28',
            'shift_type_id': 5,
            'proposed_hours': 12
        }
        
        Response: {
            'safe': bool,
            'compliant': bool,
            'violations': [...],
            'warnings': [...],
            'reason': str,
            'alternative_staff': [...]
        }
    """
    from .compliance_monitor import validate_shift_assignment
    from .models import ShiftType
    from datetime import datetime
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    # Parse request data
    try:
        import json
        data = json.loads(request.body)
        user_id = data.get('user_id')
        shift_date_str = data.get('shift_date')
        shift_type_id = data.get('shift_type_id')
        proposed_hours = data.get('proposed_hours', 12)
        
        # Validate inputs
        if not all([user_id, shift_date_str, shift_type_id]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Get user and shift type
        user = User.objects.get(pk=user_id)
        shift_type = ShiftType.objects.get(pk=shift_type_id)
        shift_date = datetime.strptime(shift_date_str, '%Y-%m-%d').date()
        
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except ShiftType.DoesNotExist:
        return JsonResponse({'error': 'Shift type not found'}, status=404)
    
    # Validate assignment
    result = validate_shift_assignment(user, shift_date, shift_type, proposed_hours)
    
    return JsonResponse(result, status=200)


@login_required
def staff_at_risk_api(request):
    """
    API endpoint for staff approaching WTD limits
    
    Returns list of staff who are approaching weekly hours limits
    
    Usage:
        GET /compliance/at-risk/?days=7&threshold=45
        Response: [
            {
                'user': {...},
                'full_name': str,
                'current_weekly_hours': float,
                'rolling_average': float,
                'risk_level': 'HIGH'|'MEDIUM'|'LOW',
                ...
            },
            ...
        ]
    """
    from .compliance_monitor import get_staff_at_risk
    
    # Get query params
    days_ahead = int(request.GET.get('days', 7))
    threshold_hours = int(request.GET.get('threshold', 45))
    
    # Get at-risk staff
    at_risk_staff = get_staff_at_risk(days_ahead, threshold_hours)
    
    return JsonResponse({
        'count': len(at_risk_staff),
        'staff': at_risk_staff
    }, status=200)


# ============================================================================
# AI-POWERED PAYROLL VALIDATOR - TASK 7 (Phase 2)
# ============================================================================

@login_required
def payroll_validation_api(request):
    """
    API endpoint for validating entire pay period
    
    Uses ML anomaly detection to flag discrepancies:
    - WTD hours mismatches (Task 6 integration)
    - Overtime anomalies
    - Agency cost validation
    - Fraud risk scoring
    
    Usage:
        GET /payroll/validate/?period_start=2025-12-01&period_end=2025-12-31
        Response: {
            'summary': {
                'total_entries': int,
                'flagged_entries': int,
                'total_discrepancy_amount': float,
                ...
            },
            'discrepancies': [...],
            'anomalies': [...],
            'fraud_alerts': [...]
        }
    """
    from .payroll_validator import validate_pay_period
    from datetime import datetime
    
    # Parse query params
    try:
        period_start_str = request.GET.get('period_start')
        period_end_str = request.GET.get('period_end')
        
        if not period_start_str or not period_end_str:
            return JsonResponse({'error': 'Missing period_start or period_end'}, status=400)
        
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d')
        period_end = datetime.strptime(period_end_str, '%Y-%m-%d')
        
    except ValueError as e:
        return JsonResponse({'error': f'Invalid date format: {str(e)}'}, status=400)
    
    # Validate pay period
    results = validate_pay_period(period_start, period_end)
    
    return JsonResponse(results, status=200)


@login_required
def payroll_entry_check_api(request):
    """
    API endpoint to check individual payroll entry
    
    Quick validation of claimed hours/amount vs scheduled
    
    Usage:
        POST /payroll/check-entry/
        Body: {
            'user_id': 123,
            'period_start': '2025-12-01',
            'period_end': '2025-12-31',
            'claimed_hours': 80.0,
            'claimed_amount': 920.00
        }
        
        Response: {
            'valid': bool,
            'issues': [str],
            'expected_hours': float,
            'expected_amount': float,
            'discrepancy': float
        }
    """
    from .payroll_validator import check_payroll_entry
    from datetime import datetime
    from decimal import Decimal
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        user_id = data.get('user_id')
        period_start_str = data.get('period_start')
        period_end_str = data.get('period_end')
        claimed_hours = Decimal(str(data.get('claimed_hours', 0)))
        claimed_amount = Decimal(str(data.get('claimed_amount', 0)))
        
        # Get user
        user = User.objects.get(pk=user_id)
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d')
        period_end = datetime.strptime(period_end_str, '%Y-%m-%d')
        
    except (json.JSONDecodeError, ValueError, User.DoesNotExist) as e:
        return JsonResponse({'error': f'Invalid request: {str(e)}'}, status=400)
    
    # Check entry
    result = check_payroll_entry(user, period_start, period_end, claimed_hours, claimed_amount)
    
    return JsonResponse(result, status=200)


@login_required
def fraud_risk_api(request, user_id):
    """
    API endpoint for fraud risk scoring
    
    Calculates fraud risk score for specific user in pay period
    
    Usage:
        GET /payroll/fraud-risk/123/?period_start=2025-12-01&period_end=2025-12-31
        Response: {
            'user_id': 123,
            'risk_score': 0.75,
            'risk_level': 'HIGH',
            'risk_factors': [...],
            'recommended_action': str
        }
    """
    from .payroll_validator import get_fraud_risk_score
    from datetime import datetime
    
    try:
        user = User.objects.get(pk=user_id)
        
        period_start_str = request.GET.get('period_start')
        period_end_str = request.GET.get('period_end')
        
        if not period_start_str or not period_end_str:
            return JsonResponse({'error': 'Missing period_start or period_end'}, status=400)
        
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d')
        period_end = datetime.strptime(period_end_str, '%Y-%m-%d')
        
    except (ValueError, User.DoesNotExist) as e:
        return JsonResponse({'error': str(e)}, status=400)
    
    # Calculate risk
    risk = get_fraud_risk_score(user, period_start, period_end)
    
    return JsonResponse(risk, status=200)


# ==============================================================================
# TASK 8: BUDGET-AWARE SMART RECOMMENDATIONS - Phase 2 API Endpoints
# ==============================================================================

@login_required
def budget_optimization_api(request):
    """
    API endpoint for budget-aware staffing optimization
    
    Finds cheapest WTD-compliant solution for staffing need
    Integrates ALL Tasks 1-7 with budget constraints
    
    Usage:
        POST /api/budget/optimize/
        Body: {
            "shift_date": "2025-12-28",
            "shift_type_id": 1,
            "unit_id": 3,
            "budget_limit": 200.00  // Optional
        }
        
        Response: {
            "recommended_option": "swap",  // or "overtime" or "agency"
            "cost": 0.00,
            "details": {...},
            "alternatives": [{type, cost, summary}],
            "budget_impact": {
                "cost": 0.00,
                "remaining_budget": 48240.50,
                "percentage_used": 3.5,
                "alert_level": "OK"
            },
            "compliance": {
                "wdt_compliant": true,
                "fraud_risk": "LOW"
            }
        }
    """
    from .budget_optimizer import get_optimal_staffing_solution
    from datetime import datetime
    from decimal import Decimal
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        shift_date_str = data.get('shift_date')
        shift_type_id = data.get('shift_type_id')
        unit_id = data.get('unit_id')
        budget_limit_str = data.get('budget_limit')
        
        if not shift_date_str or not shift_type_id or not unit_id:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        shift_date = datetime.strptime(shift_date_str, '%Y-%m-%d').date()
        shift_type = ShiftType.objects.get(pk=shift_type_id)
        unit = Unit.objects.get(pk=unit_id)
        
        budget_limit = Decimal(str(budget_limit_str)) if budget_limit_str else None
        
    except (ValueError, KeyError, ShiftType.DoesNotExist, Unit.DoesNotExist) as e:
        return JsonResponse({'error': str(e)}, status=400)
    
    # Get optimal solution (integrates Tasks 1-7)
    solution = get_optimal_staffing_solution(shift_date, shift_type, unit, budget_limit)
    
    return JsonResponse(solution, status=200)


@login_required
def budget_status_api(request):
    """
    API endpoint for current budget status
    
    Shows real-time budget tracking with spending breakdown
    
    Usage:
        GET /api/budget/status/?period_start=2025-12-01&period_end=2025-12-31
        
        Response: {
            "period": {
                "start": "2025-12-01",
                "end": "2025-12-31",
                "days_elapsed": 25,
                "days_remaining": 6
            },
            "spending": {
                "total": 38450.00,
                "regular_shifts": 24000.00,
                "overtime": 5400.00,
                "agency": 9050.00,
                "breakdown_percentage": {
                    "regular": 62.4,
                    "overtime": 14.0,
                    "agency": 23.6
                }
            },
            "budget": {
                "allocated": 50000.00,
                "spent": 38450.00,
                "remaining": 11550.00,
                "percentage_used": 76.9
            },
            "alerts": [
                {"level": "WARNING", "message": "Budget 76.9% used - monitor closely"}
            ],
            "projections": {
                "daily_burn_rate": 1538.00,
                "end_of_month": 47678.00,
                "overspend_risk": false,
                "projected_overspend": 0
            }
        }
    """
    from .budget_optimizer import get_budget_status
    from datetime import datetime
    
    period_start_str = request.GET.get('period_start')
    period_end_str = request.GET.get('period_end')
    
    try:
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date() if period_start_str else None
        period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date() if period_end_str else None
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    
    # Get budget status
    status = get_budget_status(period_start, period_end)
    
    return JsonResponse(status, status=200)


@login_required
def budget_forecast_api(request):
    """
    API endpoint for budget forecast
    
    Predicts future budget needs using ML shortage predictions
    Integrates with Task 5 (shortage_predictor)
    
    Usage:
        GET /api/budget/forecast/?days_ahead=30
        
        Response: {
            "forecast_period": {
                "start": "2025-12-25",
                "end": "2026-01-24",
                "days": 30
            },
            "predicted_shortages": 18,
            "estimated_costs": {
                "optimistic": 972.00,    // 70% swaps, 30% OT
                "realistic": 2484.00,    // 40% swaps, 40% OT, 20% agency
                "pessimistic": 4032.00   // 20% swaps, 30% OT, 50% agency
            },
            "budget_recommendations": [
                "💡 Optimize costs: Prioritize shift swaps (£0) over agency (£280/shift)."
            ]
        }
    """
    from .budget_optimizer import predict_budget_needs
    
    days_ahead_str = request.GET.get('days_ahead', '30')
    
    try:
        days_ahead = int(days_ahead_str)
        if days_ahead < 1 or days_ahead > 365:
            raise ValueError('days_ahead must be between 1 and 365')
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    
    # Get forecast (integrates with Task 5)
    forecast = predict_budget_needs(days_ahead)
    
    return JsonResponse(forecast, status=200)


# ==============================================================================
# TASK 10: NATURAL LANGUAGE QUERY INTERFACE - Phase 3 API Endpoints
# ==============================================================================

@login_required
def ai_assistant_query_api(request):
    """
    API endpoint for natural language AI assistant queries
    
    Processes plain English questions and routes to appropriate AI systems
    
    Usage:
        POST /api/ai-assistant/query/
        Body: {
            "query": "Who can work tomorrow?",
            "user_id": 123  // Optional
        }
        
        Response: {
            "query": "Who can work tomorrow?",
            "intent": "staffing_shortage",
            "confidence": 0.9,
            "entities": {
                "date": "2025-12-27",
                "shift_type": null,
                "staff_name": null,
                "unit_name": null
            },
            "response": "**Found 5 available staff for Friday, December 27:**...",
            "data": {...},
            "suggestions": [
                "Send offers to top 3 matches",
                "Check budget impact",
                "View full availability report"
            ]
        }
    """
    from .nlp_query_processor import process_natural_language_query
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=405)
    
    try:
        data = json.loads(request.body)
        query = data.get('query')
        user_id = data.get('user_id') or request.user.id
        
        if not query:
            return JsonResponse({'error': 'query field is required'}, status=400)
        
        if not isinstance(query, str) or len(query) < 3:
            return JsonResponse({'error': 'query must be at least 3 characters'}, status=400)
        
        # Process natural language query
        result = process_natural_language_query(query, user_id)
        
        # Add original query to response
        result['query'] = query
        
        return JsonResponse(result, status=200)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_login_required
def ai_assistant_suggestions_api(request):
    """
    API endpoint for getting example queries
    
    Returns list of example queries users can try
    
    Usage:
        GET /api/ai-assistant/suggestions/
        
        Response: {
            "suggestions": [
                "Who can work tomorrow?",
                "Show me the budget status",
                "Is John Smith WTD compliant?",
                ...
            ],
            "count": 10
        }
    """
    from .nlp_query_processor import get_query_suggestions
    
    suggestions = get_query_suggestions()
    
    return JsonResponse({
        'suggestions': suggestions,
        'count': len(suggestions)
    }, status=200)


@api_login_required
@require_http_methods(["POST"])
def ai_assistant_feedback_api(request):
    """
    Task 11: Submit feedback for AI assistant query
    
    POST /api/ai-assistant/feedback/
    
    Body: {
        "query_text": "Who can work tomorrow?",
        "intent_detected": "STAFF_AVAILABILITY",
        "confidence_score": 0.95,
        "response_text": "3 staff members can work tomorrow...",
        "response_data": {...},
        "rating": 4,
        "feedback_type": "HELPFUL",
        "feedback_comment": "Clear and accurate"
    }
    
    Response: {
        "feedback_id": 123,
        "preferences_updated": true,
        "user_preferences": {
            "detail_level": "STANDARD",
            "tone": "FRIENDLY",
            "avg_satisfaction": 4.2
        }
    }
    """
    from .feedback_learning import record_query_feedback, get_user_preferences
    
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['query_text', 'intent_detected', 'confidence_score', 
                          'response_text', 'rating']
        for field in required_fields:
            if field not in data:
                return JsonResponse({
                    'error': f'Missing required field: {field}'
                }, status=400)
        
        # Validate rating
        rating = int(data['rating'])
        if not 1 <= rating <= 5:
            return JsonResponse({
                'error': 'Rating must be between 1 and 5'
            }, status=400)
        
        # Record feedback
        feedback = record_query_feedback(
            user=request.user,
            query_text=data['query_text'],
            intent_detected=data['intent_detected'],
            confidence_score=float(data['confidence_score']),
            response_text=data['response_text'],
            response_data=data.get('response_data', {}),
            rating=rating,
            feedback_type=data.get('feedback_type'),
            feedback_comment=data.get('feedback_comment', '')
        )
        
        # Get updated preferences
        preferences = get_user_preferences(request.user)
        
        return JsonResponse({
            'feedback_id': feedback.id,
            'preferences_updated': True,
            'user_preferences': {
                'detail_level': preferences.preferred_detail_level,
                'tone': preferences.preferred_tone,
                'avg_satisfaction': float(preferences.avg_satisfaction_rating or 0),
                'total_queries': preferences.total_queries,
                'total_feedback': preferences.total_feedback_count
            }
        }, status=201)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_login_required
@require_http_methods(["GET"])
def ai_assistant_analytics_api(request):
    """
    Task 11: Get AI assistant performance analytics
    
    GET /api/ai-assistant/analytics/?days=30
    
    Response: {
        "period_days": 30,
        "total_queries": 150,
        "total_feedback": 45,
        "avg_rating": 4.2,
        "satisfaction_rate": 0.78,
        "positive_count": 35,
        "negative_count": 5,
        "by_intent": {
            "STAFF_AVAILABILITY": {"count": 20, "avg_rating": 4.5},
            "SHIFT_SEARCH": {"count": 15, "avg_rating": 4.0},
            ...
        },
        "by_rating": {
            "5": 20, "4": 15, "3": 5, "2": 3, "1": 2
        },
        "improvement_needed": [
            {"intent": "LEAVE_BALANCE", "count": 10, "avg_rating": 2.5},
            ...
        ]
    }
    """
    from .feedback_learning import get_feedback_analytics
    
    try:
        # Get days parameter (default 30)
        days = int(request.GET.get('days', 30))
        if days < 1:
            return JsonResponse({'error': 'days must be >= 1'}, status=400)
        
        # Get analytics (all users if senior, own data otherwise)
        user = request.user if not request.user.groups.filter(
            name__in=['Senior Staff', 'Admin']
        ).exists() else None
        
        analytics = get_feedback_analytics(user=user, days=days)
        
        return JsonResponse({
            'period_days': days,
            **analytics
        }, status=200)
        
    except ValueError:
        return JsonResponse({'error': 'Invalid days parameter'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_login_required
@require_http_methods(["GET"])
def ai_assistant_insights_api(request):
    """
    Task 11: Get AI assistant learning insights
    
    GET /api/ai-assistant/insights/?min_feedback=5
    
    Response: {
        "high_performing_intents": [
            {"intent": "STAFF_AVAILABILITY", "count": 50, "avg_rating": 4.5},
            ...
        ],
        "low_performing_intents": [
            {"intent": "LEAVE_BALANCE", "count": 20, "avg_rating": 2.3},
            ...
        ],
        "common_misclassifications": [
            {
                "detected_intent": "SHIFT_SEARCH",
                "common_feedback": "WRONG_INTENT",
                "count": 8,
                "sample_queries": ["find my shifts", "when do I work"]
            },
            ...
        ],
        "user_satisfaction_leaders": [
            {"user_id": 5, "username": "john.smith", "avg_rating": 4.8, "total": 25},
            ...
        ],
        "recommendations": [
            "Improve LEAVE_BALANCE intent accuracy (avg rating: 2.3)",
            "Review SHIFT_SEARCH misclassifications (8 reported)",
            ...
        ]
    }
    """
    from .feedback_learning import get_learning_insights
    
    # Only senior staff and admin can view insights
    if not request.user.groups.filter(name__in=['Senior Staff', 'Admin']).exists():
        return JsonResponse({
            'error': 'Insufficient permissions. Senior Staff or Admin access required.'
        }, status=403)
    
    try:
        # Get min_feedback parameter (default 5)
        min_feedback = int(request.GET.get('min_feedback', 5))
        if min_feedback < 1:
            return JsonResponse({'error': 'min_feedback must be >= 1'}, status=400)
        
        insights = get_learning_insights(min_feedback_count=min_feedback)
        
        return JsonResponse(insights, status=200)
        
    except ValueError:
        return JsonResponse({'error': 'Invalid min_feedback parameter'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
