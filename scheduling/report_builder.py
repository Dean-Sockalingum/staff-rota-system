"""
TASK 27: CUSTOM REPORT BUILDER SERVICE
Build dynamic reports with field selection, filtering, and multi-format export
"""

from django.db.models import Q, Count, Sum, Avg, F, Max, Min, Case, When, Value, CharField
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


# ============================================================================
# FIELD DEFINITIONS - Available fields for each report type
# ============================================================================

AVAILABLE_FIELDS = {
    'STAFFING': {
        'staff_name': {'label': 'Staff Name', 'model': 'User', 'field': 'full_name'},
        'sap': {'label': 'SAP Number', 'model': 'User', 'field': 'sap'},
        'role': {'label': 'Role', 'model': 'User', 'field': 'role__name'},
        'care_home': {'label': 'Care Home', 'model': 'User', 'field': 'care_home__name'},
        'unit': {'label': 'Unit', 'model': 'User', 'field': 'unit__name'},
        'contract_type': {'label': 'Contract Type', 'model': 'User', 'field': 'contract_type'},
        'hours_contracted': {'label': 'Contracted Hours', 'model': 'User', 'field': 'hours_contracted'},
        'is_active': {'label': 'Active Status', 'model': 'User', 'field': 'is_active'},
        'date_joined': {'label': 'Join Date', 'model': 'User', 'field': 'date_joined'},
    },
    'SHIFTS': {
        'shift_date': {'label': 'Shift Date', 'model': 'Shift', 'field': 'date'},
        'shift_type': {'label': 'Shift Type', 'model': 'Shift', 'field': 'shift_type__name'},
        'staff_name': {'label': 'Staff Name', 'model': 'Shift', 'field': 'staff__full_name'},
        'care_home': {'label': 'Care Home', 'model': 'Shift', 'field': 'care_home__name'},
        'unit': {'label': 'Unit', 'model': 'Shift', 'field': 'unit__name'},
        'start_time': {'label': 'Start Time', 'model': 'Shift', 'field': 'shift_type__start_time'},
        'end_time': {'label': 'End Time', 'model': 'Shift', 'field': 'shift_type__end_time'},
        'hours': {'label': 'Hours', 'model': 'Shift', 'field': 'shift_type__hours'},
        'is_agency': {'label': 'Agency Shift', 'model': 'Shift', 'field': 'is_agency'},
    },
    'LEAVE': {
        'staff_name': {'label': 'Staff Name', 'model': 'LeaveRequest', 'field': 'staff__full_name'},
        'leave_type': {'label': 'Leave Type', 'model': 'LeaveRequest', 'field': 'leave_type'},
        'start_date': {'label': 'Start Date', 'model': 'LeaveRequest', 'field': 'start_date'},
        'end_date': {'label': 'End Date', 'model': 'LeaveRequest', 'field': 'end_date'},
        'status': {'label': 'Status', 'model': 'LeaveRequest', 'field': 'status'},
        'requested_date': {'label': 'Requested Date', 'model': 'LeaveRequest', 'field': 'requested_date'},
        'approved_by': {'label': 'Approved By', 'model': 'LeaveRequest', 'field': 'approved_by__full_name'},
        'care_home': {'label': 'Care Home', 'model': 'LeaveRequest', 'field': 'staff__care_home__name'},
    },
    'TRAINING': {
        'staff_name': {'label': 'Staff Name', 'model': 'TrainingRecord', 'field': 'staff__full_name'},
        'course_name': {'label': 'Course Name', 'model': 'TrainingRecord', 'field': 'course__name'},
        'course_category': {'label': 'Category', 'model': 'TrainingRecord', 'field': 'course__category'},
        'completion_date': {'label': 'Completion Date', 'model': 'TrainingRecord', 'field': 'completion_date'},
        'expiry_date': {'label': 'Expiry Date', 'model': 'TrainingRecord', 'field': 'expiry_date'},
        'status': {'label': 'Status', 'model': 'TrainingRecord', 'field': 'status'},
        'care_home': {'label': 'Care Home', 'model': 'TrainingRecord', 'field': 'staff__care_home__name'},
    },
    'COMPLIANCE': {
        'staff_name': {'label': 'Staff Name', 'model': 'User', 'field': 'full_name'},
        'care_home': {'label': 'Care Home', 'model': 'User', 'field': 'care_home__name'},
        'pvg_expiry': {'label': 'PVG Expiry', 'model': 'User', 'field': 'pvg_expiry_date'},
        'sssc_registration': {'label': 'SSSC Number', 'model': 'User', 'field': 'sssc_registration_number'},
        'sssc_expiry': {'label': 'SSSC Expiry', 'model': 'User', 'field': 'sssc_expiry_date'},
        'contract_type': {'label': 'Contract Type', 'model': 'User', 'field': 'contract_type'},
    },
    'INCIDENTS': {
        'incident_date': {'label': 'Incident Date', 'model': 'IncidentReport', 'field': 'incident_date'},
        'incident_type': {'label': 'Incident Type', 'model': 'IncidentReport', 'field': 'incident_type'},
        'severity': {'label': 'Severity', 'model': 'IncidentReport', 'field': 'severity'},
        'reported_by': {'label': 'Reported By', 'model': 'IncidentReport', 'field': 'reported_by__full_name'},
        'care_home': {'label': 'Care Home', 'model': 'IncidentReport', 'field': 'care_home__name'},
        'status': {'label': 'Status', 'model': 'IncidentReport', 'field': 'status'},
    },
    'RESIDENTS': {
        'resident_name': {'label': 'Resident Name', 'model': 'Resident', 'field': 'full_name'},
        'care_home': {'label': 'Care Home', 'model': 'Resident', 'field': 'care_home__name'},
        'unit': {'label': 'Unit', 'model': 'Resident', 'field': 'unit__name'},
        'room_number': {'label': 'Room Number', 'model': 'Resident', 'field': 'room_number'},
        'admission_date': {'label': 'Admission Date', 'model': 'Resident', 'field': 'admission_date'},
        'care_level': {'label': 'Care Level', 'model': 'Resident', 'field': 'care_level'},
        'is_active': {'label': 'Active', 'model': 'Resident', 'field': 'is_active'},
    },
}


# ============================================================================
# QUERY BUILDER - Build dynamic queries based on selected fields and filters
# ============================================================================

def build_report_query(report_type, selected_fields, filters):
    """
    Build a dynamic query based on report type, selected fields, and filters
    
    Args:
        report_type: Type of report (STAFFING, SHIFTS, LEAVE, etc.)
        selected_fields: List of field names to include
        filters: Dictionary of filter criteria
    
    Returns:
        QuerySet with filtered data
    """
    from .models import User, Shift, LeaveRequest, TrainingRecord, IncidentReport, Resident
    
    # Determine base model
    model_map = {
        'STAFFING': User,
        'SHIFTS': Shift,
        'LEAVE': LeaveRequest,
        'TRAINING': TrainingRecord,
        'COMPLIANCE': User,
        'INCIDENTS': IncidentReport,
        'RESIDENTS': Resident,
    }
    
    model = model_map.get(report_type)
    if not model:
        return None
    
    queryset = model.objects.all()
    
    # Apply filters
    if filters:
        # Date range filters
        if 'date_from' in filters and filters['date_from']:
            date_field = get_date_field_for_model(report_type)
            if date_field:
                queryset = queryset.filter(**{f"{date_field}__gte": filters['date_from']})
        
        if 'date_to' in filters and filters['date_to']:
            date_field = get_date_field_for_model(report_type)
            if date_field:
                queryset = queryset.filter(**{f"{date_field}__lte": filters['date_to']})
        
        # Care home filter
        if 'care_home' in filters and filters['care_home']:
            care_home_field = get_care_home_field_for_model(report_type)
            if care_home_field:
                queryset = queryset.filter(**{care_home_field: filters['care_home']})
        
        # Unit filter
        if 'unit' in filters and filters['unit']:
            unit_field = get_unit_field_for_model(report_type)
            if unit_field:
                queryset = queryset.filter(**{unit_field: filters['unit']})
        
        # Staff filter
        if 'staff' in filters and filters['staff']:
            staff_field = get_staff_field_for_model(report_type)
            if staff_field:
                queryset = queryset.filter(**{staff_field: filters['staff']})
        
        # Role filter (for staff-related reports)
        if 'role' in filters and filters['role']:
            if report_type in ['STAFFING', 'COMPLIANCE']:
                queryset = queryset.filter(role__name=filters['role'])
        
        # Status filters
        if 'status' in filters and filters['status']:
            if hasattr(model, 'status'):
                queryset = queryset.filter(status=filters['status'])
        
        # Active only filter
        if 'active_only' in filters and filters['active_only']:
            if hasattr(model, 'is_active'):
                queryset = queryset.filter(is_active=True)
    
    # Select related fields for optimization
    queryset = optimize_queryset(queryset, report_type)
    
    return queryset


def get_date_field_for_model(report_type):
    """Get the primary date field for a model type"""
    date_fields = {
        'STAFFING': 'date_joined',
        'SHIFTS': 'date',
        'LEAVE': 'start_date',
        'TRAINING': 'completion_date',
        'COMPLIANCE': 'date_joined',
        'INCIDENTS': 'incident_date',
        'RESIDENTS': 'admission_date',
    }
    return date_fields.get(report_type)


def get_care_home_field_for_model(report_type):
    """Get the care home foreign key field for a model"""
    care_home_fields = {
        'STAFFING': 'care_home',
        'SHIFTS': 'care_home',
        'LEAVE': 'staff__care_home',
        'TRAINING': 'staff__care_home',
        'COMPLIANCE': 'care_home',
        'INCIDENTS': 'care_home',
        'RESIDENTS': 'care_home',
    }
    return care_home_fields.get(report_type)


def get_unit_field_for_model(report_type):
    """Get the unit foreign key field for a model"""
    unit_fields = {
        'STAFFING': 'unit',
        'SHIFTS': 'unit',
        'LEAVE': 'staff__unit',
        'TRAINING': 'staff__unit',
        'COMPLIANCE': 'unit',
        'RESIDENTS': 'unit',
    }
    return unit_fields.get(report_type)


def get_staff_field_for_model(report_type):
    """Get the staff foreign key field for a model"""
    staff_fields = {
        'SHIFTS': 'staff',
        'LEAVE': 'staff',
        'TRAINING': 'staff',
    }
    return staff_fields.get(report_type)


def optimize_queryset(queryset, report_type):
    """Add select_related and prefetch_related for performance"""
    if report_type == 'STAFFING':
        queryset = queryset.select_related('role', 'care_home', 'unit')
    elif report_type == 'SHIFTS':
        queryset = queryset.select_related('staff', 'shift_type', 'care_home', 'unit')
    elif report_type == 'LEAVE':
        queryset = queryset.select_related('staff', 'staff__care_home', 'approved_by')
    elif report_type == 'TRAINING':
        queryset = queryset.select_related('staff', 'course', 'staff__care_home')
    elif report_type == 'INCIDENTS':
        queryset = queryset.select_related('reported_by', 'care_home')
    elif report_type == 'RESIDENTS':
        queryset = queryset.select_related('care_home', 'unit')
    
    return queryset


# ============================================================================
# DATA EXTRACTION - Extract report data from queryset
# ============================================================================

def extract_report_data(queryset, selected_fields, report_type):
    """
    Extract data from queryset based on selected fields
    
    Returns:
        List of dictionaries with report data
    """
    if not queryset or not selected_fields:
        return []
    
    available_fields = AVAILABLE_FIELDS.get(report_type, {})
    data = []
    
    for obj in queryset:
        row = {}
        for field_name in selected_fields:
            field_config = available_fields.get(field_name)
            if field_config:
                # Navigate through foreign keys using double underscore notation
                field_path = field_config['field']
                value = get_nested_field_value(obj, field_path)
                row[field_config['label']] = format_field_value(value)
        data.append(row)
    
    return data


def get_nested_field_value(obj, field_path):
    """Get value from nested field path (e.g., 'staff__care_home__name')"""
    parts = field_path.split('__')
    value = obj
    
    for part in parts:
        if value is None:
            return None
        try:
            value = getattr(value, part)
            if callable(value):
                value = value()
        except AttributeError:
            return None
    
    return value


def format_field_value(value):
    """Format field value for display"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    return str(value)


# ============================================================================
# EXPORT FUNCTIONS - Generate reports in different formats
# ============================================================================

def generate_csv_report(data, filename='report.csv'):
    """Generate CSV report"""
    output = BytesIO()
    
    if not data:
        return output
    
    # Use string mode for CSV writer
    import io
    string_output = io.StringIO()
    
    # Get headers from first row
    headers = list(data[0].keys())
    
    writer = csv.DictWriter(string_output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(data)
    
    # Convert to bytes
    output.write(string_output.getvalue().encode('utf-8'))
    output.seek(0)
    
    return output


def generate_excel_report(data, report_name='Report', filename='report.xlsx'):
    """Generate Excel report with formatting"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl not installed. Cannot generate Excel reports.")
    
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = report_name[:31]  # Excel sheet name limit
    
    if not data:
        workbook.save(output)
        output.seek(0)
        return output
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Write headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    workbook.save(output)
    output.seek(0)
    
    return output


def generate_pdf_report(data, report_name='Report', filters=None, filename='report.pdf'):
    """Generate PDF report with professional formatting"""
    if not PDF_AVAILABLE:
        raise ImportError("reportlab not installed. Cannot generate PDF reports.")
    
    output = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Add title
    title = Paragraph(report_name, title_style)
    elements.append(title)
    
    # Add generation date and filters
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )
    
    generation_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    info_text = f"Generated: {generation_date}"
    
    if filters:
        filter_text = "<br/>".join([f"{k}: {v}" for k, v in filters.items() if v])
        if filter_text:
            info_text += f"<br/><b>Filters:</b><br/>{filter_text}"
    
    info = Paragraph(info_text, info_style)
    elements.append(info)
    elements.append(Spacer(1, 0.2 * inch))
    
    if not data:
        no_data = Paragraph("No data available for selected criteria.", styles['Normal'])
        elements.append(no_data)
    else:
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data:
            table_data.append([row.get(h, '') for h in headers])
        
        # Create table
        col_widths = [doc.width / len(headers)] * len(headers)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Style table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        elements.append(table)
    
    # Add footer with page numbers
    doc.build(elements)
    output.seek(0)
    
    return output


# ============================================================================
# AGGREGATION FUNCTIONS - Summary statistics for reports
# ============================================================================

def calculate_report_summary(data, report_type):
    """Calculate summary statistics for report"""
    if not data:
        return {}
    
    summary = {
        'total_records': len(data),
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }
    
    # Type-specific summaries
    if report_type == 'STAFFING':
        summary['total_staff'] = len(data)
        # Count by role if role field present
        if 'Role' in data[0]:
            role_counts = {}
            for row in data:
                role = row.get('Role', 'Unknown')
                role_counts[role] = role_counts.get(role, 0) + 1
            summary['by_role'] = role_counts
    
    elif report_type == 'SHIFTS':
        summary['total_shifts'] = len(data)
        # Calculate total hours if available
        if 'Hours' in data[0]:
            total_hours = sum([float(row.get('Hours', 0) or 0) for row in data])
            summary['total_hours'] = round(total_hours, 2)
    
    elif report_type == 'LEAVE':
        summary['total_requests'] = len(data)
        # Count by status
        if 'Status' in data[0]:
            status_counts = {}
            for row in data:
                status = row.get('Status', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            summary['by_status'] = status_counts
    
    return summary
