from collections import defaultdict

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.db import models
from django.http import JsonResponse
from django.conf import settings
from datetime import datetime, timedelta, date
from pathlib import Path
from django.utils.html import escape
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_http_methods
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date

from .models import (
    ActivityLog,
    CarePlanReview,
    LeaveRequest,
    Resident,
    Role,
    Shift,
    ShiftSwapRequest,
    ShiftType,
    StaffReallocation,
    Unit,
    User,
    IncidentReport,
)
from .models_multi_home import CareHome
from .models_feedback import DemoFeedback, FeatureRequest
from .forms_feedback import DemoFeedbackForm, FeatureRequestForm
from staff_records.models import SicknessRecord, StaffProfile

def login_view(request):
    """Login page"""
    if request.method == 'POST':
        sap = request.POST.get('sap', '').strip()
        password = request.POST.get('password', '')
        
        # Extensive debug output
        print("="*50)
        print(f"POST data: {dict(request.POST)}")
        print(f"Login attempt - SAP: '{sap}', Password: '{password}'")
        print(f"Password length: {len(password) if password else 0}")
        
        if not sap or not password:
            print("ERROR: Missing SAP or password")
            messages.error(request, 'Please provide both SAP number and password')
            return render(request, 'scheduling/login.html')
        
        user = authenticate(request, username=sap, password=password)
        print(f"Authentication result: {user}")
        
        if user:
            if not user.is_active:
                print(f"ERROR: User {sap} is inactive")
                messages.error(request, 'Your account is inactive')
                return render(request, 'scheduling/login.html')
                
            login(request, user)
            print(f"✓ Login successful for {user.sap}")
            
            if user.role and user.role.is_management:
                print(f"Redirecting to manager_dashboard")
                return redirect('manager_dashboard')
            else:
                print(f"Redirecting to staff_dashboard")
                return redirect('staff_dashboard')
        else:
            print(f"✗ Authentication failed for SAP: '{sap}'")
            messages.error(request, 'Invalid SAP number or password')
        print("="*50)
    
    return render(request, 'scheduling/login.html')

def logout_view(request):
    """Logout and redirect to login"""
    if request.method == 'POST' or request.method == 'GET':
        logout(request)
        return redirect('login')
    return redirect('login')

def offline_view(request):
    """Offline fallback page for PWA"""
    return render(request, 'scheduling/offline.html')

# ==================== MANAGER/ADMIN VIEWS ====================

@login_required
def manager_dashboard(request):
    # Detect user's home and determine if they can select different homes
    from .models_multi_home import CareHome
    
    user_home = None
    can_select_home = False
    
    # Check if user is senior management (can view any home)
    if request.user.role and request.user.role.is_senior_management_team:
        can_select_home = True
        # Allow selection via GET parameter
        selected_home_name = request.GET.get('care_home', '')
        if selected_home_name:
            try:
                user_home = CareHome.objects.get(name=selected_home_name)
            except CareHome.DoesNotExist:
                user_home = None
    else:
        # Regular managers are locked to their home
        if request.user.unit and request.user.unit.care_home:
            user_home = request.user.unit.care_home
    
    # Coverage summary for the full rota period
    rota_start = datetime(2026, 1, 4).date()
    rota_end = rota_start + timedelta(weeks=12) - timedelta(days=1)
    all_days = [(rota_start + timedelta(days=i)) for i in range((rota_end-rota_start).days+1)]
    day_shift_names = ['DAY', 'DAY_SENIOR', 'DAY_ASSISTANT']
    night_shift_names = ['NIGHT', 'NIGHT_SENIOR', 'NIGHT_ASSISTANT']
    coverage_table = []
    
    # Filter units by home if applicable
    units_qs = Unit.objects.filter(is_active=True)
    if user_home:
        units_qs = units_qs.filter(care_home=user_home)
    units = list(units_qs)
    for day in all_days:
        row = {'date': day}
        for unit in units:
            # Day shifts
            day_shifts = Shift.objects.filter(date=day, unit=unit, shift_type__name__in=day_shift_names)
            sscw_day = day_shifts.filter(user__role__name='SSCW').count()
            care_day = day_shifts.filter(user__role__name__in=['SCW', 'SCA']).count()
            # Night shifts
            night_shifts = Shift.objects.filter(date=day, unit=unit, shift_type__name__in=night_shift_names)
            sscw_night = night_shifts.filter(user__role__name='SSCWN').count()
            care_night = night_shifts.filter(user__role__name__in=['SCWN', 'SCAN']).count()
            row[f'{unit.name}_day_sscw'] = sscw_day
            row[f'{unit.name}_day_care'] = care_day
            row[f'{unit.name}_night_sscw'] = sscw_night
            row[f'{unit.name}_night_care'] = care_night
        coverage_table.append(row)
    """Dashboard for Operations Manager, Service Manager, and Admin (Screen 1.1)"""
    
    # Check if user has management permissions
    if not (request.user.role and request.user.role.is_management):
        return redirect('staff_dashboard')
    
    # Widget A: Manual Review Required
    manual_review_qs = LeaveRequest.objects.filter(
        status='MANUAL_REVIEW'
    )
    if user_home:
        manual_review_qs = manual_review_qs.filter(user__unit__care_home=user_home)
    manual_review_requests = manual_review_qs.select_related('user').order_by('created_at')

    pending_leave_qs = LeaveRequest.objects.filter(
        status='PENDING'
    )
    if user_home:
        pending_leave_qs = pending_leave_qs.filter(user__unit__care_home=user_home)
    pending_leave_requests = pending_leave_qs.select_related('user').order_by('created_at')
    
    # Widget B: Staff Reallocations Needed
    reallocations_qs = StaffReallocation.objects.filter(
        status='NEEDED'
    )
    if user_home:
        reallocations_qs = reallocations_qs.filter(target_unit__care_home=user_home)
    pending_reallocations = reallocations_qs.select_related('target_unit', 'target_shift_type').order_by('target_date')
    
    # Widget C: Today's Staffing Snapshot
    today = timezone.now().date()
    today_shifts = Shift.objects.filter(
        date=today,
        status__in=['SCHEDULED', 'CONFIRMED']
    ).select_related('unit', 'shift_type', 'user')
    
    # Calculate staffing levels by unit
    day_shift_names = ['DAY', 'DAY_SENIOR', 'DAY_ASSISTANT']
    night_shift_names = ['NIGHT', 'NIGHT_SENIOR', 'NIGHT_ASSISTANT']

    role_labels = dict(Role.ROLE_CHOICES)
    staffing_summary = {}

    def format_breakdown(counts, include_roles=None):
        if include_roles is not None:
            counts = {role: count for role, count in counts.items() if role in include_roles}

        ordered_roles = ['SSCW', 'SCW', 'SCA', 'SSCWN', 'SCWN', 'SCAN']
        added = set()
        breakdown = []

        for role_code in ordered_roles:
            count = counts.get(role_code, 0)
            if count:
                breakdown.append({
                    'code': role_code,
                    'label': role_labels.get(role_code, role_code.replace('_', ' ').title()),
                    'count': count,
                })
                added.add(role_code)

        for role_code, count in counts.items():
            if role_code in added or count == 0:
                continue
            breakdown.append({
                'code': role_code,
                'label': role_labels.get(role_code, role_code.replace('_', ' ').title()),
                'count': count,
            })

        return breakdown

    day_care_roles = {'SCW', 'SCA'}
    night_care_roles = {'SCWN', 'SCAN'}

    # Aggregate staffing by care home instead of by unit
    home_staffing_summary = {}
    
    # Get all active care homes
    from .models_multi_home import CareHome
    care_homes_qs = CareHome.objects.filter(is_active=True)
    if user_home:
        care_homes_qs = care_homes_qs.filter(id=user_home.id)
    
    for home in care_homes_qs:
        # Get all units for this home
        home_units = Unit.objects.filter(care_home=home, is_active=True)
        
        # Get all shifts for this home today
        home_shifts_today = today_shifts.filter(unit__care_home=home)
        day_shifts_qs = home_shifts_today.filter(shift_type__name__in=day_shift_names)
        night_shifts_qs = home_shifts_today.filter(shift_type__name__in=night_shift_names)

        # Count by role for day shifts
        day_role_counts = {}
        for shift in day_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if not role_code:
                continue
            day_role_counts[role_code] = day_role_counts.get(role_code, 0) + 1

        # Count by role for night shifts
        night_role_counts = {}
        for shift in night_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if not role_code:
                continue
            night_role_counts[role_code] = night_role_counts.get(role_code, 0) + 1

        # Calculate totals
        day_sscw = day_role_counts.get('SSCW', 0)
        day_staff_total = day_shifts_qs.count()
        night_sscwn = night_role_counts.get('SSCWN', 0)
        night_staff_total = night_shifts_qs.count()
        
        home_staffing_summary[home.name] = {
            'home': home,
            'day_sscw': day_sscw,
            'day_staff_total': day_staff_total,
            'night_sscwn': night_sscwn,
            'night_staff_total': night_staff_total,
        }
    
    # Keep old unit-based summary for compatibility (can be removed later)
    staffing_summary = {}
    
    # Filter units by care home if applicable
    units_for_summary_qs = Unit.objects.filter(is_active=True)
    if user_home:
        units_for_summary_qs = units_for_summary_qs.filter(care_home=user_home)
    
    for unit in units_for_summary_qs:
        unit_shifts_today = today_shifts.filter(unit=unit)
        day_shifts_qs = unit_shifts_today.filter(shift_type__name__in=day_shift_names)
        night_shifts_qs = unit_shifts_today.filter(shift_type__name__in=night_shift_names)

        day_role_counts = {}
        for shift in day_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if not role_code:
                continue
            day_role_counts[role_code] = day_role_counts.get(role_code, 0) + 1

        night_role_counts = {}
        for shift in night_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if not role_code:
                continue
            night_role_counts[role_code] = night_role_counts.get(role_code, 0) + 1

        day_care_total = sum(count for role, count in day_role_counts.items() if role in day_care_roles)
        night_care_total = sum(count for role, count in night_role_counts.items() if role in night_care_roles)

        day_supernumerary_total = day_role_counts.get('SSCW', 0)
        night_supernumerary_total = night_role_counts.get('SSCWN', 0)
        
        staffing_summary[unit.name] = {
            'unit': unit,
            'day_actual': day_care_total,
            'day_required': unit.min_day_staff,
            'night_actual': night_care_total,
            'night_required': unit.min_night_staff,
            'day_status': 'good' if day_care_total >= unit.min_day_staff else 'shortage',
            'night_status': 'good' if night_care_total >= unit.min_night_staff else 'shortage',
            'day_breakdown': format_breakdown(day_role_counts),
            'night_breakdown': format_breakdown(night_role_counts),
            'day_care_breakdown': format_breakdown(day_role_counts, include_roles=day_care_roles),
            'night_care_breakdown': format_breakdown(night_role_counts, include_roles=night_care_roles),
            'day_supernumerary': day_supernumerary_total,
            'night_supernumerary': night_supernumerary_total,
        }
    
    # Widget D: Recent Automated Approvals
    auto_approvals_qs = ActivityLog.objects.filter(
        action_type='AUTO_APPROVAL',
        automated=True
    )
    if user_home:
        auto_approvals_qs = auto_approvals_qs.filter(user__unit__care_home=user_home)
    recent_auto_approvals = auto_approvals_qs.select_related('user').order_by('-created_at')[:10]

    # Widget E: Current sickness absences
    sickness_qs = SicknessRecord.objects.filter(
        status__in=['OPEN', 'AWAITING_FIT_NOTE'],
        first_working_day__lte=today,
    ).filter(
        models.Q(actual_last_working_day__isnull=True) |
        models.Q(actual_last_working_day__gte=today)
    )
    if user_home:
        sickness_qs = sickness_qs.filter(profile__user__unit__care_home=user_home)
    current_sickness_absences = sickness_qs.select_related('profile__user').order_by('first_working_day')

    approved_leave_qs = LeaveRequest.objects.filter(
        status='APPROVED',
        end_date__gte=today,
    )
    if user_home:
        approved_leave_qs = approved_leave_qs.filter(user__unit__care_home=user_home)
    upcoming_approved_leave = approved_leave_qs.select_related('user').order_by('start_date')
    
    # Calculate staffing alerts for the next 7 days with home-specific thresholds
    home_staffing_config = {
        'HAWTHORN_HOUSE': {'day_ideal': 18, 'night_ideal': 18},
        'MEADOWBURN': {'day_ideal': 17, 'night_ideal': 17},
        'ORCHARD_GROVE': {'day_ideal': 17, 'night_ideal': 17},
        'RIVERSIDE': {'day_ideal': 17, 'night_ideal': 17},
        'VICTORIA_GARDENS': {'day_ideal': 10, 'night_ideal': 10},
    }
    
    # Determine threshold based on selected home
    if user_home:
        home_config = home_staffing_config.get(user_home.name, {'day_ideal': 17, 'night_ideal': 17})
        day_threshold = home_config['day_ideal']
        night_threshold = home_config['night_ideal']
    else:
        # Default threshold when viewing all homes
        day_threshold = 17
        night_threshold = 17
    
    staffing_alerts = []
    for i in range(7):
        check_date = today + timedelta(days=i)
        date_shifts_qs = Shift.objects.filter(
            date=check_date,
            status__in=['SCHEDULED', 'CONFIRMED']
        )
        if user_home:
            date_shifts_qs = date_shifts_qs.filter(unit__care_home=user_home)
        date_shifts = date_shifts_qs.select_related('user', 'user__role', 'shift_type')
        
        # Count day and night care staff
        day_care = date_shifts.filter(
            shift_type__name__in=day_shift_names,
            user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
        ).count()
        
        night_care = date_shifts.filter(
            shift_type__name__in=night_shift_names,
            user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
        ).count()
        
        # Check against home-specific thresholds
        if day_care < day_threshold:
            staffing_alerts.append({
                'date': check_date,
                'shift': 'Day',
                'actual': day_care,
                'required': day_threshold,
                'deficit': day_threshold - day_care,
            })
        
        if night_care < night_threshold:
            staffing_alerts.append({
                'date': check_date,
                'shift': 'Night',
                'actual': night_care,
                'required': night_threshold,
                'deficit': night_threshold - night_care,
            })
    
    # Recent incidents (last 24 hours) requiring review
    twenty_four_hours_ago = timezone.now() - timedelta(hours=24)
    incidents_qs = IncidentReport.objects.filter(
        created_at__gte=twenty_four_hours_ago
    )
    if user_home:
        incidents_qs = incidents_qs.filter(reported_by__unit__care_home=user_home)
    recent_incidents = incidents_qs.select_related('reported_by').order_by('-created_at')
    
    # Training compliance summary
    from .models import TrainingCourse, TrainingRecord
    training_compliance_summary = None
    
    mandatory_courses = TrainingCourse.objects.filter(is_mandatory=True)
    if mandatory_courses.exists() and user_home:
        # Get staff for user's home
        home_staff = User.objects.filter(
            unit__care_home=user_home,
            is_active=True
        ).distinct()
        
        total_staff = home_staff.count()
        total_required = total_staff * mandatory_courses.count()
        total_compliant = 0
        total_expiring = 0
        total_expired = 0
        
        if total_staff > 0:
            for staff in home_staff:
                for course in mandatory_courses:
                    latest_record = TrainingRecord.objects.filter(
                        staff_member=staff,
                        course=course
                    ).order_by('-completion_date').first()
                    
                    if latest_record:
                        status = latest_record.get_status()
                        if status == 'CURRENT':
                            total_compliant += 1
                        elif status == 'EXPIRING_SOON':
                            total_expiring += 1
                        elif status == 'EXPIRED':
                            total_expired += 1
            
            compliance_percentage = (total_compliant / total_required * 100) if total_required > 0 else 0
            
            training_compliance_summary = {
                'total_staff': total_staff,
                'total_required': total_required,
                'compliant': total_compliant,
                'expiring': total_expiring,
                'expired': total_expired,
                'missing': total_required - (total_compliant + total_expiring + total_expired),
                'percentage': round(compliance_percentage, 1),
            }
    
    context = {
        'manual_review_requests': manual_review_requests,
        'pending_reallocations': pending_reallocations,
        'staffing_summary': staffing_summary,
        'home_staffing_summary': home_staffing_summary,
        'recent_auto_approvals': recent_auto_approvals,
        'today': today,
        'current_sickness_absences': current_sickness_absences,
        'pending_leave_requests': pending_leave_requests,
        'upcoming_approved_leave': upcoming_approved_leave,
        'staffing_alerts': staffing_alerts,
        'coverage_table': coverage_table,
        'units': units,
        'rota_start': rota_start,
        'rota_end': rota_end,
        'recent_incidents': recent_incidents,
        'training_compliance_summary': training_compliance_summary,
        'user_home': user_home,
        'can_select_home': can_select_home,
        'all_care_homes': CareHome.objects.all().order_by('name'),
        'selected_home': user_home.name if user_home else None,
    }
    
    return render(request, 'scheduling/manager_dashboard.html', context)


def _should_auto_approve(leave_request):
    """
    Decide whether a leave request qualifies for automatic approval.
    
    Auto-approval rules:
    1. Must be ANNUAL, PERSONAL, or TRAINING leave
    2. Must not be longer than 14 days (2 weeks)
    3. Must not be in blackout period (2 weeks before/after Christmas: Dec 11 - Jan 8)
    4. Must not cause staffing shortfall (max 2 staff off per day)
    5. Must not cause minimum staffing to drop below 17
    """
    
    # Rule 1: Only auto-approve specific leave types
    auto_types = {'ANNUAL', 'PERSONAL', 'TRAINING'}
    if leave_request.leave_type not in auto_types:
        return False

    # Rule 2: Requests over 14 days require Operations Manager approval
    if leave_request.days_requested > 14:
        leave_request.status = 'MANUAL_REVIEW'
        leave_request.approval_notes = 'Request exceeds 14 days - requires Operations Manager approval'
        leave_request.save()
        return False

    # Rule 3: Check Christmas blackout period (Dec 11 - Jan 8)
    christmas_date = datetime(leave_request.start_date.year, 12, 25).date()
    blackout_start = christmas_date - timedelta(days=14)  # Dec 11
    blackout_end = christmas_date + timedelta(days=14)     # Jan 8
    
    # Adjust for requests spanning year boundary
    if leave_request.end_date.month == 1 and leave_request.start_date.month == 12:
        christmas_date_next_year = datetime(leave_request.end_date.year, 12, 25).date()
        blackout_end = christmas_date_next_year + timedelta(days=14)
    
    # Check if request overlaps with blackout period
    if (leave_request.start_date <= blackout_end and leave_request.end_date >= blackout_start):
        leave_request.is_blackout_period = True
        leave_request.status = 'MANUAL_REVIEW'
        leave_request.approval_notes = 'Request during Christmas period (Dec 11 - Jan 8) - requires management review for fairness'
        leave_request.save()
        return False
    
    # Rule 4 & 5: Check staffing levels for each day of the leave request
    current_date = leave_request.start_date
    while current_date <= leave_request.end_date:
        # Count staff already off on this date (approved or pending)
        staff_off = LeaveRequest.objects.filter(
            status__in=['APPROVED', 'PENDING'],
            start_date__lte=current_date,
            end_date__gte=current_date
        ).exclude(id=leave_request.id).count()
        
        # Rule 4: Max 2 staff off per day for auto-approval
        if staff_off >= 2:
            leave_request.causes_staffing_shortfall = True
            leave_request.status = 'MANUAL_REVIEW'
            leave_request.approval_notes = f'More than 2 staff already off on {current_date} - requires management review'
            leave_request.save()
            return False
        
        # Rule 5: Check if minimum staffing (17) would be maintained
        # Get scheduled shifts for this date
        day_shifts = Shift.objects.filter(
            date=current_date,
            shift_type__name__in=['DAY_SENIOR', 'DAY_ASSISTANT'],
            user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
        ).count()
        
        night_shifts = Shift.objects.filter(
            date=current_date,
            shift_type__name__in=['NIGHT_SENIOR', 'NIGHT_ASSISTANT'],
            user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
        ).count()
        
        # Account for this leave request and others already approved/pending
        staff_off_this_day = staff_off + 1  # +1 for this request
        
        if day_shifts - staff_off_this_day < 17 or night_shifts - staff_off_this_day < 17:
            leave_request.causes_staffing_shortfall = True
            leave_request.status = 'MANUAL_REVIEW'
            leave_request.approval_notes = f'Would cause staffing to drop below 17 on {current_date} - requires management review'
            leave_request.save()
            return False
        
        current_date += timedelta(days=1)
    
    # All checks passed - eligible for auto-approval
    return True


@login_required
def staff_guidance(request):
    """Display operational guidance and checklists for management users."""
    import markdown

    if not (request.user.role and request.user.role.is_management):
        return redirect('staff_dashboard')

    # BASE_DIR is /path/to/rotasystems, docs is at /path/to/docs
    docs_path = Path(settings.BASE_DIR).parent / 'docs' / 'staff_guidance'

    guidance_docs = [
        {
            'title': 'Staff FAQ - Frequently Asked Questions',
            'slug': 'staff-faq',
            'description': 'Comprehensive FAQ covering annual leave, sickness, shifts, and training.',
            'file': docs_path / 'STAFF_FAQ.md',
            'category': 'staff',
            'icon': 'fa-question-circle',
        },
        {
            'title': 'New Starter Guide',
            'slug': 'new-starter-guide',
            'description': 'Complete onboarding guide for new staff members.',
            'file': docs_path / 'NEW_STARTER_GUIDE.md',
            'category': 'staff',
            'icon': 'fa-user-plus',
        },
        {
            'title': 'Annual Leave - Complete Guide',
            'slug': 'annual-leave-guide',
            'description': 'Everything about requesting and managing annual leave.',
            'file': docs_path / 'ANNUAL_LEAVE_GUIDE.md',
            'category': 'staff',
            'icon': 'fa-calendar-check',
        },
        {
            'title': 'Sickness Reporting & Management',
            'slug': 'sickness-reporting-guide',
            'description': 'Complete guide to reporting sickness and absence management.',
            'file': docs_path / 'SICKNESS_REPORTING_GUIDE.md',
            'category': 'staff',
            'icon': 'fa-notes-medical',
        },
        {
            'title': 'Supporting Attendance Policy Guide',
            'slug': 'supporting-attendance-policy',
            'description': 'Official attendance policy explained in staff-friendly language - your rights, responsibilities, and support available.',
            'file': docs_path / 'SUPPORTING_ATTENDANCE_POLICY.md',
            'category': 'staff',
            'icon': 'fa-file-medical',
        },
        {
            'title': 'Manager Resources - Complete Index',
            'slug': 'manager-resources-index',
            'description': '11 comprehensive guides for managers: attendance management, OH referrals, reasonable adjustments, checklists, and more.',
            'file': docs_path / 'MANAGER_RESOURCES_INDEX.md',
            'category': 'manager',
            'icon': 'fa-users-cog',
        },
        {
            'title': 'Manager Telephone Checklist',
            'slug': 'manager-telephone-checklist',
            'description': 'Questions to cover when speaking to staff who have called in sick.',
            'file': docs_path / 'manager_telephone_checklist.md',
            'category': 'manager',
            'icon': 'fa-phone',
        },
        # === NEW COMPREHENSIVE MANAGER GUIDES ===
        {
            'title': 'Supporting Attendance - Managers Guide',
            'slug': 'managers-attendance-guide',
            'description': 'Comprehensive 42KB foundational guide: absence management, wellbeing, RTW, capability, ill-health retirement.',
            'file': docs_path / 'MANAGERS_ATTENDANCE_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-clipboard-check',
        },
        {
            'title': 'Checklist: Employee Reports Absent',
            'slug': 'checklist-employee-reports-absent',
            'description': 'Day 1 absence call checklist - essential questions, red flags, documentation.',
            'file': docs_path / 'CHECKLIST_EMPLOYEE_REPORTS_ABSENT.md',
            'category': 'manager',
            'icon': 'fa-phone-square',
        },
        {
            'title': 'Checklist: Return to Work Discussion',
            'slug': 'checklist-return-to-work',
            'description': 'Return to work interview checklist - before meeting, during meeting, adjustments, scenarios.',
            'file': docs_path / 'CHECKLIST_RETURN_TO_WORK.md',
            'category': 'manager',
            'icon': 'fa-clipboard-user',
        },
        {
            'title': 'Checklist: Attendance Review Meeting',
            'slug': 'checklist-attendance-review',
            'description': 'Formal attendance review when triggers breached - preparation, meeting structure, outcomes, scenarios.',
            'file': docs_path / 'CHECKLIST_ATTENDANCE_REVIEW.md',
            'category': 'manager',
            'icon': 'fa-file-circle-check',
        },
        {
            'title': 'OH Referrals & OHIO System Guide',
            'slug': 'managers-oh-referral-guide',
            'description': 'When/how to refer to Occupational Health, complete OHIO online system guide, understanding reports.',
            'file': docs_path / 'MANAGERS_OH_REFERRAL_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-stethoscope',
        },
        {
            'title': 'Reasonable Adjustments Guide',
            'slug': 'managers-reasonable-adjustments-guide',
            'description': '8 types of reasonable adjustments, implementation process, Equality Act 2010 compliance.',
            'file': docs_path / 'MANAGERS_REASONABLE_ADJUSTMENTS_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-hands-helping',
        },
        {
            'title': 'Disability & Mental Health Guide',
            'slug': 'managers-disability-mh-guide',
            'description': 'Equality Act 2010, disability definition, mental health support, stress management, 18 support organizations.',
            'file': docs_path / 'MANAGERS_DISABILITY_MH_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-heart',
        },
        {
            'title': 'Menopause Support Guide',
            'slug': 'managers-menopause-guide',
            'description': 'Understanding menopause, symptoms, workplace impact, reasonable adjustments, supportive conversations.',
            'file': docs_path / 'MANAGERS_MENOPAUSE_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-venus',
        },
        {
            'title': 'Absence Interview Guide',
            'slug': 'managers-absence-interview-guide',
            'description': 'Conducting long-term absence interviews - preparation, key discussions, difficult conversations, scenarios.',
            'file': docs_path / 'MANAGERS_ABSENCE_INTERVIEW_GUIDE.md',
            'category': 'manager',
            'icon': 'fa-comments',
        },
        {
            'title': 'Automated Weekly Reports',
            'slug': 'automated-weekly-reports',
            'description': 'Complete guide to automated Monday management reports - schedules, contents, and troubleshooting',
            'file': docs_path / 'AUTOMATED_WEEKLY_REPORTS.md',
            'category': 'manager',
            'icon': 'fa-calendar-alt'
        },
    ]

    selected_slug = request.GET.get('doc') or guidance_docs[0]['slug']
    selected_doc = next((doc for doc in guidance_docs if doc['slug'] == selected_slug), guidance_docs[0])

    try:
        with selected_doc['file'].open('r', encoding='utf-8') as handle:
            raw_text = handle.read()
            # Convert markdown to HTML
            md = markdown.Markdown(extensions=['tables', 'fenced_code', 'nl2br'])
            html_content = md.convert(raw_text)
            selected_doc['content'] = mark_safe(html_content)
    except FileNotFoundError:
        selected_doc['content'] = mark_safe('<div class="alert alert-warning"><i class="fas fa-exclamation-triangle"></i> Document file not found. Please check the repository.</div>')
    except Exception as e:
        selected_doc['content'] = mark_safe(f'<div class="alert alert-danger"><i class="fas fa-times-circle"></i> Error loading document: {escape(str(e))}</div>')

    context = {
        'guidance_docs': guidance_docs,
        'selected_doc': selected_doc,
    }

    return render(request, 'scheduling/staff_guidance.html', context)

@login_required 
def rota_view(request):
    """Main Rota View (Screen 1.2)"""
    
    # Check permissions
    if not (request.user.role and request.user.role.can_manage_rota):
        messages.error(request, 'You do not have permission to view the rota.')
        return redirect('staff_dashboard')
    
    # Get date range (default to current week)
    today = timezone.now().date()
    start_of_week = today - timedelta(days=(today.weekday() + 1) % 7)
    
    # Handle date navigation
    week_offset = int(request.GET.get('week_offset', 0))
    view_start_date = start_of_week + timedelta(weeks=week_offset)
    view_end_date = view_start_date + timedelta(days=6)
    
    # Handle care home filter
    selected_home = request.GET.get('care_home', 'all')
    from .models_multi_home import CareHome
    care_homes = CareHome.objects.all().order_by('name')
    
    # Handle unit filter
    selected_unit = request.GET.get('unit', 'all')
    units = Unit.objects.filter(is_active=True)
    
    # Filter units by care home if selected
    if selected_home != 'all':
        units = units.filter(care_home__name=selected_home)
    
    # Get selected home label
    if selected_home == 'all':
        selected_home_label = 'All Homes'
    else:
        home_obj = care_homes.filter(name=selected_home).first()
        selected_home_label = home_obj.get_name_display() if home_obj else selected_home

    # Handle team filter
    team_choices = [{'code': code, 'label': label} for code, label in User.TEAM_CHOICES]
    team_label_lookup = {choice['code']: choice['label'] for choice in team_choices}
    valid_team_codes = set(team_label_lookup.keys())
    team_param = request.GET.get('team', 'all') or 'all'
    selected_team = team_param.strip().upper() if team_param else 'ALL'
    if selected_team != 'ALL' and selected_team not in valid_team_codes:
        selected_team = 'all'
    elif selected_team == 'ALL':
        selected_team = 'all'
    
    # Get shifts for the week
    shifts_query = Shift.objects.filter(
        date__range=[view_start_date, view_end_date]
    ).select_related('user', 'unit', 'shift_type').order_by('date', 'shift_type__start_time')
    
    # Filter by care home first (affects both unit and non-unit filters)
    if selected_home != 'all':
        shifts_query = shifts_query.filter(unit__care_home__name=selected_home)
    
    if selected_unit != 'all':
        shifts_query = shifts_query.filter(unit__name=selected_unit)

    if selected_team != 'all':
        shifts_query = shifts_query.filter(user__team=selected_team)
    
    leave_query = LeaveRequest.objects.filter(
        status__in=['APPROVED', 'PENDING'],
        start_date__lte=view_end_date,
        end_date__gte=view_start_date,
    ).select_related('user', 'user__role', 'user__unit')
    
    # Filter leave by care home
    if selected_home != 'all':
        leave_query = leave_query.filter(user__unit__care_home__name=selected_home)

    if selected_team != 'all':
        leave_query = leave_query.filter(user__team=selected_team)

    if selected_unit == 'all':
        selected_unit_label = 'All Units'
    else:
        unit_obj = units.filter(name=selected_unit).first()
        selected_unit_label = unit_obj.get_name_display() if unit_obj else selected_unit

    selected_team_label = team_label_lookup.get(selected_team, 'All Teams') if selected_team != 'all' else 'All Teams'

    leave_by_date = defaultdict(list)

    for leave in leave_query:
        unit_name = getattr(getattr(leave.user, 'unit', None), 'name', None)
        if selected_unit != 'all' and unit_name != selected_unit:
            continue

        span_start = max(leave.start_date, view_start_date)
        span_end = min(leave.end_date, view_end_date)

        total_days = (span_end - span_start).days + 1
        for offset in range(total_days):
            current = span_start + timedelta(days=offset)
            leave_by_date[current].append(leave)

    # Organize shifts by date and user
    shifts_by_date = {}
    all_shifts_list = list(shifts_query)
    
    care_roles = {'SCW', 'SCA', 'SCWN', 'SCAN'}
    
    # Get all active units for consistent ordering across all days
    # If filtering by MGMT unit, exclude care units from display to keep it clean
    if selected_unit != 'all' and 'MGMT' in selected_unit:
        # When viewing MGMT unit, only show MGMT unit (not care units)
        all_unit_names = [selected_unit]
    else:
        all_unit_names = sorted([unit.name for unit in units if unit.is_active])

    for i in range(7):
        current_date = view_start_date + timedelta(days=i)
        date_shifts = [shift for shift in all_shifts_list if shift.date == current_date]

        day_shifts = [
            s for s in date_shifts if s.shift_type.name in ['DAY_SENIOR', 'DAY_ASSISTANT', 'ADMIN']
        ]
        night_shifts = [s for s in date_shifts if s.shift_type.name in ['NIGHT_SENIOR', 'NIGHT_ASSISTANT']]

        # Management staff (SM, OM) - always at top with pink background
        day_management = [
            s for s in day_shifts
            if getattr(getattr(s.user, 'role', None), 'name', '') in ['SM', 'OM']
        ]
        night_management = [
            s for s in night_shifts
            if getattr(getattr(s.user, 'role', None), 'name', '') in ['SM', 'OM']
        ]

        # Supernumerary duty staff (SSCW, SSCWN)
        day_supernumerary_duty = [
            s for s in day_shifts
            if getattr(getattr(s.user, 'role', None), 'name', '') in ['SSCW'] and s.shift_type.name != 'ADMIN' and s not in day_management
        ]
        day_supernumerary_admin = [
            s for s in date_shifts
            if s.shift_type.name == 'ADMIN' and getattr(getattr(s.user, 'role', None), 'name', '') in ['SSCW'] and s not in day_management
        ]
        day_supernumerary = day_supernumerary_duty + day_supernumerary_admin
        night_supernumerary = [
            s for s in night_shifts 
            if getattr(getattr(s.user, 'role', None), 'name', '') in ['SSCWN'] and s not in night_management
        ]

        day_care = [
            s for s in day_shifts
            if getattr(getattr(s.user, 'role', None), 'name', '') in care_roles and s.shift_type.name != 'ADMIN'
        ]
        night_care = [s for s in night_shifts if getattr(getattr(s.user, 'role', None), 'name', '') in care_roles]

        # Group care shifts by unit
        day_care_by_unit = defaultdict(list)
        for shift in day_care:
            unit_name = shift.unit.name if shift.unit else 'Unknown'
            day_care_by_unit[unit_name].append(shift)
        
        night_care_by_unit = defaultdict(list)
        for shift in night_care:
            unit_name = shift.unit.name if shift.unit else 'Unknown'
            night_care_by_unit[unit_name].append(shift)
        
        # Sort units alphabetically and sort staff within each unit by role then name
        for unit_name in day_care_by_unit:
            day_care_by_unit[unit_name].sort(key=lambda s: (
                s.user.role.name if s.user and s.user.role else '',
                s.user.last_name if s.user else '',
                s.user.first_name if s.user else ''
            ))
        
        for unit_name in night_care_by_unit:
            night_care_by_unit[unit_name].sort(key=lambda s: (
                s.user.role.name if s.user and s.user.role else '',
                s.user.last_name if s.user else '',
                s.user.first_name if s.user else ''
            ))

        day_other = [s for s in day_shifts if s not in day_supernumerary and s not in day_care and s not in day_management]
        night_other = [s for s in night_shifts if s not in night_supernumerary and s not in night_care and s not in night_management]

        # Ensure all units appear in the dictionary, even if empty (for horizontal alignment)
        day_care_by_unit_sorted = {}
        for unit_name in all_unit_names:
            if unit_name in day_care_by_unit:
                day_care_by_unit_sorted[unit_name] = day_care_by_unit[unit_name]
            else:
                day_care_by_unit_sorted[unit_name] = []
        
        night_care_by_unit_sorted = {}
        for unit_name in all_unit_names:
            if unit_name in night_care_by_unit:
                night_care_by_unit_sorted[unit_name] = night_care_by_unit[unit_name]
            else:
                night_care_by_unit_sorted[unit_name] = []

        shifts_by_date[current_date] = {
            'day': day_shifts,
            'night': night_shifts,
            'all': date_shifts,
            'day_management': day_management,
            'night_management': night_management,
            'day_supernumerary': day_supernumerary,
            'day_supernumerary_duty': day_supernumerary_duty,
            'day_supernumerary_admin': day_supernumerary_admin,
            'night_supernumerary': night_supernumerary,
            'day_care': day_care,
            'night_care': night_care,
            'day_care_by_unit': day_care_by_unit_sorted,
            'night_care_by_unit': night_care_by_unit_sorted,
            'day_other': day_other,
            'night_other': night_other,
            'leave': leave_by_date.get(current_date, []),
        }

    # Serialize shifts for front-end print preview consumption
    def serialize_shift(shift):
        role_name = getattr(getattr(shift.user, 'role', None), 'name', '')
        shift_type_name = shift.shift_type.name if shift.shift_type else ''
        is_sscw_admin = role_name == 'SSCW' and shift_type_name == 'ADMIN'

        return {
            'user': getattr(shift.user, 'full_name', ''),
            'role': role_name,
            'unit': shift.unit.get_name_display() if shift.unit else '',
            'unit_code': shift.unit.name if shift.unit else '',
            'start_time': (shift.shift_type.start_time.strftime('%H:%M') if shift.shift_type and shift.shift_type.start_time else ''),
            'end_time': (shift.shift_type.end_time.strftime('%H:%M') if shift.shift_type and shift.shift_type.end_time else ''),
            'shift_type': shift_type_name,
            'category': 'SSCW_ADMIN' if is_sscw_admin else '',
            'team': getattr(shift.user, 'team', ''),
        }

    shifts_serialized = {}
    for current_date, shifts_data in shifts_by_date.items():
        shifts_serialized[current_date.isoformat()] = {
            'day': [serialize_shift(shift) for shift in shifts_data['day']],
            'night': [serialize_shift(shift) for shift in shifts_data['night']],
        }
    
    # Calculate daily staffing summary
    daily_summary = {}
    staffing_alerts = []  # Track staffing shortages for dashboard
    
    # Home-specific staffing thresholds (care staff only, not including SSCW/SSCWN)
    home_staffing_config = {
        'HAWTHORN_HOUSE': {'day_ideal': 18, 'night_ideal': 18, 'day_sscw': 2, 'night_sscw': 2},
        'MEADOWBURN': {'day_ideal': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
        'ORCHARD_GROVE': {'day_ideal': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
        'RIVERSIDE': {'day_ideal': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2},
        'VICTORIA_GARDENS': {'day_ideal': 10, 'night_ideal': 10, 'day_sscw': 1, 'night_sscw': 1},
    }
    
    for current_date, shifts_data in shifts_by_date.items():
        # Break down by role groups
        day_sscw = len(shifts_data['day_supernumerary'])
        day_care = len(shifts_data['day_care'])
        night_sscw = len(shifts_data['night_supernumerary'])
        night_care = len(shifts_data['night_care'])

        # Determine staffing requirements based on selected filters
        if selected_home != 'all':
            # Use home-specific thresholds
            requirements = home_staffing_config.get(selected_home, {'day_ideal': 17, 'night_ideal': 17, 'day_sscw': 2, 'night_sscw': 2})
            required_day = requirements['day_ideal']
            required_night = requirements['night_ideal']
            required_day_sscw = requirements['day_sscw']
            required_night_sscw = requirements['night_sscw']
        elif selected_unit != 'all':
            # For individual units, use their specific requirements
            unit = units.get(name=selected_unit)
            required_day = unit.min_day_staff if unit else 17
            required_night = unit.min_night_staff if unit else 17
            required_day_sscw = 2
            required_night_sscw = 2
        else:
            # For facility-wide view (all homes) - sum up all home requirements
            required_day = sum(config['day_ideal'] for config in home_staffing_config.values())
            required_night = sum(config['night_ideal'] for config in home_staffing_config.values())
            required_day_sscw = sum(config['day_sscw'] for config in home_staffing_config.values())
            required_night_sscw = sum(config['night_sscw'] for config in home_staffing_config.values())

        # Check for staffing shortages
        day_shortage = day_care < required_day
        night_shortage = night_care < required_night
        
        # Create alerts for shortages
        if day_shortage:
            staffing_alerts.append({
                'date': current_date,
                'shift': 'Day',
                'actual': day_care,
                'required': required_day,
                'deficit': required_day - day_care,
            })
        if night_shortage:
            staffing_alerts.append({
                'date': current_date,
                'shift': 'Night',
                'actual': night_care,
                'required': required_night,
                'deficit': required_night - night_care,
            })

        daily_summary[current_date] = {
            'day_sscw': day_sscw,
            'day_sscw_required': required_day_sscw,
            'day_care': day_care,
            'day_required': required_day,
            'night_sscw': night_sscw,
            'night_sscw_required': required_night_sscw,
            'night_care': night_care,
            'night_required': required_night,
            'day_status': 'good' if not day_shortage else 'shortage',
            'night_status': 'good' if not night_shortage else 'shortage',
        }
    
    context = {
        'shifts_by_date': shifts_by_date,
        'daily_summary': daily_summary,
        'staffing_alerts': staffing_alerts,
        'view_start_date': view_start_date,
        'view_end_date': view_end_date,
        'week_offset': week_offset,
        'care_homes': care_homes,
        'selected_home': selected_home,
        'selected_home_label': selected_home_label,
        'selected_unit': selected_unit,
        'selected_unit_label': selected_unit_label,
        'units': units,
        'team_choices': team_choices,
        'selected_team': selected_team,
        'selected_team_label': selected_team_label,
        'date_range': [view_start_date + timedelta(days=i) for i in range(7)],
        'roles': Role.objects.all(),
        'today': timezone.now().date(),
        'shifts_serialized': shifts_serialized,
    }
    
    return render(request, 'scheduling/rota_view.html', context)

@login_required
def reports_dashboard(request):
    """Reports Dashboard (Screen 1.3)"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        return redirect('staff_dashboard')
    
    # Calculate some basic statistics for the dashboard
    today = timezone.now().date()
    
    # Total active staff
    total_staff = User.objects.filter(is_active=True, role__is_management=False).count()
    
    # Staff on duty today
    on_duty_today = Shift.objects.filter(
        date=today,
        status__in=['SCHEDULED', 'CONFIRMED']
    ).count()
    
    # Pending leave requests
    pending_requests = LeaveRequest.objects.filter(
        status__in=['PENDING', 'MANUAL_REVIEW']
    ).count()
    
    # Leave days taken this month
    first_of_month = today.replace(day=1)
    month_leave_days = LeaveRequest.objects.filter(
        status='APPROVED',
        start_date__gte=first_of_month,
        start_date__lte=today
    ).aggregate(total_days=models.Sum('days_requested'))['total_days'] or 0
    
    # Get all units for filters
    from .models_multi_home import CareHome
    units = Unit.objects.filter(is_active=True).select_related('care_home').order_by('care_home__name', 'name')
    care_homes = CareHome.objects.filter(is_active=True).order_by('name')
    
    context = {
        'available_reports': [
            {
                'name': 'Overtime & Agency Usage', 
                'description': 'Comprehensive OT and agency breakdown by home, role, hours, and costs', 
                'url': 'ot_agency_report', 
                'icon': 'fa-chart-line',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
            {
                'name': 'Staff Vacancies', 
                'description': 'Current and upcoming staff vacancies by care home with leaving reasons', 
                'url': 'staff_vacancies_report', 
                'icon': 'fa-user-times',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
            {
                'name': 'Leave Usage Targets', 
                'description': '40-week strategy dashboard showing staff progress against leave targets', 
                'url': 'leave_usage_targets', 
                'icon': 'fa-bullseye',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
            {
                'name': 'Rota Cost Analysis', 
                'description': 'Financial report on wage costs by care home with trend charts and export', 
                'url': 'rota_cost_analysis', 
                'icon': 'fa-pound-sign',
                'export_pdf': 'export_cost_analysis_pdf',
                'export_excel': 'export_cost_analysis_excel',
                'export_csv': 'export_cost_analysis_csv'
            },
            {
                'name': 'Staff Sickness Report', 
                'description': 'Shows sickness rates and trends',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
            {
                'name': 'Annual Leave Report', 
                'description': 'Summary of leave allowances, taken, and remaining',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
            {
                'name': 'Staff Reallocation Report', 
                'description': 'Fairness report showing which staff are moved most often',
                'export_pdf': None,
                'export_excel': None,
                'export_csv': None
            },
        ],
        'total_staff': total_staff,
        'on_duty_today': on_duty_today,
        'pending_requests': pending_requests,
        'month_leave_days': month_leave_days,
        'units': units,
        'care_homes': care_homes,
    }
    
    return render(request, 'scheduling/reports_dashboard.html', context)

@login_required
def get_annual_leave_report(request):
    """Generate real-time annual leave report with comprehensive statistics"""
    from staff_records.models import AnnualLeaveEntitlement
    from decimal import Decimal
    from datetime import date
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get current leave year (2025)
    year = int(request.GET.get('year', date.today().year))
    
    # Get all entitlements for the specified year
    entitlements = AnnualLeaveEntitlement.objects.filter(
        leave_year_start__year=year
    ).select_related('profile__user__role').order_by('profile__user__last_name')
    
    if not entitlements.exists():
        return JsonResponse({'error': f'No leave entitlements found for {year}'}, status=404)
    
    # Calculate overall statistics (including carryover)
    total_staff = entitlements.count()
    total_entitlement_hours = sum(e.total_entitlement_hours for e in entitlements)
    total_carryover_hours = sum(e.carryover_hours for e in entitlements)
    total_available_hours = sum(e.total_available_hours for e in entitlements)  # includes carryover
    total_used_hours = sum(e.hours_used for e in entitlements)
    total_pending_hours = sum(e.hours_pending for e in entitlements)
    total_remaining_hours = sum(e.hours_remaining for e in entitlements)
    
    # Split by day/night based on role
    day_staff_entitlements = []
    night_staff_entitlements = []
    
    for ent in entitlements:
        role_name = ent.profile.user.role.name if ent.profile.user.role else ''
        is_night = 'NIGHT' in role_name.upper() or role_name.upper() in ['SCWN', 'SCAN', 'SSCWN']
        
        if is_night:
            night_staff_entitlements.append(ent)
        else:
            day_staff_entitlements.append(ent)
    
    # Calculate day shift totals
    day_total_hours = sum(e.total_entitlement_hours for e in day_staff_entitlements)
    day_used_hours = sum(e.hours_used for e in day_staff_entitlements)
    day_remaining_hours = sum(e.hours_remaining for e in day_staff_entitlements)
    
    # Calculate night shift totals
    night_total_hours = sum(e.total_entitlement_hours for e in night_staff_entitlements)
    night_used_hours = sum(e.hours_used for e in night_staff_entitlements)
    night_remaining_hours = sum(e.hours_remaining for e in night_staff_entitlements)
    
    # Calculate end of year date
    end_of_year = date(year, 12, 31)
    
    # Build staff list with details
    staff_list = []
    for ent in entitlements:
        role_name = ent.profile.user.role.name if ent.profile.user.role else 'N/A'
        is_night = 'NIGHT' in role_name.upper() or role_name.upper() in ['SCWN', 'SCAN', 'SSCWN']
        
        staff_list.append({
            'sap': ent.profile.user.sap,
            'name': ent.profile.user.full_name,
            'role': role_name,
            'shift_type': 'Night' if is_night else 'Day',
            'total_hours': float(ent.total_entitlement_hours),
            'total_days': float(ent.days_entitlement),
            'used_hours': float(ent.hours_used),
            'used_days': float(ent.days_used),
            'pending_hours': float(ent.hours_pending),
            'remaining_hours': float(ent.hours_remaining),
            'remaining_days': float(ent.days_remaining),
            'carryover_hours': float(ent.carryover_hours),
            'leave_year': f"{ent.leave_year_start} to {ent.leave_year_end}",
            'usage_percent': round((float(ent.hours_used) / float(ent.total_entitlement_hours) * 100) if ent.total_entitlement_hours > 0 else 0, 1)
        })
    
    # Build response
    response_data = {
        'year': year,
        'end_date': end_of_year.isoformat(),
        'summary': {
            'total_staff': total_staff,
            'total_entitlement_hours': round(float(total_entitlement_hours), 1),
            'total_carryover_hours': round(float(total_carryover_hours), 1),
            'total_available_hours': round(float(total_available_hours), 1),
            'total_used_hours': round(float(total_used_hours), 1),
            'total_pending_hours': round(float(total_pending_hours), 1),
            'total_remaining_hours': round(float(total_remaining_hours), 1),
            'overall_usage_percent': round((float(total_used_hours) / float(total_available_hours) * 100) if total_available_hours > 0 else 0, 1),
            'overall_remaining_percent': round((float(total_remaining_hours) / float(total_available_hours) * 100) if total_available_hours > 0 else 0, 1)
        },
        'day_shifts': {
            'count': len(day_staff_entitlements),
            'total_hours': round(float(day_total_hours), 1),
            'used_hours': round(float(day_used_hours), 1),
            'remaining_hours': round(float(day_remaining_hours), 1),
            'usage_percent': round((float(day_used_hours) / float(day_total_hours) * 100) if day_total_hours > 0 else 0, 1)
        },
        'night_shifts': {
            'count': len(night_staff_entitlements),
            'total_hours': round(float(night_total_hours), 1),
            'used_hours': round(float(night_used_hours), 1),
            'remaining_hours': round(float(night_remaining_hours), 1),
            'usage_percent': round((float(night_used_hours) / float(night_total_hours) * 100) if night_total_hours > 0 else 0, 1)
        },
        'staff': staff_list
    }
    
    return JsonResponse(response_data)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@login_required
def staff_search_rota(request):
    """Search for staff member and show their off-duty periods for 6 weeks"""
    
    # Check permissions
    if not (request.user.role and request.user.role.can_manage_rota):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'GET':
        try:
            search_term = request.GET.get('search', '').strip()
            
            if not search_term:
                return JsonResponse({'staff': []})
                
            # Search for staff by SAP number, first name, or last name
            staff = User.objects.filter(
                Q(sap__icontains=search_term) |
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term),
                is_active=True
            ).select_related('role', 'unit').order_by('sap')[:10]
            
            staff_data = []
            for member in staff:
                staff_data.append({
                    'sap': member.sap,
                    'name': member.full_name,
                    'role': member.role.get_name_display() if member.role else 'No Role',
                    'unit': member.unit.get_name_display() if member.unit else 'No Unit'
                })
                
            response = JsonResponse({'staff': staff_data})
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken'
            return response
        
        except Exception as e:
            return JsonResponse({'error': f'Search failed: {str(e)}'}, status=500)
    
    elif request.method == 'POST':
        # Get detailed schedule for selected staff member
        sap = request.POST.get('sap')
        if not sap:
            return JsonResponse({'error': 'Staff member not specified'}, status=400)
            
        try:
            staff_member = User.objects.get(sap=sap, is_active=True)
        except User.DoesNotExist:
            return JsonResponse({'error': 'Staff member not found'}, status=404)
            
        # Get 6-week period starting from current week
        today = timezone.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_date = start_of_week + timedelta(weeks=6, days=-1)
        
        # Get all shifts for this staff member in the 6-week period
        shifts = Shift.objects.filter(
            user=staff_member,
            date__range=[start_of_week, end_date]
        ).select_related('unit', 'shift_type').order_by('date')
        
        # Get leave requests for this period
        leave_requests = LeaveRequest.objects.filter(
            user=staff_member,
            start_date__lte=end_date,
            end_date__gte=start_of_week,
            status__in=['APPROVED', 'PENDING']
        ).order_by('start_date')
        
        # Create schedule data
        schedule_data = {
            'staff_info': {
                'sap': staff_member.sap,
                'name': staff_member.full_name,
                'role': staff_member.role.get_name_display() if staff_member.role else 'No Role',
                'unit': staff_member.unit.get_name_display() if staff_member.unit else 'No Unit',
                'shifts_per_week': round(shifts.count() / 6, 1) if shifts.count() > 0 else 0  # Actual average
            },
            'period': {
                'start': start_of_week.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'weeks': []
        }
        
        # Process 6 weeks
        for week_num in range(6):
            week_start = start_of_week + timedelta(weeks=week_num)
            week_end = week_start + timedelta(days=6)
            
            week_shifts = [s for s in shifts if week_start <= s.date <= week_end]
            week_leave = [l for l in leave_requests if not (l.end_date < week_start or l.start_date > week_end)]
            
            # Create daily data
            days = []
            for day_offset in range(7):
                current_date = week_start + timedelta(days=day_offset)
                day_shifts = [s for s in week_shifts if s.date == current_date]
                day_leave = [l for l in week_leave if l.start_date <= current_date <= l.end_date]
                
                day_data = {
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day_name': current_date.strftime('%A'),
                    'is_off_duty': len(day_shifts) == 0 and len(day_leave) == 0,
                    'shifts': [],
                    'leave': []
                }
                
                for shift in day_shifts:
                    day_data['shifts'].append({
                        'id': shift.id,
                        'shift_type': shift.shift_type.get_name_display(),
                        'unit': shift.unit.get_name_display(),
                        'start_time': shift.shift_type.start_time.strftime('%H:%M'),
                        'end_time': shift.shift_type.end_time.strftime('%H:%M')
                    })
                    
                for leave in day_leave:
                    day_data['leave'].append({
                        'type': leave.get_leave_type_display(),
                        'status': leave.get_status_display()
                    })
                    
                days.append(day_data)
            
            week_data = {
                'week_number': week_num + 1,
                'start_date': week_start.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'total_shifts': len(week_shifts),
                'off_duty_days': len([d for d in days if d['is_off_duty']]),
                'days': days
            }
            
            schedule_data['weeks'].append(week_data)
            
        return JsonResponse(schedule_data)
        
@login_required
def edit_shift(request):
    """Edit or delete a shift with proper CSRF protection"""
    
    # Check permissions
    if not (request.user.role and request.user.role.can_manage_rota):
        return JsonResponse({'error': 'Permission denied'}, status=403)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            shift_id = data.get('shift_id')
            action = data.get('action')  # 'edit' or 'delete'
            
            if not shift_id:
                return JsonResponse({'error': 'Shift ID required'}, status=400)
                
            try:
                shift = Shift.objects.get(id=shift_id)
            except Shift.DoesNotExist:
                return JsonResponse({'error': 'Shift not found'}, status=404)
                
            if action == 'delete':
                # Log the deletion
                ActivityLog.objects.create(
                    user=request.user,
                    action='SHIFT_DELETED',
                    description=f'Deleted shift for {shift.user.full_name} on {shift.date} ({shift.shift_type.get_name_display()})'
                )
                
                shift.delete()
                return JsonResponse({'success': True, 'message': 'Shift deleted successfully'})
                
            elif action == 'edit':
                # Get new values
                new_unit_name = data.get('unit')
                new_shift_type_name = data.get('shift_type')
                new_user_sap = data.get('user_sap')
                
                old_values = {
                    'user': shift.user.full_name,
                    'unit': shift.unit.name,
                    'shift_type': shift.shift_type.name
                }
                
                # Update unit if provided
                if new_unit_name:
                    try:
                        new_unit = Unit.objects.get(name=new_unit_name, is_active=True)
                        shift.unit = new_unit
                    except Unit.DoesNotExist:
                        return JsonResponse({'error': f'Unit {new_unit_name} not found'}, status=400)
                        
                # Update shift type if provided
                if new_shift_type_name:
                    try:
                        new_shift_type = ShiftType.objects.get(name=new_shift_type_name)
                        shift.shift_type = new_shift_type
                    except ShiftType.DoesNotExist:
                        return JsonResponse({'error': f'Shift type {new_shift_type_name} not found'}, status=400)
                        
                # Update user if provided
                if new_user_sap:
                    try:
                        new_user = User.objects.get(sap=new_user_sap, is_active=True)
                        shift.user = new_user
                    except User.DoesNotExist:
                        return JsonResponse({'error': f'Staff member {new_user_sap} not found'}, status=400)
                
                shift.save()
                
                # Log the change
                changes = []
                if new_unit_name and old_values['unit'] != new_unit_name:
                    changes.append(f'unit: {old_values["unit"]} → {new_unit_name}')
                if new_shift_type_name and old_values['shift_type'] != new_shift_type_name:
                    changes.append(f'shift type: {old_values["shift_type"]} → {new_shift_type_name}')
                if new_user_sap and old_values['user'] != shift.user.full_name:
                    changes.append(f'staff: {old_values["user"]} → {shift.user.full_name}')
                    
                if changes:
                    ActivityLog.objects.create(
                        user=request.user,
                        action='SHIFT_UPDATED',
                        description=f'Updated shift on {shift.date}: {"; ".join(changes)}'
                    )
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Shift updated successfully',
                    'shift': {
                        'id': shift.id,
                        'user': shift.user.full_name,
                        'unit': shift.unit.get_name_display(),
                        'shift_type': shift.shift_type.get_name_display(),
                        'date': shift.date.strftime('%Y-%m-%d')
                    }
                })
                
            else:
                return JsonResponse({'error': 'Invalid action'}, status=400)
                
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def add_shift(request):
    """Add a new shift (for agency staff or overtime)"""
    
    # Check permissions
    if not (request.user.role and request.user.role.can_manage_rota):
        return JsonResponse({'error': 'Permission denied'}, status=403)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_sap = data.get('user_sap')
            shift_date_str = data.get('shift_date')
            unit_name = data.get('unit')
            shift_type_code = data.get('shift_type')
            notes = data.get('notes', '')
            
            # New fields for enhanced tracking
            shift_classification = data.get('shift_classification', 'REGULAR')  # REGULAR, OVERTIME, or AGENCY
            shift_pattern = data.get('shift_pattern', 'DAY_0800_2000')  # Standard patterns or CUSTOM
            custom_start_time = data.get('custom_start_time')
            custom_end_time = data.get('custom_end_time')
            agency_company_id = data.get('agency_company_id')
            agency_staff_name = data.get('agency_staff_name', '')
            agency_hourly_rate = data.get('agency_hourly_rate')
            
            # Validation
            if not all([user_sap, shift_date_str, unit_name, shift_type_code]):
                return JsonResponse({
                    'error': 'Missing required fields. Please provide SAP number, date, unit, and shift type.'
                }, status=400)
            
            # Validate shift pattern and custom times
            if shift_pattern == 'CUSTOM':
                if not custom_start_time or not custom_end_time:
                    return JsonResponse({
                        'error': 'Custom start and end times are required when using custom shift pattern.'
                    }, status=400)
            
            # Validate agency requirements
            if shift_classification == 'AGENCY':
                if not agency_company_id:
                    return JsonResponse({
                        'error': 'Agency company is required for agency shifts.'
                    }, status=400)
            
            # Parse date
            try:
                shift_date = datetime.strptime(shift_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)
            
            # Parse custom times if provided
            parsed_start_time = None
            parsed_end_time = None
            if shift_pattern == 'CUSTOM':
                try:
                    parsed_start_time = datetime.strptime(custom_start_time, '%H:%M').time()
                    parsed_end_time = datetime.strptime(custom_end_time, '%H:%M').time()
                except ValueError:
                    return JsonResponse({'error': 'Invalid time format. Use HH:MM (24-hour).'}, status=400)
            
            # Get user
            try:
                user = User.objects.get(sap=user_sap, is_active=True)
            except User.DoesNotExist:
                return JsonResponse({
                    'error': f'Staff member with SAP {user_sap} not found or inactive.'
                }, status=404)
            
            # Get unit
            try:
                unit = Unit.objects.get(name=unit_name, is_active=True)
            except Unit.DoesNotExist:
                return JsonResponse({
                    'error': f'Unit {unit_name} not found or inactive.'
                }, status=404)
            
            # Get agency company if needed
            agency_company = None
            if shift_classification == 'AGENCY':
                try:
                    from .models import AgencyCompany
                    agency_company = AgencyCompany.objects.get(id=agency_company_id, is_active=True)
                except AgencyCompany.DoesNotExist:
                    return JsonResponse({
                        'error': f'Agency company not found or inactive.'
                    }, status=404)
            
            # Map shift type code to database shift type
            shift_type_mapping = {
                'DAY': 'DAY_SENIOR',
                'NIGHT': 'NIGHT_SENIOR',
                'TWILIGHT': 'TWILIGHT',
                'EARLY': 'EARLY',
                'LATE': 'LATE'
            }
            
            shift_type_name = shift_type_mapping.get(shift_type_code, shift_type_code)
            
            # Get shift type
            try:
                shift_type = ShiftType.objects.get(name=shift_type_name)
            except ShiftType.DoesNotExist:
                # Try to find a suitable default shift type
                if 'NIGHT' in shift_type_code:
                    shift_type = ShiftType.objects.filter(name__icontains='NIGHT').first()
                else:
                    shift_type = ShiftType.objects.filter(name__icontains='DAY').first()
                
                if not shift_type:
                    return JsonResponse({
                        'error': f'Shift type {shift_type_code} not found in system.'
                    }, status=404)
            
            # Check for duplicate shift
            existing_shift = Shift.objects.filter(
                user=user,
                date=shift_date,
                unit=unit
            ).first()
            
            if existing_shift:
                return JsonResponse({
                    'error': f'{user.full_name} already has a shift on {shift_date} in {unit.get_name_display()}.'
                }, status=400)
            
            # Create the shift with enhanced fields
            shift = Shift.objects.create(
                user=user,
                date=shift_date,
                unit=unit,
                shift_type=shift_type,
                shift_classification=shift_classification,
                shift_pattern=shift_pattern,
                custom_start_time=parsed_start_time,
                custom_end_time=parsed_end_time,
                agency_company=agency_company,
                agency_staff_name=agency_staff_name,
                agency_hourly_rate=agency_hourly_rate,
                notes=notes,
                created_by=request.user
            )
            
            # Log the creation
            log_description = f'Added {shift_classification.lower()} shift for {user.full_name} on {shift_date}'
            log_description += f' ({shift_type.get_name_display()}) in {unit.get_name_display()}'
            
            if shift_classification == 'AGENCY':
                log_description += f' - Agency: {agency_company.name}'
                if agency_staff_name:
                    log_description += f' ({agency_staff_name})'
            
            if shift_pattern == 'CUSTOM':
                log_description += f' - Custom times: {parsed_start_time.strftime("%H:%M")} to {parsed_end_time.strftime("%H:%M")}'
            else:
                pattern_display = dict(Shift.SHIFT_PATTERN_CHOICES).get(shift_pattern, shift_pattern)
                log_description += f' - Pattern: {pattern_display}'
            
            if notes:
                log_description += f' - Note: {notes}'
            
            ActivityLog.objects.create(
                user=request.user,
                action='SHIFT_CREATED',
                description=log_description
            )
            
            return JsonResponse({
                'success': True,
                'message': f'{shift_classification.title()} shift added successfully for {user.full_name} on {shift_date.strftime("%d/%m/%Y")}',
                'shift': {
                    'id': shift.id,
                    'user': user.full_name,
                    'user_sap': user.sap,
                    'unit': unit.get_name_display(),
                    'shift_type': shift_type.get_name_display(),
                    'date': shift_date.strftime('%Y-%m-%d'),
                    'classification': shift_classification,
                    'pattern': shift_pattern,
                    'start_time': shift.start_time.strftime('%H:%M'),
                    'end_time': shift.end_time.strftime('%H:%M'),
                    'duration_hours': shift.duration_hours,
                    'agency_company': agency_company.name if agency_company else None,
                    'agency_staff_name': agency_staff_name if agency_staff_name else None
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)
            
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# ==================== STAFF VIEWS ====================

@login_required
def staff_dashboard(request):
    """My Rota - Personal dashboard for general staff (Screen 2.1)"""
    
    # Get current date and date range for display
    today = timezone.now().date()
    start_date = today
    end_date = today + timedelta(weeks=6) - timedelta(days=1)  # Look ahead six full weeks
    
    # Get user's personal shifts and approved leave
    user_shifts = Shift.objects.filter(
        user=request.user,
        date__range=[start_date, end_date]
    ).select_related('unit', 'shift_type').order_by('date')
    
    user_leave = LeaveRequest.objects.filter(
        user=request.user,
        status='APPROVED',
        start_date__lte=end_date,
        end_date__gte=start_date
    ).order_by('start_date')
    
    # Organize by date for calendar display
    calendar_data = {}
    for i in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(days=i)
        
        # Get shifts for this date
        date_shifts = user_shifts.filter(date=current_date)
        
        # Check if user has leave on this date
        date_leave = None
        for leave in user_leave:
            if leave.start_date <= current_date <= leave.end_date:
                date_leave = leave
                break
        
        calendar_data[current_date] = {
            'shifts': date_shifts,
            'leave': date_leave,
            'is_today': current_date == today,
            'is_past': current_date < today,
        }
    
    context = {
        'calendar_data': calendar_data,
        'annual_leave_remaining': request.user.annual_leave_remaining,
    'start_date': start_date,
    'end_date': end_date,
        'today': today,
    }
    
    return render(request, 'scheduling/staff_dashboard.html', context)

@login_required
def request_annual_leave(request):
    """Annual leave request form using AnnualLeaveEntitlement system with hours-based tracking"""
    from staff_records.models import AnnualLeaveEntitlement, AnnualLeaveTransaction
    from decimal import Decimal
    
    # Get user's current leave entitlement
    try:
        entitlement = AnnualLeaveEntitlement.objects.get(
            profile__user=request.user,
            leave_year_start__lte=timezone.now().date(),
            leave_year_end__gte=timezone.now().date()
        )
    except AnnualLeaveEntitlement.DoesNotExist:
        messages.error(request, 'No leave entitlement found for the current year. Please contact HR.')
        return redirect('staff_dashboard')
    
    if request.method == 'POST':
        leave_type = request.POST.get('leave_type', 'ANNUAL')
        start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
        reason = request.POST.get('reason', '')
        
        # Count actual scheduled working days during the leave period
        working_days_count = Shift.objects.filter(
            user=request.user,
            date__gte=start_date,
            date__lte=end_date
        ).count()
        
        # If no shifts scheduled, use calendar days (for staff not yet rostered)
        if working_days_count == 0:
            # Fallback: count weekdays only (Mon-Fri)
            current_date = start_date
            working_days_count = 0
            while current_date <= end_date:
                if current_date.weekday() < 5:  # Monday=0, Friday=4
                    working_days_count += 1
                current_date += timedelta(days=1)
        
        days_requested = working_days_count
        
        # Calculate hours to deduct based on work pattern
        # Management/Admin: 7 hours per day (5-day week, 35hr contract)
        # 35hr shift staff (34.5 actual): 11.66 hours per day (12-hour shifts)
        # 24hr shift staff (23.0 actual): 12 hours per day (12-hour shifts)
        
        if request.user.role and request.user.role.is_management:
            # Management staff work standard 5-day week
            hours_per_day = Decimal('7.00')
        elif entitlement.contracted_hours_per_week >= Decimal('30.00'):
            # 35hr shift workers (12-hour shifts)
            hours_per_day = Decimal('11.66')
        else:
            # 24hr shift workers (12-hour shifts)
            hours_per_day = Decimal('12.00')
        
        hours_requested = Decimal(str(working_days_count)) * hours_per_day
        
        # Basic validation
        if start_date > end_date:
            messages.error(request, 'End date must be after start date')
        elif start_date < timezone.now().date():
            messages.error(request, 'Cannot request leave for past dates')
        elif leave_type == 'ANNUAL' and hours_requested > entitlement.hours_remaining:
            messages.error(request, 
                f'Insufficient leave balance. You have {entitlement.hours_remaining} hours '
                f'({entitlement.days_remaining} days) remaining. '
                f'You are requesting {hours_requested} hours ({days_requested} days).')
        else:
            # Create leave request with pending status
            leave_request = LeaveRequest.objects.create(
                user=request.user,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                days_requested=days_requested,
                reason=reason,
                status='PENDING'
            )
            
            # Create a pending transaction (will be confirmed on approval)
            pending_transaction = AnnualLeaveTransaction.objects.create(
                entitlement=entitlement,
                transaction_type='DEDUCTION',
                hours=-hours_requested,
                balance_after=entitlement.hours_remaining - hours_requested,
                related_request=leave_request,
                description=f'Annual leave request: {start_date} to {end_date} ({days_requested} days, {hours_requested} hrs) - PENDING',
                created_by=request.user
            )
            
            # Update pending hours
            entitlement.hours_pending += hours_requested
            entitlement.recalculate_balance()
            entitlement.save()
            
            # Process through auto-approval logic
            _process_auto_approval(leave_request, request.user)

            # Recalculate balance after approval decision
            entitlement.refresh_from_db()
            
            # Provide feedback based on outcome
            if leave_request.status == 'APPROVED':
                messages.success(request, 
                    f'✓ Leave automatically approved! {days_requested} days ({hours_requested} hrs) from {start_date.strftime("%d/%m/%Y")} '
                    f'to {end_date.strftime("%d/%m/%Y")}. You have {entitlement.hours_remaining} hours '
                    f'({entitlement.days_remaining} days) remaining.')
            elif leave_request.status == 'MANUAL_REVIEW':
                messages.warning(request,
                    f'Leave request submitted for management review. Reason: {leave_request.approval_notes}. '
                    f'{hours_requested} hours marked as pending.')
            else:
                messages.info(request,
                    f'Leave request submitted. {days_requested} days ({hours_requested} hrs) requested. '
                    f'You will have {entitlement.hours_remaining} hours ({entitlement.days_remaining} days) remaining if approved.')
            
            return redirect('staff_dashboard')
    
    # Get user's leave history
    leave_history = request.user.leave_requests.filter(
        leave_type='ANNUAL'
    ).order_by('-start_date')[:10]
    
    context = {
        'leave_types': LeaveRequest.LEAVE_TYPES,
        'entitlement': entitlement,
        'total_hours': entitlement.total_entitlement_hours,
        'total_days': entitlement.days_entitlement,
        'hours_used': entitlement.hours_used,
        'days_used': entitlement.days_used,
        'hours_pending': entitlement.hours_pending,
        'hours_remaining': entitlement.hours_remaining,
        'days_remaining': entitlement.days_remaining,
        'leave_year': f"{entitlement.leave_year_start.strftime('%d/%m/%Y')} - {entitlement.leave_year_end.strftime('%d/%m/%Y')}",
        'leave_history': leave_history,
    }

    
    return render(request, 'scheduling/request_leave.html', context)


def _process_auto_approval(leave_request, acting_user):
    """
    Run auto-approval checks and update the request if eligible.
    Creates activity log entries for transparency.
    """

    if leave_request.status != 'PENDING':
        return

    # Run auto-approval checks
    if not _should_auto_approve(leave_request):
        # _should_auto_approve already updated status to MANUAL_REVIEW if needed
        if leave_request.status == 'MANUAL_REVIEW':
            ActivityLog.objects.create(
                user=leave_request.user,
                action_type='LEAVE_REQUEST',
                description=f"Leave request flagged for manual review: {leave_request.approval_notes}",
                automated=True,
                created_by=acting_user,
            )
        return

    # Auto-approve the request
    now = timezone.now()
    leave_request.status = 'APPROVED'
    leave_request.automated_decision = True
    leave_request.approval_date = now
    leave_request.approved_by = acting_user
    leave_request.approval_notes = 'Auto-approved: Met all criteria (≤2 staff off, ≥17 staff on duty, not in blackout period)'
    leave_request.save(update_fields=['status', 'automated_decision', 'approval_date', 'approved_by', 'approval_notes', 'updated_at'])

    ActivityLog.objects.create(
        user=leave_request.user,
        action_type='AUTO_APPROVAL',
        description=f"Leave from {leave_request.start_date} to {leave_request.end_date} auto-approved ({leave_request.days_requested} days)",
        automated=True,
        created_by=acting_user,
    )

@login_required
def request_shift_swap(request):
    """Shift swap request form"""
    
    # Get user's upcoming shifts
    today = timezone.now().date()
    user_shifts = Shift.objects.filter(
        user=request.user,
        date__gte=today,
        status='SCHEDULED'
    ).select_related('unit', 'shift_type').order_by('date')
    
    if request.method == 'POST':
        requesting_shift_id = request.POST.get('requesting_shift')
        target_shift_id = request.POST.get('target_shift')
        reason = request.POST.get('reason', '')
        
        try:
            requesting_shift = Shift.objects.get(id=requesting_shift_id, user=request.user)
            target_shift = Shift.objects.get(id=target_shift_id)
            
            # Create swap request
            swap_request = ShiftSwapRequest.objects.create(
                requesting_user=request.user,
                target_user=target_shift.user,
                requesting_shift=requesting_shift,
                target_shift=target_shift,
                reason=reason,
                status='PENDING'
            )
            
            messages.success(request, 'Shift swap request submitted successfully')
            return redirect('staff_dashboard')
            
        except Shift.DoesNotExist:
            messages.error(request, 'Invalid shift selection')
    
    # Get all available shifts for swapping (same role, different users)
    if request.user.role:
        available_shifts = Shift.objects.filter(
            date__gte=today,
            status='SCHEDULED',
            user__role=request.user.role
        ).exclude(user=request.user).select_related('user', 'unit', 'shift_type').order_by('date')
    else:
        available_shifts = Shift.objects.none()
    
    context = {
        'user_shifts': user_shifts,
        'available_shifts': available_shifts,
    }
    
@login_required
def staff_management(request):
    """Staff Management Dashboard showing home unit assignments with SSCW managers"""
    from .models_multi_home import CareHome
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to manage staff.')
        return redirect('staff_dashboard')
    
    # Get care home filter parameter
    care_home_filter = request.GET.get('care_home', '')
    
    # Get all active care units (excluding ADMIN)
    care_units = Unit.objects.filter(is_active=True).exclude(name='ADMIN')
    
    # Apply care home filter if selected
    if care_home_filter:
        care_units = care_units.filter(care_home__name=care_home_filter)
    
    care_units = care_units.order_by('name')
    
    # Get all staff members
    all_staff = User.objects.filter(is_active=True).select_related('role', 'home_unit', 'unit').order_by('home_unit__name', 'role__name', 'last_name')
    
    # Get staff statistics
    total_staff = all_staff.count()
    management_staff = all_staff.filter(role__is_management=True).count()
    care_staff = all_staff.filter(role__is_management=False).count()
    
    # Build unit structure with home assignments
    units_with_home_staff = []
    
    for unit in care_units:
        # Check if this is a MGMT unit
        is_mgmt_unit = 'MGMT' in unit.name
        
        # For MGMT units, show SM/OM/SSCW/SSCWN (all supernumerary staff)
        if is_mgmt_unit:
            # Get SM, OM, SSCW, and SSCWN staff for this unit (all supernumerary)
            mgmt_staff = all_staff.filter(
                home_unit=unit,
                role__name__in=['SM', 'OM', 'SSCW', 'SSCWN']
            ).order_by('role__name')
            
            units_with_home_staff.append({
                'unit': unit,
                'is_mgmt_unit': True,
                'mgmt_staff': mgmt_staff,
                'static_sscw_managers': {'day_sscw': None, 'night_sscw': None},
                'day_staff': {},
                'night_staff': {},
                'day_staff_count': 0,
                'night_staff_count': 0,
                'total_home_staff': mgmt_staff.count()
            })
            continue  # Skip normal processing for MGMT units
        
        # Normal processing for care units
        from datetime import timedelta
        today = timezone.now().date()
        six_weeks_ago = today - timedelta(weeks=6)
        
        # Get the most frequently assigned day SSCW for this unit (same care home only)
        day_sscw = User.objects.filter(
            role__name='SSCW',
            is_active=True,
            shift_preference='DAY_SENIOR',
            unit__care_home=unit.care_home,  # Same care home
            shifts__unit=unit,
            shifts__date__gte=six_weeks_ago,
            shifts__shift_type__name='DAY_SENIOR'
        ).annotate(
            shift_count=models.Count('shifts')
        ).order_by('-shift_count').first()
        
        # Get the most frequently assigned night SSCWN for this unit (same care home only)
        night_sscw = User.objects.filter(
            role__name='SSCWN',
            is_active=True,
            shift_preference='NIGHT_SENIOR',
            unit__care_home=unit.care_home,  # Same care home
            shifts__unit=unit,
            shifts__date__gte=six_weeks_ago,
            shifts__shift_type__name='NIGHT_SENIOR'
        ).annotate(
            shift_count=models.Count('shifts')
        ).order_by('-shift_count').first()
        
        # Fallback if no SSCW found with recent shifts - get first available from same care home
        if not day_sscw:
            day_sscw = User.objects.filter(
                role__name='SSCW',
                is_active=True,
                shift_preference='DAY_SENIOR',
                unit__care_home=unit.care_home  # Same care home
            ).first()
            
        if not night_sscw:
            night_sscw = User.objects.filter(
                role__name='SSCWN',
                is_active=True,
                shift_preference='NIGHT_SENIOR',
                unit__care_home=unit.care_home  # Same care home
            ).first()
        
        # Static SSCW assignments
        static_sscw_managers = {
            'day_sscw': day_sscw,
            'night_sscw': night_sscw
        }
        
        # Get staff with this unit as their home/base unit
        home_staff = all_staff.filter(
            home_unit=unit,
            role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']  # All care staff have home units
        )
        
        # Separate by day and night shifts based on role
        day_staff = home_staff.filter(
            role__name__in=['SCW', 'SCA']  # Day roles
        ).order_by('role__name', 'team', 'last_name')
        
        night_staff = home_staff.filter(
            role__name__in=['SCWN', 'SCAN']  # Night roles
        ).order_by('role__name', 'team', 'last_name')
        
        # Group day staff by team, then by role
        day_staff_teams = {}
        for team in ['A', 'B', 'C']:
            team_day_staff = day_staff.filter(team=team)
            if team_day_staff.exists():
                roles_data = {}
                total_team_shifts = 0
                
                for role_name in ['SCW', 'SCA']:
                    role_staff = team_day_staff.filter(role__name=role_name)
                    if role_staff.exists():
                        staff_list = []
                        for staff in role_staff:
                            staff_list.append({
                                'staff': staff,
                                'contracted_shifts': staff.shifts_per_week_override or staff.shifts_per_week,
                                'team': staff.team
                            })
                            total_team_shifts += (staff.shifts_per_week_override or staff.shifts_per_week)
                        roles_data[role_name] = staff_list
                
                if roles_data:
                    day_staff_teams[team] = {
                        'roles': roles_data,
                        'total': team_day_staff.count(),
                        'total_shifts': total_team_shifts
                    }
        
        # Group night staff by team, then by role
        night_staff_teams = {}
        for team in ['A', 'B', 'C']:
            team_night_staff = night_staff.filter(team=team)
            if team_night_staff.exists():
                roles_data = {}
                total_team_shifts = 0
                
                for role_name in ['SCWN', 'SCAN']:  # Night roles
                    role_staff = team_night_staff.filter(role__name=role_name)
                    if role_staff.exists():
                        staff_list = []
                        for staff in role_staff:
                            staff_list.append({
                                'staff': staff,
                                'contracted_shifts': staff.shifts_per_week_override or staff.shifts_per_week,
                                'team': staff.team
                            })
                            total_team_shifts += (staff.shifts_per_week_override or staff.shifts_per_week)
                        roles_data[role_name] = staff_list
                
                if roles_data:
                    night_staff_teams[team] = {
                        'roles': roles_data,
                        'total': team_night_staff.count(),
                        'total_shifts': total_team_shifts
                    }
        
        # Count totals
        total_day_staff = day_staff.count()
        total_night_staff = night_staff.count()
        total_home_staff = total_day_staff + total_night_staff
        
        units_with_home_staff.append({
            'unit': unit,
            'static_sscw_managers': static_sscw_managers,
            'day_staff': day_staff_teams,
            'night_staff': night_staff_teams,
            'day_staff_count': total_day_staff,
            'night_staff_count': total_night_staff,
            'total_home_staff': total_home_staff,
            'has_staff': total_home_staff > 0 or static_sscw_managers['day_sscw'] or static_sscw_managers['night_sscw']
        })
    
    # Get overall statistics
    staff_with_home_units = all_staff.filter(home_unit__isnull=False, role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']).count()
    staff_without_home_units = all_staff.filter(home_unit__isnull=True, role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']).count()
    
    # Get all care homes for filter dropdown
    all_care_homes = CareHome.objects.all().order_by('name')
    
    context = {
        'units_with_home_staff': units_with_home_staff,
        'total_staff': total_staff,
        'management_staff': management_staff,
        'care_staff': care_staff,
        'staff_with_home_units': staff_with_home_units,
        'staff_without_home_units': staff_without_home_units,
        'all_care_homes': all_care_homes,
        'selected_care_home': care_home_filter,
    }
    
    return render(request, 'scheduling/staff_management.html', context)

@login_required
def staff_detail(request, sap):
    """Individual staff member details and management"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to view staff details.')
        return redirect('staff_dashboard')
    
    staff_member = get_object_or_404(User, sap=sap)
    
    # Get recent shifts
    today = timezone.now().date()
    recent_shifts = Shift.objects.filter(
        user=staff_member,
        date__gte=today - timedelta(days=30)
    ).select_related('unit', 'shift_type').order_by('-date')[:10]
    
    # Get upcoming shifts
    upcoming_shifts = Shift.objects.filter(
        user=staff_member,
        date__gte=today
    ).select_related('unit', 'shift_type').order_by('date')[:10]
    
    # Get leave requests
    recent_leave = LeaveRequest.objects.filter(
        user=staff_member
    ).order_by('-created_at')[:5]
    
    # Calculate stats
    shifts_this_month = Shift.objects.filter(
        user=staff_member,
        date__gte=today.replace(day=1)
    ).count()
    
    context = {
        'staff_member': staff_member,
        'recent_shifts': recent_shifts,
        'upcoming_shifts': upcoming_shifts,
        'recent_leave': recent_leave,
        'shifts_this_month': shifts_this_month,
    }
    
    return render(request, 'scheduling/staff_detail.html', context)

@login_required
def add_staff(request):
    """Add new staff member"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to add staff.')
        return redirect('staff_dashboard')
    
    if request.method == 'POST':
        sap = request.POST.get('sap')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        role_id = request.POST.get('role')
        annual_leave_allowance = int(request.POST.get('annual_leave_allowance', 28))
        
        # Validation
        if User.objects.filter(sap=sap).exists():
            messages.error(request, f'Staff member with SAP {sap} already exists.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, f'Staff member with email {email} already exists.')
        else:
            try:
                role = Role.objects.get(id=role_id)
                
                # Create new staff member
                new_staff = User.objects.create(
                    sap=sap,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone_number=phone_number,
                    role=role,
                    annual_leave_allowance=annual_leave_allowance,
                    is_active=True
                )
                # Set default password (user should change on first login)
                new_staff.set_password('password123')
                new_staff.save()
                
                messages.success(request, f'Staff member {new_staff.full_name} added successfully!')
                return redirect('staff_management')
                
            except Role.DoesNotExist:
                messages.error(request, 'Selected role does not exist.')
            except Exception as e:
                messages.error(request, f'Error creating staff member: {str(e)}')
    
    context = {
        'roles': Role.objects.all(),
    }
    
    return render(request, 'scheduling/add_staff.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def leave_approval_dashboard(request):
    """Dedicated dashboard for reviewing and approving leave requests."""

    if not (request.user.role and request.user.role.can_approve_leave):
        messages.error(request, 'You do not have permission to approve leave requests.')
        return redirect('staff_dashboard')

    # Filters
    status_filter = request.GET.get('status', 'pending').lower()
    unit_filter = request.GET.get('unit', 'all')

    base_queryset = LeaveRequest.objects.select_related('user', 'user__unit', 'approved_by').order_by('-created_at')

    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        action = request.POST.get('action')
        leave_request = get_object_or_404(base_queryset, id=request_id)

        if leave_request.status not in ['PENDING', 'MANUAL_REVIEW']:
            messages.warning(request, 'That request has already been processed.')
            return redirect(request.path)

        if action == 'approve':
            leave_request.status = 'APPROVED'
            leave_request.approval_date = timezone.now()
            leave_request.approved_by = request.user
            leave_request.automated_decision = False
            leave_request.save(update_fields=['status', 'approval_date', 'approved_by', 'automated_decision', 'updated_at'])

            ActivityLog.objects.create(
                user=leave_request.user,
                action_type='LEAVE_APPROVED',
                description=f"Leave approved manually for {leave_request.start_date} to {leave_request.end_date}",
                automated=False,
                created_by=request.user,
            )

            messages.success(request, f"Approved leave for {leave_request.user.full_name}.")
        elif action == 'deny':
            leave_request.status = 'DENIED'
            leave_request.approval_date = timezone.now()
            leave_request.approved_by = request.user
            leave_request.automated_decision = False
            leave_request.save(update_fields=['status', 'approval_date', 'approved_by', 'automated_decision', 'updated_at'])

            ActivityLog.objects.create(
                user=leave_request.user,
                action_type='LEAVE_DENIED',
                description=f"Leave denied for {leave_request.start_date} to {leave_request.end_date}",
                automated=False,
                created_by=request.user,
            )

            messages.info(request, f"Denied leave for {leave_request.user.full_name}.")

        return redirect(request.path)

    # Re-run auto approval checks for any pending requests using the requester context
    for pending in LeaveRequest.objects.filter(status='PENDING'):
        _process_auto_approval(pending, pending.user)

    filtered_queryset = base_queryset

    if status_filter == 'pending':
        filtered_queryset = filtered_queryset.filter(status__in=['PENDING', 'MANUAL_REVIEW'])
    elif status_filter == 'manual':
        filtered_queryset = filtered_queryset.filter(status='MANUAL_REVIEW')
    elif status_filter == 'approved':
        filtered_queryset = filtered_queryset.filter(status='APPROVED')
    elif status_filter == 'denied':
        filtered_queryset = filtered_queryset.filter(status='DENIED')

    if unit_filter != 'all':
        filtered_queryset = filtered_queryset.filter(user__unit__name=unit_filter)

    pending_requests = filtered_queryset.filter(status__in=['PENDING', 'MANUAL_REVIEW'])

    recent_approved = base_queryset.filter(status='APPROVED').exclude(approval_date__isnull=True).order_by('-approval_date')[:20]

    today = timezone.now().date()
    counts = {
        'pending': LeaveRequest.objects.filter(status='PENDING').count(),
        'review': LeaveRequest.objects.filter(status='MANUAL_REVIEW').count(),
        'recent_approved': LeaveRequest.objects.filter(status='APPROVED', approval_date__date__gte=today - timedelta(days=7)).count(),
        'auto_approved_last_30': LeaveRequest.objects.filter(
            status='APPROVED',
            automated_decision=True,
            approval_date__date__gte=today - timedelta(days=30)
        ).count(),
    }

    context = {
        'pending_requests': pending_requests,
        'recent_approved': recent_approved,
        'counts': counts,
        'units': Unit.objects.filter(is_active=True),
        'status_filter': status_filter,
        'unit_filter': unit_filter,
    }

    return render(request, 'scheduling/leave/leave_approvals.html', context)
@login_required
def auto_assign_teams(request):
    """Automatically assign staff to teams across units"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get all care units (excluding admin)
        care_units = Unit.objects.exclude(name='ADMIN')
        
        # Get all care staff (SCW and SCA)
        care_staff = User.objects.filter(
            is_active=True,
            role__name__in=['SCW', 'SCA']
        ).order_by('role__name', 'sap')
        
        teams = ['A', 'B', 'C']
        assigned_count = 0
        
        # Assign staff to units and teams
        for unit in care_units:
            unit_care_staff = []
            
            # Calculate how many staff needed per team for this unit
            # Target: Each team should have mix of SCWs and SCAs
            scws_needed = 6  # 2 SCWs per team * 3 teams = 6 SCWs per unit
            scas_needed = 9  # 3 SCAs per team * 3 teams = 9 SCAs per unit (more for dementia)
            
            if unit.name == 'DEMENTIA':
                scas_needed = 12  # 4 SCAs per team for dementia unit
            
            # Get available SCWs and SCAs
            available_scws = care_staff.filter(role__name='SCW', unit__isnull=True)[:scws_needed]
            available_scas = care_staff.filter(role__name='SCA', unit__isnull=True)[:scas_needed]
            
            # Assign SCWs to teams (2 per team)
            scw_list = list(available_scws)
            for i, scw in enumerate(scw_list):
                team = teams[i % 3]  # Rotate through teams A, B, C
                scw.unit = unit
                scw.team = team
                scw.save()
                assigned_count += 1
            
            # Assign SCAs to teams
            sca_list = list(available_scas)
            for i, sca in enumerate(sca_list):
                team = teams[i % 3]  # Rotate through teams A, B, C
                sca.unit = unit
                sca.team = team
                sca.save()
                assigned_count += 1
        
        # Assign remaining management and support staff to admin unit
        remaining_staff = User.objects.filter(
            is_active=True,
            unit__isnull=True
        )
        
        admin_unit = Unit.objects.get(name='ADMIN')
        for staff in remaining_staff:
            staff.unit = admin_unit
            if staff.role and staff.role.name in ['SCW', 'SCA']:
                staff.team = 'A'  # Default team for any remaining care staff
            staff.save()
            assigned_count += 1
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully assigned {assigned_count} staff members to teams and units'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def team_management(request):
    """Team Management Dashboard showing balanced team structures"""
    
    # Check permissions
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, 'You do not have permission to view team management.')
        return redirect('staff_dashboard')
    
    teams = ['A', 'B', 'C']
    team_data = []
    
    for team in teams:
        # Day shift staff
        day_scw = User.objects.filter(
            team=team, 
            role__name='SCW', 
            shift_preference='DAY_SENIOR', 
            is_active=True
        ).select_related('role')
        
        day_sca = User.objects.filter(
            team=team, 
            role__name='SCA', 
            shift_preference='DAY_ASSISTANT', 
            is_active=True
        ).select_related('role')
        
        day_sscw = User.objects.filter(
            team=team, 
            role__name='SSCW', 
            shift_preference='DAY_SENIOR', 
            is_active=True
        ).select_related('role')
        
        # Night shift staff
        night_scw = User.objects.filter(
            team=team, 
            role__name='SCW', 
            shift_preference='NIGHT_SENIOR', 
            is_active=True
        ).select_related('role')
        
        night_sca = User.objects.filter(
            team=team, 
            role__name='SCA', 
            shift_preference='NIGHT_ASSISTANT', 
            is_active=True
        ).select_related('role')
        
        night_sscw = User.objects.filter(
            team=team, 
            role__name='SSCW', 
            shift_preference='NIGHT_SENIOR', 
            is_active=True
        ).select_related('role')
        
        team_info = {
            'name': team,
            'day_staff': {
                'scw': day_scw,
                'sca': day_sca,
                'sscw': day_sscw,
                'total': day_scw.count() + day_sca.count() + day_sscw.count()
            },
            'night_staff': {
                'scw': night_scw,
                'sca': night_sca,
                'sscw': night_sscw,
                'total': night_scw.count() + night_sca.count() + night_sscw.count()
            }
        }
        
        team_data.append(team_info)
    
    # Overall statistics
    total_day_staff = User.objects.filter(
        shift_preference__in=['DAY_SENIOR', 'DAY_ASSISTANT'],
        is_active=True,
        team__isnull=False
    ).count()
    
    total_night_staff = User.objects.filter(
        shift_preference__in=['NIGHT_SENIOR', 'NIGHT_ASSISTANT'],
        is_active=True,
        team__isnull=False
    ).count()
    
    unassigned_staff = User.objects.filter(
        team__isnull=True,
        is_active=True,
        role__name__in=['SCW', 'SCA', 'SSCW']
    ).count()
    
    today = timezone.now().date()
    default_start = (
        Shift.objects.filter(date__gte=today)
        .order_by('date')
        .values_list('date', flat=True)
        .first()
    )
    if not default_start:
        default_start = today
    default_weeks = 3
    default_end = default_start + timedelta(days=7 * default_weeks - 1)

    context = {
        'teams': team_data,
        'total_day_staff': total_day_staff,
        'total_night_staff': total_night_staff,
        'unassigned_staff': unassigned_staff,
        'default_start_date': default_start,
        'default_end_date': default_end,
        'default_weeks': default_weeks,
    }
    
    return render(request, 'scheduling/team_management.html', context)


@login_required
def team_shift_summary(request):
    """Return per-team shift coverage metrics for the requested window."""

    if not (request.user.role and request.user.role.is_management):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    weeks_param = request.GET.get('weeks')
    try:
        weeks = int(weeks_param) if weeks_param else 3
    except ValueError:
        weeks = 3
    weeks = max(1, min(weeks, 12))

    start_param = request.GET.get('start_date')
    start_date = parse_date(start_param) if start_param else None
    today = timezone.now().date()
    if start_date is None:
        start_date = (
            Shift.objects.filter(date__gte=today)
            .order_by('date')
            .values_list('date', flat=True)
            .first()
        ) or today

    end_date = start_date + timedelta(days=7 * weeks - 1)

    relevant_users = (
        User.objects.filter(
            is_active=True,
            role__name__in=['SSCW', 'SCW', 'SCA'],
            team__in=['A', 'B', 'C'],
        )
        .select_related('role')
        .order_by('team', 'role__name', 'last_name', 'first_name')
    )

    shift_qs = Shift.objects.filter(
        user__in=relevant_users,
        date__range=(start_date, end_date),
        status__in=['SCHEDULED', 'CONFIRMED'],
    )

    shift_totals = {
        row['user__sap']: row['total']
        for row in shift_qs.values('user__sap').annotate(total=Count('id'))
    }

    shift_type_breakdown = defaultdict(dict)
    for row in shift_qs.values('user__sap', 'shift_type__name').annotate(total=Count('id')):
        shift_type_breakdown[row['user__sap']][row['shift_type__name']] = row['total']

    teams_payload = {
        name: {
            'name': name,
            'members': [],
            'scheduled_total': 0,
            'expected_total': 0,
            'scheduled_day_total': 0,
            'scheduled_night_total': 0,
            'headcount': 0,
        }
        for name in ['A', 'B', 'C']
    }

    for user in relevant_users:
        scheduled_total = shift_totals.get(user.sap, 0)
        breakdown = shift_type_breakdown.get(user.sap, {})
        day_count = (
            breakdown.get('DAY_SENIOR', 0)
            + breakdown.get('DAY_ASSISTANT', 0)
            + breakdown.get('ADMIN', 0)
        )
        night_count = breakdown.get('NIGHT_SENIOR', 0) + breakdown.get('NIGHT_ASSISTANT', 0)
        shifts_per_week = user.shifts_per_week
        expected_total = shifts_per_week * weeks

        team_bucket = teams_payload.get(user.team)
        if not team_bucket:
            continue

        team_bucket['members'].append({
            'sap': user.sap,
            'name': user.full_name,
            'role': user.role.get_name_display() if user.role else 'Unknown',
            'shift_preference': user.get_shift_preference_display() if user.shift_preference else 'Not set',
            'team': user.team,
            'shifts_per_week': shifts_per_week,
            'override': user.shifts_per_week_override,
            'expected_total': expected_total,
            'scheduled_total': scheduled_total,
            'scheduled_day': day_count,
            'scheduled_night': night_count,
        })

        team_bucket['scheduled_total'] += scheduled_total
        team_bucket['scheduled_day_total'] += day_count
        team_bucket['scheduled_night_total'] += night_count
        team_bucket['expected_total'] += expected_total
        team_bucket['headcount'] += 1

    teams_ordered = [teams_payload[name] for name in ['A', 'B', 'C']]

    grand_totals = {
        'scheduled_total': sum(team['scheduled_total'] for team in teams_ordered),
        'expected_total': sum(team['expected_total'] for team in teams_ordered),
        'scheduled_day_total': sum(team['scheduled_day_total'] for team in teams_ordered),
        'scheduled_night_total': sum(team['scheduled_night_total'] for team in teams_ordered),
        'headcount': sum(team['headcount'] for team in teams_ordered),
    }

    return JsonResponse(
        {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'weeks': weeks,
            'teams': teams_ordered,
            'totals': grand_totals,
        }
    )


@login_required
@require_http_methods(["POST"])
def update_team_assignment(request):
    """Update a staff member's team or shift override from the team dashboard."""

    if not (request.user.role and request.user.role.is_management):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON payload'}, status=400)

    sap = payload.get('sap')
    if not sap:
        return JsonResponse({'error': 'SAP identifier required'}, status=400)

    user = get_object_or_404(User, sap=sap)

    updates = []

    if 'team' in payload:
        team_value = payload['team']
        if team_value not in [None, '', 'A', 'B', 'C']:
            return JsonResponse({'error': 'Invalid team selection'}, status=400)
        user.team = team_value or None
        updates.append('team')

    if 'shifts_per_week_override' in payload:
        override_value = payload['shifts_per_week_override']
        if override_value in [None, '', 'null']:
            user.shifts_per_week_override = None
        else:
            try:
                override_int = int(override_value)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'Override must be an integer value'}, status=400)
            if override_int < 0 or override_int > 7:
                return JsonResponse({'error': 'Override must be between 0 and 7 shifts per week'}, status=400)
            user.shifts_per_week_override = override_int
        updates.append('shifts_per_week_override')

    if not updates:
        return JsonResponse({'error': 'No fields provided for update'}, status=400)

    user.save(update_fields=updates + ['updated_at'])

    return JsonResponse({
        'success': True,
        'sap': user.sap,
        'team': user.team,
        'shifts_per_week': user.shifts_per_week,
        'override': user.shifts_per_week_override,
    })


# ============================================================================
# AUDIT & COMPLIANCE VIEWS
# ============================================================================

@login_required
def audit_dashboard(request):
    """Main audit dashboard showing recent activity and quick stats"""
    from .models import (
        DataChangeLog,
        SystemAccessLog,
        ComplianceViolation,
        ComplianceCheck,
        AuditReport
    )
    
    # Get date range for filtering (default: last 7 days)
    days = int(request.GET.get('days', 7))
    start_date = timezone.now() - timedelta(days=days)
    
    # Recent data changes
    recent_changes = DataChangeLog.objects.filter(
        timestamp__gte=start_date
    ).select_related('user', 'content_type').order_by('-timestamp')[:50]
    
    # Recent access logs
    recent_access = SystemAccessLog.objects.filter(
        timestamp__gte=start_date
    ).select_related('user').order_by('-timestamp')[:50]
    
    # Open compliance violations
    open_violations = ComplianceViolation.objects.filter(
        status='OPEN'
    ).select_related('compliance_check__rule', 'affected_user').order_by('-detected_at')[:20]
    
    # Recent compliance checks
    recent_checks = ComplianceCheck.objects.filter(
        started_at__gte=start_date
    ).select_related('rule').order_by('-started_at')[:20]
    
    # Stats
    stats = {
        'total_changes': DataChangeLog.objects.filter(timestamp__gte=start_date).count(),
        'total_access_events': SystemAccessLog.objects.filter(timestamp__gte=start_date).count(),
        'open_violations': ComplianceViolation.objects.filter(status='OPEN').count(),
        'critical_violations': ComplianceViolation.objects.filter(
            status='OPEN', 
            compliance_check__rule__severity='CRITICAL'
        ).count(),
        'failed_logins': SystemAccessLog.objects.filter(
            timestamp__gte=start_date,
            access_type='LOGIN_FAILED'
        ).count(),
        'recent_checks': ComplianceCheck.objects.filter(
            started_at__gte=start_date,
            status='COMPLETED'
        ).count(),
    }
    
    # Change breakdown by action
    change_breakdown = DataChangeLog.objects.filter(
        timestamp__gte=start_date
    ).values('action').annotate(count=Count('id')).order_by('-count')
    
    context = {
        'recent_changes': recent_changes,
        'recent_access': recent_access,
        'open_violations': open_violations,
        'recent_checks': recent_checks,
        'stats': stats,
        'change_breakdown': change_breakdown,
        'days': days,
        'start_date': start_date,
    }
    
    return render(request, 'scheduling/audit_dashboard.html', context)


@login_required
def compliance_dashboard(request):
    """Compliance monitoring dashboard"""
    from .models import (
        ComplianceRule,
        ComplianceCheck,
        ComplianceViolation
    )
    
    # Get all active compliance rules
    rules = ComplianceRule.objects.filter(is_active=True).order_by('category', 'severity')
    
    # Violation stats by status
    violation_stats = {}
    for status_code, status_name in ComplianceViolation.STATUS_CHOICES:
        violation_stats[status_code] = ComplianceViolation.objects.filter(status=status_code).count()
    
    # Violation breakdown by severity
    severity_breakdown = ComplianceViolation.objects.filter(
        status='OPEN'
    ).values('compliance_check__rule__severity').annotate(count=Count('id')).order_by('-count')
    
    # Recent violations by category
    category_breakdown = ComplianceViolation.objects.filter(
        status='OPEN'
    ).values('compliance_check__rule__category').annotate(count=Count('id')).order_by('-count')
    
    # Recent compliance checks
    recent_checks = ComplianceCheck.objects.select_related('rule').order_by('-started_at')[:20]
    
    # Critical open violations
    critical_violations = ComplianceViolation.objects.filter(
        status='OPEN',
        compliance_check__rule__severity='CRITICAL'
    ).select_related('compliance_check__rule', 'affected_user').order_by('-detected_at')[:10]
    
    context = {
        'rules': rules,
        'violation_stats': violation_stats,
        'severity_breakdown': severity_breakdown,
        'category_breakdown': category_breakdown,
        'recent_checks': recent_checks,
        'critical_violations': critical_violations,
    }
    
    return render(request, 'scheduling/compliance_dashboard.html', context)


@login_required
def data_change_log_list(request):
    """List and filter data change logs"""
    from .models import DataChangeLog
    from django.contrib.contenttypes.models import ContentType
    
    # Base queryset
    logs = DataChangeLog.objects.select_related('user', 'content_type').order_by('-timestamp')
    
    # Filters
    action = request.GET.get('action')
    if action:
        logs = logs.filter(action=action)
    
    user_sap = request.GET.get('user')
    if user_sap:
        logs = logs.filter(user__sap=user_sap)
    
    content_type_id = request.GET.get('content_type')
    if content_type_id:
        logs = logs.filter(content_type_id=content_type_id)
    
    days = request.GET.get('days')
    if days:
        start_date = timezone.now() - timedelta(days=int(days))
        logs = logs.filter(timestamp__gte=start_date)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Available filters
    actions = DataChangeLog.objects.values_list('action', flat=True).distinct()
    content_types = ContentType.objects.filter(
        id__in=DataChangeLog.objects.values_list('content_type_id', flat=True).distinct()
    )
    
    context = {
        'page_obj': page_obj,
        'actions': actions,
        'content_types': content_types,
        'current_filters': {
            'action': action,
            'user': user_sap,
            'content_type': content_type_id,
            'days': days,
        }
    }
    
    return render(request, 'scheduling/data_change_log.html', context)


@login_required
def system_access_log_list(request):
    """List and filter system access logs"""
    from .models import SystemAccessLog
    
    # Base queryset
    logs = SystemAccessLog.objects.select_related('user').order_by('-timestamp')
    
    # Filters
    access_type = request.GET.get('access_type')
    if access_type:
        logs = logs.filter(access_type=access_type)
    
    user_sap = request.GET.get('user')
    if user_sap:
        logs = logs.filter(user__sap=user_sap)
    
    success = request.GET.get('success')
    if success:
        logs = logs.filter(success=(success == 'true'))
    
    days = request.GET.get('days')
    if days:
        start_date = timezone.now() - timedelta(days=int(days))
        logs = logs.filter(timestamp__gte=start_date)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Available filters
    access_types = SystemAccessLog.ACCESS_TYPE_CHOICES
    
    context = {
        'page_obj': page_obj,
        'access_types': access_types,
        'current_filters': {
            'access_type': access_type,
            'user': user_sap,
            'success': success,
            'days': days,
        }
    }
    
    return render(request, 'scheduling/system_access_log.html', context)


@login_required
def compliance_violation_list(request):
    """List and manage compliance violations"""
    from .models import ComplianceViolation
    
    # Base queryset
    violations = ComplianceViolation.objects.select_related(
        'compliance_check__rule', 'affected_user', 'acknowledged_by', 'resolved_by'
    ).order_by('-detected_at')
    
    # Filters
    status = request.GET.get('status')
    if status:
        violations = violations.filter(status=status)
    
    severity = request.GET.get('severity')
    if severity:
        violations = violations.filter(compliance_check__rule__severity=severity)
    
    category = request.GET.get('category')
    if category:
        violations = violations.filter(compliance_check__rule__category=category)
    
    affected_user_sap = request.GET.get('affected_user')
    if affected_user_sap:
        violations = violations.filter(affected_user__sap=affected_user_sap)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(violations, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Available filters
    from .models import ComplianceRule
    statuses = ComplianceViolation.STATUS_CHOICES
    severities = ComplianceRule.SEVERITY_CHOICES
    categories = ComplianceRule.RULE_CATEGORY_CHOICES
    
    context = {
        'page_obj': page_obj,
        'statuses': statuses,
        'severities': severities,
        'categories': categories,
        'current_filters': {
            'status': status,
            'severity': severity,
            'category': category,
            'affected_user': affected_user_sap,
        }
    }
    
    return render(request, 'scheduling/compliance_violation_list.html', context)


@login_required
def compliance_violation_detail(request, violation_id):
    """View and update a specific compliance violation"""
    from .models import ComplianceViolation
    
    violation = get_object_or_404(
        ComplianceViolation.objects.select_related(
            'compliance_check__rule', 'affected_user', 'acknowledged_by', 'resolved_by'
        ),
        id=violation_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'acknowledge':
            violation.status = 'ACKNOWLEDGED'
            violation.acknowledged_by = request.user
            violation.acknowledged_at = timezone.now()
            violation.save()
            messages.success(request, 'Violation acknowledged')
        
        elif action == 'resolve':
            violation.status = 'RESOLVED'
            violation.resolved_by = request.user
            violation.resolved_at = timezone.now()
            violation.resolution_notes = request.POST.get('resolution_notes', '')
            violation.save()
            messages.success(request, 'Violation resolved')
        
        elif action == 'accept_risk':
            violation.status = 'ACCEPTED_RISK'
            violation.resolved_by = request.user
            violation.resolved_at = timezone.now()
            violation.resolution_notes = request.POST.get('resolution_notes', '')
            violation.save()
            messages.success(request, 'Risk accepted')
        
        elif action == 'false_positive':
            violation.status = 'FALSE_POSITIVE'
            violation.resolved_by = request.user
            violation.resolved_at = timezone.now()
            violation.resolution_notes = request.POST.get('resolution_notes', '')
            violation.save()
            messages.success(request, 'Marked as false positive')
        
        return redirect('compliance_violation_detail', violation_id=violation_id)
    
    context = {
        'violation': violation,
    }
    
    return render(request, 'scheduling/compliance_violation_detail.html', context)


@login_required
def audit_report_list(request):
    """List generated audit reports"""
    from .models import AuditReport
    
    reports = AuditReport.objects.select_related('generated_by').order_by('-generated_at')
    
    # Filters
    report_type = request.GET.get('report_type')
    if report_type:
        reports = reports.filter(report_type=report_type)
    
    status = request.GET.get('status')
    if status:
        reports = reports.filter(status=status)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(reports, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Available filters
    report_types = AuditReport.REPORT_TYPE_CHOICES
    statuses = AuditReport.STATUS_CHOICES
    
    context = {
        'page_obj': page_obj,
        'report_types': report_types,
        'statuses': statuses,
        'current_filters': {
            'report_type': report_type,
            'status': status,
        }
    }
    
    return render(request, 'scheduling/audit_report_list.html', context)


@login_required
def audit_report_detail(request, report_id):
    """View a specific audit report"""
    from .models import AuditReport
    
    report = get_object_or_404(
        AuditReport.objects.select_related('generated_by'),
        id=report_id
    )
    
    context = {
        'report': report,
    }
    
    return render(request, 'scheduling/audit_report_detail.html', context)


@login_required
def generate_audit_report(request):
    """Generate a new audit report"""
    from .models_audit import AuditReport
    from .utils.reports import REPORT_GENERATORS
    from django.utils import timezone
    import json
    
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        title = request.POST.get('title')
        period_start = parse_date(request.POST.get('period_start'))
        period_end = parse_date(request.POST.get('period_end'))
        
        # Create the report
        report = AuditReport.objects.create(
            report_type=report_type,
            title=title,
            period_start=period_start,
            period_end=period_end,
            generated_by=request.user,
            status='GENERATING'
        )
        
        # Generate report data using appropriate generator
        try:
            if report_type in REPORT_GENERATORS:
                generator = REPORT_GENERATORS[report_type]
                
                # Convert dates to datetime with timezone
                start_dt = timezone.make_aware(
                    timezone.datetime.combine(period_start, timezone.datetime.min.time())
                )
                end_dt = timezone.make_aware(
                    timezone.datetime.combine(period_end, timezone.datetime.max.time())
                )
                
                # Generate the data
                report_data = generator(start_dt, end_dt)
                
                # Save the data
                report.report_data = report_data
                report.status = 'COMPLETED'
                report.save()
                
                messages.success(request, f'Report "{title}" generated successfully')
            else:
                report.status = 'FAILED'
                report.save()
                messages.error(request, f'Unknown report type: {report_type}')
        except Exception as e:
            report.status = 'FAILED'
            report.save()
            messages.error(request, f'Error generating report: {str(e)}')
        
        return redirect('audit_report_detail', report_id=report.id)
    
    # GET request - show form
    report_types = AuditReport.REPORT_TYPE_CHOICES
    
    context = {
        'report_types': report_types,
    }
    
    return render(request, 'scheduling/generate_audit_report.html', context)


# ============================================================================
# AI ASSISTANT API
# ============================================================================

# Home name variations for natural language processing
HOME_NAME_VARIATIONS = {
    'orchard grove': ['orchard grove', 'og', 'orchard', 'grove'],
    'meadowburn': ['meadowburn', 'meadow', 'meadowburn house'],
    'hawthorn house': ['hawthorn house', 'hawthorn', 'hh'],
    'riverside': ['riverside', 'riverside house'],
    'victoria gardens': ['victoria gardens', 'vg', 'victoria', 'gardens']
}

def normalize_home_name(query):
    """
    Extract and normalize care home name from query text
    Phase 3: Enhanced with fuzzy matching for typos
    """
    query_lower = query.lower()
    
    # First try exact matching (fastest)
    for canonical_name, variations in HOME_NAME_VARIATIONS.items():
        for variation in variations:
            if variation in query_lower:
                return canonical_name
    
    # Phase 3: If no exact match, try fuzzy matching
    # Extract potential home name from query (simple word extraction)
    words = query_lower.split()
    for word in words:
        if len(word) >= 3:  # Only check words with 3+ chars
            fuzzy_matches = fuzzy_match_home(word, threshold=0.6)
            if fuzzy_matches:
                # Return the best match's canonical name
                best_home, similarity = fuzzy_matches[0]
                return best_home.name
    
    # Also try multi-word combinations (e.g., "orchard grove")
    for i in range(len(words) - 1):
        two_word = f"{words[i]} {words[i+1]}"
        fuzzy_matches = fuzzy_match_home(two_word, threshold=0.6)
        if fuzzy_matches:
            best_home, similarity = fuzzy_matches[0]
            return best_home.name
    
    return None

def get_home_performance(home_name, date=None):
    """
    Get comprehensive performance metrics for a specific care home
    Returns: occupancy, staffing, quality metrics, fiscal status, care plan compliance
    """
    if date is None:
        date = timezone.now().date()
    
    try:
        # Try exact match first, then case-insensitive
        home = CareHome.objects.get(name=home_name.upper().replace(' ', '_'))
    except CareHome.DoesNotExist:
        # Try matching against display names
        for h in CareHome.objects.all():
            if h.get_name_display().lower() == home_name.lower():
                home = h
                break
        else:
            return None
    
    # Get units for this home
    units = home.units.filter(is_active=True)
    unit_ids = list(units.values_list('id', flat=True))
    
    # === OCCUPANCY METRICS ===
    total_beds = home.bed_capacity  # Use CareHome bed_capacity
    residents = Resident.objects.filter(unit__in=units, is_active=True).count()
    occupancy_rate = (residents / total_beds * 100) if total_beds > 0 else 0
    
    # === STAFFING METRICS (TODAY) ===
    today_shifts = Shift.objects.filter(
        unit__in=units,
        date=date,
        status__in=['SCHEDULED', 'CONFIRMED']
    ).select_related('user', 'shift_type')
    
    staffing_today = {
        'total_shifts': today_shifts.count(),
        'day_shifts': today_shifts.filter(shift_type__name__icontains='DAY').count(),
        'night_shifts': today_shifts.filter(shift_type__name__icontains='NIGHT').count(),
        'unfilled': today_shifts.filter(user__isnull=True).count()
    }
    
    # === QUALITY METRICS (30 DAYS) ===
    thirty_days_ago = date - timedelta(days=30)
    
    # Note: IncidentReport doesn't have unit/home field, so we'll use a simplified approach
    # In production, you'd add a care_home foreign key to IncidentReport model
    # For now, get all incidents (across all homes)
    all_incidents_30d = IncidentReport.objects.filter(
        created_at__gte=timezone.make_aware(datetime.combine(thirty_days_ago, datetime.min.time()))
    )
    
    # Count incidents (simplified - actual implementation should filter by home)
    quality_30d = {
        'total_incidents': 0,  # Placeholder - needs care_home field on IncidentReport
        'major_harm': 0,
        'deaths': 0,
        'ci_notifications': 0
    }
    
    # === FISCAL STATUS (THIS MONTH) ===
    month_start = date.replace(day=1)
    if date.month == 12:
        month_end = date.replace(year=date.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = date.replace(month=date.month + 1, day=1) - timedelta(days=1)
    
    month_shifts = Shift.objects.filter(
        unit__in=units,
        date__gte=month_start,
        date__lte=month_end
    )
    
    fiscal_status = {
        'agency_shifts': month_shifts.filter(shift_classification='AGENCY').count(),
        'overtime_shifts': month_shifts.filter(shift_classification='OVERTIME').count(),
        'total_shifts': month_shifts.count()
    }
    
    agency_pct = (fiscal_status['agency_shifts'] / fiscal_status['total_shifts'] * 100) if fiscal_status['total_shifts'] > 0 else 0
    overtime_pct = (fiscal_status['overtime_shifts'] / fiscal_status['total_shifts'] * 100) if fiscal_status['total_shifts'] > 0 else 0
    
    fiscal_status['agency_percentage'] = agency_pct
    fiscal_status['overtime_percentage'] = overtime_pct
    
    # === CARE PLAN COMPLIANCE ===
    # CarePlanReview links directly to resident, not through a care_plan object
    care_plans = CarePlanReview.objects.filter(resident__unit__in=units)
    total_plans = care_plans.count()
    
    overdue_reviews = care_plans.filter(
        due_date__lt=date,
        status__in=['PENDING', 'STARTED']  # Not completed
    ).count()
    
    upcoming_reviews = care_plans.filter(
        due_date__gte=date,
        due_date__lte=date + timedelta(days=7),
        status__in=['PENDING', 'STARTED']
    ).count()
    
    care_plan_compliance = {
        'total_plans': total_plans,
        'overdue_reviews': overdue_reviews,
        'upcoming_7days': upcoming_reviews,
        'compliance_rate': ((total_plans - overdue_reviews) / total_plans * 100) if total_plans > 0 else 100
    }
    
    return {
        'home_name': home.name,
        'display_name': home.get_name_display(),  # Use get_name_display() for readable name
        'occupancy': {
            'residents': residents,
            'beds': total_beds,
            'rate': occupancy_rate
        },
        'staffing_today': staffing_today,
        'quality_30d': quality_30d,
        'fiscal_status': fiscal_status,
        'care_plans': care_plan_compliance
    }

def compare_homes(metric='overall'):
    """
    Compare all homes across specified metrics
    Metrics: overall, quality, compliance, occupancy, fiscal
    """
    homes = CareHome.objects.all().order_by('name')
    comparison = []
    
    for home in homes:
        perf = get_home_performance(home.name)
        if perf:
            comparison.append(perf)
    
    # Sort by specified metric
    if metric == 'quality':
        comparison.sort(key=lambda x: x['quality_30d']['total_incidents'])
    elif metric == 'compliance':
        comparison.sort(key=lambda x: x['care_plans']['compliance_rate'], reverse=True)
    elif metric == 'occupancy':
        comparison.sort(key=lambda x: x['occupancy']['rate'], reverse=True)
    elif metric == 'fiscal':
        comparison.sort(key=lambda x: x['fiscal_status']['agency_percentage'])
    
    return comparison

def _process_home_performance_query(query):
    """
    Process queries about specific home performance, quality audits, or comparisons
    Examples:
    - "Show me Orchard Grove's performance"
    - "How is Hawthorn House doing?"
    - "Quality audit for Victoria Gardens"
    - "Compare all homes"
    - "Which home has the best compliance?"
    """
    query_lower = query.lower()
    
    # === COMPARISON QUERIES ===
    if any(word in query_lower for word in ['compare', 'comparison', 'which home', 'best', 'worst', 'all homes']):
        # Determine comparison metric
        if 'quality' in query_lower or 'incident' in query_lower:
            comparison = compare_homes('quality')
            metric_name = 'Quality (Incidents)'
        elif 'compliance' in query_lower or 'care plan' in query_lower:
            comparison = compare_homes('compliance')
            metric_name = 'Care Plan Compliance'
        elif 'occupancy' in query_lower or 'bed' in query_lower:
            comparison = compare_homes('occupancy')
            metric_name = 'Occupancy Rate'
        elif 'fiscal' in query_lower or 'agency' in query_lower or 'cost' in query_lower:
            comparison = compare_homes('fiscal')
            metric_name = 'Fiscal Performance (Agency Usage)'
        else:
            comparison = compare_homes('overall')
            metric_name = 'Overall Performance'
        
        # Format comparison response
        answer = f"**🏆 Multi-Home Comparison: {metric_name}**\n\n"
        
        for i, home_data in enumerate(comparison, 1):
            medal = ['🥇', '🥈', '🥉'][i-1] if i <= 3 else f"{i}."
            answer += f"{medal} **{home_data['display_name']}**\n"
            answer += f"   • Occupancy: {home_data['occupancy']['rate']:.1f}% ({home_data['occupancy']['residents']}/{home_data['occupancy']['beds']} beds)\n"
            answer += f"   • Staffing Today: {home_data['staffing_today']['total_shifts']} shifts"
            if home_data['staffing_today']['unfilled'] > 0:
                answer += f" ⚠️ {home_data['staffing_today']['unfilled']} unfilled"
            answer += "\n"
            answer += f"   • Quality (30d): {home_data['quality_30d']['total_incidents']} incidents"
            if home_data['quality_30d']['ci_notifications'] > 0:
                answer += f" 🔴 {home_data['quality_30d']['ci_notifications']} CI notifications"
            answer += "\n"
            answer += f"   • Agency Usage: {home_data['fiscal_status']['agency_percentage']:.1f}% ({home_data['fiscal_status']['agency_shifts']} shifts)\n"
            answer += f"   • Care Plan Compliance: {home_data['care_plans']['compliance_rate']:.1f}%"
            if home_data['care_plans']['overdue_reviews'] > 0:
                answer += f" ⚠️ {home_data['care_plans']['overdue_reviews']} overdue"
            answer += "\n\n"
        
        answer += "**📊 Quick Links:**\n"
        answer += "• [View Senior Dashboard](/senior-dashboard/)\n"
        answer += "• [Generate Custom Report](/reports/custom/)\n"
        answer += "• [Export Comparison Data](/reports/export/)\n"
        
        home_perf_score = match_intent_keywords(query, 'home_performance')
        
        return {
            'answer': answer,
            'related': ['Senior Dashboard', 'Quality Reports', 'Fiscal Analysis', 'Export Data'],
            'category': 'home_comparison',
            'comparison_data': comparison,
            'confidence': calculate_confidence_score(query, {
                'answer': answer,
                'category': 'home_comparison',
                'comparison_data': comparison
            }, home_perf_score)
        }
    
    # === HOME-SPECIFIC QUERIES ===
    home_name = normalize_home_name(query)
    if not home_name:
        return None
    
    # Get performance data
    perf = get_home_performance(home_name)
    if not perf:
        return {
            'answer': f"❌ Could not find data for '{home_name}'. Available homes: Orchard Grove, Meadowburn, Hawthorn House, Riverside, Victoria Gardens.",
            'related': ['View All Homes', 'Senior Dashboard'],
            'category': 'error'
        }
    
    # === QUALITY AUDIT QUERY ===
    if 'audit' in query_lower or 'quality' in query_lower:
        answer = f"**🔍 Quality Audit: {perf['display_name']}**\n\n"
        answer += f"**📊 30-Day Quality Metrics**\n"
        answer += f"• Total Incidents: {perf['quality_30d']['total_incidents']}\n"
        
        if perf['quality_30d']['deaths'] > 0:
            answer += f"• ☠️ Deaths: {perf['quality_30d']['deaths']} (requires immediate review)\n"
        if perf['quality_30d']['major_harm'] > 0:
            answer += f"• 🔴 Major Harm: {perf['quality_30d']['major_harm']}\n"
        
        answer += f"• CI Notifications Required: {perf['quality_30d']['ci_notifications']}\n\n"
        
        # Quality status indicator
        if perf['quality_30d']['deaths'] > 0 or perf['quality_30d']['major_harm'] > 0:
            answer += "**Status:** 🔴 CRITICAL - Immediate management attention required\n\n"
        elif perf['quality_30d']['total_incidents'] > 10:
            answer += "**Status:** 🟡 MODERATE - Enhanced monitoring recommended\n\n"
        else:
            answer += "**Status:** 🟢 GOOD - Within acceptable range\n\n"
        
        answer += f"**📋 Care Plan Compliance**\n"
        answer += f"• Total Plans: {perf['care_plans']['total_plans']}\n"
        answer += f"• Compliance Rate: {perf['care_plans']['compliance_rate']:.1f}%\n"
        if perf['care_plans']['overdue_reviews'] > 0:
            answer += f"• ⚠️ Overdue Reviews: {perf['care_plans']['overdue_reviews']} (action required)\n"
        if perf['care_plans']['upcoming_7days'] > 0:
            answer += f"• 📅 Due Next 7 Days: {perf['care_plans']['upcoming_7days']}\n"
        
        answer += "\n**🔗 Actions:**\n"
        answer += f"• [View Full Incident Log](/incidents/?home={perf['home_name']})\n"
        answer += f"• [Care Plan Reviews](/care-plans/?home={perf['home_name']})\n"
        answer += "• [Generate Quality Report](/reports/quality/)\n"
        
        home_perf_score = match_intent_keywords(query, 'home_performance')
        
        return {
            'answer': answer,
            'related': ['Incident Reports', 'Care Plans', 'Quality Metrics'],
            'category': 'quality_audit',
            'audit_data': perf,
            'confidence': calculate_confidence_score(query, {
                'answer': answer,
                'category': 'quality_audit',
                'audit_data': perf
            }, home_perf_score)
        }
    
    # === GENERAL PERFORMANCE QUERY ===
    answer = f"**📊 Performance Overview: {perf['display_name']}**\n\n"
    
    # Occupancy
    answer += f"**🏠 Occupancy**\n"
    answer += f"• Current: {perf['occupancy']['residents']} residents / {perf['occupancy']['beds']} beds\n"
    answer += f"• Rate: {perf['occupancy']['rate']:.1f}%"
    if perf['occupancy']['rate'] < 85:
        answer += " 🟡 Below target"
    elif perf['occupancy']['rate'] >= 95:
        answer += " 🟢 Excellent"
    answer += "\n\n"
    
    # Staffing
    answer += f"**👥 Staffing (Today)**\n"
    answer += f"• Total Shifts: {perf['staffing_today']['total_shifts']}\n"
    answer += f"• Day Shifts: {perf['staffing_today']['day_shifts']} | Night Shifts: {perf['staffing_today']['night_shifts']}\n"
    if perf['staffing_today']['unfilled'] > 0:
        answer += f"• ⚠️ Unfilled: {perf['staffing_today']['unfilled']} (urgent action needed)\n"
    else:
        answer += "• ✅ Fully staffed\n"
    answer += "\n"
    
    # Quality
    answer += f"**⭐ Quality (30 Days)**\n"
    answer += f"• Incidents: {perf['quality_30d']['total_incidents']}"
    if perf['quality_30d']['total_incidents'] > 10:
        answer += " 🟡 Above average"
    answer += "\n"
    if perf['quality_30d']['ci_notifications'] > 0:
        answer += f"• 🔴 CI Notifications: {perf['quality_30d']['ci_notifications']}\n"
    answer += "\n"
    
    # Fiscal
    answer += f"**💰 Fiscal Status (This Month)**\n"
    answer += f"• Agency Usage: {perf['fiscal_status']['agency_percentage']:.1f}% ({perf['fiscal_status']['agency_shifts']} shifts)"
    if perf['fiscal_status']['agency_percentage'] > 15:
        answer += " 🔴 High"
    elif perf['fiscal_status']['agency_percentage'] > 8:
        answer += " 🟡 Elevated"
    else:
        answer += " 🟢 Good"
    answer += "\n"
    answer += f"• Overtime: {perf['fiscal_status']['overtime_percentage']:.1f}% ({perf['fiscal_status']['overtime_shifts']} shifts)\n"
    answer += "\n"
    
    # Care Plans
    answer += f"**📋 Care Plan Compliance**\n"
    answer += f"• Rate: {perf['care_plans']['compliance_rate']:.1f}%"
    if perf['care_plans']['compliance_rate'] < 90:
        answer += " 🔴 Below target"
    elif perf['care_plans']['compliance_rate'] >= 98:
        answer += " 🟢 Excellent"
    answer += "\n"
    if perf['care_plans']['overdue_reviews'] > 0:
        answer += f"• ⚠️ Overdue: {perf['care_plans']['overdue_reviews']} reviews\n"
    
    answer += "\n**🔗 Quick Actions:**\n"
    answer += f"• [View Home Details](/homes/{perf['home_name']}/)\n"
    answer += f"• [Quality Audit](/quality-audit/?home={perf['home_name']})\n"
    answer += f"• [Staffing Report](/reports/staffing/?home={perf['home_name']})\n"
    answer += "• [Senior Dashboard](/senior-dashboard/)\n"
    
    home_perf_score = match_intent_keywords(query, 'home_performance')
    
    return {
        'answer': answer,
        'related': ['Senior Dashboard', 'Home Details', 'Quality Audit', 'Staffing Report'],
        'category': 'home_performance',
        'performance_data': perf,
        'confidence': calculate_confidence_score(query, {
            'answer': answer,
            'category': 'home_performance',
            'performance_data': perf
        }, home_perf_score)
    }

class ReportGenerator:
    """Generate intelligent reports based on natural language queries"""
    
    @staticmethod
    def generate_staffing_summary():
        """Generate comprehensive staffing summary"""
        total_staff = User.objects.filter(is_active=True).count()
        
        # Breakdown by role
        roles = Role.objects.all()
        by_role = {}
        for role in roles:
            count = User.objects.filter(role=role, is_active=True).count()
            if count > 0:
                by_role[f"{role.name}"] = count
        
        # Breakdown by grade (same as role in this system)
        by_grade = by_role.copy()  # Grades are represented by roles
        
        summary = f"You currently have {total_staff} active staff members across {len(by_role)} different roles."
        if by_grade:
            summary += f" Breakdown by grade: {len(by_grade)} grade levels."
        
        return {
            'summary': summary,
            'total': total_staff,
            'by_role': by_role,
            'by_grade': by_grade
        }
    
    @staticmethod
    def generate_staff_by_grade_report(grade_query=None):
        """Generate staff report by specific grade/role or show all grades"""
        active_staff = User.objects.filter(
            is_active=True
        ).select_related('role', 'unit')
        
        # Count by role (which serves as grade)
        by_grade = {}
        staff_details = {}
        
        for user in active_staff:
            if user.role:
                grade = user.role.name.strip().upper()
                by_grade[grade] = by_grade.get(grade, 0) + 1
                
                if grade not in staff_details:
                    staff_details[grade] = []
                
                staff_details[grade].append({
                    'name': user.full_name,
                    'sap': user.sap,
                    'role': user.role.name,
                    'unit': user.home_unit.name if user.home_unit else 'Unassigned'
                })
        
        # If specific grade queried, filter results
        if grade_query:
            grade_upper = grade_query.strip().upper()
            
            # ONLY exact match - no partial matching to avoid confusion
            if grade_upper in by_grade:
                count = by_grade[grade_upper]
                staff_list = staff_details[grade_upper]
                summary = f"You have {count} staff member{'s' if count != 1 else ''} at grade {grade_upper}."
                
                return {
                    'summary': summary,
                    'grade_queried': grade_upper,
                    'count': count,
                    'staff': staff_list,
                    'all_grades': by_grade
                }
            else:
                # No match found
                summary = f"No staff found at grade '{grade_query}'. Available grades: {', '.join(sorted(by_grade.keys()))}."
                
                return {
                    'summary': summary,
                    'grade_queried': grade_query,
                    'count': 0,
                    'staff': [],
                    'all_grades': by_grade
                }
        
        # No specific grade - return all
        total_with_grade = sum(by_grade.values())
        summary = f"You have {total_with_grade} staff members assigned across {len(by_grade)} different grades."
        
        return {
            'summary': summary,
            'total': total_with_grade,
            'by_grade': by_grade,
            'staff_details': staff_details,
            'all_grades': by_grade
        }
    
    @staticmethod
    def generate_sickness_report(days=7):
        """Generate sickness absence report"""
        from django.utils import timezone
        from datetime import timedelta
        
        cutoff = timezone.now() - timedelta(days=days)
        
        # Get recent sickness records
        recent_records = SicknessRecord.objects.filter(
            reported_at__gte=cutoff
        ).select_related('profile__user', 'profile__user__role')
        
        # Get currently off sick (no return date or return date in future)
        currently_sick = recent_records.filter(
            Q(actual_last_working_day__isnull=True) | 
            Q(estimated_return_to_work__gt=timezone.now().date())
        )
        
        sick_staff = []
        for record in currently_sick:
            if record.profile and record.profile.user:
                days_off = (timezone.now().date() - record.first_working_day).days if record.first_working_day else 0
                sick_staff.append({
                    'name': record.profile.user.full_name,
                    'sap': record.profile.user.sap,
                    'days': days_off,
                    'status': record.status or 'Open'
                })
        
        summary = f"Currently {len(sick_staff)} staff off sick. {recent_records.count()} new sickness records in the last {days} days."
        
        return {
            'summary': summary,
            'currently_sick_count': len(sick_staff),
            'recent_cases': recent_records.count(),
            'sick_staff': sick_staff,
            'days_analyzed': days
        }
    
    @staticmethod
    def generate_incident_report(days=7):
        """Generate incident report with severity breakdown"""
        from django.utils import timezone
        from datetime import timedelta
        
        cutoff = timezone.now() - timedelta(days=days)
        
        # Get recent incidents
        incidents = IncidentReport.objects.filter(created_at__gte=cutoff)
        
        # Breakdown by severity
        by_severity = {}
        for incident in incidents:
            severity = incident.severity or 'Unknown'
            by_severity[severity] = by_severity.get(severity, 0) + 1
        
        # Breakdown by type
        by_type = {}
        for incident in incidents:
            incident_type = incident.get_incident_type_display() if hasattr(incident, 'get_incident_type_display') else (incident.incident_type or 'Unknown')
            by_type[incident_type] = by_type.get(incident_type, 0) + 1
        
        # Count CI notifications required
        ci_required = sum(1 for inc in incidents if inc.requires_care_inspectorate_notification())
        
        summary = f"{incidents.count()} total incidents in the last {days} days. {ci_required} require Care Inspectorate notification."
        
        # Check for critical incidents
        critical_alerts = []
        if by_severity.get('Death', 0) > 0:
            critical_alerts.append(f"☠️ {by_severity['Death']} death(s) reported")
        if by_severity.get('Major Harm', 0) > 0:
            critical_alerts.append(f"🔴 {by_severity['Major Harm']} major harm incident(s)")
        
        return {
            'summary': summary,
            'total': incidents.count(),
            'by_severity': by_severity,
            'by_type': by_type,
            'ci_notifications_required': ci_required,
            'critical_alerts': critical_alerts,
            'days_analyzed': days
        }
    
    @staticmethod
    def generate_shift_coverage_report(date_str=None):
        """Generate shift coverage report for a specific date with staff names"""
        from datetime import datetime
        
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                target_date = timezone.now().date()
        else:
            target_date = timezone.now().date()
        
        # Get all shifts for the date
        shifts = Shift.objects.filter(date=target_date).select_related('unit', 'shift_type', 'user')
        
        # Coverage by unit with staff names
        units = Unit.objects.filter(is_active=True)
        by_unit = {}
        staff_by_unit = {}
        
        for unit in units:
            unit_shifts = shifts.filter(unit=unit)
            day_shifts = unit_shifts.filter(shift_type__name__icontains='DAY')
            night_shifts = unit_shifts.filter(shift_type__name__icontains='NIGHT')
            
            by_unit[unit.name] = {
                'day': day_shifts.count(),
                'night': night_shifts.count(),
                'total': unit_shifts.count()
            }
            
            # Get staff names
            staff_by_unit[unit.name] = {
                'day_staff': [f"{s.user.full_name} ({s.shift_type.get_name_display()})" for s in day_shifts if s.user],
                'night_staff': [f"{s.user.full_name} ({s.shift_type.get_name_display()})" for s in night_shifts if s.user],
            }
        
        # Coverage by shift type
        by_shift_type = {}
        staff_by_shift_type = {}
        for shift_type in ShiftType.objects.all():
            shift_list = shifts.filter(shift_type=shift_type)
            count = shift_list.count()
            if count > 0:
                by_shift_type[shift_type.name] = count
                staff_by_shift_type[shift_type.name] = [s.user.full_name for s in shift_list if s.user]
        
        summary = f"On {target_date.strftime('%A, %d %B %Y')}: {shifts.count()} shifts scheduled across {len([u for u in by_unit.values() if u['total'] > 0])} units."
        
        return {
            'summary': summary,
            'date': target_date.strftime('%Y-%m-%d'),
            'total_shifts': shifts.count(),
            'by_unit': by_unit,
            'by_shift_type': by_shift_type,
            'staff_by_unit': staff_by_unit,
            'staff_by_shift_type': staff_by_shift_type
        }
    
    @staticmethod
    def generate_staffing_shortage_report(days=14):
        """
        Analyze staffing levels for next N days and identify TRUE shortages.
        
        Logic:
        1. Check if home has < 17 total staff = CRITICAL shortage (needs agency/OT)
        2. If home has >= 17 staff but units unbalanced = Reallocation needed (not a shortage)
        3. Only report as "shortage" if home-level staffing is inadequate
        """
        from datetime import timedelta
        from collections import defaultdict
        
        today = timezone.now().date()
        end_date = today + timedelta(days=days)
        
        # CRITICAL: Minimum safe staffing level across entire home
        MINIMUM_SAFE_STAFFING = 17
        
        # Expected coverage per unit per shift type (for reallocation guidance only)
        EXPECTED_COVERAGE = {
            'DAY': 4,  # 4 day staff per unit
            'NIGHT': 4,  # 4 night staff per unit
        }
        
        # Analyze each day
        shortage_days = []  # True shortages (below 17 total)
        reallocation_days = []  # Adequate staff but needs reallocation
        critical_below_minimum_days = []
        units = Unit.objects.filter(is_active=True)
        
        for day_offset in range(days):
            check_date = today + timedelta(days=day_offset)
            day_shortages = []
            
            # Calculate TOTAL staff on duty for this day (across all units)
            total_day_staff = Shift.objects.filter(
                date=check_date,
                shift_type__name__icontains='DAY'
            ).count()
            
            total_night_staff = Shift.objects.filter(
                date=check_date,
                shift_type__name__icontains='NIGHT'
            ).count()
            
            # CRITICAL CHECK: Is home below minimum safe staffing?
            day_below_minimum = total_day_staff < MINIMUM_SAFE_STAFFING
            night_below_minimum = total_night_staff < MINIMUM_SAFE_STAFFING
            
            # If home has adequate staff, check for unit imbalances (reallocation opportunities)
            unit_imbalances = []
            for unit in units:
                # Get shifts for this unit on this date
                unit_shifts = Shift.objects.filter(
                    date=check_date,
                    unit=unit
                ).select_related('shift_type')
                
                # Count by shift type
                day_count = unit_shifts.filter(shift_type__name__icontains='DAY').count()
                night_count = unit_shifts.filter(shift_type__name__icontains='NIGHT').count()
                
                # Check for unit-level gaps
                day_gap = max(0, EXPECTED_COVERAGE['DAY'] - day_count)
                night_gap = max(0, EXPECTED_COVERAGE['NIGHT'] - night_count)
                
                if day_gap > 0 or night_gap > 0:
                    unit_imbalances.append({
                        'unit': unit.name,
                        'day_gap': day_gap,
                        'night_gap': night_gap,
                        'day_current': day_count,
                        'night_current': night_count
                    })
            
            # DECISION LOGIC:
            # If below minimum (17 staff) = TRUE SHORTAGE
            # If >= minimum but units unbalanced = REALLOCATION NEEDED
            
            if day_below_minimum or night_below_minimum:
                # TRUE SHORTAGE - needs agency/OT/additional staff
                day_info = {
                    'date': check_date,
                    'day_name': check_date.strftime('%A'),
                    'shortages': unit_imbalances,
                    'total_gaps': sum(s['day_gap'] + s['night_gap'] for s in unit_imbalances),
                    'total_day_staff': total_day_staff,
                    'total_night_staff': total_night_staff,
                    'day_below_minimum': day_below_minimum,
                    'night_below_minimum': night_below_minimum,
                    'critical': True,
                    'reason': 'Below minimum safe staffing'
                }
                shortage_days.append(day_info)
                critical_below_minimum_days.append(day_info)
                
            elif unit_imbalances and (total_day_staff >= MINIMUM_SAFE_STAFFING or total_night_staff >= MINIMUM_SAFE_STAFFING):
                # REALLOCATION NEEDED - home has adequate staff but units unbalanced
                # Calculate automated fair reallocation
                reallocation_plan = ReportGenerator._calculate_fair_reallocation(
                    check_date, units, total_day_staff, total_night_staff, EXPECTED_COVERAGE
                )
                
                realloc_info = {
                    'date': check_date,
                    'day_name': check_date.strftime('%A'),
                    'imbalances': unit_imbalances,
                    'total_day_staff': total_day_staff,
                    'total_night_staff': total_night_staff,
                    'reason': 'Staff reallocation needed between units',
                    'reallocation_plan': reallocation_plan  # NEW: Specific staff moves
                }
                reallocation_days.append(realloc_info)
        
        # Calculate totals for TRUE shortages only
        total_shortage_days = len(shortage_days)
        total_critical_gaps = sum(day['total_gaps'] for day in shortage_days)
        
        # Build summary
        if shortage_days:
            summary = f"🚨 CRITICAL: {total_shortage_days} days BELOW MINIMUM of {MINIMUM_SAFE_STAFFING} staff! "
            summary += f"{total_critical_gaps} additional staff needed (agency/OT)."
        elif reallocation_days:
            summary = f"✅ Adequate staffing levels maintained. {len(reallocation_days)} days need staff reallocation between units."
        else:
            summary = f"✅ No shortages detected for next {days} days - all shifts fully covered with balanced distribution."
        
        return {
            'summary': summary,
            'days_analyzed': days,
            'shortage_days': shortage_days,  # TRUE shortages (below 17)
            'reallocation_days': reallocation_days,  # Adequate staff but imbalanced
            'total_shortage_days': total_shortage_days,
            'total_gaps': total_critical_gaps,
            'critical_below_minimum_days': critical_below_minimum_days,
            'minimum_safe_staffing': MINIMUM_SAFE_STAFFING,
            'expected_coverage': EXPECTED_COVERAGE,
            'has_true_shortages': len(shortage_days) > 0,
            'has_reallocation_needs': len(reallocation_days) > 0
        }
    
    @staticmethod
    def _calculate_fair_reallocation(date, units, total_day_staff, total_night_staff, expected_coverage):
        """
        Calculate automated fair reallocation between units.
        Returns specific staff moves to balance units fairly.
        
        Algorithm:
        1. Identify units with excess staff (above expected)
        2. Identify units with gaps (below expected)
        3. Suggest specific staff moves from excess to gap units
        4. Prioritize minimal disruption
        """
        from scheduling.models import Shift
        
        reallocation_moves = {'day': [], 'night': []}
        
        for shift_period in ['DAY', 'NIGHT']:
            # Get all units with their current staffing
            unit_staffing = {}
            unit_staff_objects = {}  # Store actual shift objects for suggestions
            
            for unit in units:
                shifts = Shift.objects.filter(
                    date=date,
                    unit=unit,
                    shift_type__name__icontains=shift_period
                ).select_related('user', 'shift_type')
                
                count = shifts.count()
                unit_staffing[unit.name] = count
                unit_staff_objects[unit.name] = list(shifts)
            
            # Calculate excess and gaps
            excess_units = {}
            gap_units = {}
            
            for unit_name, count in unit_staffing.items():
                expected = expected_coverage[shift_period]
                if count > expected:
                    excess_units[unit_name] = count - expected
                elif count < expected:
                    gap_units[unit_name] = expected - count
            
            # Generate reallocation suggestions
            for gap_unit, gap_count in gap_units.items():
                moves_for_unit = []
                remaining_gap = gap_count
                
                # Try to fill from excess units
                for excess_unit, excess_count in list(excess_units.items()):
                    if remaining_gap <= 0:
                        break
                    
                    # Calculate how many we can move from this unit
                    can_move = min(excess_count, remaining_gap)
                    
                    # Get specific staff to move
                    staff_to_move = unit_staff_objects[excess_unit][:can_move]
                    
                    for shift in staff_to_move:
                        if shift.user:  # Only suggest moves for assigned staff
                            moves_for_unit.append({
                                'from_unit': excess_unit,
                                'to_unit': gap_unit,
                                'staff_name': shift.user.full_name,
                                'staff_sap': shift.user.sap,
                                'role': shift.user.role.get_name_display() if shift.user.role else 'Staff',
                                'shift_id': shift.id
                            })
                    
                    # Update remaining
                    excess_units[excess_unit] -= can_move
                    remaining_gap -= can_move
                    
                    if excess_units[excess_unit] <= 0:
                        del excess_units[excess_unit]
                
                if moves_for_unit:
                    period_key = 'day' if shift_period == 'DAY' else 'night'
                    reallocation_moves[period_key].extend(moves_for_unit)
        
        return reallocation_moves
    
    @staticmethod
    def generate_shortage_text_message(shortage_data):
        """Generate SMS/text message for staff shortage alert - ONLY for critical days below 17 staff"""
        shortage_days = shortage_data.get('shortage_days', [])
        critical_below_minimum = shortage_data.get('critical_below_minimum_days', [])
        minimum_safe = shortage_data.get('minimum_safe_staffing', 17)
        
        if not shortage_days:
            return {
                'message': None,
                'reason': 'No shortages detected - no message needed',
                'should_send_sms': False
            }
        
        # CRITICAL CHECK: Only send SMS if days are below minimum safe staffing
        if not critical_below_minimum:
            return {
                'message': None,
                'reason': f'Shortages exist but all days have {minimum_safe}+ staff. SMS only sent when below minimum.',
                'should_send_sms': False,
                'shortages_detected': len(shortage_days),
                'note': 'Use email/manual alerts for non-critical shortages'
            }
        
        # Focus on CRITICAL days below minimum (next 7 days)
        next_week_critical = [day for day in critical_below_minimum if 
                             (day['date'] - timezone.now().date()).days <= 7]
        
        if not next_week_critical:
            # Critical days exist but beyond 7 days - still send but different urgency
            next_week_critical = critical_below_minimum[:3]
        
        # Build message
        message_parts = []
        
        # Header - make it URGENT
        message_parts.append(f"🚨 CRITICAL STAFFING SHORTAGE")
        message_parts.append(f"\n⚠️ Below minimum safe staffing of {minimum_safe}")
        message_parts.append(f"\nURGENT: We need staff for:")
        
        # List critical days (max 3)
        for day in next_week_critical[:3]:
            date_str = day['date'].strftime('%a %d %b')
            
            # Show which shifts are critical
            critical_info = []
            if day.get('day_below_minimum'):
                critical_info.append(f"DAY ({day.get('total_day_staff', 0)}/{minimum_safe})")
            if day.get('night_below_minimum'):
                critical_info.append(f"NIGHT ({day.get('total_night_staff', 0)}/{minimum_safe})")
            
            shifts_info = ', '.join(critical_info) if critical_info else f"{day['total_gaps']} shifts"
            message_parts.append(f"\n• {date_str}: {shifts_info}")
        
        # Call to action
        message_parts.append(f"\n\n💰 ENHANCED RATES available")
        message_parts.append(f"📱 Reply YES or call office ASAP")
        message_parts.append(f"\nThank you - Management")
        
        full_message = ''.join(message_parts)
        
        return {
            'message': full_message,
            'should_send_sms': True,
            'critical_days': len(critical_below_minimum),
            'days_below_minimum': len(next_week_critical),
            'character_count': len(full_message),
            'estimated_sms_segments': (len(full_message) // 160) + 1,
            'urgency': 'CRITICAL'
        }
        
        full_message = ''.join(message_parts)
        
        # Also create a more detailed email version
        email_message = f"""Dear Team,

We have identified staffing shortages for the upcoming period and need your help.

SHORTAGES IDENTIFIED:
"""
        for day in shortage_days[:7]:  # Full week
            date_str = day['date'].strftime('%A, %d %B %Y')
            email_message += f"\n{date_str} - {day['total_gaps']} shifts needed:\n"
            for shortage in day['shortages']:
                if shortage['day_shortage'] > 0:
                    email_message += f"  • {shortage['unit']}: {shortage['day_shortage']} day shift(s)\n"
                if shortage['night_shortage'] > 0:
                    email_message += f"  • {shortage['unit']}: {shortage['night_shortage']} night shift(s)\n"
        
        email_message += f"""
WHAT WE'RE OFFERING:
• Enhanced pay rates for these shifts
• Flexibility with shift times where possible
• Appreciation and recognition for helping out

HOW TO RESPOND:
1. Reply to this message with the date(s) you can work
2. Call the office on [PHONE NUMBER]
3. Use the staff portal to pick up shifts

We really appreciate your flexibility and support during this time.

Thank you,
Management Team
"""
        
        return {
            'sms_message': full_message,
            'sms_length': len(full_message),
            'sms_count': (len(full_message) // 160) + 1,  # SMS segments
            'email_message': email_message,
            'recipient_count': User.objects.filter(is_staff=False, is_active=True).count(),
            # 'critical_dates': [day['date'].strftime('%Y-%m-%d') for day in next_week[:3]],
            # 'total_gaps': total_gaps
        }
    
    @staticmethod
    def generate_leave_summary():
        """Generate annual leave summary"""
        from staff_records.models import AnnualLeaveEntitlement
        
        # Pending leave requests
        pending = LeaveRequest.objects.filter(status='PENDING').count()
        
        # Approved future leave
        approved = LeaveRequest.objects.filter(
            status='APPROVED',
            start_date__gte=timezone.now().date()
        ).count()
        
        # Staff with low leave balance (<40 hours)
        low_balance_staff = []
        entitlements = AnnualLeaveEntitlement.objects.select_related('profile__user')
        for ent in entitlements:
            if float(ent.hours_remaining) < 40:
                low_balance_staff.append({
                    'name': ent.profile.user.full_name if ent.profile and ent.profile.user else 'Unknown',
                    'hours': float(ent.hours_remaining)
                })
        
        # Sort by lowest balance
        low_balance_staff.sort(key=lambda x: x['hours'])
        
        summary = f"{pending} pending leave requests. {approved} approved future leave bookings. {len(low_balance_staff)} staff with low leave balance."
        
        return {
            'summary': summary,
            'pending_requests': pending,
            'approved_future': approved,
            'low_balance_count': len(low_balance_staff),
            'low_balance_staff': low_balance_staff[:5]  # Top 5
        }
    
    @staticmethod
    def interpret_query(query):
        """Interpret natural language query and return appropriate report"""
        query_lower = query.lower()
        
        # Staffing queries
        if any(word in query_lower for word in ['how many staff', 'staff count', 'total staff', 'staffing levels']):
            return {
                'type': 'staffing_summary',
                'data': ReportGenerator.generate_staffing_summary()
            }
        
        # Grade-specific staffing queries (e.g., "how many SSCW do we have?")
        grade_patterns = ['grade', 'sscw', 'sscwn', 'scw', 'scwn', 'sca', 'scan', 'sm', 'om', 'senior', 'nurse', 'rn', 'team leader', 'manager', 'supernumerary']
        if any(word in query_lower for word in grade_patterns):
            # Try to extract the grade from the query
            import re
            
            # Check for "supernumerary" special case FIRST
            if 'supernumerary' in query_lower:
                # Return combined results for supernumerary roles
                all_results = {
                    'summary': '',
                    'count': 0,
                    'staff': [],
                    'matching_grades': {},
                    'all_grades': {}
                }
                
                supernumerary_roles = ['SSCW', 'SSCWN', 'SM', 'OM']
                for role in supernumerary_roles:
                    role_result = ReportGenerator.generate_staff_by_grade_report(role)
                    if role_result.get('count', 0) > 0:
                        all_results['count'] += role_result['count']
                        all_results['staff'].extend(role_result.get('staff', []))
                        all_results['matching_grades'][role] = role_result['count']
                
                all_results['summary'] = f"Found {all_results['count']} supernumerary staff members across {len(all_results['matching_grades'])} roles: {', '.join(all_results['matching_grades'].keys())}."
                
                return {
                    'type': 'staff_by_grade',
                    'data': all_results
                }
            
            # Common grade patterns - check for exact word boundaries to avoid "sscw" matching "scw"
            # Use word boundaries (\b) for exact matching
            grade_query = None
            
            # Check for specific role codes (exact word matches with word boundaries)
            role_codes = [
                ('sscwn', 'SSCWN'),   # Must check before SSCW
                ('sscw', 'SSCW'),     # Must check before SCW
                ('scwn', 'SCWN'),     # Must check before SCW
                ('scan', 'SCAN'),     # Must check before SCA
                ('scw', 'SCW'),
                ('sca', 'SCA'),
                ('sm', 'SM'),
                ('om', 'OM'),
                ('rn', 'RN'),
            ]
            
            for code, role in role_codes:
                # Use word boundary regex for exact match
                if re.search(r'\b' + code + r'\b', query_lower):
                    grade_query = role
                    break
            
            # If no role code matched, try descriptive names
            if not grade_query:
                descriptive_patterns = {
                    'senior support care worker night': 'SSCWN',
                    'senior support worker night': 'SSCWN',
                    'senior scw night': 'SSCWN',
                    'senior support care worker': 'SSCW',
                    'senior support worker': 'SSCW',
                    'senior scw': 'SSCW',
                    'support care worker night': 'SCWN',
                    'support worker night': 'SCWN',
                    'scw night': 'SCWN',
                    'support care assistant night': 'SCAN',
                    'sca night': 'SCAN',
                    'support care worker': 'SCW',
                    'support worker': 'SCW',
                    'support care assistant': 'SCA',
                    'service manager': 'SM',
                    'operations manager': 'OM',
                    'registered nurse': 'RN',
                    'team leader': 'Team Leader',
                }
                
                # Check longer phrases first
                for phrase, role in sorted(descriptive_patterns.items(), key=lambda x: len(x[0]), reverse=True):
                    if phrase in query_lower:
                        grade_query = role
                        break
            
            # If still no match and query mentions "grade", extract it
            if not grade_query and 'grade' in query_lower:
                # Try to find "grade X" pattern
                match = re.search(r'grade\s+([a-z0-9]+)', query_lower)
                if match:
                    grade_query = match.group(1).upper()
            
            return {
                'type': 'staff_by_grade',
                'data': ReportGenerator.generate_staff_by_grade_report(grade_query)
            }
        
        # Text message generation for shortages (check BEFORE shortage detection)
        if any(word in query_lower for word in ['generate text', 'send message', 'send sms', 'sms staff', 'text all staff', 'message staff', 'text message']):
            # First check for shortages
            shortage_data = ReportGenerator.generate_staffing_shortage_report(14)
            
            return {
                'type': 'shortage_text_message',
                'shortage_data': shortage_data,
                'message_data': ReportGenerator.generate_shortage_text_message(shortage_data)
            }
        
        # Staffing shortage queries
        if any(word in query_lower for word in ['shortage', 'short staffed', 'staff gaps', 'understaffed', 'need staff']):
            # Determine timeframe
            days = 14  # Default 2 weeks
            if 'next week' in query_lower or '7 day' in query_lower:
                days = 7
            elif 'next month' in query_lower or '30 day' in query_lower:
                days = 30
            elif '14 day' in query_lower or 'two week' in query_lower:
                days = 14
            
            return {
                'type': 'staffing_shortage',
                'data': ReportGenerator.generate_staffing_shortage_report(days)
            }
        
        # Sickness queries
        if any(word in query_lower for word in ['sick', 'sickness', 'absence', 'off sick']):
            # Determine timeframe
            days = 7
            if 'today' in query_lower or 'now' in query_lower:
                days = 1
            elif 'week' in query_lower:
                days = 7
            elif 'month' in query_lower:
                days = 30
            
            return {
                'type': 'sickness_report',
                'data': ReportGenerator.generate_sickness_report(days)
            }
        
        # Incident queries
        if any(word in query_lower for word in ['incident', 'accident', 'fall', 'injury', 'injuries']):
            # Determine timeframe
            days = 7
            if 'today' in query_lower or 'now' in query_lower:
                days = 1
            elif 'week' in query_lower:
                days = 7
            elif 'month' in query_lower:
                days = 30
            
            return {
                'type': 'incident_report',
                'data': ReportGenerator.generate_incident_report(days)
            }
        
        # Coverage queries
        if any(word in query_lower for word in ['coverage', 'shifts', 'working', 'rota', 'schedule']):
            # Extract date if mentioned
            date_str = None
            if 'today' in query_lower:
                date_str = timezone.now().date().strftime('%Y-%m-%d')
            elif 'tomorrow' in query_lower:
                date_str = (timezone.now().date() + timedelta(days=1)).strftime('%Y-%m-%d')
            
            return {
                'type': 'shift_coverage',
                'data': ReportGenerator.generate_shift_coverage_report(date_str)
            }
        
        # Leave queries
        if any(word in query_lower for word in ['annual leave', 'leave request', 'holiday', 'vacation', 'leave balance']):
            return {
                'type': 'leave_summary',
                'data': ReportGenerator.generate_leave_summary()
            }
        
        return None




def _process_careplan_query(query):
    """Process care plan review queries like 'when is DEM01 review due?'"""
    from .models import Resident, CarePlanReview
    from datetime import date, timedelta
    import re
    
    query_lower = query.lower()
    
    # Check for general review statistics queries
    if any(phrase in query_lower for phrase in ['how many reviews', 'reviews this month', 'reviews this week', 'reviews due', 'overdue reviews']):
        today = date.today()
        
        # Determine time range
        if 'this month' in query_lower or 'month' in query_lower:
            start_date = today.replace(day=1)
            # Get last day of month
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            period_name = "this month"
        elif 'this week' in query_lower or 'week' in query_lower:
            # Week starts on Monday
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            period_name = "this week"
        elif 'today' in query_lower:
            start_date = today
            end_date = today
            period_name = "today"
        else:
            # Default to this month
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            period_name = "this month"
        
        # Get reviews in the date range
        reviews_in_period = CarePlanReview.objects.filter(
            due_date__gte=start_date,
            due_date__lte=end_date
        )
        
        total_reviews = reviews_in_period.count()
        completed = reviews_in_period.filter(status='COMPLETED').count()
        overdue = reviews_in_period.filter(status='OVERDUE').count()
        due_soon = reviews_in_period.filter(status='DUE').count()
        upcoming = reviews_in_period.filter(status='UPCOMING').count()
        
        # Build answer
        answer = f"""**📋 Care Plan Reviews {period_name.title()}**

**Period:** {start_date.strftime('%B %d')} - {end_date.strftime('%B %d, %Y')}

**Total Reviews Due:** {total_reviews}
• ✅ Completed: {completed} ({(completed/total_reviews*100) if total_reviews > 0 else 0:.0f}%)
• ❌ Overdue: {overdue}
• ⚠️ Due Soon: {due_soon}
• 📅 Upcoming: {upcoming}

**Compliance Rate:** {(completed/(completed+overdue+due_soon)*100) if (completed+overdue+due_soon) > 0 else 0:.0f}%"""
        
        if overdue > 0:
            answer += f"\n\n⚠️ **Action Required:** {overdue} review{'s' if overdue != 1 else ''} overdue!"
        
        answer += f"""

**Quick Actions:**
• [View All Reviews](/careplan/)
• [Compliance Report](/careplan/reports/)
• [Overdue Reviews](/careplan/?status=OVERDUE)"""
        
        careplan_score = match_intent_keywords(query, 'careplan')
        
        return {
            'answer': answer,
            'related': ['View Compliance Report', 'Care Plan Overview', 'Generate Reviews'],
            'category': 'careplan_stats',
            'data': {
                'total': total_reviews,
                'completed': completed,
                'overdue': overdue,
                'due_soon': due_soon
            },
            'confidence': calculate_confidence_score(query, {
                'answer': answer,
                'category': 'careplan_stats',
                'data': {'total': total_reviews}
            }, careplan_score)
        }
    
    # Check for overdue reviews query
    if 'overdue' in query_lower and 'review' in query_lower:
        overdue_reviews = CarePlanReview.objects.filter(status='OVERDUE').select_related('resident')
        count = overdue_reviews.count()
        
        if count == 0:
            careplan_score = match_intent_keywords(query, 'careplan')
            return {
                'answer': "✅ **Great news!** No overdue care plan reviews.\n\nAll reviews are up to date! [View All Reviews](/careplan/)",
                'related': ['Care Plan Overview', 'Compliance Report'],
                'category': 'careplan_stats',
                'confidence': calculate_confidence_score(query, {
                    'answer': "No overdue reviews",
                    'category': 'careplan_stats'
                }, careplan_score)
            }
        
        # List up to 10 overdue reviews
        answer = f"❌ **{count} Overdue Care Plan Review{'s' if count != 1 else ''}**\n\n"
        for review in overdue_reviews[:10]:
            answer += f"• **{review.resident.resident_id}** - {review.resident.full_name} ({review.resident.unit.name})\n"
            answer += f"  Due: {review.due_date.strftime('%b %d, %Y')} - **{review.days_overdue} days overdue**\n"
        
        if count > 10:
            answer += f"\n...and {count - 10} more\n"
        
        answer += f"\n[View All Overdue Reviews](/careplan/?status=OVERDUE)"
        
        careplan_score = match_intent_keywords(query, 'careplan')
        
        return {
            'answer': answer,
            'related': ['View Compliance Report', 'Complete Reviews'],
            'category': 'careplan_stats',
            'data': {'overdue_count': count},
            'confidence': calculate_confidence_score(query, {
                'answer': answer,
                'category': 'careplan_stats',
                'data': {'overdue_count': count}
            }, careplan_score)
        }
    
    # Pattern: "when is [resident ID] review due?" or "show review for [resident ID]"
    review_patterns = [
        r"(?:when\s+is\s+)?([A-Z]{3}\d{2})\s*(?:'s)?\s*(?:review|care\s*plan)",
        r"(?:review\s+for\s+|show\s+review\s+)?([A-Z]{3}\d{2})",
        r"([A-Z]{3}\d{2})\s+(?:review|due|status)",
    ]
    
    for pattern in review_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match and any(word in query_lower for word in ['review', 'due', 'care plan', 'status']):
            resident_id = match.group(1).upper()
            
            try:
                resident = Resident.objects.get(resident_id=resident_id, is_active=True)
                latest_review = resident.care_plan_reviews.order_by('-due_date').first()
                
                if latest_review:
                    # Status emoji
                    if latest_review.status == 'COMPLETED':
                        status_emoji = "✅"
                    elif latest_review.status == 'OVERDUE':
                        status_emoji = "❌"
                    elif latest_review.status == 'DUE':
                        status_emoji = "⚠️"
                    else:
                        status_emoji = "📅"
                    
                    # Build answer
                    answer = f"""**{status_emoji} Care Plan Review for {resident.resident_id}**

**Resident:** {resident.full_name}
**Unit:** {resident.unit.name} - Room {resident.room_number}
**Keyworker:** {resident.keyworker.full_name if resident.keyworker else 'Not assigned'}

**Current Review:**
• Type: {latest_review.get_review_type_display()}
• Due Date: {latest_review.due_date.strftime('%B %d, %Y')}
• Status: {latest_review.get_status_display()}"""
                    
                    if latest_review.status == 'OVERDUE':
                        answer += f"\n• **⚠️ OVERDUE by {latest_review.days_overdue} days!**"
                    elif latest_review.status == 'DUE':
                        answer += f"\n• Due in {latest_review.days_until_due} days"
                    elif latest_review.status == 'UPCOMING':
                        answer += f"\n• Due in {latest_review.days_until_due} days"
                    elif latest_review.status == 'COMPLETED':
                        answer += f"\n• Completed: {latest_review.completed_date.strftime('%B %d, %Y')}"
                        # Check for next review
                        next_review = resident.care_plan_reviews.filter(
                            status__in=['UPCOMING', 'DUE', 'OVERDUE']
                        ).order_by('due_date').first()
                        if next_review:
                            answer += f"\n• Next Review: {next_review.due_date.strftime('%B %d, %Y')}"
                    
                    answer += f"""

**Admission Date:** {resident.admission_date.strftime('%B %d, %Y')} ({resident.days_since_admission} days ago)

**Quick Actions:**
• [View Unit Dashboard](/careplan/unit/{resident.unit.name}/)
• [Complete Review](/careplan/review/{latest_review.id}/)
• [View All Reviews](/careplan/)"""
                    
                    careplan_score = match_intent_keywords(query, 'careplan')
                    
                    return {
                        'answer': answer,
                        'related': ['Complete Review', 'View Compliance Report', 'Unit Dashboard'],
                        'category': 'careplan_query',
                        'data': {
                            'resident_id': resident.resident_id,
                            'unit': resident.unit.name,
                            'review_status': latest_review.status,
                            'due_date': latest_review.due_date.isoformat()
                        },
                        'confidence': calculate_confidence_score(query, {
                            'answer': answer,
                            'category': 'careplan_query',
                            'data': {'resident_id': resident.resident_id}
                        }, careplan_score)
                    }
                else:
                    return {
                        'answer': f"**{resident.resident_id} - {resident.full_name}**\n\n❌ No care plan review found for this resident.\n\nRun: `python3 manage.py generate_careplan_reviews` to create review schedule.",
                        'related': ['Generate Reviews', 'View All Residents'],
                        'category': 'careplan_query'
                    }
            except Resident.DoesNotExist:
                return {
                    'answer': f"❌ Could not find resident: **{resident_id}**\n\nResident IDs follow the format: **XXX##** (e.g., DEM01, BLU15)\n\n[View All Residents](/careplan/)",
                    'related': ['View Residents', 'Care Plan Overview'],
                    'category': 'careplan_query'
                }
            except Exception as e:
                return None
    
    return None

# ============================================================================
# PHASE 1: AI ASSISTANT IMPROVEMENTS - Query Analytics & Better Understanding
# ============================================================================

# Synonym mapping for better intent detection
INTENT_KEYWORDS = {
    'vacancy': {
        'primary': ['vacancy', 'vacancies', 'vacant', 'unfilled'],
        'secondary': ['leaving', 'leaver', 'resigned', 'quit', 'departed', 'resignation'],
        'context': ['position', 'role', 'staff', 'post']
    },
    'staffing': {
        'primary': ['staff', 'staffing', 'working', 'on duty', 'roster', 'rota'],
        'secondary': ['who', 'schedule', 'shift', 'today', 'tonight', 'tomorrow', 'scw', 'sscw', 'senior care worker', 'care worker', 'rn', 'nurse', 'registered nurse', 'carer', 'assistant'],
        'context': ['count', 'number', 'how many', 'list', 'at', 'in']
    },
    'leave': {
        'primary': ['leave', 'holiday', 'vacation', 'time off', 'annual leave'],
        'secondary': ['balance', 'remaining', 'used', 'taken', 'booked'],
        'context': ['days', 'hours', 'allowance', 'entitlement']
    },
    'sickness': {
        'primary': ['sick', 'sickness', 'ill', 'illness', 'unwell'],
        'secondary': ['absence', 'off sick', 'absent', 'medical'],
        'context': ['certificate', 'note', 'days', 'self-certify']
    },
    'careplan': {
        'primary': ['care plan', 'careplan', 'review', 'assessment'],
        'secondary': ['resident', 'due', 'overdue', 'completed'],
        'context': ['compliance', 'rate', 'percentage']
    },
    'shortage': {
        'primary': ['shortage', 'short', 'gap', 'understaffed', 'uncovered'],
        'secondary': ['need', 'needed', 'require', 'missing'],
        'context': ['shift', 'cover', 'staff', 'rota']
    },
    'agency': {
        'primary': ['agency', 'temp', 'temporary', 'cover'],
        'secondary': ['cost', 'hours', 'booking', 'external'],
        'context': ['staff', 'worker', 'company']
    },
    'home_performance': {
        'primary': ['performance', 'quality', 'audit', 'compliance'],
        'secondary': ['orchard grove', 'riverside', 'meadowburn', 'hawthorn', 'victoria gardens'],
        'context': ['compare', 'comparison', 'dashboard', 'report']
    }
}

# Query templates with regex patterns for entity extraction
QUERY_TEMPLATES = {
    'staff_count': [
        r'^how\s+many\s+(?P<role>[\w\s]+?)\s+(?:are\s+)?(?:at|in)\s+(?P<location>[\w\s]+)',
        r'^count\s+(?P<role>[\w\s]+?)\s+(?:at|in)\s+(?P<location>[\w\s]+)',
        r'^(?P<role>[\w\s]+?)\s+count\s+(?:at|in)\s+(?P<location>[\w\s]+)',
    ],
    'staff_list': [
        r'^(?:who\s+is|who\s+are|list|show)\s+(?:the\s+)?(?P<role>[\w\s]+?)\s+(?:at|in)\s+(?P<location>[\w\s]+)',
        r'^show\s+me\s+(?:all\s+)?(?P<role>[\w\s]+)',
        r'^list\s+(?:all\s+)?(?P<role>[\w\s]+)',
    ],
    'sickness': [
        r'^(?:what\s+is|show|get)\s+(?:the\s+)?sickness\s+(?:in|at|for)\s+(?P<location>[\w\s]+)',
        r'^sickness\s+(?:report|data|stats)\s+(?:for\s+)?(?P<location>[\w\s]+)',
    ],
    'vacancy': [
        r'^how\s+many\s+(?:vacancies|vacant\s+positions)',
        r'^(?:show|list)\s+(?:all\s+)?vacancies',
    ],
    'staff_info': [
        r'^when\s+did\s+(?P<name>[\w\s]+?)\s+(?:start|commence|join)',
        r'^what\s+date\s+did\s+(?P<name>[\w\s]+?)\s+(?:start|commence)',
    ],
}

def match_query_template(query):
    """
    Match query against templates and extract entities
    Returns (template_type, confidence, entities) or (None, 0, {})
    """
    import re
    query_lower = query.lower().strip()
    
    for template_type, patterns in QUERY_TEMPLATES.items():
        for pattern in patterns:
            match = re.match(pattern, query_lower, re.IGNORECASE)
            if match:
                entities = match.groupdict()
                # Confidence based on how specific the match is
                confidence = 0.7 + (0.1 * len(entities))  # Base 0.7, +0.1 per entity
                confidence = min(confidence, 1.0)  # Cap at 1.0
                return (template_type, confidence, entities)
    
    return (None, 0.0, {})

def calculate_confidence_score(query, response_data, intent_score=None):
    """
    Calculate confidence score for a query response
    Returns float between 0.0 and 1.0
    """
    score = 0.0
    factors = []
    
    # Factor 1: Intent keyword matching (0-0.4)
    if intent_score is not None:
        intent_confidence = min(intent_score / 3.0, 0.4)  # Max 0.4 for strong match
        score += intent_confidence
        factors.append(f"intent:{intent_confidence:.2f}")
    
    # Factor 2: Entity extraction (0-0.3)
    if 'count' in response_data or 'home' in response_data or 'role' in response_data:
        entity_confidence = 0.3
        score += entity_confidence
        factors.append(f"entity:{entity_confidence:.2f}")
    
    # Factor 3: Result ambiguity (0-0.3)
    if 'answer' in response_data:
        # Check for multiple matches or clarification needed
        if 'multiple' in response_data.get('answer', '').lower():
            ambiguity_confidence = 0.1  # Low confidence for multiple matches
        else:
            ambiguity_confidence = 0.3  # High confidence for single result
        score += ambiguity_confidence
        factors.append(f"ambiguity:{ambiguity_confidence:.2f}")
    
    # Ensure score is between 0 and 1
    score = max(0.0, min(1.0, score))
    
    return score

# ============================================================================
# PHASE 3: AI ASSISTANT IMPROVEMENTS - Fuzzy Matching & Conversation Context
# ============================================================================

def fuzzy_match_staff(search_term, threshold=0.6, max_results=5):
    """
    Fuzzy match staff names using similarity scoring
    Returns list of (User, similarity_score) tuples
    """
    from difflib import SequenceMatcher
    
    search_term_lower = search_term.lower().strip()
    if not search_term_lower or len(search_term_lower) < 2:
        return []
    
    # Get all active staff
    all_staff = User.objects.filter(is_active=True).select_related('role', 'unit')
    
    matches = []
    for staff in all_staff:
        full_name = f"{staff.first_name} {staff.last_name}".lower()
        
        # Calculate similarity scores for different combinations
        full_similarity = SequenceMatcher(None, search_term_lower, full_name).ratio()
        first_similarity = SequenceMatcher(None, search_term_lower, staff.first_name.lower()).ratio()
        last_similarity = SequenceMatcher(None, search_term_lower, staff.last_name.lower()).ratio()
        
        # Take the best match
        best_similarity = max(full_similarity, first_similarity, last_similarity)
        
        if best_similarity >= threshold:
            matches.append((staff, best_similarity))
    
    # Sort by similarity (highest first) and return top results
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches[:max_results]

def fuzzy_match_home(search_term, threshold=0.6):
    """
    Fuzzy match care home names
    Returns list of (CareHome, similarity_score) tuples
    """
    from difflib import SequenceMatcher
    
    search_term_lower = search_term.lower().strip()
    if not search_term_lower or len(search_term_lower) < 2:
        return []
    
    # Get all homes
    all_homes = CareHome.objects.all()
    
    # Common abbreviations and variations
    home_variations = {
        'og': 'orchard grove',
        'hh': 'hawthorn house',
        'vg': 'victoria gardens',
        'rs': 'riverside',
        'mb': 'meadowburn',
        'orchard': 'orchard grove',
        'hawthorn': 'hawthorn house',
        'victoria': 'victoria gardens',
    }
    
    # Expand search term if it's an abbreviation
    expanded_search = home_variations.get(search_term_lower, search_term_lower)
    
    matches = []
    for home in all_homes:
        home_name = home.name.replace('_', ' ').lower()
        display_name = home.get_name_display().lower()
        
        # Calculate similarity with both name formats
        name_similarity = SequenceMatcher(None, expanded_search, home_name).ratio()
        display_similarity = SequenceMatcher(None, expanded_search, display_name).ratio()
        
        best_similarity = max(name_similarity, display_similarity)
        
        if best_similarity >= threshold:
            matches.append((home, best_similarity))
    
    # Sort by similarity (highest first)
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches

def fuzzy_match_resident(search_term, threshold=0.6, max_results=5):
    """
    Fuzzy match resident names and IDs
    Returns list of (Resident, similarity_score) tuples
    """
    from difflib import SequenceMatcher
    from scheduling.models import Resident
    
    search_term_lower = search_term.lower().strip()
    if not search_term_lower or len(search_term_lower) < 2:
        return []
    
    # Get all active residents
    all_residents = Resident.objects.filter(is_active=True).select_related('unit')
    
    matches = []
    for resident in all_residents:
        # Check resident ID (exact match gets bonus)
        if search_term_lower.upper() == resident.resident_id:
            matches.append((resident, 1.0))
            continue
        
        full_name = resident.full_name.lower()
        resident_id = resident.resident_id.lower()
        
        # Calculate similarity scores
        name_similarity = SequenceMatcher(None, search_term_lower, full_name).ratio()
        id_similarity = SequenceMatcher(None, search_term_lower, resident_id).ratio()
        
        best_similarity = max(name_similarity, id_similarity)
        
        if best_similarity >= threshold:
            matches.append((resident, best_similarity))
    
    # Sort by similarity (highest first) and return top results
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches[:max_results]

def get_conversation_context(request):
    """
    Get conversation context from session
    Returns list of previous queries with their results
    """
    if not hasattr(request, 'session'):
        return []
    
    context = request.session.get('ai_conversation_context', [])
    return context

def update_conversation_context(request, query, intent_type, entities, result):
    """
    Update conversation context in session
    Keeps last 5 queries for context
    """
    if not hasattr(request, 'session'):
        return
    
    from datetime import datetime
    
    context = request.session.get('ai_conversation_context', [])
    
    # Add new entry
    context.append({
        'query': query,
        'intent': intent_type,
        'entities': entities,
        'result_category': result.get('category', 'unknown'),
        'timestamp': datetime.now().isoformat(),
    })
    
    # Keep only last 5 queries
    context = context[-5:]
    
    request.session['ai_conversation_context'] = context
    request.session.modified = True

def resolve_context_reference(query, context):
    """
    Resolve contextual references in follow-up queries
    Returns (resolved_query, context_used) or (None, None) if no context detected
    """
    query_lower = query.lower().strip()
    
    if not context:
        return (None, None)
    
    # Get most recent context
    last_context = context[-1]
    
    # Pattern 1: "tell me more" or "more details"
    if any(phrase in query_lower for phrase in ['tell me more', 'more details', 'more info', 'elaborate', 'expand']):
        # Repeat the last query with added detail request
        last_query = last_context.get('query', '')
        return (f"{last_query} (detailed)", last_context)
    
    # Pattern 2: "what about X" or "how about X"
    import re
    what_about_match = re.search(r'(?:what|how)\s+about\s+(.+)', query_lower)
    if what_about_match and last_context.get('intent'):
        new_entity = what_about_match.group(1).strip()
        last_intent = last_context.get('intent')
        
        # Reconstruct query with new entity
        if last_intent == 'staffing' or last_intent == 'staff_count':
            # Get role from last query if present
            last_entities = last_context.get('entities', {})
            role = last_entities.get('role', 'staff')
            return (f"how many {role} at {new_entity}", last_context)
        elif last_intent == 'home_performance':
            return (f"show me {new_entity} performance", last_context)
        elif last_intent == 'sickness':
            return (f"what is sickness in {new_entity}", last_context)
    
    # Pattern 3: "and X?" - continuing from previous
    and_match = re.search(r'^and\s+(.+)', query_lower)
    if and_match and last_context.get('intent'):
        additional = and_match.group(1).strip()
        last_intent = last_context.get('intent')
        
        if last_intent in ['staffing', 'staff_count', 'home_performance']:
            # Treat as new location
            return (f"what about {additional}", last_context)
    
    # Pattern 4: Single word that might be a home name
    words = query_lower.split()
    if len(words) == 1 and len(words[0]) > 2:
        last_intent = last_context.get('intent')
        if last_intent in ['staffing', 'staff_count', 'home_performance', 'sickness']:
            # Might be asking about different home
            last_entities = last_context.get('entities', {})
            role = last_entities.get('role', 'staff')
            
            if last_intent == 'staff_count':
                return (f"how many {role} at {words[0]}", last_context)
            elif last_intent == 'home_performance':
                return (f"show me {words[0]} performance", last_context)
            elif last_intent == 'sickness':
                return (f"what is sickness in {words[0]}", last_context)
    
    return (None, None)

def match_intent_keywords(query, intent_type):
    """Score how well a query matches an intent using synonym mapping"""
    query_lower = query.lower()
    keywords = INTENT_KEYWORDS.get(intent_type, {})
    
    score = 0.0
    
    # Primary keywords = high weight
    for keyword in keywords.get('primary', []):
        if keyword in query_lower:
            score += 1.0
    
    # Secondary keywords = medium weight
    for keyword in keywords.get('secondary', []):
        if keyword in query_lower:
            score += 0.5
    
    # Context keywords = low weight (confirms intent)
    for keyword in keywords.get('context', []):
        if keyword in query_lower:
            score += 0.2
    
    return score

def generate_helpful_suggestions(query, failed_type=None):
    """Generate contextual suggestions when queries fail"""
    query_lower = query.lower()
    suggestions = []
    
    # Detect what user might be asking about based on keywords
    if any(word in query_lower for word in ['staff', 'worker', 'carer', 'employee']):
        suggestions.extend([
            "Who is working today?",
            "Show me all senior carers",
            "How many staff do we have?"
        ])
    
    if any(word in query_lower for word in ['leave', 'holiday', 'vacation', 'time off']):
        suggestions.extend([
            "Show leave requests for this week",
            "How much leave does [Name] have?",
            "List approved leave for December"
        ])
    
    if any(word in query_lower for word in ['sick', 'illness', 'absent', 'unwell']):
        suggestions.extend([
            "Who is off sick?",
            "Show sickness report for this month",
            "List current sickness absences"
        ])
    
    if any(word in query_lower for word in ['vacancy', 'vacant', 'leaving', 'leaver']):
        suggestions.extend([
            "How many staff vacancies?",
            "Show upcoming leavers",
            "List current vacancies"
        ])
    
    if any(word in query_lower for word in ['care plan', 'review', 'resident']):
        suggestions.extend([
            "Show care plan compliance",
            "When is [ResidentID] review due?",
            "List overdue care plan reviews"
        ])
    
    # If no specific topic detected, show general options
    if not suggestions:
        suggestions = [
            "Who is working today?",
            "Show staffing summary",
            "How many staff vacancies?",
            "Show care plan compliance",
            "List leave requests"
        ]
    
    # Limit to 5 suggestions
    return suggestions[:5]

def log_ai_query(query, success, response_type=None, error_message=None, user=None, response_time_ms=None):
    """Log AI assistant query for analytics"""
    from .models import AIQueryLog
    try:
        AIQueryLog.objects.create(
            query=query,
            success=success,
            response_type=response_type,
            error_message=error_message,
            user=user,
            response_time_ms=response_time_ms
        )
    except Exception as e:
        # Don't let logging errors break the assistant
        print(f"Warning: Failed to log AI query: {e}")

def _process_staff_count_by_role_query(query):
    """
    Process "how many [role] at [home]" queries
    Returns staff count by role and home, or None if not this type of query
    """
    from scheduling.models import User, CareHome
    import re
    
    query_lower = query.lower()
    
    # Check if this is a "how many [role]" query
    if not any(phrase in query_lower for phrase in ['how many', 'count', 'number of']):
        return None
    
    # Map common role terms to database role values
    role_map = {
        # Day shift roles
        'scw': 'SCW',
        'social care worker': 'SCW',
        'care worker': 'SCW',
        'sscw': 'SSCW',
        'senior social care worker': 'SSCW',
        'senior care worker': 'SSCW',
        'sca': 'SCA',
        'social care assistant': 'SCA',
        'care assistant': 'SCA',
        'sm': 'SM',
        'service manager': 'SM',
        'manager': 'SM',
        'om': 'OM',
        'operations manager': 'OM',
        # Night shift roles
        'scwn': 'SCWN',
        'night care worker': 'SCWN',
        'night scw': 'SCWN',
        'sscwn': 'SSCWN',
        'night senior care worker': 'SSCWN',
        'senior night care worker': 'SSCWN',
        'night sscw': 'SSCWN',
        'scan': 'SCAN',
        'night care assistant': 'SCAN',
        'night sca': 'SCAN',
        # Generic terms
        'carer': 'SCW',
        'senior carer': 'SSCW',
        'nursing assistant': 'SCA',
        'hca': 'SCA',
        'healthcare assistant': 'SCA'
    }
    
    # Map home names
    home_map = {
        'orchard grove': 'ORCHARD_GROVE',
        'og': 'ORCHARD_GROVE',
        'riverside': 'RIVERSIDE',
        'meadowburn': 'MEADOWBURN',
        'hawthorn': 'HAWTHORN_HOUSE',
        'hawthorn house': 'HAWTHORN_HOUSE',
        'victoria gardens': 'VICTORIA_GARDENS',
        'vg': 'VICTORIA_GARDENS'
    }
    
    # Try to extract role and home from query
    found_role = None
    found_home = None
    is_night_shift = False
    
    # Check for night shift indicators
    if any(term in query_lower for term in ['night', 'nightshift', 'night shift', 'nights']):
        is_night_shift = True
    
    # Sort role terms by length (longest first) to avoid substring matches
    # e.g., check "sscw" before "scw" so "sscw" doesn't match "scw"
    sorted_role_terms = sorted(role_map.items(), key=lambda x: len(x[0]), reverse=True)
    
    for role_term, role_code in sorted_role_terms:
        if role_term in query_lower:
            found_role = role_code
            break
    
    # If night shift is specified, convert to night role variant
    if is_night_shift and found_role:
        night_role_map = {
            'SCW': 'SCWN',
            'SSCW': 'SSCWN',
            'SCA': 'SCAN'
        }
        found_role = night_role_map.get(found_role, found_role)
    
    for home_term, home_code in home_map.items():
        if home_term in query_lower:
            found_home = home_code
            break
    
    # If we found a role, count staff
    if found_role:
        query_filter = {'role__name': found_role, 'is_active': True}
        if found_home:
            query_filter['unit__care_home__name'] = found_home
        
        staff_count = User.objects.filter(**query_filter).count()
        
        # Get home display name
        if found_home:
            try:
                home_obj = CareHome.objects.get(name=found_home)
                home_display = home_obj.get_name_display()
            except:
                home_display = found_home.replace('_', ' ').title()
            location_text = f" at {home_display}"
        else:
            location_text = " across all homes"
        
        # Get role display name  
        role_display = found_role
        
        result = {
            'answer': f'There are **{staff_count} {role_display}** staff{location_text}.',
            'type': 'staff_count',
            'count': staff_count,
            'role': found_role,
            'home': found_home
        }
        
        # Calculate confidence score
        intent_score = match_intent_keywords(query, 'staffing')
        result['confidence'] = calculate_confidence_score(query, result, intent_score)
        
        return result
    
    return None

def _process_staff_list_by_role_query(query):
    """
    Process "who is [role]" or "list [role]" queries
    Returns list of staff names by role and home, or None if not this type of query
    """
    from scheduling.models import User, CareHome
    
    query_lower = query.lower()
    
    # Check if this is a "who is" or "list" query
    if not any(phrase in query_lower for phrase in ['who is', 'who are', 'list', 'show me']):
        return None
    
    # Map common role terms to database role values (same as count function)
    role_map = {
        # Day shift roles
        'scw': 'SCW',
        'social care worker': 'SCW',
        'care worker': 'SCW',
        'sscw': 'SSCW',
        'senior social care worker': 'SSCW',
        'senior care worker': 'SSCW',
        'sca': 'SCA',
        'social care assistant': 'SCA',
        'care assistant': 'SCA',
        'sm': 'SM',
        'service manager': 'SM',
        'manager': 'SM',
        'om': 'OM',
        'operations manager': 'OM',
        # Night shift roles
        'scwn': 'SCWN',
        'night care worker': 'SCWN',
        'night scw': 'SCWN',
        'sscwn': 'SSCWN',
        'night senior care worker': 'SSCWN',
        'senior night care worker': 'SSCWN',
        'night sscw': 'SSCWN',
        'scan': 'SCAN',
        'night care assistant': 'SCAN',
        'night sca': 'SCAN',
        # Generic terms
        'carer': 'SCW',
        'senior carer': 'SSCW',
        'nursing assistant': 'SCA',
        'hca': 'SCA',
        'healthcare assistant': 'SCA'
    }
    
    # Map home names
    home_map = {
        'orchard grove': 'ORCHARD_GROVE',
        'og': 'ORCHARD_GROVE',
        'riverside': 'RIVERSIDE',
        'meadowburn': 'MEADOWBURN',
        'hawthorn': 'HAWTHORN_HOUSE',
        'hawthorn house': 'HAWTHORN_HOUSE',
        'victoria gardens': 'VICTORIA_GARDENS',
        'vg': 'VICTORIA_GARDENS'
    }
    
    # Try to extract role and home from query
    found_role = None
    found_home = None
    is_night_shift = False
    
    # Check for night shift indicators
    if any(term in query_lower for term in ['night', 'nightshift', 'night shift', 'nights']):
        is_night_shift = True
    
    # Sort role terms by length (longest first) to avoid substring matches
    sorted_role_terms = sorted(role_map.items(), key=lambda x: len(x[0]), reverse=True)
    
    for role_term, role_code in sorted_role_terms:
        if role_term in query_lower:
            found_role = role_code
            break
    
    # If night shift is specified, convert to night role variant
    if is_night_shift and found_role:
        night_role_map = {
            'SCW': 'SCWN',
            'SSCW': 'SSCWN',
            'SCA': 'SCAN'
        }
        found_role = night_role_map.get(found_role, found_role)
    
    for home_term, home_code in home_map.items():
        if home_term in query_lower:
            found_home = home_code
            break
    
    # If we found a role, get staff list
    if found_role:
        query_filter = {'role__name': found_role, 'is_active': True}
        if found_home:
            query_filter['unit__care_home__name'] = found_home
        
        staff_list = User.objects.filter(**query_filter).values_list('first_name', 'last_name', 'sap')
        
        if not staff_list:
            return None
        
        # Get home display name
        if found_home:
            try:
                home_obj = CareHome.objects.get(name=found_home)
                home_display = home_obj.get_name_display()
            except:
                home_display = found_home.replace('_', ' ').title()
            location_text = f" at {home_display}"
        else:
            location_text = " across all homes"
        
        # Format staff names
        staff_names = [f"**{first} {last}** ({sap})" for first, last, sap in staff_list]
        
        # Get role display name
        role_display = found_role
        
        if len(staff_names) == 1:
            answer = f"The {role_display}{location_text} is:\n\n{staff_names[0]}"
        else:
            answer = f"There are **{len(staff_names)} {role_display}** staff{location_text}:\n\n" + "\n".join([f"{i+1}. {name}" for i, name in enumerate(staff_names)])
        
        result = {
            'answer': answer,
            'type': 'staff_list',
            'count': len(staff_names),
            'role': found_role,
            'home': found_home,
            'staff': [{'name': f"{first} {last}", 'sap': sap} for first, last, sap in staff_list]
        }
        
        # Calculate confidence score
        intent_score = match_intent_keywords(query, 'staffing')
        result['confidence'] = calculate_confidence_score(query, result, intent_score)
        
        return result
    
    return None

def _process_sickness_query(query):
    """
    Process sickness/absence queries for specific homes or overall
    Returns sickness data by home, or None if not a sickness query
    """
    from staff_records.models import SicknessRecord
    from scheduling.models import CareHome
    from datetime import timedelta
    
    query_lower = query.lower()
    
    # Check if this is a sickness query using synonym matching
    sickness_score = match_intent_keywords(query, 'sickness')
    
    if sickness_score < 0.5:
        return None
    
    # Map home names
    home_map = {
        'orchard grove': 'ORCHARD_GROVE',
        'og': 'ORCHARD_GROVE',
        'riverside': 'RIVERSIDE',
        'meadowburn': 'MEADOWBURN',
        'hawthorn': 'HAWTHORN_HOUSE',
        'hawthorn house': 'HAWTHORN_HOUSE',
        'victoria gardens': 'VICTORIA_GARDENS',
        'vg': 'VICTORIA_GARDENS'
    }
    
    # Try to extract home from query
    found_home = None
    for home_term, home_code in home_map.items():
        if home_term in query_lower:
            found_home = home_code
            break
    
    # Get sickness data for last 30 days
    from django.utils import timezone
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    if found_home:
        # Specific home sickness data
        try:
            home_obj = CareHome.objects.get(name=found_home)
            home_display = home_obj.get_name_display()
        except:
            home_display = found_home.replace('_', ' ').title()
        
        sickness_records = SicknessRecord.objects.filter(
            profile__user__unit__care_home__name=found_home,
            first_working_day__lte=end_date,
            status__in=['OPEN', 'AWAITING_FIT_NOTE']
        ).select_related('profile', 'profile__user', 'profile__user__role')
        
        # Count open records and total days
        total_records = sickness_records.count()
        total_days = sum(r.total_working_days_sick for r in sickness_records)
        current_sick = sickness_records.filter(status='OPEN').count()
        
        answer = f"**🏥 Sickness Absence - {home_display}**\n\n"
        answer += f"**Last 30 days ({start_date.strftime('%d %b')} - {end_date.strftime('%d %b %Y')}):**\n"
        answer += f"• Open sickness records: {total_records}\n"
        answer += f"• Total working days lost: {total_days}\n"
        answer += f"• Currently off sick (OPEN status): {current_sick} staff\n"
        
        if current_sick > 0:
            current_list = sickness_records.filter(status='OPEN')
            answer += f"\n**Currently absent:**\n"
            for record in current_list[:10]:
                user = record.profile.user
                role_name = user.role.name if user.role else 'N/A'
                answer += f"• {user.full_name} ({role_name}) - Since {record.first_working_day.strftime('%d %b')}\n"
        
        result = {
            'answer': answer,
            'type': 'sickness',
            'home': found_home,
            'total_records': total_records,
            'total_days': total_days,
            'current_sick': current_sick
        }
        
        # Calculate confidence score
        sickness_score = match_intent_keywords(query, 'sickness')
        result['confidence'] = calculate_confidence_score(query, result, sickness_score)
        
        return result
    else:
        # All homes sickness summary
        all_homes = CareHome.objects.all()
        results = []
        
        for home in all_homes:
            sickness_records = SicknessRecord.objects.filter(
                profile__user__unit__care_home=home,
                first_working_day__lte=end_date,
                status__in=['OPEN', 'AWAITING_FIT_NOTE']
            )
            
            total_days = sum(r.total_working_days_sick for r in sickness_records)
            current_sick = sickness_records.filter(status='OPEN').count()
            
            results.append(f"• **{home.get_name_display()}**: {sickness_records.count()} records, {total_days} days lost, {current_sick} currently sick")
        
        answer = f"**🏥 Sickness Absence - All Homes**\n\n"
        answer += f"**Last 30 days ({start_date.strftime('%d %b')} - {end_date.strftime('%d %b %Y')}):**\n\n"
        answer += "\n".join(results)
        
        return {
            'answer': answer,
            'type': 'sickness_all',
        }

def _process_vacancy_query(query):
    """Process queries about staff vacancies (leavers)"""
    from staff_records.models import StaffProfile
    from datetime import date, timedelta
    
    query_lower = query.lower()
    
    # Use synonym matching instead of hard-coded keywords
    vacancy_score = match_intent_keywords(query, 'vacancy')
    if vacancy_score < 0.5:  # Threshold for intent match
        return None
    
    # Get all vacancies (employment_status='LEAVER')
    all_vacancies = StaffProfile.objects.filter(
        employment_status='LEAVER'
    ).select_related('user__role', 'user__unit__care_home').order_by('end_date')
    
    today = date.today()
    
    # Split into current vacancies and upcoming leavers
    current_vacancies = []
    upcoming_leavers = []
    
    for profile in all_vacancies:
        if profile.end_date:
            if profile.end_date < today:
                days_vacant = (today - profile.end_date).days
                current_vacancies.append({
                    'name': f"{profile.user.first_name} {profile.user.last_name}",
                    'sap': profile.user.sap,
                    'role': profile.user.role.name if profile.user.role else 'N/A',
                    'home': profile.user.unit.care_home.name if profile.user.unit else 'N/A',
                    'end_date': profile.end_date,
                    'days_vacant': days_vacant,
                    'hours': getattr(profile.user, 'hours_per_week', 37.5)
                })
            else:
                days_until = (profile.end_date - today).days
                upcoming_leavers.append({
                    'name': f"{profile.user.first_name} {profile.user.last_name}",
                    'sap': profile.user.sap,
                    'role': profile.user.role.name if profile.user.role else 'N/A',
                    'home': profile.user.unit.care_home.name if profile.user.unit else 'N/A',
                    'end_date': profile.end_date,
                    'days_until': days_until,
                    'hours': getattr(profile.user, 'hours_per_week', 37.5)
                })
    
    # Build response
    total = len(current_vacancies) + len(upcoming_leavers)
    
    if total == 0:
        answer = "✅ **No Staff Vacancies**\n\nThere are currently no vacant positions or upcoming leavers recorded in the system."
    else:
        answer = f"**📊 Staff Vacancies Report**\n\n"
        answer += f"**Total: {total}** ({len(current_vacancies)} current + {len(upcoming_leavers)} upcoming)\n\n"
        
        if current_vacancies:
            answer += f"**🚨 Current Vacancies ({len(current_vacancies)}):**\n"
            for v in current_vacancies[:10]:
                severity = "🔴 HIGH" if v['days_vacant'] > 30 else "🟡 MEDIUM" if v['days_vacant'] > 14 else "🟢 LOW"
                answer += f"\n• **{v['name']}** (SAP: {v['sap']})\n"
                answer += f"  Role: {v['role']} | Home: {v['home']}\n"
                answer += f"  Left: {v['end_date'].strftime('%d %b %Y')} ({v['days_vacant']} days ago) | {severity}\n"
            
            if len(current_vacancies) > 10:
                answer += f"\n... and {len(current_vacancies) - 10} more\n"
        
        if upcoming_leavers:
            answer += f"\n**📅 Upcoming Leavers ({len(upcoming_leavers)}):**\n"
            for v in upcoming_leavers[:10]:
                answer += f"\n• **{v['name']}** (SAP: {v['sap']})\n"
                answer += f"  Role: {v['role']} | Home: {v['home']}\n"
                answer += f"  Leaving: {v['end_date'].strftime('%d %b %Y')} (in {v['days_until']} days)\n"
            
            if len(upcoming_leavers) > 10:
                answer += f"\n... and {len(upcoming_leavers) - 10} more\n"
    
    return {
        'answer': answer,
        'related': ['View Senior Dashboard', 'Recruitment Status', 'Staff Records'],
        'category': 'vacancy_report',
        'report_data': {
            'total_vacancies': total,
            'current_vacant': len(current_vacancies),
            'upcoming_leavers': len(upcoming_leavers),
            'vacancies': current_vacancies,
            'leavers': upcoming_leavers
        },
        'confidence': calculate_confidence_score(query, {
            'answer': answer,
            'category': 'vacancy_report',
            'report_data': {'total_vacancies': total}
        }, vacancy_score)
    }

def _process_staff_query(query):
    """Process staff-specific queries like leave balance, staff search, etc."""
    from staff_records.models import AnnualLeaveEntitlement, StaffProfile
    import re
    
    query_lower = query.lower()
    
    # Pattern -1: "What date did [name] commence/start/join" - staff commencement date queries
    commence_patterns = [
        r"(?:what\s+date\s+did|when\s+did)\s+([A-Za-z0-9\s]+?)\s+(?:commence|start|join|begin)",
        r"([A-Za-z0-9\s]+?)\s+(?:commencement|start|join)\s+date",
        r"(?:show|get|find)\s+(?:commencement|start|join)\s+date\s+for\s+([A-Za-z0-9\s]+)",
    ]
    
    for pattern in commence_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            identifier = match.group(1).strip()
            
            # Check if identifier contains a SAP number (numeric code)
            sap_match = re.search(r'\b(\d{6,})\b', identifier)
            
            if sap_match:
                # SAP number found - use it for exact match
                sap = sap_match.group(1)
                matching_staff = User.objects.filter(sap=sap, is_active=True).select_related('role', 'unit')
            else:
                # No SAP - search by name only
                name_parts = identifier.split()
                q = Q()
                for part in name_parts:
                    if len(part) >= 2:  # At least 2 characters
                        q |= Q(first_name__icontains=part) | Q(last_name__icontains=part)
                
                matching_staff = User.objects.filter(q, is_active=True).select_related('role', 'unit')
                
                # PHASE 3: If no exact matches, try fuzzy matching
                if matching_staff.count() == 0:
                    fuzzy_matches = fuzzy_match_staff(identifier, threshold=0.65, max_results=5)
                    if fuzzy_matches:
                        # Build suggestion message with fuzzy matches
                        suggestions = [f"• **{staff.full_name}** ({staff.sap}) - {staff.role.name if staff.role else 'N/A'}" for staff, similarity in fuzzy_matches]
                        answer = f"**🔍 No exact match for '{identifier}', did you mean:**\n\n" + "\n".join(suggestions)
                        answer += "\n\n💡 Try using one of these names or their SAP number."
                        
                        staff_score = match_intent_keywords(query, 'staff')
                        return {
                            'answer': answer,
                            'related': ['Staff Directory', 'View All Staff'],
                            'category': 'staff_search',
                            'confidence': calculate_confidence_score(query, {'answer': answer, 'category': 'staff_search'}, staff_score * 0.7)  # Lower confidence for fuzzy matches
                        }
            
            if matching_staff.count() == 1:
                staff = matching_staff.first()
                
                # Try to get commencement date from StaffProfile
                try:
                    profile = StaffProfile.objects.get(user_id=staff.sap)
                    if profile.created_at:
                        commence_date = profile.created_at.strftime('%d %B %Y')
                        answer = f"**📅 Commencement Date for {staff.full_name}**\n\n"
                        answer += f"**Name:** {staff.full_name}\n"
                        answer += f"**SAP:** {staff.sap}\n"
                        answer += f"**Role:** {staff.role.name if staff.role else 'N/A'}\n"
                        answer += f"**Unit:** {staff.unit.get_name_display() if staff.unit else 'N/A'}\n"
                        answer += f"**Commenced:** {commence_date}\n"
                        
                        staff_score = match_intent_keywords(query, 'staff')
                        
                        return {
                            'answer': answer,
                            'related': ['View Staff Profile', 'Staff Directory'],
                            'category': 'staff_info',
                            'data': {'sap': staff.sap, 'name': staff.full_name, 'commenced': commence_date},
                            'confidence': calculate_confidence_score(query, {
                                'answer': answer,
                                'category': 'staff_info',
                                'data': {'sap': staff.sap}
                            }, staff_score)
                        }
                    else:
                        return {
                            'answer': f"**ℹ️ Commencement date not available for {staff.full_name}** ({staff.sap})\n\nPlease check their staff profile or contact HR.",
                            'related': ['View Staff Profile', 'Contact HR'],
                            'category': 'staff_info'
                        }
                except StaffProfile.DoesNotExist:
                    staff_score = match_intent_keywords(query, 'staff')
                    return {
                        'answer': f"**ℹ️ No profile found for {staff.full_name}** ({staff.sap})\n\nPlease contact HR for commencement details.",
                        'related': ['Contact HR', 'Staff Directory'],
                        'category': 'staff_info',
                        'confidence': calculate_confidence_score(query, {'answer': 'No profile', 'category': 'staff_info'}, staff_score)
                    }
            elif matching_staff.count() > 1:
                # Multiple matches - ask for clarification
                results = [f"• **{s.full_name}** ({s.sap}) - {s.role.name if s.role else 'N/A'}" for s in matching_staff[:5]]
                answer = f"**🔍 Multiple staff members found matching '{identifier}':**\n\n" + "\n".join(results)
                answer += "\n\nPlease be more specific or use their SAP number."
                
                staff_score = match_intent_keywords(query, 'staff')
                
                return {
                    'answer': answer,
                    'related': ['Staff Directory'],
                    'category': 'staff_search',
                    'confidence': calculate_confidence_score(query, {'answer': answer, 'category': 'staff_search'}, staff_score)
                }
            else:
                staff_score = match_intent_keywords(query, 'staff')
                return {
                    'answer': f"**❌ No staff member found matching '{identifier}'**\n\nTry:\n• Checking the spelling\n• Using their full name\n• Using their SAP number",
                    'related': ['Staff Directory', 'View All Staff'],
                    'category': 'staff_search',
                    'confidence': calculate_confidence_score(query, {'answer': 'Not found', 'category': 'staff_search'}, staff_score)
                }
    
    # Pattern 0: "List all X" or "Show all X" where X is a role
    role_list_patterns = [
        r"(?:list|show)\s+(?:all\s+)?(?:the\s+)?(senior\s+(?:social\s+)?care\s+workers?|social\s+care\s+workers?|social\s+care\s+assistants?|senior\s+carers?|care\s+workers?|care\s+assistants?|night\s+staff|day\s+staff|managers?|admins?|support\s+workers?|assistants?)",
        r"(?:who\s+are\s+(?:the\s+)?|get\s+(?:all\s+)?)(senior\s+(?:social\s+)?care\s+workers?|social\s+care\s+workers?|social\s+care\s+assistants?|senior\s+carers?|care\s+workers?|care\s+assistants?|night\s+staff|day\s+staff|managers?|admins?|support\s+workers?)",
    ]
    
    for pattern in role_list_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            role_term = match.group(1).lower()
            
            # Map common terms to actual role codes/names in database
            # Database codes: SCA, SCW, SSCW (Day shift) and SCAN, SCWN, SSCWN (Night shift - "N" suffix)
            # OM=Operations Manager, SM=Service Manager, ADMIN=Admin
            role_mapping = {
                # Senior roles (day and night)
                'senior social care worker': ['SSCW', 'SSCWN'],  # Both day and night
                'senior social care workers': ['SSCW', 'SSCWN'],
                'senior carer': ['SSCW', 'SSCWN'],
                'senior carers': ['SSCW', 'SSCWN'],
                
                # Regular care workers (day and night)
                'social care worker': ['SCW', 'SCWN'],  # Both day and night
                'social care workers': ['SCW', 'SCWN'],
                'care worker': ['SCW', 'SCWN', 'SSCW', 'SSCWN'],  # All care worker types
                'care workers': ['SCW', 'SCWN', 'SSCW', 'SSCWN'],
                
                # Care assistants (day and night)
                'social care assistant': ['SCA', 'SCAN'],  # Both day and night
                'social care assistants': ['SCA', 'SCAN'],
                'care assistant': ['SCA', 'SCAN'],
                'care assistants': ['SCA', 'SCAN'],
                'assistant': ['SCA', 'SCAN'],
                'assistants': ['SCA', 'SCAN'],
                
                # Night staff (all night shift roles)
                'night staff': ['SCAN', 'SCWN', 'SSCWN'],  # All night shift variants
                
                # Day staff (all day shift roles)
                'day staff': ['SCA', 'SCW', 'SSCW'],  # All day shift variants
                
                # Management
                'manager': ['OM', 'SM', 'OPERATIONS_MANAGER'],
                'managers': ['OM', 'SM', 'OPERATIONS_MANAGER'],
                
                # Admin
                'admin': ['ADMIN'],
                'admins': ['ADMIN'],
                
                # Support workers
                'support worker': ['SCW', 'SCWN', 'SSCW', 'SSCWN'],
                'support workers': ['SCW', 'SCWN', 'SSCW', 'SSCWN'],
            }
            
            role_names = role_mapping.get(role_term, [])
            
            if role_names:
                # Find staff with these roles
                staff = User.objects.filter(
                    role__name__in=role_names,
                    is_active=True
                ).select_related('role', 'unit').order_by('last_name', 'first_name')
                
                if staff.exists():
                    results = []
                    for s in staff:
                        unit_name = s.unit.name if s.unit else 'No Unit'
                        results.append(f"• **{s.full_name}** ({s.sap}) - {unit_name}")
                    
                    answer = f"**👥 {role_term.title()} ({staff.count()} total)**\n\n" + "\n".join(results)
                    
                    staffing_score = match_intent_keywords(query, 'staffing')
                    
                    return {
                        'answer': answer,
                        'related': ['View All Staff', 'Staff by Unit'],
                        'category': 'staff_list',
                        'data': {'role': role_term, 'count': staff.count()},
                        'confidence': calculate_confidence_score(query, {'answer': answer, 'category': 'staff_list', 'data': {'count': staff.count()}}, staffing_score)
                    }
                else:
                    staffing_score = match_intent_keywords(query, 'staffing')
                    return {
                        'answer': f"❌ No active {role_term} found.\n\nTry:\n• Different role name\n• View all staff grades",
                        'related': ['View All Staff', 'Show All Grades'],
                        'category': 'staff_list',
                        'confidence': calculate_confidence_score(query, {'answer': 'None found', 'category': 'staff_list'}, staffing_score)
                    }
    
    # Pattern 0.5: "Who is on X shift today/tonight/tomorrow?" or "Who is working X shift?"
    shift_patterns = [
        r"who\s+(?:is|are)\s+(?:on|working)\s+(?:the\s+)?(day|night|morning|evening|late)\s+shift\s+(?:today|tonight|tomorrow)?",
        r"(?:show|list)\s+(?:all\s+)?(?:the\s+)?(day|night|morning|evening|late)\s+shift\s+(?:staff|workers?)?",
        r"who\s+(?:is|are)\s+working\s+(?:today|tonight|tomorrow)\s+(?:on\s+)?(day|night|morning|evening|late)?",
    ]
    
    for pattern in shift_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            shift_type = match.group(1).lower() if match.lastindex >= 1 else 'day'
            
            # Determine date
            from datetime import timedelta
            today = timezone.now().date()
            
            if 'tomorrow' in query_lower:
                target_date = today + timedelta(days=1)
                day_name = "Tomorrow"
            elif 'tonight' in query_lower or 'today' in query_lower:
                target_date = today
                day_name = "Today"
            else:
                target_date = today
                day_name = "Today"
            
            # Map shift type to ShiftType.name values
            # ShiftType.name choices: DAY_SENIOR, DAY_ASSISTANT, NIGHT_SENIOR, NIGHT_ASSISTANT, ADMIN
            if shift_type in ['night', 'tonight']:
                # Query for NIGHT_SENIOR and NIGHT_ASSISTANT
                shift_names = ['NIGHT_SENIOR', 'NIGHT_ASSISTANT']
                shift_display = 'Night'
            else:  # day, morning, evening, late
                # Query for DAY_SENIOR and DAY_ASSISTANT
                shift_names = ['DAY_SENIOR', 'DAY_ASSISTANT']
                shift_display = 'Day'
            
            # Query shifts for this date and type
            from scheduling.models import Shift
            shifts = Shift.objects.filter(
                date=target_date,
                shift_type__name__in=shift_names
            ).select_related('user', 'user__role', 'user__unit', 'unit', 'shift_type').order_by('unit__name', 'user__last_name')
            
            if shifts.exists():
                results = []
                current_unit = None
                
                for shift in shifts:
                    unit_name = shift.unit.name if shift.unit else 'No Unit'
                    
                    # Group by unit
                    if unit_name != current_unit:
                        if current_unit is not None:
                            results.append("")  # Add blank line between units
                        results.append(f"**{unit_name}:**")
                        current_unit = unit_name
                    
                    user = shift.user
                    role_name = user.role.name if user.role else 'No Role'
                    time_range = f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}" if shift.start_time and shift.end_time else ''
                    
                    results.append(f"• {user.full_name} ({role_name}) {time_range}")
                
                answer = f"""**🗓️ {shift_display} Shift - {day_name} ({target_date.strftime('%d %B %Y')})**

**{shifts.count()} staff member(s) scheduled:**

""" + "\n".join(results)
                
                staffing_score = match_intent_keywords(query, 'staffing')
                
                return {
                    'answer': answer,
                    'related': ['Coverage Report', 'View All Shifts', 'Staffing Levels'],
                    'category': 'shift_query',
                    'data': {'date': str(target_date), 'shift_type': shift_display, 'count': shifts.count()},
                    'confidence': calculate_confidence_score(query, {'answer': answer, 'category': 'shift_query', 'data': {'count': shifts.count()}}, staffing_score)
                }
            else:
                return {
                    'answer': f"""**ℹ️ No {shift_display} Shift Staff - {day_name}**

**Date:** {target_date.strftime('%d %B %Y')}

No staff scheduled for {shift_display.lower()} shift on this date.

**Possible reasons:**
• Rota not yet generated for this date
• No staff allocated to {shift_display.lower()} shift
• Date is outside current rota period

**Next steps:**
• [View Coverage Report](/reports/coverage/)
• [Generate New Rota](/rota/generate/)
• [Check Staffing Levels](/manager-dashboard/)""",
                    'related': ['Generate Rota', 'Coverage Report', 'View Staff'],
                    'category': 'shift_query'
                }
    
    # Pattern 1: "How much leave does X have?" or "X leave remaining" or "Show my leave balance"
    # BUT NOT "Who is on leave" or "Show all on leave" or "Annual leave pending approvals"
    leave_patterns = [
        r"(?:how\s+much\s+)?leave\s+(?:does\s+)?([A-Z0-9]+|[A-Za-z\s]+?)\s+have",
        r"([A-Z0-9]+|[A-Za-z\s]+?)\s+(?:leave\s+)?(?:remaining|left|balance)",
        r"(?:show\s+)?(?:my|check\s+my)\s+leave\s+(?:balance|remaining|left)",
        r"(?:check\s+)?(?:leave\s+for\s+)?([A-Z0-9]{3,})",  # SAP IDs are at least 3 chars
    ]
    
    # Skip if query is about who is on leave, pending approvals, or other leave-related reports
    skip_phrases = [
        'who is on leave', 'on leave this', 'on leave next', 'people on leave',
        'pending approval', 'pending leave', 'leave approval', 'annual leave pending',
        'leave request', 'approve leave'
    ]
    
    if any(phrase in query_lower for phrase in skip_phrases):
        pass  # Skip leave balance patterns
    else:
        for pattern in leave_patterns:
            match = re.search(pattern, query, re.IGNORECASE)
            if match and any(word in query_lower for word in ['leave', 'remaining', 'balance', 'holiday']):
                # For "show my leave" pattern, we need to identify the current user
                if 'my leave' in query_lower:
                    identifier = 'current_user'  # Placeholder - in real implementation would get from request.user
                else:
                    identifier = match.group(1).strip() if match.lastindex >= 1 else None
                
                if not identifier:
                    continue
                    
                # Try to find staff by SAP or name
                try:
                    if identifier == 'current_user':
                        # This would need request.user passed to this function
                        # For now, return helpful message
                        return {
                            'answer': "**ℹ️ Personal Leave Balance**\n\nTo check your leave balance, please:\n• Log in to the system\n• Visit your staff profile\n• Or ask: \"How much leave does [YOUR NAME] have?\"",
                            'related': ['My Profile', 'Request Leave'],
                            'category': 'staff_query'
                        }
                    elif identifier.isupper() and len(identifier) < 10:  # Likely SAP
                        user = User.objects.filter(sap__iexact=identifier).first()
                    else:  # Likely name
                        # Search by name parts
                        name_parts = identifier.split()
                        q = Q()
                        for part in name_parts:
                            q |= Q(first_name__icontains=part) | Q(last_name__icontains=part)
                        user = User.objects.filter(q).first()
                    
                    if user:
                        try:
                            profile = StaffProfile.objects.get(user=user)
                            entitlement = AnnualLeaveEntitlement.objects.get(
                                profile=profile,
                                leave_year_start__year=timezone.now().year
                            )
                            
                            days = float(entitlement.days_remaining)
                            hours = float(entitlement.hours_remaining)
                            used_hours = float(entitlement.hours_used)
                            total_hours = float(entitlement.total_entitlement_hours)
                            pending = float(entitlement.hours_pending)
                            
                            # Determine urgency
                            if days < 5:
                                urgency = "🔴 **URGENT**"
                            elif days < 10:
                                urgency = "🟡 **MODERATE**"
                            else:
                                urgency = "🟢 **LOW**"
                            
                            answer = f"""**📊 Leave Balance for {user.full_name} ({user.sap})**

**Remaining Leave:** {hours:.1f} hours ({days:.1f} days)

**Full Breakdown:**
• Total Entitlement: {total_hours:.1f} hours
• Used: {used_hours:.1f} hours
• Pending Approval: {pending:.1f} hours
• Available: {hours:.1f} hours ({days:.1f} days)

**Urgency Level:** {urgency}
{'⚠️ Staff should book leave soon!' if days < 10 else '✅ Plenty of time remaining'}

**Quick Actions:**
• [View Staff File](/staff-records/{user.sap}/)
• [View Leave History](/admin/staff_records/annualleavetransaction/?profile__user__sap={user.sap})
• [Annual Leave Report](/reports-dashboard/)"""
                            
                            return {
                                'answer': answer,
                                'related': ['Request Leave', 'View All Staff', 'Annual Leave Report'],
                                'category': 'staff_query',
                                'data': {
                                    'sap': user.sap,
                                    'name': user.full_name,
                                    'hours_remaining': hours,
                                    'days_remaining': days
                                }
                            }
                        except StaffProfile.DoesNotExist:
                            return {
                                'answer': f"**{user.full_name} ({user.sap})** found, but no staff profile exists.",
                                'related': ['Setup Staff Profiles'],
                                'category': 'staff_query'
                            }
                        except AnnualLeaveEntitlement.DoesNotExist:
                            return {
                                'answer': f"**{user.full_name} ({user.sap})** found, but no leave entitlement record exists for {timezone.now().year}.",
                                'related': ['Setup Leave Entitlements'],
                                'category': 'staff_query'
                            }
                    else:
                        return {
                            'answer': f"❌ Could not find staff member: **{identifier}**\n\nTry:\n• Using their SAP ID (e.g., ADMIN001)\n• Using their full name\n• Searching in the staff list",
                            'related': ['Search Staff', 'View All Staff'],
                            'category': 'staff_query'
                        }
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    return {
                        'answer': f"⚠️ Error looking up staff: {str(e)}",
                        'category': 'error'
                    }
    
    # Pattern 1.5: "Who is on leave this week/today/next week" or "People on leave"
    leave_list_patterns = [
        r"who\s+(?:is|are)\s+on\s+leave\s+(?:this\s+)?(week|today|tomorrow|next\s+week|month)",
        r"(?:show|list)\s+(?:all\s+)?(?:people|staff)\s+on\s+leave",
        r"(?:people|staff)\s+on\s+leave\s+(?:this\s+)?(week|today|tomorrow)",
    ]
    
    for pattern in leave_list_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            from scheduling.models import LeaveRequest
            from datetime import timedelta
            
            time_period = match.group(1) if match.lastindex >= 1 else 'week'
            today = timezone.now().date()
            
            # Determine date range
            if time_period == 'today':
                start_date = end_date = today
                period_name = "Today"
            elif time_period == 'tomorrow':
                start_date = end_date = today + timedelta(days=1)
                period_name = "Tomorrow"
            elif time_period == 'next week':
                # Next Monday to next Sunday
                days_to_monday = (7 - today.weekday()) % 7
                if days_to_monday == 0:
                    days_to_monday = 7
                start_date = today + timedelta(days=days_to_monday)
                end_date = start_date + timedelta(days=6)
                period_name = "Next Week"
            elif time_period == 'month':
                # This month
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(day=31)
                else:
                    end_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
                period_name = "This Month"
            else:  # week or default
                # This week (Monday to Sunday)
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                period_name = "This Week"
            
            # Find approved leave requests that overlap with this period
            leave_requests = LeaveRequest.objects.filter(
                status='APPROVED',
                leave_type='ANNUAL',  # Focus on annual leave
                start_date__lte=end_date,
                end_date__gte=start_date
            ).select_related('user', 'user__unit').order_by('start_date', 'user__last_name')
            
            if leave_requests.exists():
                results = []
                for req in leave_requests:
                    user = req.user
                    unit_name = user.unit.name if user.unit else 'No Unit'
                    duration = req.days_requested
                    results.append(
                        f"• **{user.full_name}** ({user.sap}) - {unit_name}\n"
                        f"  📅 {req.start_date.strftime('%d %b')} to {req.end_date.strftime('%d %b')} ({duration} days)"
                    )
                
                answer = f"""**🏖️ Staff on Leave - {period_name}**
**Period:** {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}

**{leave_requests.count()} staff member(s) on leave:**

""" + "\n\n".join(results)
                
                return {
                    'answer': answer,
                    'related': ['Annual Leave Report', 'View Staff', 'Coverage Report'],
                    'category': 'leave_list',
                    'data': {'period': period_name, 'count': leave_requests.count()}
                }
            else:
                return {
                    'answer': f"**✅ No staff on leave - {period_name}**\n\n**Period:** {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}\n\nAll staff available for duty.",
                    'related': ['Annual Leave Report', 'View Staff', 'Request Leave'],
                    'category': 'leave_list'
                }
    
    # Pattern 1.55: "Annual leave pending approvals" or "Pending leave requests"
    pending_leave_patterns = [
        r"(?:annual\s+)?leave\s+pending\s+(?:approval|review)",
        r"pending\s+(?:annual\s+)?leave\s+(?:request|approval)",
        r"(?:show|list)\s+pending\s+leave",
    ]
    
    for pattern in pending_leave_patterns:
        if re.search(pattern, query, re.IGNORECASE):
            from scheduling.models import LeaveRequest
            
            pending = LeaveRequest.objects.filter(
                status='PENDING',
                leave_type='ANNUAL'
            ).select_related('user', 'user__unit', 'user__role').order_by('created_at')
            
            if pending.exists():
                results = []
                for req in pending:
                    user = req.user
                    unit_name = user.unit.name if user.unit else 'No Unit'
                    role_name = user.role.name if user.role else 'No Role'
                    duration = req.days_requested
                    days_waiting = (timezone.now().date() - req.created_at.date()).days
                    
                    results.append(
                        f"• **{user.full_name}** ({user.sap}) - {role_name}, {unit_name}\n"
                        f"  📅 {req.start_date.strftime('%d %b')} to {req.end_date.strftime('%d %b')} ({duration} days)\n"
                        f"  ⏰ Requested {days_waiting} day(s) ago"
                    )
                
                answer = f"""**⏳ Annual Leave Pending Approval ({pending.count()} request(s))**

""" + "\n\n".join(results) + f"""

**Quick Actions:**
• [Review All Pending Requests](/admin/scheduling/leaverequest/?status__exact=PENDING)
• [Approve/Deny Requests](/reports-dashboard/)
• [Check Coverage Impact](/reports/coverage/)"""
                
                return {
                    'answer': answer,
                    'related': ['Annual Leave Report', 'Approve Leave', 'Coverage Report'],
                    'category': 'leave_pending',
                    'data': {'count': pending.count()}
                }
            else:
                return {
                    'answer': "**✅ No Pending Leave Requests**\n\nAll annual leave requests have been processed.\n\n**Next steps:**\n• [View Approved Leave](/admin/scheduling/leaverequest/?status__exact=APPROVED)\n• [View All Leave Requests](/admin/scheduling/leaverequest/)",
                    'related': ['View Leave Requests', 'Annual Leave Report'],
                    'category': 'leave_pending'
                }
    
    # Pattern 1.57: "Agency usage this month" or "Overtime costs"
    agency_usage_patterns = [
        r"agency\s+(?:usage|use|staff|shifts?)\s+(?:this\s+)?(month|week|today)",
        r"(?:show|get)\s+agency\s+(?:usage|hours|costs?)",
        r"overtime\s+costs?\s+(?:this\s+)?(month|week)",
        r"(?:show|get)\s+overtime\s+(?:usage|hours|costs?)",
    ]
    
    for pattern in agency_usage_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            from scheduling.models import Shift
            from datetime import timedelta
            from decimal import Decimal
            
            time_period = match.group(1) if match.lastindex >= 1 else 'month'
            today = timezone.now().date()
            
            # Determine if it's agency or overtime query
            is_agency = 'agency' in query_lower
            is_overtime = 'overtime' in query_lower
            
            # Determine date range
            if time_period == 'today':
                start_date = end_date = today
                period_name = "Today"
            elif time_period == 'week':
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                period_name = "This Week"
            else:  # month
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(day=31)
                else:
                    end_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
                period_name = "This Month"
            
            # Query for agency or overtime shifts
            if is_agency:
                shifts = Shift.objects.filter(
                    date__gte=start_date,
                    date__lte=end_date,
                    shift_classification='AGENCY'
                ).select_related('agency_company', 'unit')
                
                query_type = "Agency"
            elif is_overtime:
                shifts = Shift.objects.filter(
                    date__gte=start_date,
                    date__lte=end_date,
                    shift_classification='OVERTIME'
                ).select_related('user', 'user__role', 'unit')
                
                query_type = "Overtime"
            else:
                # Default to both
                shifts = Shift.objects.filter(
                    date__gte=start_date,
                    date__lte=end_date,
                    shift_classification__in=['AGENCY', 'OVERTIME']
                ).select_related('user', 'agency_company', 'unit')
                
                query_type = "Additional Staffing (Agency + Overtime)"
            
            if shifts.exists():
                # Calculate statistics
                total_shifts = shifts.count()
                total_cost = Decimal('0.00')
                total_hours = Decimal('0.00')
                
                # Break down by unit or company
                by_unit = {}
                by_company = {}
                
                for shift in shifts:
                    # Calculate hours and cost - convert to Decimal to avoid type errors
                    if hasattr(shift, 'duration_hours') and shift.duration_hours:
                        hours = Decimal(str(shift.duration_hours))
                    else:
                        hours = Decimal('12.00')  # Default shift duration
                    
                    total_hours += hours
                    
                    if shift.agency_hourly_rate:
                        cost = Decimal(str(shift.agency_hourly_rate)) * hours
                        total_cost += cost
                    
                    # Track by unit
                    unit_name = shift.unit.name if shift.unit else 'No Unit'
                    if unit_name not in by_unit:
                        by_unit[unit_name] = {'count': 0, 'hours': Decimal('0.00')}
                    by_unit[unit_name]['count'] += 1
                    by_unit[unit_name]['hours'] += hours
                    
                    # Track by company (for agency)
                    if is_agency and shift.agency_company:
                        company_name = shift.agency_company.name
                        if company_name not in by_company:
                            by_company[company_name] = {'count': 0, 'hours': Decimal('0.00')}
                        by_company[company_name]['count'] += 1
                        by_company[company_name]['hours'] += hours
                
                answer = f"""**📊 {query_type} Usage - {period_name}**
**Period:** {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}

**Summary:**
• Total Shifts: {total_shifts}
• Total Hours: {total_hours:.1f}
• Estimated Cost: £{total_cost:,.2f}

**Breakdown by Unit:**
"""
                for unit, data in sorted(by_unit.items()):
                    answer += f"• {unit}: {data['count']} shifts ({data['hours']:.1f} hours)\n"
                
                if by_company:
                    answer += "\n**Breakdown by Agency:**\n"
                    for company, data in sorted(by_company.items()):
                        answer += f"• {company}: {data['count']} shifts ({data['hours']:.1f} hours)\n"
                
                answer += f"""

**Quick Actions:**
• [View Detailed Report](/reports/additional-staffing/)
• [Export to Excel](/reports/export/)
• [Manage Agency Companies](/additional-staffing/agencies/)"""
                
                return {
                    'answer': answer,
                    'related': ['Additional Staffing Report', 'Agency Companies', 'Cost Analysis'],
                    'category': 'agency_usage',
                    'data': {
                        'period': period_name,
                        'total_shifts': total_shifts,
                        'total_hours': float(total_hours),
                        'total_cost': float(total_cost)
                    }
                }
            else:
                return {
                    'answer': f"**✅ No {query_type} - {period_name}**\n\n**Period:** {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}\n\nNo {query_type.lower()} shifts recorded for this period.",
                    'related': ['Add Agency Shift', 'Add Overtime', 'View Reports'],
                    'category': 'agency_usage'
                }
    
    # Pattern 1.6: "List agency companies" or "Show agencies"
    agency_patterns = [
        r"(?:list|show)\s+(?:all\s+)?(?:agency|agencies)\s+(?:companies|providers)?",
        r"(?:agency|agencies)\s+(?:companies|list)",
        r"(?:show|get)\s+(?:the\s+)?agencies",
    ]
    
    for pattern in agency_patterns:
        if re.search(pattern, query, re.IGNORECASE):
            from scheduling.models import AgencyCompany
            
            agencies = AgencyCompany.objects.filter(is_active=True).order_by('name')
            
            if agencies.exists():
                results = []
                for agency in agencies:
                    contact = agency.contact_name if hasattr(agency, 'contact_name') and agency.contact_name else 'N/A'
                    phone = agency.phone if hasattr(agency, 'phone') and agency.phone else 'N/A'
                    email = agency.email if hasattr(agency, 'email') and agency.email else 'N/A'
                    
                    results.append(
                        f"**{agency.name}**\n"
                        f"• Contact: {contact}\n"
                        f"• Phone: {phone}\n"
                        f"• Email: {email}"
                    )
                
                answer = f"""**🏢 Active Agency Companies ({agencies.count()} total)**

""" + "\n\n".join(results) + f"""

**Quick Actions:**
• [Add New Agency](/additional-staffing/agency/add/)
• [View Agency Staff](/additional-staffing/agency-staff/)
• [Agency Usage Report](javascript:void(0);) _(Ask: "Show agency usage this month")_"""
                
                return {
                    'answer': answer,
                    'related': ['Agency Usage Report', 'Add Agency', 'Additional Staffing'],
                    'category': 'agency_list',
                    'data': {'count': agencies.count()}
                }
            else:
                return {
                    'answer': "**ℹ️ No Agency Companies Registered**\n\nNo active agency companies found in the system.\n\n**Next Steps:**\n• [Add New Agency Company](/additional-staffing/agency/add/)\n• Set up agency staff database\n• Configure agency rates",
                    'related': ['Add Agency', 'Setup Additional Staffing'],
                    'category': 'agency_list'
                }
    
    # Pattern 2: "Search for X" or "Find staff X" or "Show me X details"
    search_patterns = [
        r"(?:search|find|look\s+for|locate)\s+(?:staff\s+)?(.+?)(?:\s+details?)?$",
        r"(?:who\s+is|show\s+me)\s+([A-Za-z\s]+?)(?:\s+details?)?$",
        r"(?:info|information)\s+(?:about|on|for)\s+(.+?)$",
    ]
    
    for pattern in search_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            search_term = match.group(1).strip()
            
            # Skip if search term contains generic words that shouldn't trigger staff search
            if search_term.lower() in ['all', 'staff', 'grades', 'roles', 'units', 'everyone']:
                continue
            
            # Search staff by name parts and SAP
            name_parts = search_term.split()
            q = Q(sap__icontains=search_term)
            for part in name_parts:
                if len(part) > 1:  # Skip single letters
                    q |= Q(first_name__icontains=part) | Q(last_name__icontains=part)
            
            staff = User.objects.filter(q)
            
            if staff.exists():
                # If only one match, show detailed profile
                if staff.count() == 1:
                    user = staff.first()
                    role_name = user.role.name if user.role else 'No Role'
                    unit_name = user.unit.name if user.unit else 'No Unit'
                    team = user.team if user.team else 'N/A'
                    
                    # Get contracted hours
                    try:
                        profile = StaffProfile.objects.get(user=user)
                        contracted_hours = profile.contracted_hours_per_week if hasattr(profile, 'contracted_hours_per_week') else 'N/A'
                    except StaffProfile.DoesNotExist:
                        contracted_hours = 'N/A'
                    
                    answer = f"""**👤 {user.full_name}**

**Basic Information:**
• SAP ID: {user.sap}
• Email: {user.email}
• Role: {role_name}
• Unit: {unit_name}
• Team: {team}
• Contracted Hours: {contracted_hours}/week
• Status: {'✅ Active' if user.is_active else '❌ Inactive'}

**Quick Links:**
• [View Full Profile](/staff/{user.sap}/)
• [Check Leave Balance](javascript:void(0);) _(Ask: "How much leave does {user.sap} have?")_

**What would you like to know about {user.first_name}?**"""
                    
                    return {
                        'answer': answer,
                        'related': ['Check Leave Balance', 'View Shifts', 'Staff Profile'],
                        'category': 'staff_profile',
                        'data': {'sap': user.sap, 'name': user.full_name}
                    }
                else:
                    # Multiple matches - show list (limit to 10)
                    staff_list = staff[:10]
                    results = []
                    for s in staff_list:
                        role_name = s.role.name if s.role else 'No Role'
                        unit_name = s.unit.name if s.unit else 'No Unit'
                        results.append(f"• **{s.full_name}** ({s.sap}) - {role_name}, {unit_name}")
                    
                    answer = f"**🔍 Search Results for '{search_term}':**\n\n" + "\n".join(results)
                    if len(staff_list) == 10:
                        answer += "\n\n*Showing first 10 results. Be more specific for better results.*"
                    
                    return {
                        'answer': answer,
                        'related': ['View Full Staff List'],
                        'category': 'staff_search',
                        'data': {'count': staff.count(), 'term': search_term}
                    }
            else:
                return {
                    'answer': f"❌ No staff found matching: **{search_term}**\n\nTry:\n• Different spelling\n• SAP ID\n• First or last name only",
                    'related': ['View All Staff', 'Add New Staff'],
                    'category': 'staff_search'
                }
    
    # Pattern 3: "Open staff file for X" or "Show X's profile"
    file_patterns = [
        r"(?:open|show|view)\s+(?:staff\s+file|profile|record)\s+(?:for\s+)?([A-Z0-9]+|[A-Za-z\s]+)",
        r"([A-Z0-9]+)'?s?\s+(?:file|profile|record)",
    ]
    
    for pattern in file_patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            identifier = match.group(1).strip()
            
            try:
                if identifier.isupper() and len(identifier) < 10:
                    user = User.objects.filter(sap__iexact=identifier).first()
                else:
                    name_parts = identifier.split()
                    q = Q()
                    for part in name_parts:
                        q |= Q(first_name__icontains=part) | Q(last_name__icontains=part)
                    user = User.objects.filter(q).first()
                
                if user:
                    role_name = user.role.name if user.role else 'No Role'
                    unit_name = user.unit.name if user.unit else 'No Unit'
                    team = user.team if user.team else 'N/A'
                    
                    # Get contracted hours from staff profile
                    try:
                        profile = StaffProfile.objects.get(user=user)
                        contracted_hours = profile.contracted_hours_per_week if hasattr(profile, 'contracted_hours_per_week') else 'N/A'
                    except StaffProfile.DoesNotExist:
                        contracted_hours = 'N/A'
                    
                    answer = f"""**👤 Staff Profile: {user.full_name}**

**Basic Information:**
• SAP ID: {user.sap}
• Email: {user.email}
• Role: {role_name}
• Unit: {unit_name}
• Team: {team}
• Contracted Hours: {contracted_hours}/week

**Quick Links:**
• [View Full Profile](/admin/scheduling/user/{user.sap}/change/)
• [Check Leave Balance](javascript:void(0);) _(Ask: "How much leave does {user.sap} have?")_
• [View Shift History](/admin/scheduling/shift/?assigned_to__sap={user.sap})
• [View Sickness Records](/admin/staff_records/sicknessrecord/?staff__sap={user.sap})

**What would you like to know about {user.first_name}?**"""
                    
                    return {
                        'answer': answer,
                        'related': ['Check Leave Balance', 'View Shifts', 'Edit Profile'],
                        'category': 'staff_profile',
                        'data': {'sap': user.sap, 'name': user.full_name}
                    }
                else:
                    return {
                        'answer': f"❌ Could not find staff member: **{identifier}**",
                        'related': ['Search Staff'],
                        'category': 'staff_profile'
                    }
            except Exception as e:
                return {
                    'answer': f"⚠️ Error: {str(e)}",
                    'category': 'error'
                }
    
    return None  # No staff query detected


# ==================== AGENCY & ADDITIONAL STAFFING APIs ====================

@login_required
@login_required
def agency_companies_api(request):
    """API endpoint to fetch active agency companies"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Only GET requests allowed'}, status=405)
    
    try:
        from .models import AgencyCompany
        agencies = AgencyCompany.objects.filter(is_active=True).order_by('name')
        
        data = {
            'agencies': [
                {
                    'id': agency.id,
                    'name': agency.name,
                    'hourly_rate_day': float(agency.hourly_rate_day),
                    'hourly_rate_night': float(agency.hourly_rate_night),
                    'contact_person': agency.contact_person,
                    'contact_phone': agency.contact_phone
                }
                for agency in agencies
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def daily_additional_staffing_report(request):
    """Generate daily report of overtime and agency usage"""
    
    # Check permissions
    if not (request.user.role and (request.user.role.can_manage_rota or request.user.role.is_management)):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    target_date_str = request.GET.get('date')
    if target_date_str:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    else:
        target_date = timezone.now().date()
    
    # Get all additional staffing shifts for the day
    additional_shifts = Shift.objects.filter(
        date=target_date,
        shift_classification__in=['OVERTIME', 'AGENCY']
    ).select_related('user', 'unit', 'shift_type', 'agency_company')
    
    # Organize by type
    overtime_shifts = []
    agency_shifts_by_company = defaultdict(list)
    
    total_overtime_hours = 0
    total_agency_hours = 0
    total_agency_cost = 0
    
    for shift in additional_shifts:
        shift_data = {
            'staff_name': shift.user.full_name,
            'sap': shift.user.sap,
            'unit': shift.unit.get_name_display(),
            'pattern': shift.get_shift_pattern_display(),
            'start_time': shift.start_time.strftime('%H:%M'),
            'end_time': shift.end_time.strftime('%H:%M'),
            'hours': shift.duration_hours,
            'notes': shift.notes or ''
        }
        
        if shift.shift_classification == 'OVERTIME':
            overtime_shifts.append(shift_data)
            total_overtime_hours += shift.duration_hours
        elif shift.shift_classification == 'AGENCY':
            shift_data['agency_staff_name'] = shift.agency_staff_name
            shift_data['hourly_rate'] = float(shift.agency_hourly_rate) if shift.agency_hourly_rate else 0
            shift_data['total_cost'] = shift_data['hourly_rate'] * shift.duration_hours
            
            company_name = shift.agency_company.name if shift.agency_company else 'Unknown Agency'
            agency_shifts_by_company[company_name].append(shift_data)
            
            total_agency_hours += shift.duration_hours
            total_agency_cost += shift_data['total_cost']
    
    report_data = {
        'date': target_date.strftime('%Y-%m-%d'),
        'date_display': target_date.strftime('%A, %d %B %Y'),
        'overtime': {
            'shifts': overtime_shifts,
            'total_hours': round(total_overtime_hours, 2),
            'shift_count': len(overtime_shifts)
        },
        'agency': {
            'by_company': dict(agency_shifts_by_company),
            'total_hours': round(total_agency_hours, 2),
            'total_cost': round(total_agency_cost, 2),
            'shift_count': sum(len(shifts) for shifts in agency_shifts_by_company.values())
        },
        'summary': {
            'total_additional_hours': round(total_overtime_hours + total_agency_hours, 2),
            'total_shifts': len(additional_shifts),
            'total_cost': round(total_agency_cost, 2)
        }
    }
    
    return JsonResponse(report_data)


@login_required
def weekly_additional_staffing_report(request):
    """Generate weekly report (Sunday to Saturday) of overtime and agency usage"""
    
    # Check permissions
    if not (request.user.role and (request.user.role.can_manage_rota or request.user.role.is_management)):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get target week - default to previous week (for Monday morning reports)
    target_date_str = request.GET.get('date')
    if target_date_str:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    else:
        # Default to previous week (Sunday to Saturday)
        today = timezone.now().date()
        days_since_sunday = (today.weekday() + 1) % 7
        last_sunday = today - timedelta(days=days_since_sunday + 7)
        target_date = last_sunday
    
    # Calculate week range (Sunday to Saturday)
    days_since_sunday = (target_date.weekday() + 1) % 7
    week_start = target_date - timedelta(days=days_since_sunday)
    week_end = week_start + timedelta(days=6)
    
    # Get all additional staffing shifts for the week
    additional_shifts = Shift.objects.filter(
        date__range=[week_start, week_end],
        shift_classification__in=['OVERTIME', 'AGENCY']
    ).select_related('user', 'unit', 'shift_type', 'agency_company').order_by('date')
    
    # Organize by type and day
    daily_breakdown = {}
    overtime_totals = {'hours': 0, 'shifts': 0}
    agency_totals_by_company = defaultdict(lambda: {'hours': 0, 'cost': 0, 'shifts': 0})
    
    current_day = week_start
    while current_day <= week_end:
        day_str = current_day.strftime('%Y-%m-%d')
        daily_breakdown[day_str] = {
            'date_display': current_day.strftime('%A, %d %B'),
            'overtime_hours': 0,
            'overtime_shifts': 0,
            'agency_hours': 0,
            'agency_cost': 0,
            'agency_shifts': 0,
            'agency_by_company': defaultdict(lambda: {'hours': 0, 'cost': 0, 'shifts': 0})
        }
        current_day += timedelta(days=1)
    
    # Process shifts
    for shift in additional_shifts:
        day_str = shift.date.strftime('%Y-%m-%d')
        day_data = daily_breakdown[day_str]
        
        if shift.shift_classification == 'OVERTIME':
            day_data['overtime_hours'] += shift.duration_hours
            day_data['overtime_shifts'] += 1
            overtime_totals['hours'] += shift.duration_hours
            overtime_totals['shifts'] += 1
            
        elif shift.shift_classification == 'AGENCY':
            company_name = shift.agency_company.name if shift.agency_company else 'Unknown Agency'
            shift_cost = (float(shift.agency_hourly_rate) if shift.agency_hourly_rate else 0) * shift.duration_hours
            
            day_data['agency_hours'] += shift.duration_hours
            day_data['agency_cost'] += shift_cost
            day_data['agency_shifts'] += 1
            
            day_data['agency_by_company'][company_name]['hours'] += shift.duration_hours
            day_data['agency_by_company'][company_name]['cost'] += shift_cost
            day_data['agency_by_company'][company_name]['shifts'] += 1
            
            agency_totals_by_company[company_name]['hours'] += shift.duration_hours
            agency_totals_by_company[company_name]['cost'] += shift_cost
            agency_totals_by_company[company_name]['shifts'] += 1
    
    # Convert defaultdicts to regular dicts for JSON
    for day_data in daily_breakdown.values():
        day_data['agency_by_company'] = dict(day_data['agency_by_company'])
    
    report_data = {
        'week_start': week_start.strftime('%Y-%m-%d'),
        'week_end': week_end.strftime('%Y-%m-%d'),
        'week_display': f"{week_start.strftime('%d %B')} - {week_end.strftime('%d %B %Y')}",
        'daily_breakdown': daily_breakdown,
        'overtime_totals': {
            'hours': round(overtime_totals['hours'], 2),
            'shifts': overtime_totals['shifts']
        },
        'agency_totals': {
            'by_company': {
                company: {
                    'hours': round(data['hours'], 2),
                    'cost': round(data['cost'], 2),
                    'shifts': data['shifts']
                }
                for company, data in agency_totals_by_company.items()
            },
            'total_hours': round(sum(data['hours'] for data in agency_totals_by_company.values()), 2),
            'total_cost': round(sum(data['cost'] for data in agency_totals_by_company.values()), 2),
            'total_shifts': sum(data['shifts'] for data in agency_totals_by_company.values())
        },
        'grand_totals': {
            'total_additional_hours': round(
                overtime_totals['hours'] + sum(data['hours'] for data in agency_totals_by_company.values()), 
                2
            ),
            'total_shifts': overtime_totals['shifts'] + sum(data['shifts'] for data in agency_totals_by_company.values()),
            'total_cost': round(sum(data['cost'] for data in agency_totals_by_company.values()), 2)
        }
    }
    
    return JsonResponse(report_data)


@login_required
def ot_agency_report(request):
    """
    Comprehensive Overtime and Agency Usage Report with breakdown by home, role, and reasons
    """
    from decimal import Decimal
    
    # Check permissions
    if not (request.user.role and (request.user.role.can_manage_rota or request.user.role.is_management)):
        messages.error(request, 'You do not have permission to view this report.')
        return redirect('manager_dashboard')
    
    # Get date range from query params
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    home_filter = request.GET.get('home_filter', '')
    
    # Default to current month if no dates provided
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Last day of current month
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
    
    # Get all care homes
    care_homes = CareHome.objects.all().order_by('name')
    
    # Initialize report data
    report_data = None
    
    if start_date and end_date:
        # Get all OT and Agency shifts in date range
        ot_agency_shifts = Shift.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            shift_classification__in=['OVERTIME', 'AGENCY']
        ).select_related('user', 'user__role', 'unit', 'unit__care_home', 'agency_company', 'shift_type')
        
        # Filter by home if specified
        if home_filter:
            ot_agency_shifts = ot_agency_shifts.filter(unit__care_home__name=home_filter)
        
        # Organize data by home
        homes_data = []
        grand_ot_hours = 0
        grand_ot_cost = 0
        grand_ot_shifts = 0
        grand_agency_hours = 0
        grand_agency_cost = 0
        grand_agency_shifts = 0
        
        for home in care_homes:
            # Filter by home if no home_filter, or include only selected home
            if home_filter and home.name != home_filter:
                continue
                
            home_shifts = [s for s in ot_agency_shifts if s.unit.care_home == home]
            
            if not home_shifts:
                continue
            
            # Breakdown by role for this home
            ot_by_role = defaultdict(lambda: {'hours': 0, 'count': 0, 'cost': 0, 'reasons': set()})
            agency_by_role = defaultdict(lambda: {'hours': 0, 'count': 0, 'cost': 0, 'reasons': set()})
            
            for shift in home_shifts:
                role = shift.user.role.name if shift.user and shift.user.role else 'Unknown'
                hours = shift.duration_hours or 12.5
                
                if shift.shift_classification == 'OVERTIME':
                    # OT cost = 1.5x base rate
                    base_rate = 15.0  # Average base rate
                    cost = hours * base_rate * 1.5
                    ot_by_role[role]['hours'] += hours
                    ot_by_role[role]['count'] += 1
                    ot_by_role[role]['cost'] += cost
                    if shift.notes:
                        ot_by_role[role]['reasons'].add(shift.notes[:50])  # Truncate long notes
                
                elif shift.shift_classification == 'AGENCY':
                    # Agency cost from hourly rate
                    rate = float(shift.agency_hourly_rate) if shift.agency_hourly_rate else 25.0
                    cost = hours * rate
                    agency_by_role[role]['hours'] += hours
                    agency_by_role[role]['count'] += 1
                    agency_by_role[role]['cost'] += cost
                    if shift.notes:
                        agency_by_role[role]['reasons'].add(shift.notes[:50])
            
            # Convert reasons sets to lists
            for role_data in ot_by_role.values():
                role_data['reasons'] = list(role_data['reasons'])[:3]  # Top 3 reasons
            for role_data in agency_by_role.values():
                role_data['reasons'] = list(role_data['reasons'])[:3]
            
            # Calculate home totals
            ot_total_hours = sum(d['hours'] for d in ot_by_role.values())
            ot_total_cost = sum(d['cost'] for d in ot_by_role.values())
            ot_total_shifts = sum(d['count'] for d in ot_by_role.values())
            
            agency_total_hours = sum(d['hours'] for d in agency_by_role.values())
            agency_total_cost = sum(d['cost'] for d in agency_by_role.values())
            agency_total_shifts = sum(d['count'] for d in agency_by_role.values())
            
            # Add to grand totals
            grand_ot_hours += ot_total_hours
            grand_ot_cost += ot_total_cost
            grand_ot_shifts += ot_total_shifts
            grand_agency_hours += agency_total_hours
            grand_agency_cost += agency_total_cost
            grand_agency_shifts += agency_total_shifts
            
            homes_data.append({
                'home_name': home.get_name_display(),
                'home_code': home.name,
                'overtime_by_role': dict(ot_by_role),
                'agency_by_role': dict(agency_by_role),
                'ot_total_hours': ot_total_hours,
                'ot_total_cost': ot_total_cost,
                'ot_total_shifts': ot_total_shifts,
                'agency_total_hours': agency_total_hours,
                'agency_total_cost': agency_total_cost,
                'agency_total_shifts': agency_total_shifts,
                'total_hours': ot_total_hours + agency_total_hours,
                'total_cost': ot_total_cost + agency_total_cost,
            })
        
        report_data = {
            'by_home': homes_data,
            'grand_totals': {
                'total_ot_hours': grand_ot_hours,
                'total_ot_cost': grand_ot_cost,
                'total_ot_shifts': grand_ot_shifts,
                'total_agency_hours': grand_agency_hours,
                'total_agency_cost': grand_agency_cost,
                'total_agency_shifts': grand_agency_shifts,
                'combined_hours': grand_ot_hours + grand_agency_hours,
                'combined_cost': grand_ot_cost + grand_agency_cost,
                'combined_shifts': grand_ot_shifts + grand_agency_shifts,
            }
        }
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'selected_home': home_filter,
        'care_homes': care_homes,
        'report_data': report_data,
    }
    
    return render(request, 'scheduling/ot_agency_report.html', context)


@login_required
def ot_agency_report_csv(request):
    """Export OT and Agency report as detailed CSV with shift-level breakdown"""
    import csv
    from django.http import HttpResponse
    
    # Get same parameters as main report
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    home_filter = request.GET.get('home_filter', '')
    detail_level = request.GET.get('detail', 'detailed')  # 'summary' or 'detailed'
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date().replace(day=1)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename_suffix = 'detailed' if detail_level == 'detailed' else 'summary'
    response['Content-Disposition'] = f'attachment; filename="ot_agency_report_{filename_suffix}_{start_date}_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(['Overtime and Agency Usage Report - Detailed Breakdown'])
    writer.writerow([f'Period: {start_date.strftime("%d %B %Y")} - {end_date.strftime("%d %B %Y")}'])
    writer.writerow(['Generated: ' + timezone.now().strftime("%d %B %Y %H:%M")])
    writer.writerow(['Generated: ' + timezone.now().strftime("%d %B %Y %H:%M")])
    writer.writerow([])
    
    # Get data
    ot_agency_shifts = Shift.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        shift_classification__in=['OVERTIME', 'AGENCY']
    ).select_related('user', 'user__role', 'unit', 'unit__care_home', 'shift_type', 'agency_company').order_by('unit__care_home__name', 'date', 'shift_classification')
    
    if home_filter:
        ot_agency_shifts = ot_agency_shifts.filter(unit__care_home__name=home_filter)
    
    care_homes = CareHome.objects.all().order_by('name')
    
    # Grand totals trackers
    grand_ot_hours = grand_ot_cost = grand_ot_count = 0
    grand_agency_hours = grand_agency_cost = grand_agency_count = 0
    
    for home in care_homes:
        if home_filter and home.name != home_filter:
            continue
            
        home_shifts = [s for s in ot_agency_shifts if s.unit.care_home == home]
        
        if not home_shifts:
            continue
        
        # Home header
        writer.writerow([])
        writer.writerow(['=' * 100])
        writer.writerow([f'CARE HOME: {home.get_name_display()}'])
        writer.writerow(['=' * 100])
        writer.writerow([])
        
        # === OVERTIME SECTION ===
        ot_shifts = [s for s in home_shifts if s.shift_classification == 'OVERTIME']
        
        if ot_shifts:
            writer.writerow(['OVERTIME USAGE - DETAILED BREAKDOWN'])
            writer.writerow(['-' * 100])
            writer.writerow(['Date', 'Day', 'Staff Name', 'SAP ID', 'Grade/Role', 'Unit', 'Shift Type', 'Hours', 'Base Rate', 'OT Rate (1.5x)', 'Cost', 'Reason/Notes'])
            
            ot_total_hours = ot_total_cost = 0
            ot_by_role = defaultdict(lambda: {'count': 0, 'hours': 0, 'cost': 0})
            
            for shift in ot_shifts:
                staff_name = f"{shift.user.first_name} {shift.user.last_name}" if shift.user else 'Unknown'
                sap_id = shift.user.sap_id if shift.user else 'N/A'
                role = shift.user.role.name if shift.user and shift.user.role else 'Unknown'
                unit = shift.unit.name if shift.unit else 'N/A'
                shift_type = shift.shift_type.name if shift.shift_type else 'N/A'
                hours = shift.duration_hours or 12.5
                base_rate = 15.0  # Could be pulled from user profile if available
                ot_rate = base_rate * 1.5
                cost = hours * ot_rate
                reason = (shift.notes or 'No reason provided').replace('\n', ' ').replace('\r', ' ')
                day_name = shift.date.strftime('%A')
                
                writer.writerow([
                    shift.date.strftime('%d/%m/%Y'),
                    day_name,
                    staff_name,
                    sap_id,
                    role,
                    unit,
                    shift_type,
                    f'{hours:.1f}',
                    f'£{base_rate:.2f}',
                    f'£{ot_rate:.2f}',
                    f'£{cost:.2f}',
                    reason
                ])
                
                ot_total_hours += hours
                ot_total_cost += cost
                ot_by_role[role]['count'] += 1
                ot_by_role[role]['hours'] += hours
                ot_by_role[role]['cost'] += cost
            
            # OT Summary for this home
            writer.writerow([])
            writer.writerow(['OVERTIME SUMMARY BY GRADE'])
            writer.writerow(['Grade/Role', 'Number of Shifts', 'Total Hours', 'Total Cost'])
            
            for role in sorted(ot_by_role.keys()):
                data = ot_by_role[role]
                writer.writerow([
                    role,
                    data['count'],
                    f"{data['hours']:.1f}",
                    f"£{data['cost']:.2f}"
                ])
            
            writer.writerow([])
            writer.writerow(['OVERTIME TOTAL', len(ot_shifts), f'{ot_total_hours:.1f}', f'£{ot_total_cost:.2f}'])
            writer.writerow([])
            
            grand_ot_hours += ot_total_hours
            grand_ot_cost += ot_total_cost
            grand_ot_count += len(ot_shifts)
        else:
            writer.writerow(['OVERTIME USAGE - No overtime shifts in this period'])
            writer.writerow([])
        
        # === AGENCY SECTION ===
        agency_shifts = [s for s in home_shifts if s.shift_classification == 'AGENCY']
        
        if agency_shifts:
            writer.writerow(['AGENCY USAGE - DETAILED BREAKDOWN'])
            writer.writerow(['-' * 100])
            writer.writerow(['Date', 'Day', 'Staff Name', 'SAP ID', 'Grade/Role', 'Unit', 'Shift Type', 'Agency Company', 'Hours', 'Hourly Rate', 'Cost', 'Reason/Notes'])
            
            agency_total_hours = agency_total_cost = 0
            agency_by_role = defaultdict(lambda: {'count': 0, 'hours': 0, 'cost': 0})
            
            for shift in agency_shifts:
                staff_name = f"{shift.user.first_name} {shift.user.last_name}" if shift.user else 'Unknown'
                sap_id = shift.user.sap_id if shift.user else 'N/A'
                role = shift.user.role.name if shift.user and shift.user.role else 'Unknown'
                unit = shift.unit.name if shift.unit else 'N/A'
                shift_type = shift.shift_type.name if shift.shift_type else 'N/A'
                agency_name = shift.agency_company.name if shift.agency_company else 'Unknown Agency'
                hours = shift.duration_hours or 12.5
                hourly_rate = float(shift.agency_hourly_rate) if shift.agency_hourly_rate else 25.0
                cost = hours * hourly_rate
                reason = (shift.notes or 'No reason provided').replace('\n', ' ').replace('\r', ' ')
                day_name = shift.date.strftime('%A')
                
                writer.writerow([
                    shift.date.strftime('%d/%m/%Y'),
                    day_name,
                    staff_name,
                    sap_id,
                    role,
                    unit,
                    shift_type,
                    agency_name,
                    f'{hours:.1f}',
                    f'£{hourly_rate:.2f}',
                    f'£{cost:.2f}',
                    reason
                ])
                
                agency_total_hours += hours
                agency_total_cost += cost
                agency_by_role[role]['count'] += 1
                agency_by_role[role]['hours'] += hours
                agency_by_role[role]['cost'] += cost
            
            # Agency Summary for this home
            writer.writerow([])
            writer.writerow(['AGENCY SUMMARY BY GRADE'])
            writer.writerow(['Grade/Role', 'Number of Shifts', 'Total Hours', 'Total Cost'])
            
            for role in sorted(agency_by_role.keys()):
                data = agency_by_role[role]
                writer.writerow([
                    role,
                    data['count'],
                    f"{data['hours']:.1f}",
                    f"£{data['cost']:.2f}"
                ])
            
            writer.writerow([])
            writer.writerow(['AGENCY TOTAL', len(agency_shifts), f'{agency_total_hours:.1f}', f'£{agency_total_cost:.2f}'])
            writer.writerow([])
            
            grand_agency_hours += agency_total_hours
            grand_agency_cost += agency_total_cost
            grand_agency_count += len(agency_shifts)
        else:
            writer.writerow(['AGENCY USAGE - No agency shifts in this period'])
            writer.writerow([])
    
    # Grand Total Summary
    writer.writerow([])
    writer.writerow(['=' * 100])
    writer.writerow(['GRAND TOTALS ACROSS ALL HOMES'])
    writer.writerow(['=' * 100])
    writer.writerow([])
    writer.writerow(['Category', 'Number of Shifts', 'Total Hours', 'Total Cost'])
    writer.writerow(['Overtime', grand_ot_count, f'{grand_ot_hours:.1f}', f'£{grand_ot_cost:.2f}'])
    writer.writerow(['Agency', grand_agency_count, f'{grand_agency_hours:.1f}', f'£{grand_agency_cost:.2f}'])
    writer.writerow(['COMBINED TOTAL', grand_ot_count + grand_agency_count, f'{grand_ot_hours + grand_agency_hours:.1f}', f'£{grand_ot_cost + grand_agency_cost:.2f}'])
    
    return response


@login_required
def staff_vacancies_report(request):
    """
    Staff Vacancies Report showing current and upcoming vacancies by care home
    """
    from staff_records.models import StaffProfile
    
    # Check permissions
    if not (request.user.role and (request.user.role.can_manage_rota or request.user.role.is_management)):
        messages.error(request, 'You do not have permission to view this report.')
        return redirect('manager_dashboard')
    
    # Get filter parameters
    home_filter = request.GET.get('home_filter', '')
    status_filter = request.GET.get('status_filter', 'all')  # all, current, upcoming
    
    # Get all care homes for the filter dropdown
    care_homes = CareHome.objects.all().order_by('name')
    
    # Get all leavers (staff with employment_status = 'LEAVER' or end_date set)
    today = timezone.now().date()
    
    # Current vacancies (already left)
    current_vacancies_query = StaffProfile.objects.filter(
        employment_status='LEAVER',
        end_date__lte=today
    ).select_related('user', 'user__role', 'user__unit', 'user__unit__care_home')
    
    # Upcoming leavers (future end date)
    upcoming_vacancies_query = StaffProfile.objects.filter(
        employment_status='LEAVER',
        end_date__gt=today
    ).select_related('user', 'user__role', 'user__unit', 'user__unit__care_home')
    
    # Filter by home if specified
    if home_filter:
        current_vacancies_query = current_vacancies_query.filter(user__unit__care_home__name=home_filter)
        upcoming_vacancies_query = upcoming_vacancies_query.filter(user__unit__care_home__name=home_filter)
    
    # Apply status filter
    if status_filter == 'current':
        upcoming_vacancies_query = StaffProfile.objects.none()
    elif status_filter == 'upcoming':
        current_vacancies_query = StaffProfile.objects.none()
    
    # Organize vacancies by care home
    vacancies_by_home = defaultdict(lambda: {
        'current': [],
        'upcoming': [],
        'total_current': 0,
        'total_upcoming': 0,
        'total_hours': 0
    })
    
    # Process current vacancies
    for profile in current_vacancies_query:
        if not profile.user or not profile.user.unit or not profile.user.unit.care_home:
            continue
            
        home_name = profile.user.unit.care_home.name
        days_ago = (today - profile.end_date).days if profile.end_date else 0
        hours = getattr(profile.user, 'hours_per_week', 37.5)
        
        vacancy_data = {
            'name': profile.user.full_name if profile.user else 'Unknown',
            'sap_id': profile.user.sap_id if profile.user else 'N/A',
            'role': profile.user.role.name if profile.user and profile.user.role else 'Unknown',
            'unit': profile.user.unit.name if profile.user and profile.user.unit else 'N/A',
            'end_date': profile.end_date,
            'days_ago': days_ago,
            'hours_per_week': hours
        }
        
        vacancies_by_home[home_name]['current'].append(vacancy_data)
        vacancies_by_home[home_name]['total_current'] += 1
        vacancies_by_home[home_name]['total_hours'] += hours
    
    # Process upcoming leavers
    for profile in upcoming_vacancies_query:
        if not profile.user or not profile.user.unit or not profile.user.unit.care_home:
            continue
            
        home_name = profile.user.unit.care_home.name
        days_until = (profile.end_date - today).days if profile.end_date else 0
        hours = getattr(profile.user, 'hours_per_week', 37.5)
        
        vacancy_data = {
            'name': profile.user.full_name if profile.user else 'Unknown',
            'sap_id': profile.user.sap_id if profile.user else 'N/A',
            'role': profile.user.role.name if profile.user and profile.user.role else 'Unknown',
            'unit': profile.user.unit.name if profile.user and profile.user.unit else 'N/A',
            'end_date': profile.end_date,
            'days_until': days_until,
            'hours_per_week': hours
        }
        
        vacancies_by_home[home_name]['upcoming'].append(vacancy_data)
        vacancies_by_home[home_name]['total_upcoming'] += 1
        vacancies_by_home[home_name]['total_hours'] += hours
    
    # Calculate grand totals
    grand_total_current = sum(home['total_current'] for home in vacancies_by_home.values())
    grand_total_upcoming = sum(home['total_upcoming'] for home in vacancies_by_home.values())
    grand_total_hours = sum(home['total_hours'] for home in vacancies_by_home.values())
    
    context = {
        'vacancies_by_home': dict(vacancies_by_home),
        'care_homes': care_homes,
        'home_filter': home_filter,
        'status_filter': status_filter,
        'grand_total_current': grand_total_current,
        'grand_total_upcoming': grand_total_upcoming,
        'grand_total_hours': grand_total_hours,
        'total_vacancies': grand_total_current + grand_total_upcoming,
    }
    
    return render(request, 'scheduling/staff_vacancies_report.html', context)


@login_required
def staff_vacancies_report_csv(request):
    """
    Export staff vacancies report to CSV
    """
    from staff_records.models import StaffProfile
    import csv
    
    # Check permissions
    if not (request.user.role and (request.user.role.can_manage_rota or request.user.role.is_management)):
        messages.error(request, 'You do not have permission to export this report.')
        return redirect('manager_dashboard')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_vacancies_report.csv"'
    
    writer = csv.writer(response)
    
    # Get filter parameters
    home_filter = request.GET.get('home_filter', '')
    status_filter = request.GET.get('status_filter', 'all')
    
    today = timezone.now().date()
    
    # Title
    writer.writerow(['STAFF VACANCIES REPORT'])
    writer.writerow([f'Generated: {today.strftime("%d %B %Y")}'])
    writer.writerow([])
    
    # Get vacancies
    current_vacancies = StaffProfile.objects.filter(
        employment_status='LEAVER',
        end_date__lte=today
    ).select_related('user', 'user__role', 'user__unit', 'user__unit__care_home')
    
    upcoming_vacancies = StaffProfile.objects.filter(
        employment_status='LEAVER',
        end_date__gt=today
    ).select_related('user', 'user__role', 'user__unit', 'user__unit__care_home')
    
    # Apply filters
    if home_filter:
        current_vacancies = current_vacancies.filter(user__unit__care_home__name=home_filter)
        upcoming_vacancies = upcoming_vacancies.filter(user__unit__care_home__name=home_filter)
    
    # Organize by home
    homes_dict = defaultdict(lambda: {'current': [], 'upcoming': []})
    
    for profile in current_vacancies:
        if profile.user and profile.user.unit and profile.user.unit.care_home:
            home_name = profile.user.unit.care_home.name
            homes_dict[home_name]['current'].append(profile)
    
    for profile in upcoming_vacancies:
        if profile.user and profile.user.unit and profile.user.unit.care_home:
            home_name = profile.user.unit.care_home.name
            homes_dict[home_name]['upcoming'].append(profile)
    
    grand_current = 0
    grand_upcoming = 0
    grand_hours = 0
    
    # Output by home
    for home_name in sorted(homes_dict.keys()):
        home_data = homes_dict[home_name]
        
        writer.writerow(['=' * 100])
        writer.writerow([home_name.replace('_', ' ').title()])
        writer.writerow(['=' * 100])
        writer.writerow([])
        
        # Current vacancies
        if home_data['current'] or status_filter != 'upcoming':
            writer.writerow(['CURRENT VACANCIES (Already Left)'])
            writer.writerow(['-' * 100])
            writer.writerow(['Staff Name', 'SAP ID', 'Role', 'Unit', 'End Date', 'Days Ago', 'Hours/Week', 'Leaving Reason'])
            
            home_current_hours = 0
            for profile in home_data['current']:
                hours = getattr(profile.user, 'hours_per_week', 37.5)
                days_ago = (today - profile.end_date).days if profile.end_date else 0
                
                writer.writerow([
                    profile.user.full_name if profile.user else 'Unknown',
                    profile.user.sap_id if profile.user else 'N/A',
                    profile.user.role.name if profile.user and profile.user.role else 'Unknown',
                    profile.user.unit.name if profile.user and profile.user.unit else 'N/A',
                    profile.end_date.strftime('%d/%m/%Y') if profile.end_date else 'N/A',
                    days_ago,
                    f'{hours:.1f}',
                    profile.leaving_reason or 'Not specified'
                ])
                
                home_current_hours += hours
                grand_current += 1
                grand_hours += hours
            
            writer.writerow([])
            writer.writerow(['Subtotal', '', '', '', '', '', f'{home_current_hours:.1f} hours/week', f'{len(home_data["current"])} positions'])
            writer.writerow([])
        
        # Upcoming leavers
        if home_data['upcoming'] or status_filter != 'current':
            writer.writerow(['UPCOMING LEAVERS (Future End Dates)'])
            writer.writerow(['-' * 100])
            writer.writerow(['Staff Name', 'SAP ID', 'Role', 'Unit', 'End Date', 'Days Until', 'Hours/Week', 'Leaving Reason'])
            
            home_upcoming_hours = 0
            for profile in home_data['upcoming']:
                hours = getattr(profile.user, 'hours_per_week', 37.5)
                days_until = (profile.end_date - today).days if profile.end_date else 0
                
                writer.writerow([
                    profile.user.full_name if profile.user else 'Unknown',
                    profile.user.sap_id if profile.user else 'N/A',
                    profile.user.role.name if profile.user and profile.user.role else 'Unknown',
                    profile.user.unit.name if profile.user and profile.user.unit else 'N/A',
                    profile.end_date.strftime('%d/%m/%Y') if profile.end_date else 'N/A',
                    days_until,
                    f'{hours:.1f}',
                    profile.leaving_reason or 'Not specified'
                ])
                
                home_upcoming_hours += hours
                grand_upcoming += 1
                grand_hours += hours
            
            writer.writerow([])
            writer.writerow(['Subtotal', '', '', '', '', '', f'{home_upcoming_hours:.1f} hours/week', f'{len(home_data["upcoming"])} positions'])
            writer.writerow([])
        
        writer.writerow([])
    
    # Grand totals
    writer.writerow(['=' * 100])
    writer.writerow(['GRAND TOTALS ACROSS ALL HOMES'])
    writer.writerow(['=' * 100])
    writer.writerow([])
    writer.writerow(['Category', 'Number of Positions', 'Total Hours/Week'])
    writer.writerow(['Current Vacancies', grand_current, f'{sum(getattr(p.user, "hours_per_week", 37.5) for p in current_vacancies):.1f}'])
    writer.writerow(['Upcoming Leavers', grand_upcoming, f'{sum(getattr(p.user, "hours_per_week", 37.5) for p in upcoming_vacancies):.1f}'])
    writer.writerow(['TOTAL', grand_current + grand_upcoming, f'{grand_hours:.1f}'])
    
    return response


@login_required
def ai_assistant_page(request):
    """
    Dedicated AI Assistant demo page for showcasing the chatbot
    """
    return render(request, 'scheduling/ai_assistant_page.html')


@login_required
@require_http_methods(["POST"])
def ai_assistant_api(request):
    """
    API endpoint for AI assistant queries with enhanced report generation
    Requires authentication and CSRF protection for security
    
    Phase 3 Features:
    - Conversation context tracking (last 5 queries)
    - Fuzzy name matching for typos and variations
    - Context-aware follow-up question resolution
    """
    import time
    start_time = time.time()
    
    try:
        data = json.loads(request.body)
        query = data.get('query', '').strip()
        
        if not query:
            log_ai_query('', False, error_message='Empty query', user=request.user)
            return JsonResponse({'error': 'No query provided'}, status=400)
        
        # PHASE 3: Get conversation context
        context = get_conversation_context(request)
        
        # PHASE 3: Try to resolve contextual references (follow-up questions)
        resolved_query, context_used = resolve_context_reference(query, context)
        if resolved_query:
            # Add context hint to the response
            original_query = query
            query = resolved_query
            context_hint = f"💡 *Understood as: '{query}' (based on previous context)*\n\n"
        else:
            context_hint = ""
            original_query = query
        
        # Extract entities from query for context tracking
        template_type, template_conf, entities = match_query_template(query)
        
        # PRIORITY 1: Try to process as staff list by role query (who is/list staff)
        # This catches: "Who is SM at Orchard Grove?", "List all SCW", "Show me nurses"
        staff_list_result = _process_staff_list_by_role_query(query)
        if staff_list_result:
            # Add context hint if applicable
            if context_hint:
                staff_list_result['answer'] = context_hint + staff_list_result['answer']
            
            # Update conversation context
            update_conversation_context(request, original_query, 'staffing', entities, staff_list_result)
            
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'staff_list', user=request.user, response_time_ms=response_time)
            return JsonResponse(staff_list_result)
        
        # PRIORITY 2: Try to process as staff count by role query (specific counts)
        # This catches: "How many SCW at Hawthorn?", "Count nurses at Victoria Gardens"
        staff_count_result = _process_staff_count_by_role_query(query)
        if staff_count_result:
            if context_hint:
                staff_count_result['answer'] = context_hint + staff_count_result['answer']
            update_conversation_context(request, original_query, 'staff_count', entities, staff_count_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'staff_count', user=request.user, response_time_ms=response_time)
            return JsonResponse(staff_count_result)
        
        # PRIORITY 3: Try to process as sickness/absence query
        # This catches: "What is the sickness in Orchard Grove?", "Show sickness absence", "How many staff off sick?"
        sickness_result = _process_sickness_query(query)
        if sickness_result:
            if context_hint:
                sickness_result['answer'] = context_hint + sickness_result['answer']
            update_conversation_context(request, original_query, 'sickness', entities, sickness_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'sickness', user=request.user, response_time_ms=response_time)
            return JsonResponse(sickness_result)
        
        # PRIORITY 4: Try to process as home performance/comparison query (Head of Service)
        # This catches: "Show me Orchard Grove's performance", "Compare all homes", "Quality audit for Victoria Gardens"
        home_result = _process_home_performance_query(query)
        if home_result:
            if context_hint:
                home_result['answer'] = context_hint + home_result['answer']
            update_conversation_context(request, original_query, 'home_performance', entities, home_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'home_performance', user=request.user, response_time_ms=response_time)
            return JsonResponse(home_result)
        
        # PRIORITY 5: Try to process as vacancy query
        # This catches: "How many vacancies?", "Show staff leaving", "List leavers"
        vacancy_result = _process_vacancy_query(query)
        if vacancy_result:
            if context_hint:
                vacancy_result['answer'] = context_hint + vacancy_result['answer']
            update_conversation_context(request, original_query, 'vacancy', entities, vacancy_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'vacancy', user=request.user, response_time_ms=response_time)
            return JsonResponse(vacancy_result)
        
        # PRIORITY 6: Try to process as specific staff query (names, SAPs, roles)
        # This catches: "Show me Jane Smith", "How much leave does X have", "List all senior carers"
        staff_result = _process_staff_query(query)
        if staff_result:
            if context_hint:
                staff_result['answer'] = context_hint + staff_result['answer']
            update_conversation_context(request, original_query, 'staff_query', entities, staff_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'staff_query', user=request.user, response_time_ms=response_time)
            return JsonResponse(staff_result)
        
        # PRIORITY 7: Try to process as care plan review query
        # This catches: "When is DEM01 review due?", "How many reviews this month?"
        careplan_result = _process_careplan_query(query)
        if careplan_result:
            if context_hint:
                careplan_result['answer'] = context_hint + careplan_result['answer']
            update_conversation_context(request, original_query, 'careplan', entities, careplan_result)
            response_time = int((time.time() - start_time) * 1000)
            log_ai_query(query, True, 'careplan', user=request.user, response_time_ms=response_time)
            return JsonResponse(careplan_result)
        
        # PRIORITY 4: Try to interpret as report query (more general queries)
        # This catches: "Show staffing summary", "Who is working today?", etc.
        report_result = ReportGenerator.interpret_query(query)
        if report_result:
            report_type = report_result['type']
            
            # shortage_text_message has different structure
            if report_type == 'shortage_text_message':
                shortage_data = report_result.get('shortage_data', {})
                message_data = report_result.get('message_data', {})
                
                if message_data.get('sms_message'):
                    answer = f"**📱 Staff Shortage Alert Message Generated**\n\n"
                    answer += f"**SMS Version** ({message_data['sms_count']} message{'s' if message_data['sms_count'] > 1 else ''}, {message_data['sms_length']} characters):\n"
                    answer += f"```\n{message_data['sms_message']}\n```\n\n"
                    
                    answer += f"**📧 Email Version:**\n```\n{message_data['email_message'][:500]}...\n```\n\n"
                    
                    answer += f"**📊 Message Details:**\n"
                    answer += f"• Will be sent to: {message_data['recipient_count']} active staff\n"
                    answer += f"• Critical dates: {', '.join(message_data['critical_dates'])}\n"
                    answer += f"• Total shifts needed: {message_data['total_gaps']}\n\n"
                    
                    answer += f"**⚠️ Next Steps:**\n"
                    answer += f"1. Review the message content above\n"
                    answer += f"2. Copy the SMS text to your messaging system\n"
                    answer += f"3. Send to all active staff OR use bulk SMS service\n"
                    answer += f"4. Monitor responses and update rota accordingly\n\n"
                    
                    answer += f"💡 **Tip:** For mass SMS, export staff phone numbers and use your SMS provider's bulk send feature."
                    
                    return JsonResponse({
                        'answer': answer,
                        'related': ['Export Staff Contacts', 'View Shortages', 'Update Rota', 'Send Now'],
                        'category': 'report',
                        'report_type': report_type,
                        'report_data': {
                            'shortage_summary': shortage_data,
                            'message': message_data
                        }
                    })
                else:
                    answer = f"**No Message Needed**\n\n"
                    answer += f"✅ {message_data.get('reason', 'No staffing shortages detected')}\n\n"
                    answer += f"All shifts for the next 14 days are fully covered. Great job!"
                    
                    return JsonResponse({
                        'answer': answer,
                        'related': ['View Rota', 'Check Coverage'],
                        'category': 'report',
                        'report_type': report_type,
                        'report_data': message_data
                    })
            
            # All other report types have 'data' key
            report_data = report_result['data']
            
            # Format the response based on report type
            if report_type == 'staffing_summary':
                answer = f"**Staffing Summary**\n\n{report_data['summary']}\n\n**Breakdown by Role:**\n"
                for role, count in report_data['by_role'].items():
                    answer += f"• {role}: {count}\n"
                
                if report_data.get('by_grade'):
                    answer += f"\n**Breakdown by Grade:**\n"
                    for grade, count in sorted(report_data['by_grade'].items()):
                        answer += f"• {grade}: {count}\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View Dashboard', 'Generate Report', 'Export Data'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
            
            elif report_type == 'staff_by_grade':
                answer = f"**Staff by Grade Report**\n\n{report_data['summary']}\n\n"
                
                if report_data.get('grade_queried'):
                    # Specific grade query
                    if report_data['count'] > 0:
                        answer += f"**Staff Members:**\n"
                        for i, staff in enumerate(report_data.get('staff', [])[:10], 1):  # Limit to 10
                            answer += f"{i}. {staff['name']} (SAP: {staff['sap']}) - {staff['role']}\n"
                            if staff.get('unit'):
                                answer += f"   Unit: {staff['unit']}\n"
                        
                        if len(report_data.get('staff', [])) > 10:
                            answer += f"\n... and {len(report_data['staff']) - 10} more\n"
                        
                        if report_data.get('matching_grades'):
                            answer += f"\n**Breakdown by Matching Grades:**\n"
                            for grade, count in sorted(report_data['matching_grades'].items()):
                                answer += f"• {grade}: {count}\n"
                else:
                    # All grades query
                    answer += f"**Breakdown by Grade:**\n"
                    for grade, count in sorted(report_data.get('by_grade', {}).items()):
                        answer += f"• {grade}: {count}\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View Staff List', 'Export Report', 'Staff Details'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
            
            elif report_type == 'sickness_report':
                answer = f"**Sickness Report**\n\n{report_data['summary']}\n\n"
                if report_data['sick_staff']:
                    answer += "**Currently Off Sick:**\n"
                    for staff in report_data['sick_staff']:
                        answer += f"• {staff['name']} ({staff['sap']}) - {staff['days']} days - {staff['status']}\n"
                else:
                    answer += "✅ No staff currently off sick\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View Sickness Records', 'Send Reminders', 'Generate Report'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
            
            elif report_type == 'staffing_shortage':
                answer = f"**Staffing Shortage Analysis**\n\n{report_data['summary']}\n\n"
                
                # Collect all recommendations for actionable buttons
                all_recommendations = []
                recommendation_date = None
                recommendation_reason = None
                
                # Show TRUE SHORTAGES (below minimum) if any exist
                if report_data.get('shortage_days'):
                    answer += "🚨 **CRITICAL SHORTAGES - Below Minimum Safe Staffing:**\n"
                    for day in report_data['shortage_days'][:7]:
                        date_str = day['date'].strftime('%a %d %b')
                        
                        # Highlight which shifts are critical
                        critical_shifts = []
                        if day.get('day_below_minimum'):
                            critical_shifts.append(f"DAY: {day['total_day_staff']}/17 staff")
                        if day.get('night_below_minimum'):
                            critical_shifts.append(f"NIGHT: {day['total_night_staff']}/17 staff")
                        
                        answer += f"\n📅 **{date_str}** - {' | '.join(critical_shifts)} 🚨\n"
                        
                        # Show unit breakdown for context
                        for shortage in day['shortages'][:3]:
                            if shortage.get('day_gap', 0) > 0:
                                answer += f"  • {shortage['unit']}: needs {shortage['day_gap']} day staff (current: {shortage['day_current']})\n"
                            if shortage.get('night_gap', 0) > 0:
                                answer += f"  • {shortage['unit']}: needs {shortage['night_gap']} night staff (current: {shortage['night_current']})\n"
                    
                    answer += f"\n🚨 **ACTION REQUIRED:** Contact agency staff or arrange overtime immediately!\n"
                    answer += f"\n💡 **Tip:** Ask me to 'generate text message for staff shortage' to alert all staff.\n"
                
                # Show REALLOCATION NEEDS (adequate staff but imbalanced) separately
                elif report_data.get('reallocation_days'):
                    answer += "ℹ️ **Automated Staff Reallocation Plan:**\n\n"
                    answer += f"✅ **Good news:** All days have adequate total staffing (17+ staff)\n"
                    answer += f"📋 **Action needed:** Reallocate staff between units as suggested below\n\n"
                    
                    for day in report_data['reallocation_days'][:7]:
                        date_str = day['date'].strftime('%a %d %b')
                        answer += f"\n📅 **{date_str}** - {day['total_day_staff']} day staff, {day['total_night_staff']} night staff ✅\n"
                        
                        # Show automated reallocation plan
                        realloc_plan = day.get('reallocation_plan', {})
                        
                        # Collect moves for first day's recommendations (most urgent)
                        if not all_recommendations and not recommendation_date:
                            recommendation_date = day['date'].strftime('%Y-%m-%d')
                            recommendation_reason = f"Balance staffing for {day['day_name']}"
                            
                            for move in realloc_plan.get('day', []):
                                all_recommendations.append(move)
                            for move in realloc_plan.get('night', []):
                                all_recommendations.append(move)
                        
                        if realloc_plan.get('day'):
                            answer += f"\n**DAY SHIFT REALLOCATIONS:**\n"
                            for move in realloc_plan['day']:
                                answer += f"  ➡️ Move {move['staff_name']} ({move['role']})\n"
                                answer += f"     FROM: {move['from_unit']} → TO: {move['to_unit']}\n"
                        
                        if realloc_plan.get('night'):
                            answer += f"\n**NIGHT SHIFT REALLOCATIONS:**\n"
                            for move in realloc_plan['night']:
                                answer += f"  ➡️ Move {move['staff_name']} ({move['role']})\n"
                                answer += f"     FROM: {move['from_unit']} → TO: {move['to_unit']}\n"
                        
                        if not realloc_plan.get('day') and not realloc_plan.get('night'):
                            answer += f"  **Current imbalances:**\n"
                            for imbalance in day['imbalances'][:5]:
                                if imbalance.get('day_gap', 0) > 0:
                                    answer += f"  • {imbalance['unit']}: {imbalance['day_gap']} more day staff needed (current: {imbalance['day_current']})\n"
                                if imbalance.get('night_gap', 0) > 0:
                                    answer += f"  • {imbalance['unit']}: {imbalance['night_gap']} more night staff needed (current: {imbalance['night_current']})\n"
                
                # Build response with actionable recommendations if available
                response_data = {
                    'answer': answer,
                    'related': ['View Detailed Report', 'Generate Alert Message', 'View Rota'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                }
                
                # Add recommendations if we have moves to suggest
                if all_recommendations:
                    import uuid
                    response_data['recommendations'] = all_recommendations
                    response_data['recommendation_id'] = str(uuid.uuid4())
                    response_data['date'] = recommendation_date
                    response_data['reason'] = recommendation_reason
                else:
                    # No recommendations but check if we need to add helpful tips
                    if report_data.get('reallocation_days'):
                        answer += f"\n💡 **How to apply:** Go to Rota View → Edit shifts to reassign staff as suggested above.\n"
                        answer += f"**No agency/OT required** - you have enough staff, just redistribute them!\n"
                        response_data['answer'] = answer
                    elif not report_data.get('shortage_days'):
                        answer += f"✅ **No issues detected!**\n\n"
                        answer += f"• All days have minimum 17 staff ✅\n"
                        answer += f"• All units are properly balanced ✅\n"
                        answer += f"• Next {report_data.get('days_analyzed', 14)} days fully covered!\n"
                        response_data['answer'] = answer
                
                return JsonResponse(response_data)
            
            elif report_type == 'incident_report':
                answer = f"**Incident Report**\n\n{report_data['summary']}\n\n"
                if report_data['by_severity']:
                    answer += "**By Severity:**\n"
                    severity_emoji = {
                        'Death': '☠️',
                        'Major Harm': '🔴',
                        'Moderate Harm': '🟠',
                        'Low Harm': '🟡',
                        'No Harm': '🟢'
                    }
                    for severity, count in report_data['by_severity'].items():
                        emoji = severity_emoji.get(severity, '•')
                        answer += f"{emoji} {severity}: {count}\n"
                
                if report_data['critical_alerts']:
                    answer += f"\n**⚠️ Critical Alerts:**\n"
                    for alert in report_data['critical_alerts']:
                        answer += f"{alert}\n"
                
                if report_data['ci_notifications_required'] > 0:
                    answer += f"\n⚠️ {report_data['ci_notifications_required']} incidents require Care Inspectorate notification\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View All Incidents', 'Notify CI', 'Generate Report'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
            
            elif report_type == 'shift_coverage':
                answer = f"**Shift Coverage**\n\n{report_data['summary']}\n\n"
                if report_data['by_unit']:
                    answer += "**Coverage by Unit:**\n"
                    for unit, coverage in report_data['by_unit'].items():
                        if coverage['total'] > 0:
                            answer += f"\n**{unit}:** Day {coverage['day']} | Night {coverage['night']}\n"
                            
                            # Show staff names
                            staff_data = report_data.get('staff_by_unit', {}).get(unit, {})
                            if staff_data.get('day_staff'):
                                answer += f"  Day shift: {', '.join(staff_data['day_staff'][:10])}\n"
                                if len(staff_data['day_staff']) > 10:
                                    answer += f"  ... and {len(staff_data['day_staff']) - 10} more\n"
                            if staff_data.get('night_staff'):
                                answer += f"  Night shift: {', '.join(staff_data['night_staff'][:10])}\n"
                                if len(staff_data['night_staff']) > 10:
                                    answer += f"  ... and {len(staff_data['night_staff']) - 10} more\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View Rota', 'Edit Shifts', 'Generate Report'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
            
            elif report_type == 'leave_summary':
                answer = f"**Leave Summary**\n\n{report_data['summary']}\n\n"
                if report_data['low_balance_staff']:
                    answer += f"**Staff with Low Leave Balance (<40 hours):**\n"
                    for staff in report_data['low_balance_staff']:
                        answer += f"• {staff['name']}: {staff['hours']:.1f} hours remaining\n"
                
                return JsonResponse({
                    'answer': answer,
                    'related': ['View Leave Requests', 'Approve Leave', 'Send Reminders'],
                    'category': 'report',
                    'report_type': report_type,
                    'report_data': report_data
                })
        
        # PRIORITY 4: Import the HelpAssistant for general help (fallback)
        from .management.commands.help_assistant import HelpAssistant
        assistant = HelpAssistant()
        
        # Special handling for "show all topics"
        if 'show all topics' in query.lower() or 'list topics' in query.lower():
            topics_list = []
            for category, items in assistant.knowledge_base.items():
                for key, data in items.items():
                    title = data.get('question', [''])[0].title()
                    topics_list.append(f"• {title}")
            
            answer = "**Available Help Topics:**\n\n" + "\n".join(topics_list[:20])
            if len(topics_list) > 20:
                answer += f"\n\n...and {len(topics_list) - 20} more topics!"
            
            return JsonResponse({
                'answer': answer,
                'related': [],
                'category': 'topics'
            })
        
        # Find the answer in knowledge base
        result = assistant.find_answer(query)
        
        if result:
            return JsonResponse({
                'answer': result['answer'],
                'related': result.get('related', []),
                'category': result.get('category', '')
            })
        else:
            # No exact match - provide smart suggestions based on query content
            query_lower = query.lower()
            
            # Detect what user might be asking about
            suggestions = []
            category_detected = None
            
            if any(word in query_lower for word in ['confidence', 'score', 'low', 'percentage', '%']):
                category_detected = "**About Confidence Scores:**\n\nThe AI calculates confidence based on how well it understands your question.\n\n"
                suggestions = [
                    "💡 Try asking: 'Why is confidence low?' or 'What is confidence score?'",
                    "• Be more specific with names, dates, and care homes",
                    "• Use proper terminology: SCW, OM, Hawthorn House, etc.",
                    "• Try the quick action buttons below for common queries"
                ]
            elif any(word in query_lower for word in ['help', 'how', 'what can', 'capabilities', 'do']):
                category_detected = "**AI Assistant Help:**\n\n"
                suggestions = [
                    "💡 Ask: 'What can you do?' to see all my capabilities",
                    "💡 Ask: 'How to ask questions?' for query tips and examples",
                    "💡 Ask: 'Show all topics' to see everything I can answer",
                    "• Try the example queries shown above",
                    "• Use the quick action buttons for instant results"
                ]
            elif any(word in query_lower for word in ['staff', 'worker', 'employee', 'carer', 'nurse']):
                category_detected = "**Staff Queries - Try These:**\n\n"
                suggestions = [
                    "✅ 'Show me [Name] details' - View specific staff member",
                    "✅ 'List all [Role] at [Home]' - e.g., 'List all SCW at Hawthorn House'",
                    "✅ 'How many [Role] at [Home]?' - Count staff by role",
                    "✅ 'Who is working today?' - Today's staff roster",
                    "✅ 'Search for [Name]' - Find staff by name"
                ]
            elif any(word in query_lower for word in ['coverage', 'shortage', 'short staffed', 'rota', 'roster']):
                category_detected = "**Coverage & Shortages - Try These:**\n\n"
                suggestions = [
                    "✅ 'What's the coverage today?' - Today's staffing levels",
                    "✅ 'Are we short staffed next week?' - Future shortage check",
                    "✅ 'Show staffing shortage' - Detailed shortage analysis",
                    "✅ 'Coverage for [date]' - Specific date coverage",
                    "✅ 'Check shortages at [Home]' - Home-specific check"
                ]
            elif any(word in query_lower for word in ['sick', 'sickness', 'absence', 'ill', 'off']):
                category_detected = "**Sickness & Absence - Try These:**\n\n"
                suggestions = [
                    "✅ 'Who is off sick today?' - Current sickness list",
                    "✅ 'Sickness report for [Home]' - Home-specific sickness",
                    "✅ 'Show me all sickness absence' - Full sickness overview",
                    "✅ 'How many staff off sick?' - Sickness count"
                ]
            elif any(word in query_lower for word in ['leave', 'holiday', 'annual leave', 'time off']):
                category_detected = "**Annual Leave - Try These:**\n\n"
                suggestions = [
                    "✅ 'How much leave does [Name/SAP] have?' - Leave balance check",
                    "✅ 'Show leave balance for [Name]' - Specific staff leave",
                    "✅ 'List approved leave this week' - Weekly leave schedule",
                    "✅ 'Annual leave summary' - Overview report"
                ]
            elif any(word in query_lower for word in ['care plan', 'review', 'resident', 'chi']):
                category_detected = "**Care Plan Reviews - Try These:**\n\n"
                suggestions = [
                    "✅ 'When is [Resident ID] review due?' - Specific resident review",
                    "✅ 'Show overdue care plan reviews' - Overdue reviews list",
                    "✅ 'How many reviews this month?' - Monthly review count",
                    "✅ 'Care plan compliance status' - Compliance overview"
                ]
            elif any(word in query_lower for word in ['training', 'course', 'compliance']):
                category_detected = "**Training Compliance - Try These:**\n\n"
                suggestions = [
                    "✅ 'Training compliance breakdown' - Full training report",
                    "✅ 'Show training by person' - Person-view training matrix",
                    "✅ 'Training report for [Home]' - Home-specific training",
                    "✅ 'Who needs [course name] training?' - Course-specific check"
                ]
            elif any(word in query_lower for word in ['home', 'performance', 'quality', 'orchard', 'hawthorn', 'victoria', 'riverside', 'meadowburn']):
                category_detected = "**Home Performance - Try These:**\n\n"
                suggestions = [
                    "✅ 'Show me [Home] performance' - Specific home dashboard",
                    "✅ 'Compare all care homes' - Multi-home comparison",
                    "✅ 'Quality audit for [Home]' - Quality metrics",
                    "✅ 'Performance dashboard' - Overall performance"
                ]
            else:
                # Generic fallback
                category_detected = "**I can help with many things! Try these:**\n\n"
                suggestions = [
                    "💡 Ask: 'What can you do?' - See all my capabilities",
                    "💡 Ask: 'How to ask questions?' - Get query tips and examples",
                    "💡 Ask: 'Show all topics' - See complete help topics",
                    "• Use specific names, dates, and care homes in your questions",
                    "• Try the quick action buttons below for instant results"
                ]
            
            # Build comprehensive fallback response
            fallback_answer = f"""❓ I didn't quite understand: **"{query}"**

{category_detected}
{chr(10).join(suggestions)}

**Common Query Examples:**

📋 **Staff:** "Show me Jane Smith details" | "List all SCW at Hawthorn House"
📊 **Coverage:** "What's the coverage today?" | "Are we short staffed next week?"
🤒 **Sickness:** "Who is off sick today?" | "Sickness report for Orchard Grove"
💼 **Leave:** "How much leave does ADMIN001 have?" | "Show leave balance"
📝 **Reviews:** "When is CHI0101451001AC review due?" | "Show overdue reviews"
🏥 **Performance:** "Show me Orchard Grove's performance" | "Compare all homes"
🎓 **Training:** "Training compliance breakdown" | "Show training by person"

**Need Help?**
• Ask: "What can you do?" for complete capabilities
• Ask: "How to ask?" for query tips and best practices
• Use the quick action buttons below for instant results

**💡 Tip:** Be specific! Include names, dates, and locations for best results.
"""
            
            return JsonResponse({
                'answer': fallback_answer,
                'related': ['What Can You Do?', 'How To Ask Questions', 'Show All Topics', 'Staff Queries', 'Coverage Reports'],
                'category': 'help'
            })
        
        # If nothing matched, return helpful error with suggestions
        suggestions = generate_helpful_suggestions(query)
        error_msg = f"❓ I couldn't understand: **{query}**\n\n**Try asking:**\n"
        error_msg += "\n".join(f"• {s}" for s in suggestions)
        error_msg += "\n\n💡 **Tip:** Be specific - mention staff names, dates, or care homes."
        
        response_time = int((time.time() - start_time) * 1000)
        log_ai_query(query, False, 'error', error_message='No handler matched', user=request.user, response_time_ms=response_time)
        
        return JsonResponse({
            'answer': error_msg,
            'suggestions': suggestions,
            'related': ['Staffing Report', 'Care Plan Reviews', 'Staff Vacancies', 'Leave Requests'],
            'category': 'help'
        })
    
    except json.JSONDecodeError:
        log_ai_query('', False, 'error', error_message='Invalid JSON', user=request.user)
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        error_detail = str(e)
        log_ai_query(query if 'query' in locals() else '', False, 'error', error_message=error_detail, user=request.user)
        return JsonResponse({'error': f'Server error: {error_detail}'}, status=500)


@login_required
def leave_usage_targets(request):
    """
    Leave Usage Targets Dashboard - shows how staff are tracking against
    the 40-week leave usage target to prevent year-end build-up
    """
    from staff_records.models import AnnualLeaveEntitlement, StaffProfile
    from datetime import date, timedelta
    from decimal import Decimal
    
    # Check permissions - management only
    if not (request.user.role and request.user.role.is_management):
        return redirect('staff_dashboard')
    
    # Get year parameter
    year = int(request.GET.get('year', date.today().year))
    unit_filter = request.GET.get('unit', '')
    
    # Calculate weeks elapsed in the year
    year_start = date(year, 1, 1)
    today = date.today()
    
    if year == today.year:
        days_elapsed = (today - year_start).days
    elif year < today.year:
        # Past year - use full year
        days_elapsed = 365
    else:
        # Future year - use 0
        days_elapsed = 0
    
    weeks_elapsed = days_elapsed / 7
    weeks_remaining = max(0, 52 - weeks_elapsed)
    
    # 40-week strategy constants
    TARGET_WEEKS = 40
    
    # Get all entitlements for the year
    entitlements = AnnualLeaveEntitlement.objects.filter(
        leave_year_start__year=year
    ).select_related('profile__user__role', 'profile__user__unit')
    
    # Apply unit filter if specified
    if unit_filter:
        entitlements = entitlements.filter(profile__user__unit__name=unit_filter)
    
    # Process each entitlement and calculate status
    staff_data = []
    status_counts = {'green': 0, 'amber': 0, 'red': 0}
    
    for ent in entitlements:
        user = ent.profile.user
        
        # Skip inactive staff
        if not user.is_active:
            continue
        
        # Calculate total hours (including carryover)
        total_hours = float(ent.total_entitlement_hours)
        used_hours = float(ent.hours_used)
        remaining_hours = float(ent.hours_remaining)
        pending_hours = float(ent.hours_pending)
        
        # Convert to days (assuming 7.5 hour day)
        total_days = total_hours / 7.5
        used_days = used_hours / 7.5
        remaining_days = remaining_hours / 7.5
        pending_days = pending_hours / 7.5
        
        # Calculate 40-week targets
        target_hours_per_week = total_hours / TARGET_WEEKS
        target_days_per_week = total_days / TARGET_WEEKS
        
        # Calculate expected usage based on weeks elapsed
        if weeks_elapsed > 0:
            # Use the lesser of weeks_elapsed or TARGET_WEEKS for calculation
            effective_weeks = min(weeks_elapsed, TARGET_WEEKS)
            expected_hours = target_hours_per_week * effective_weeks
            expected_days = target_days_per_week * effective_weeks
        else:
            expected_hours = 0
            expected_days = 0
        
        # Calculate actual vs expected
        variance_hours = used_hours - expected_hours
        variance_days = used_days - expected_days
        
        # Calculate percentage of expected
        if expected_hours > 0:
            percent_of_expected = (used_hours / expected_hours) * 100
        else:
            percent_of_expected = 0
        
        # Determine status (traffic light)
        # Green: within ±10% of expected or ahead
        # Amber: 10-25% behind expected
        # Red: more than 25% behind expected
        if percent_of_expected >= 90:
            status = 'green'
            status_label = 'On Target'
            status_icon = '✅'
        elif percent_of_expected >= 75:
            status = 'amber'
            status_label = 'Slightly Behind'
            status_icon = '⚠️'
        else:
            status = 'red'
            status_label = 'Significantly Behind'
            status_icon = '🔴'
        
        status_counts[status] += 1
        
        # Calculate monthly target
        target_days_per_month = target_days_per_week * 4.33
        
        # Calculate what's needed for rest of year
        if weeks_remaining > 0:
            hours_per_week_needed = remaining_hours / weeks_remaining
            days_per_week_needed = remaining_days / weeks_remaining
        else:
            hours_per_week_needed = 0
            days_per_week_needed = 0
        
        staff_data.append({
            'user': user,
            'sap': user.sap,
            'name': user.full_name,
            'unit': user.unit.name if user.unit else 'No Unit',
            'role': user.role.name if user.role else 'No Role',
            'total_days': round(total_days, 1),
            'used_days': round(used_days, 1),
            'remaining_days': round(remaining_days, 1),
            'pending_days': round(pending_days, 1),
            'expected_days': round(expected_days, 1),
            'variance_days': round(variance_days, 1),
            'percent_of_expected': round(percent_of_expected, 1),
            'status': status,
            'status_label': status_label,
            'status_icon': status_icon,
            'target_days_per_week': round(target_days_per_week, 2),
            'target_days_per_month': round(target_days_per_month, 1),
            'days_per_week_needed': round(days_per_week_needed, 2),
        })
    
    # Sort by status (red first, then amber, then green) and name
    status_order = {'red': 0, 'amber': 1, 'green': 2}
    staff_data.sort(key=lambda x: (status_order[x['status']], x['name']))
    
    # Calculate team totals
    total_entitled_days = sum(s['total_days'] for s in staff_data)
    total_used_days = sum(s['used_days'] for s in staff_data)
    total_remaining_days = sum(s['remaining_days'] for s in staff_data)
    total_expected_days = sum(s['expected_days'] for s in staff_data)
    
    # Get unique units for filter dropdown
    all_units = User.objects.filter(
        is_active=True, 
        unit__isnull=False
    ).values_list('unit__name', flat=True).distinct().order_by('unit__name')
    
    context = {
        'year': year,
        'current_year': date.today().year,
        'weeks_elapsed': round(weeks_elapsed, 1),
        'weeks_remaining': round(weeks_remaining, 1),
        'target_weeks': TARGET_WEEKS,
        'staff_data': staff_data,
        'total_staff': len(staff_data),
        'status_counts': status_counts,
        'total_entitled_days': round(total_entitled_days, 1),
        'total_used_days': round(total_used_days, 1),
        'total_remaining_days': round(total_remaining_days, 1),
        'total_expected_days': round(total_expected_days, 1),
        'team_percent_of_expected': round((total_used_days / total_expected_days * 100) if total_expected_days > 0 else 0, 1),
        'all_units': all_units,
        'selected_unit': unit_filter,
    }
    
    return render(request, 'scheduling/leave_usage_targets.html', context)


# ============================================================================
# CARE PLAN REVIEW VIEWS
# ============================================================================

@login_required
def careplan_overview(request):
    """Quick overview of all residents and their care plan review status"""
    from .models import Resident, CarePlanReview
    from .models_multi_home import CareHome
    
    # Get filter parameters
    care_home_filter = request.GET.get('care_home', '')
    unit_filter = request.GET.get('unit', '')
    status_filter = request.GET.get('status', '')
    
    # Get all residents with their latest review
    residents = Resident.objects.filter(is_active=True).select_related(
        'unit', 'unit__care_home', 'keyworker', 'unit_manager'
    ).prefetch_related('care_plan_reviews')
    
    # Apply filters
    if care_home_filter:
        residents = residents.filter(unit__care_home__name=care_home_filter)
    
    if unit_filter:
        residents = residents.filter(unit__name=unit_filter)
    
    # Build resident data
    resident_data = []
    for resident in residents.order_by('unit__name', 'room_number'):
        # Get latest review
        latest_review = resident.care_plan_reviews.order_by('-due_date').first()
        
        # Apply status filter
        if status_filter and latest_review:
            if latest_review.status != status_filter:
                continue
        elif status_filter:
            continue
        
        resident_data.append({
            'resident': resident,
            'review': latest_review,
        })
    
    # Get units for filter - respect care home filter
    if care_home_filter:
        all_units = Unit.objects.filter(care_home__name=care_home_filter).order_by('name')
    else:
        all_units = Unit.objects.all().order_by('name')
    
    # Get all care homes for filter dropdown
    all_care_homes = CareHome.objects.all().order_by('name')
    
    # Calculate statistics
    total_residents = len(resident_data)
    reviews = [r['review'] for r in resident_data if r['review']]
    
    completed_count = sum(1 for r in reviews if r.status == 'COMPLETED')
    overdue_count = sum(1 for r in reviews if r.status == 'OVERDUE')
    due_count = sum(1 for r in reviews if r.status == 'DUE')
    upcoming_count = sum(1 for r in reviews if r.status == 'UPCOMING')
    
    compliance_rate = round((completed_count / len(reviews) * 100) if reviews else 0, 1)
    
    # Calculate unit-level statistics
    unit_stats = []
    for unit in all_units:
        unit_residents = Resident.objects.filter(unit=unit, is_active=True)
        unit_resident_count = unit_residents.count()
        
        if unit_resident_count > 0:
            unit_reviews = []
            for res in unit_residents:
                latest = res.care_plan_reviews.order_by('-due_date').first()
                if latest:
                    unit_reviews.append(latest)
            
            unit_completed = sum(1 for r in unit_reviews if r.status == 'COMPLETED')
            unit_overdue = sum(1 for r in unit_reviews if r.status == 'OVERDUE')
            unit_compliance = round((unit_completed / len(unit_reviews) * 100) if unit_reviews else 0, 1)
            
            unit_stats.append({
                'unit': unit,
                'total': unit_resident_count,
                'completed': unit_completed,
                'overdue': unit_overdue,
                'compliance_rate': unit_compliance,
            })
    
    context = {
        'residents': resident_data,
        'total_residents': total_residents,
        'completed_count': completed_count,
        'overdue_count': overdue_count,
        'due_count': due_count,
        'upcoming_count': upcoming_count,
        'compliance_rate': compliance_rate,
        'all_units': all_units,
        'all_care_homes': all_care_homes,
        'unit_stats': unit_stats,
        'selected_care_home': care_home_filter,
        'selected_unit': unit_filter,
        'selected_status': status_filter,
        'status_choices': [
            ('COMPLETED', 'Completed'),
            ('OVERDUE', 'Overdue'),
            ('DUE', 'Due Soon'),
            ('UPCOMING', 'Upcoming'),
            ('IN_PROGRESS', 'In Progress'),
            ('PENDING_APPROVAL', 'Pending Approval'),
        ],
    }
    
    return render(request, 'scheduling/careplan_overview.html', context)


@login_required
def careplan_unit_view(request, unit_name):
    """Detailed view of residents in a specific unit"""
    from .models import Resident, CarePlanReview
    
    unit = get_object_or_404(Unit, name=unit_name)
    
    # Get all residents without ordering first
    residents = Resident.objects.filter(
        unit=unit,
        is_active=True
    ).select_related('keyworker', 'unit_manager').prefetch_related('care_plan_reviews')
    
    # Build resident data with reviews
    resident_data = []
    for resident in residents:
        latest_review = resident.care_plan_reviews.order_by('-due_date').first()
        resident_data.append({
            'resident': resident,
            'review': latest_review,
        })
    
    # Sort by room number naturally (handles numbers properly: 1, 2, 10, 11 instead of 1, 10, 11, 2)
    def natural_sort_key(item):
        room = item['resident'].room_number or ''
        # Extract numeric part for proper sorting
        import re
        parts = re.split(r'(\d+)', room)
        return [int(part) if part.isdigit() else part.lower() for part in parts]
    
    resident_data.sort(key=natural_sort_key)
    
    # Calculate unit statistics
    reviews = [r['review'] for r in resident_data if r['review']]
    completed_count = sum(1 for r in reviews if r.status == 'COMPLETED')
    overdue_count = sum(1 for r in reviews if r.status == 'OVERDUE')
    due_count = sum(1 for r in reviews if r.status == 'DUE')
    
    context = {
        'unit': unit,
        'residents': resident_data,
        'total_residents': len(resident_data),
        'completed_count': completed_count,
        'overdue_count': overdue_count,
        'due_count': due_count,
        'compliance_rate': round((completed_count / len(reviews) * 100) if reviews else 0, 1),
    }
    
    return render(request, 'scheduling/careplan_unit_view.html', context)


@login_required
def careplan_review_detail(request, review_id):
    """View and edit a specific care plan review"""
    from .models import CarePlanReview
    
    review = get_object_or_404(CarePlanReview, id=review_id)
    
    if request.method == 'POST':
        # Handle form submission
        review.care_needs_assessment = request.POST.get('care_needs_assessment', '')
        review.goals_progress = request.POST.get('goals_progress', '')
        review.changes_required = request.POST.get('changes_required', '')
        review.family_involvement = request.POST.get('family_involvement', '')
        
        action = request.POST.get('action')
        
        if action == 'save_draft':
            review.status = 'IN_PROGRESS'
            messages.success(request, 'Review saved as draft')
        elif action == 'submit':
            review.status = 'PENDING_APPROVAL'
            review.completed_by = request.user
            review.completed_date = date.today()
            messages.success(request, 'Review submitted for manager approval')
        elif action == 'approve' and request.user == review.unit_manager:
            review.status = 'COMPLETED'
            review.manager_approved = True
            review.manager_approval_date = date.today()
            review.manager_approved_by = request.user
            review.manager_comments = request.POST.get('manager_comments', '')
            review.save()
            
            # Auto-generate next review
            from datetime import timedelta
            next_due_date = review.completed_date + timedelta(days=183)  # 6 months
            
            CarePlanReview.objects.create(
                resident=review.resident,
                review_type='SIX_MONTH',
                due_date=next_due_date,
                keyworker=review.resident.keyworker,
                unit_manager=review.resident.unit_manager,
                status='UPCOMING'
            )
            
            messages.success(request, f'Review approved and completed. Next review scheduled for {next_due_date.strftime("%B %d, %Y")}')
            return redirect('careplan_review_detail', review_id=review.id)
        
        review.save()
        return redirect('careplan_review_detail', review_id=review.id)
    
    context = {
        'review': review,
        'resident': review.resident,
    }
    
    return render(request, 'scheduling/careplan_review_detail.html', context)


@login_required
def careplan_compliance_report(request):
    """Generate compliance report for care plan reviews between two dates"""
    from .models import CarePlanReview
    from datetime import datetime, timedelta
    
    # Get date range from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    unit_filter = request.GET.get('unit', '')
    
    # Default to last 30 days if no dates provided
    if not start_date_str or not end_date_str:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Get reviews completed in date range
    reviews = CarePlanReview.objects.filter(
        completed_date__gte=start_date,
        completed_date__lte=end_date,
        status='COMPLETED'
    ).select_related('resident', 'resident__unit', 'keyworker', 'completed_by')
    
    # Apply unit filter
    if unit_filter:
        reviews = reviews.filter(resident__unit__name=unit_filter)
    
    # Calculate statistics
    total_reviews = reviews.count()
    on_time_reviews = 0
    overdue_reviews = 0
    days_overdue_total = 0
    
    review_details = []
    for review in reviews:
        # Check if it was completed on time
        was_on_time = review.completed_date <= review.due_date
        
        if was_on_time:
            on_time_reviews += 1
            days_late = 0
        else:
            overdue_reviews += 1
            days_late = (review.completed_date - review.due_date).days
            days_overdue_total += days_late
        
        review_details.append({
            'review': review,
            'was_on_time': was_on_time,
            'days_late': days_late,
        })
    
    # Calculate percentages
    on_time_percentage = round((on_time_reviews / total_reviews * 100) if total_reviews > 0 else 0, 1)
    overdue_percentage = round((overdue_reviews / total_reviews * 100) if total_reviews > 0 else 0, 1)
    avg_days_late = round((days_overdue_total / overdue_reviews) if overdue_reviews > 0 else 0, 1)
    
    # Get all units for filter
    all_units = Unit.objects.all().order_by('name')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_reviews': total_reviews,
        'on_time_reviews': on_time_reviews,
        'overdue_reviews': overdue_reviews,
        'on_time_percentage': on_time_percentage,
        'overdue_percentage': overdue_percentage,
        'avg_days_late': avg_days_late,
        'review_details': review_details,
        'all_units': all_units,
        'selected_unit': unit_filter,
    }
    
    return render(request, 'scheduling/careplan_compliance_report.html', context)


@login_required
def careplan_manager_dashboard(request):
    """
    Manager Dashboard for Care Plan Reviews
    - Approval queue (pending manager approval)
    - Compliance overview by unit
    - Overdue alerts
    - Quick statistics
    """
    # Check if user is a manager
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, "Access denied. This page is for managers only.")
        return redirect('careplan_overview')
    
    # Unit and care home filters
    unit_filter = request.GET.get('unit', '')
    care_home_filter = request.GET.get('care_home', '')
    
    # Base queryset
    all_reviews = CarePlanReview.objects.select_related('resident', 'resident__unit', 'resident__unit__care_home', 'keyworker', 'completed_by')
    
    # Apply care home filter first
    if care_home_filter:
        all_reviews = all_reviews.filter(resident__unit__care_home__name=care_home_filter)
    
    # Apply unit filter
    if unit_filter:
        all_reviews = all_reviews.filter(resident__unit__name=unit_filter)
    
    # Pending Approvals - Reviews submitted by keyworkers awaiting manager approval
    pending_approvals = all_reviews.filter(
        status='PENDING_APPROVAL'
    ).order_by('due_date')
    
    # Overdue Reviews - Critical items needing immediate attention
    overdue_reviews = all_reviews.filter(
        status='OVERDUE'
    ).order_by('due_date')
    
    # Due Soon - Reviews due within 7 days
    today = timezone.now().date()
    due_soon_date = today + timedelta(days=7)
    due_soon_reviews = all_reviews.filter(
        status__in=['UPCOMING', 'DUE'],
        due_date__lte=due_soon_date,
        due_date__gte=today
    ).order_by('due_date')
    
    # Compliance Statistics by Unit
    units_stats = []
    all_units = Unit.objects.filter(is_active=True).select_related('care_home').order_by('name')
    
    # Filter units by care home if selected
    if care_home_filter:
        all_units = all_units.filter(care_home__name=care_home_filter)
    
    for unit in all_units:
        unit_reviews = CarePlanReview.objects.filter(resident__unit=unit)
        
        total = unit_reviews.count()
        completed = unit_reviews.filter(status='COMPLETED').count()
        overdue = unit_reviews.filter(status='OVERDUE').count()
        pending_approval = unit_reviews.filter(status='PENDING_APPROVAL').count()
        due_this_month = unit_reviews.filter(
            due_date__year=today.year,
            due_date__month=today.month
        ).count()
        
        # Calculate compliance rate (completed on time in last 6 months)
        six_months_ago = today - timedelta(days=180)
        recent_reviews = unit_reviews.filter(
            completed_date__gte=six_months_ago
        )
        total_recent = recent_reviews.count()
        on_time_recent = 0
        
        for review in recent_reviews:
            if review.completed_date and review.due_date:
                if review.completed_date <= review.due_date:
                    on_time_recent += 1
        
        compliance_rate = round((on_time_recent / total_recent * 100) if total_recent > 0 else 0, 1)
        
        # Status indicator
        if compliance_rate >= 90:
            status = 'excellent'
            status_color = 'success'
        elif compliance_rate >= 75:
            status = 'good'
            status_color = 'info'
        elif compliance_rate >= 60:
            status = 'warning'
            status_color = 'warning'
        else:
            status = 'critical'
            status_color = 'danger'
        
        units_stats.append({
            'unit': unit,
            'total': total,
            'completed': completed,
            'overdue': overdue,
            'pending_approval': pending_approval,
            'due_this_month': due_this_month,
            'compliance_rate': compliance_rate,
            'status': status,
            'status_color': status_color,
        })
    
    # Overall Statistics
    total_reviews = all_reviews.count()
    total_pending_approvals = pending_approvals.count()
    total_overdue = overdue_reviews.count()
    total_due_soon = due_soon_reviews.count()
    total_completed = all_reviews.filter(status='COMPLETED').count()
    
    # Calculate overall compliance rate
    recent_all = all_reviews.filter(completed_date__gte=today - timedelta(days=180))
    total_recent_all = recent_all.count()
    on_time_all = 0
    for review in recent_all:
        if review.completed_date and review.due_date and review.completed_date <= review.due_date:
            on_time_all += 1
    
    overall_compliance = round((on_time_all / total_recent_all * 100) if total_recent_all > 0 else 0, 1)
    
    # Recent Activity - Last 10 completed reviews
    recent_activity = all_reviews.filter(status='COMPLETED').order_by('-completed_date')[:10]
    
    # Get all care homes for filter dropdown
    from .models_multi_home import CareHome
    all_care_homes = CareHome.objects.all().order_by('name')
    
    context = {
        'pending_approvals': pending_approvals,
        'overdue_reviews': overdue_reviews,
        'due_soon_reviews': due_soon_reviews,
        'units_stats': units_stats,
        'total_reviews': total_reviews,
        'total_pending_approvals': total_pending_approvals,
        'total_overdue': total_overdue,
        'total_due_soon': total_due_soon,
        'total_completed': total_completed,
        'overall_compliance': overall_compliance,
        'recent_activity': recent_activity,
        'all_units': all_units,
        'selected_unit': unit_filter,
        'all_care_homes': all_care_homes,
        'selected_care_home': care_home_filter,
        'today': today,
    }
    
    return render(request, 'scheduling/careplan_manager_dashboard.html', context)


@login_required
def careplan_approve_review(request, review_id):
    """
    Manager approves a care plan review
    """
    if not (request.user.role and request.user.role.is_management):
        messages.error(request, "Access denied. Only managers can approve reviews.")
        return redirect('careplan_overview')
    
    review = get_object_or_404(CarePlanReview, id=review_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        manager_comments = request.POST.get('manager_comments', '')
        
        if action == 'approve':
            review.manager_approved = True
            review.manager_approval_date = timezone.now().date()
            review.manager_approved_by = request.user
            review.manager_comments = manager_comments
            review.status = 'COMPLETED'
            review.save()
            
            # Log the approval
            ActivityLog.objects.create(
                user=request.user,
                action_type='LEAVE_APPROVED',  # Using existing action type
                description=f'Approved care plan review for {review.resident.full_name} (ID: {review.resident.resident_id})',
                automated=False
            )
            
            messages.success(request, f'Care plan review for {review.resident.full_name} approved successfully.')
            
        elif action == 'reject':
            review.status = 'IN_PROGRESS'  # Send back to keyworker
            review.manager_comments = manager_comments
            review.save()
            
            # Log the rejection
            ActivityLog.objects.create(
                user=request.user,
                action_type='LEAVE_DENIED',  # Using existing action type
                description=f'Returned care plan review for {review.resident.full_name} (ID: {review.resident.resident_id}) for revision',
                automated=False
            )
            
            messages.warning(request, f'Care plan review for {review.resident.full_name} returned for revision.')
        
        return redirect('careplan_manager_dashboard')
    
    # GET request - show approval form
    context = {
        'review': review,
    }
    return render(request, 'scheduling/careplan_approve.html', context)


# ========================================
# STAFFING ALERT VIEWS
# ========================================

@login_required
@require_http_methods(["GET", "POST"])
def staffing_alert_respond(request, token, action):
    """
    Handle one-click responses from email/SMS links
    Token-based authentication allows responding without logging in
    
    SAFEGUARDS AGAINST OVERSTAFFING:
    1. Checks alert status (FILLED alerts reject new acceptances)
    2. Database locking in accept_shift() prevents race conditions
    3. User-friendly error messages explain why acceptance failed
    4. Redirects to dashboard for situational awareness
    """
    from scheduling.models import StaffingAlertResponse
    
    try:
        response = StaffingAlertResponse.objects.select_related(
            'alert', 'alert__unit', 'alert__shift_type', 'user'
        ).get(response_token=token)
    except StaffingAlertResponse.DoesNotExist:
        messages.error(request, '❌ Invalid or expired response link.')
        return redirect('dashboard')
    
    # SAFEGUARD: Check if alert is still active and accepting responses
    if response.alert.status == 'FILLED':
        messages.warning(
            request, 
            f'⚠️ This alert has already been filled. '
            f'{response.alert.shortage} positions were needed and all have been accepted. '
            f'Thank you for your willingness to help!'
        )
        return redirect('staffing_my_alerts')
    
    if response.alert.status in ['EXPIRED', 'CANCELLED']:
        status_msg = 'expired' if response.alert.status == 'EXPIRED' else 'cancelled'
        messages.warning(request, f'⚠️ This alert has been {status_msg} and is no longer accepting responses.')
        return redirect('staffing_my_alerts')
    
    # SAFEGUARD: Check if user already responded
    if response.status != 'PENDING':
        messages.info(request, f'ℹ️ You have already {response.status.lower()} this alert.')
        return redirect('staffing_my_alerts')
    
    if action == 'accept':
        try:
            # CRITICAL: accept_shift() uses database locking to prevent race conditions
            shift = response.accept_shift(method='EMAIL_LINK')
            
            # Check if alert is now filled
            remaining = response.alert.positions_remaining
            if remaining == 0:
                extra_msg = " This was the last position needed - alert is now fully staffed!"
            else:
                extra_msg = f" ({remaining} position{'s' if remaining != 1 else ''} still needed)"
            
            messages.success(
                request,
                f'✅ Shift accepted! You are now scheduled for {response.alert.shift_type.get_name_display()} '
                f'on {response.alert.shift_date} at {response.alert.unit.get_name_display()}.{extra_msg}'
            )
            
            # Log the acceptance
            ActivityLog.objects.create(
                user=request.user,
                action_type='SHIFT_CREATED',
                description=f'Accepted staffing alert #{response.alert.id} and created shift for {response.alert.shift_date}',
                automated=False
            )
            
        except Exception as e:
            error_msg = str(e)
            
            # User-friendly error messages
            if "already been filled" in error_msg:
                messages.error(
                    request, 
                    f'❌ Sorry, this alert was just filled by another staff member. '
                    f'All {response.alert.shortage} positions have been accepted. Thank you for responding!'
                )
            elif "already scheduled" in error_msg:
                messages.error(request, f'❌ {error_msg}')
            else:
                messages.error(request, f'❌ Unable to accept shift: {error_msg}')
            
    elif action == 'decline':
        reason = request.POST.get('reason', 'Not specified') if request.method == 'POST' else None
        response.decline_shift(reason=reason)
        messages.info(request, 'You have declined this shift request.')
        
        # Log the decline
        ActivityLog.objects.create(
            user=request.user,
            action_type='LEAVE_DENIED',  # Using existing action type
            description=f'Declined staffing alert for {response.alert.shift_date}',
            automated=False
        )
    
    return redirect('staffing_my_alerts')


@login_required
def staffing_my_alerts(request):
    """
    Show all active alerts for the current user
    """
    from scheduling.models import StaffingAlert, StaffingAlertResponse
    from django.db.models import Q
    
    # Get user's responses
    my_responses = StaffingAlertResponse.objects.filter(
        user=request.user
    ).select_related('alert', 'alert__unit', 'alert__shift_type', 'shift_created').order_by(
        '-alert__shift_date', '-alert__created_at'
    )
    
    # Separate by status
    pending_responses = my_responses.filter(status='PENDING', alert__status__in=['PENDING', 'PARTIALLY_FILLED'])
    accepted_responses = my_responses.filter(status='ACCEPTED')
    declined_responses = my_responses.filter(status='DECLINED')
    
    # Get all active alerts (for awareness)
    active_alerts = StaffingAlert.objects.filter(
        status__in=['PENDING', 'PARTIALLY_FILLED'],
        expires_at__gt=timezone.now()
    ).select_related('unit', 'shift_type').order_by('shift_date', '-priority')
    
    context = {
        'pending_responses': pending_responses,
        'accepted_responses': accepted_responses,
        'declined_responses': declined_responses,
        'active_alerts': active_alerts,
    }
    
    return render(request, 'scheduling/staffing_my_alerts.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff or u.role.name in ['MANAGER', 'SENIOR_MANAGER'])
def staffing_create_alert(request):
    """
    Manager interface to manually create staffing alerts
    """
    from scheduling.models import StaffingAlert
    
    if request.method == 'POST':
        try:
            unit_id = request.POST.get('unit')
            shift_type_id = request.POST.get('shift_type')
            shift_date = request.POST.get('shift_date')
            required_staff = int(request.POST.get('required_staff', 17))
            message = request.POST.get('message', '')
            
            # Get current staff count
            current_staff = Shift.objects.filter(
                unit_id=unit_id,
                shift_type_id=shift_type_id,
                date=shift_date,
                status='SCHEDULED'
            ).count()
            
            shortage = required_staff - current_staff
            
            if shortage <= 0:
                messages.warning(request, f'No shortage detected. Currently scheduled: {current_staff}/{required_staff}')
            else:
                # Create alert
                alert = StaffingAlert.objects.create(
                    alert_type='SHORTAGE',
                    unit_id=unit_id,
                    shift_type_id=shift_type_id,
                    shift_date=shift_date,
                    required_staff=required_staff,
                    current_staff=current_staff,
                    shortage=shortage,
                    created_by=request.user,
                    message=message or f'We need {shortage} additional staff. Can you help?',
                    expires_at=timezone.now() + timedelta(hours=12),
                    priority=min(10, shortage * 2)
                )
                
                messages.success(
                    request,
                    f'✅ Alert created for {shortage} staff on {shift_date}. '
                    f'Run "python manage.py send_staffing_alerts" to notify staff.'
                )
                
                # Log the alert creation
                ActivityLog.objects.create(
                    user=request.user,
                    action_type='SHIFT_CREATED',
                    description=f'Created staffing alert for {shift_date} ({shortage} positions)',
                    automated=False
                )
                
                return redirect('staffing_dashboard')
                
        except Exception as e:
            messages.error(request, f'Error creating alert: {str(e)}')
    
    # GET request - show form
    units = Unit.objects.filter(is_active=True).order_by('name')
    shift_types = ShiftType.objects.all().order_by('name')
    
    context = {
        'units': units,
        'shift_types': shift_types,
        'default_date': (date.today() + timedelta(days=1)).isoformat(),
    }
    
    return render(request, 'scheduling/staffing_create_alert.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff or u.role.name in ['MANAGER', 'SENIOR_MANAGER'])
def staffing_dashboard(request):
    """
    Manager dashboard for monitoring staffing alerts
    """
    from scheduling.models import StaffingAlert
    from django.db.models import Count, Q
    
    # Active alerts
    active_alerts = StaffingAlert.objects.filter(
        status__in=['PENDING', 'PARTIALLY_FILLED'],
        expires_at__gt=timezone.now()
    ).select_related('unit', 'shift_type', 'created_by').prefetch_related(
        'responses'
    ).order_by('shift_date', '-priority')
    
    # Recent completed alerts
    completed_alerts = StaffingAlert.objects.filter(
        status__in=['FILLED', 'EXPIRED']
    ).select_related('unit', 'shift_type').order_by('-updated_at')[:10]
    
    # Statistics
    total_active = active_alerts.count()
    total_positions = sum(a.shortage for a in active_alerts)
    total_accepted = sum(a.responses_accepted for a in active_alerts)
    total_pending = sum(a.responses_pending for a in active_alerts)
    
    context = {
        'active_alerts': active_alerts,
        'completed_alerts': completed_alerts,
        'stats': {
            'total_active': total_active,
            'total_positions': total_positions,
            'total_accepted': total_accepted,
            'total_pending': total_pending,
            'fill_rate': round((total_accepted / total_positions * 100) if total_positions > 0 else 0, 1)
        }
    }
    
    return render(request, 'scheduling/staffing_dashboard.html', context)


@login_required
def home_dashboard(request, home_slug=None):
    """
    Unified home-specific dashboard with role-based access control.
    Combines functionality from manager dashboard with automatic home detection.
    
    URL patterns:
        /dashboard/ - Auto-detects user's home
        /dashboard/hawthorn-house/ - Specific home view
        /dashboard/meadowburn/
        /dashboard/orchard-grove/
        /dashboard/riverside/
        /dashboard/victoria-gardens/
    
    Access levels:
        - FULL (SM/OM): See all data, can approve, manage rotas
        - MOST (SSCW): View schedules, team data, submit requests
        - LIMITED (Staff): View own info, submit requests only
    """
    from .models_multi_home import CareHome
    from .decorators import require_management
    
    # Determine which home to display
    care_home = None
    can_select_home = False
    
    # Check if user is senior management (can view any home)
    if request.user.role and request.user.role.is_senior_management_team:
        can_select_home = True
        
        # Try to get home from URL slug or GET parameter
        if home_slug:
            # Convert slug to home name (hawthorn-house -> HAWTHORN_HOUSE)
            home_name = home_slug.upper().replace('-', '_')
            try:
                care_home = CareHome.objects.get(name=home_name)
            except CareHome.DoesNotExist:
                messages.error(request, f"Care home not found: {home_slug}")
                return redirect('home_dashboard')
        else:
            # Check GET parameter
            home_param = request.GET.get('care_home')
            if home_param:
                try:
                    care_home = CareHome.objects.get(name=home_param)
                except CareHome.DoesNotExist:
                    pass
    else:
        # Regular staff - lock to their assigned home
        if request.user.assigned_care_home:
            care_home = request.user.assigned_care_home
            # Redirect to proper URL if viewing generic /dashboard/
            if not home_slug:
                home_url_slug = care_home.name.lower().replace('_', '-')
                return redirect('home_dashboard_specific', home_slug=home_url_slug)
        else:
            messages.error(request, "You are not assigned to a care home.")
            return redirect('staff_dashboard')
    
    # If no home determined yet, redirect to user's home or show error
    if not care_home:
        if request.user.assigned_care_home:
            home_url_slug = request.user.assigned_care_home.name.lower().replace('_', '-')
            return redirect('home_dashboard_specific', home_slug=home_url_slug)
        else:
            messages.error(request, "Please select a care home to view.")
            return redirect('senior_dashboard')
    
    # Get user's permission level
    permission_level = request.user.role.permission_level if request.user.role else 'LIMITED'
    
    # Base date ranges
    today = timezone.now().date()
    rota_start = datetime(2026, 1, 4).date()
    rota_end = rota_start + timedelta(weeks=12) - timedelta(days=1)
    
    # Filter units by home
    units_qs = Unit.objects.filter(is_active=True, care_home=care_home)
    units = list(units_qs)
    
    # Define shift categories
    day_shift_names = ['DAY', 'DAY_SENIOR', 'DAY_ASSISTANT']
    night_shift_names = ['NIGHT', 'NIGHT_SENIOR', 'NIGHT_ASSISTANT']
    day_care_roles = {'SCW', 'SCA'}
    night_care_roles = {'SCWN', 'SCAN'}
    role_labels = dict(Role.ROLE_CHOICES)
    
    # === Widget Data (filtered by permission level) ===
    
    # FULL and MOST access: Leave requests requiring action
    manual_review_requests = []
    pending_leave_requests = []
    if permission_level in ['FULL', 'MOST']:
        manual_review_qs = LeaveRequest.objects.filter(
            status='MANUAL_REVIEW',
            user__unit__care_home=care_home
        ).select_related('user').order_by('created_at')
        manual_review_requests = list(manual_review_qs)
        
        pending_qs = LeaveRequest.objects.filter(
            status='PENDING',
            user__unit__care_home=care_home
        ).select_related('user').order_by('created_at')
        
        # FULL can see all, MOST only sees requests needing SSCW input
        if permission_level == 'MOST':
            pending_qs = pending_qs.filter(user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN'])
        
        pending_leave_requests = list(pending_qs)
    
    # FULL access: Staff reallocations
    pending_reallocations = []
    if permission_level == 'FULL':
        reallocations_qs = StaffReallocation.objects.filter(
            status='NEEDED',
            target_unit__care_home=care_home
        ).select_related('target_unit', 'target_shift_type').order_by('target_date')
        pending_reallocations = list(reallocations_qs)
    
    # ALL levels: Today's staffing snapshot
    today_shifts = Shift.objects.filter(
        date=today,
        unit__care_home=care_home,
        status__in=['SCHEDULED', 'CONFIRMED']
    ).select_related('unit', 'shift_type', 'user', 'user__role')
    
    def format_breakdown(counts, include_roles=None):
        if include_roles is not None:
            counts = {role: count for role, count in counts.items() if role in include_roles}
        
        ordered_roles = ['SSCW', 'SCW', 'SCA', 'SSCWN', 'SCWN', 'SCAN']
        added = set()
        breakdown = []
        
        for role_code in ordered_roles:
            count = counts.get(role_code, 0)
            if count:
                breakdown.append({
                    'code': role_code,
                    'label': role_labels.get(role_code, role_code.replace('_', ' ').title()),
                    'count': count,
                })
                added.add(role_code)
        
        for role_code, count in counts.items():
            if role_code in added or count == 0:
                continue
            breakdown.append({
                'code': role_code,
                'label': role_labels.get(role_code, role_code.replace('_', ' ').title()),
                'count': count,
            })
        
        return breakdown
    
    staffing_summary = {}
    for unit in units:
        unit_shifts_today = today_shifts.filter(unit=unit)
        day_shifts_qs = unit_shifts_today.filter(shift_type__name__in=day_shift_names)
        night_shifts_qs = unit_shifts_today.filter(shift_type__name__in=night_shift_names)
        
        day_role_counts = {}
        for shift in day_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if role_code:
                day_role_counts[role_code] = day_role_counts.get(role_code, 0) + 1
        
        night_role_counts = {}
        for shift in night_shifts_qs:
            role_code = getattr(getattr(shift.user, 'role', None), 'name', None)
            if role_code:
                night_role_counts[role_code] = night_role_counts.get(role_code, 0) + 1
        
        day_care_total = sum(count for role, count in day_role_counts.items() if role in day_care_roles)
        night_care_total = sum(count for role, count in night_role_counts.items() if role in night_care_roles)
        
        staffing_summary[unit.name] = {
            'unit': unit,
            'day_actual': day_care_total,
            'day_required': unit.min_day_staff,
            'night_actual': night_care_total,
            'night_required': unit.min_night_staff,
            'day_status': 'good' if day_care_total >= unit.min_day_staff else 'shortage',
            'night_status': 'good' if night_care_total >= unit.min_night_staff else 'shortage',
            'day_breakdown': format_breakdown(day_role_counts),
            'night_breakdown': format_breakdown(night_role_counts),
            'day_supernumerary': day_role_counts.get('SSCW', 0),
            'night_supernumerary': night_role_counts.get('SSCWN', 0),
        }
    
    # FULL and MOST: Current sickness absences
    current_sickness_absences = []
    if permission_level in ['FULL', 'MOST']:
        sickness_qs = SicknessRecord.objects.filter(
            status__in=['OPEN', 'AWAITING_FIT_NOTE'],
            first_working_day__lte=today,
            profile__user__unit__care_home=care_home
        ).filter(
            models.Q(actual_last_working_day__isnull=True) |
            models.Q(actual_last_working_day__gte=today)
        ).select_related('profile__user').order_by('first_working_day')
        current_sickness_absences = list(sickness_qs)
    
    # ALL levels: Upcoming approved leave (LIMITED sees only own)
    approved_leave_qs = LeaveRequest.objects.filter(
        status='APPROVED',
        end_date__gte=today,
        user__unit__care_home=care_home
    )
    if permission_level == 'LIMITED':
        approved_leave_qs = approved_leave_qs.filter(user=request.user)
    
    upcoming_approved_leave = list(approved_leave_qs.select_related('user').order_by('start_date'))
    
    # FULL and MOST: Staffing alerts (7-day forecast)
    staffing_alerts = []
    if permission_level in ['FULL', 'MOST']:
        home_staffing_config = {
            'HAWTHORN_HOUSE': {'day_ideal': 18, 'night_ideal': 18},
            'MEADOWBURN': {'day_ideal': 17, 'night_ideal': 17},
            'ORCHARD_GROVE': {'day_ideal': 17, 'night_ideal': 17},
            'RIVERSIDE': {'day_ideal': 17, 'night_ideal': 17},
            'VICTORIA_GARDENS': {'day_ideal': 10, 'night_ideal': 10},
        }
        
        home_config = home_staffing_config.get(care_home.name, {'day_ideal': 17, 'night_ideal': 17})
        day_threshold = home_config['day_ideal']
        night_threshold = home_config['night_ideal']
        
        for i in range(7):
            check_date = today + timedelta(days=i)
            date_shifts = Shift.objects.filter(
                date=check_date,
                unit__care_home=care_home,
                status__in=['SCHEDULED', 'CONFIRMED']
            ).select_related('user', 'user__role', 'shift_type')
            
            day_care = date_shifts.filter(
                shift_type__name__in=day_shift_names,
                user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
            ).count()
            
            night_care = date_shifts.filter(
                shift_type__name__in=night_shift_names,
                user__role__name__in=['SCW', 'SCA', 'SCWN', 'SCAN']
            ).count()
            
            if day_care < day_threshold:
                staffing_alerts.append({
                    'date': check_date,
                    'shift': 'Day',
                    'actual': day_care,
                    'required': day_threshold,
                    'deficit': day_threshold - day_care,
                })
            
            if night_care < night_threshold:
                staffing_alerts.append({
                    'date': check_date,
                    'shift': 'Night',
                    'actual': night_care,
                    'required': night_threshold,
                    'deficit': night_threshold - night_care,
                })
    
    # FULL and MOST: Recent incidents (last 24 hours)
    recent_incidents = []
    if permission_level in ['FULL', 'MOST']:
        twenty_four_hours_ago = timezone.now() - timedelta(hours=24)
        incidents_qs = IncidentReport.objects.filter(
            created_at__gte=twenty_four_hours_ago,
            unit__care_home=care_home
        ).select_related('reported_by').order_by('-created_at')
        recent_incidents = list(incidents_qs)
    
    # LIMITED: User's own leave requests only
    user_leave_requests = []
    if permission_level == 'LIMITED':
        user_leave_requests = list(
            LeaveRequest.objects.filter(user=request.user)
            .order_by('-created_at')[:10]
        )
    
    context = {
        'care_home': care_home,
        'can_select_home': can_select_home,
        'permission_level': permission_level,
        'all_care_homes': CareHome.objects.all().order_by('name'),
        
        # Dashboard data
        'manual_review_requests': manual_review_requests,
        'pending_leave_requests': pending_leave_requests,
        'pending_reallocations': pending_reallocations,
        'staffing_summary': staffing_summary,
        'current_sickness_absences': current_sickness_absences,
        'upcoming_approved_leave': upcoming_approved_leave,
        'staffing_alerts': staffing_alerts,
        'recent_incidents': recent_incidents,
        'user_leave_requests': user_leave_requests,
        
        # Metadata
        'today': today,
        'units': units,
        'rota_start': rota_start,
        'rota_end': rota_end,
    }
    
    return render(request, 'scheduling/home_dashboard.html', context)


# ============================================================================
# DEMO FEEDBACK SYSTEM
# ============================================================================

@login_required
def demo_feedback(request):
    """
    Collect structured feedback from demo system testing
    Helps gather iteration requirements and user insights
    """
    if request.method == 'POST':
        form = DemoFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.submitted_by = request.user
            
            # Capture IP and user agent for analytics
            feedback.ip_address = request.META.get('REMOTE_ADDR')
            feedback.user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
            
            feedback.save()
            
            messages.success(request, 'Thank you for your feedback! Your insights will help us improve the system.')
            
            # Redirect back to dashboard or feedback thank you page
            return redirect('demo_feedback_thanks')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-populate user role if available
        initial_data = {}
        if request.user.role:
            role_mapping = {
                'SSCW': 'SENIOR',
                'SCW': 'STAFF',
                'SM': 'MANAGER',
                'HOS': 'HOS',
                'OM': 'MANAGER',
                'IDI': 'ADMIN',
            }
            initial_data['user_role'] = role_mapping.get(request.user.role.name, 'STAFF')
        
        if request.user.unit:
            initial_data['care_home'] = request.user.unit.care_home.get_name_display()
        
        form = DemoFeedbackForm(initial=initial_data)
    
    return render(request, 'scheduling/demo_feedback.html', {
        'form': form,
        'page_title': 'Demo Feedback'
    })


@login_required
def demo_feedback_thanks(request):
    """Thank you page after feedback submission"""
    return render(request, 'scheduling/demo_feedback_thanks.html', {
        'page_title': 'Thank You'
    })


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def view_feedback_results(request):
    """
    Management view to see all feedback responses
    Restricted to management team
    """
    feedbacks = DemoFeedback.objects.all().select_related('submitted_by', 'submitted_by__role')
    
    # Calculate aggregate statistics
    total_responses = feedbacks.count()
    
    if total_responses > 0:
        avg_overall = round(feedbacks.aggregate(avg=models.Avg('overall_rating'))['avg'] or 0, 1)
        avg_ease = round(feedbacks.aggregate(avg=models.Avg('ease_of_use'))['avg'] or 0, 1)
        would_recommend_pct = round((feedbacks.filter(would_recommend=True).count() / total_responses) * 100, 1)
        ready_for_daily_pct = round((feedbacks.filter(ready_to_use_daily=True).count() / total_responses) * 100, 1)
        
        # Feature ratings
        feature_ratings = {}
        for field in ['rota_viewing_rating', 'shift_swapping_rating', 'leave_request_rating', 
                     'ai_assistant_rating', 'dashboard_rating', 'mobile_experience_rating']:
            avg = feedbacks.filter(**{f'{field}__isnull': False}).aggregate(avg=models.Avg(field))['avg']
            feature_ratings[field] = round(avg, 1) if avg else None
        
        # Feedback needing attention
        needs_attention = feedbacks.filter(
            Q(overall_rating__lte=2) | 
            Q(ease_of_use__lte=2) |
            Q(ready_to_use_daily=False)
        ).count()
    else:
        avg_overall = 0
        avg_ease = 0
        would_recommend_pct = 0
        ready_for_daily_pct = 0
        feature_ratings = {}
        needs_attention = 0
    
    # Group by role
    by_role = feedbacks.values('user_role').annotate(count=Count('id')).order_by('-count')
    
    context = {
        'feedbacks': feedbacks,
        'total_responses': total_responses,
        'avg_overall': avg_overall,
        'avg_ease': avg_ease,
        'would_recommend_pct': would_recommend_pct,
        'ready_for_daily_pct': ready_for_daily_pct,
        'feature_ratings': feature_ratings,
        'needs_attention': needs_attention,
        'by_role': by_role,
        'page_title': 'Feedback Results'
    }
    
    return render(request, 'scheduling/feedback_results.html', context)


@login_required
def submit_feature_request(request):
    """
    Allow users to submit feature requests
    """
    if request.method == 'POST':
        form = FeatureRequestForm(request.POST)
        if form.is_valid():
            feature_request = form.save(commit=False)
            feature_request.requested_by = request.user
            feature_request.save()
            
            messages.success(request, 'Feature request submitted! We\'ll review it and update you on the status.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FeatureRequestForm()
    
    return render(request, 'scheduling/feature_request.html', {
        'form': form,
        'page_title': 'Request a Feature'
    })


# ============================================================================
# SMART STAFF MATCHING API (Task 1 - Phase 1)
# ============================================================================

@login_required
def smart_matching_test_page(request):
    """Render the Smart Staff Matching test interface"""
    return render(request, 'scheduling/smart_matching_test.html', {
        'page_title': 'Smart Staff Matching Test'
    })


@login_required
@require_http_methods(["GET"])
def smart_staff_matching_api(request, shift_id):
    """
    API endpoint: Get smart staff recommendations for a shift
    
    GET /api/smart-matching/<shift_id>/
    
    Response:
    {
        "success": true,
        "shift_id": 12345,
        "shift_date": "2025-12-13",
        "shift_time": "07:00:00 - 19:00:00",
        "unit": "Orchard Grove",
        "required_role": "SCW",
        "recommendations": [
            {
                "staff_sap": "123456",
                "staff_name": "John Doe",
                "staff_role": "SCW",
                "total_score": 87.5,
                "distance_score": 95.0,
                "overtime_score": 80.0,
                "skill_score": 100.0,
                "preference_score": 75.0,
                "fatigue_score": 100.0,
                "wdt_compliant": true,
                "recommended": true,
                "breakdown": {...}
            },
            ...
        ],
        "total_available": 45,
        "timestamp": "2025-12-13T10:30:00Z"
    }
    """
    from .staff_matching import get_smart_staff_recommendations
    
    try:
        shift = get_object_or_404(Shift, id=shift_id)
        
        # Check permissions (managers/admins only)
        if not (request.user.is_superuser or request.user.is_operational_manager or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only managers can access staff matching.'
            }, status=403)
        
        # Get recommendations
        limit = int(request.GET.get('limit', 10))
        recommendations = get_smart_staff_recommendations(shift, max_recommendations=limit)
        
        return JsonResponse({
            'success': True,
            **recommendations
        })
        
    except Shift.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Shift with ID {shift_id} not found'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Smart matching error for shift {shift_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error generating recommendations: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def auto_send_smart_offers_api(request, shift_id):
    """
    API endpoint: Automatically send OT offers to top-matched staff
    
    POST /api/smart-matching/<shift_id>/send-offers/
    
    Body:
    {
        "auto_send_count": 3  // Optional, default 3
    }
    
    Response:
    {
        "success": true,
        "offers_sent": 3,
        "batch_id": 456,
        "deadline": "2025-12-13T11:00:00Z",
        "top_recommendations": [...]
    }
    """
    from .staff_matching import auto_send_smart_offers
    
    try:
        shift = get_object_or_404(Shift, id=shift_id)
        
        # Check permissions (managers/admins only)
        if not (request.user.is_superuser or request.user.is_operational_manager or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only managers can send OT offers.'
            }, status=403)
        
        # Parse request body
        data = json.loads(request.body) if request.body else {}
        auto_send_count = int(data.get('auto_send_count', 3))
        
        # Validate count
        if auto_send_count < 1 or auto_send_count > 20:
            return JsonResponse({
                'success': False,
                'error': 'auto_send_count must be between 1 and 20'
            }, status=400)
        
        # Auto-send offers
        result = auto_send_smart_offers(shift, auto_send_count=auto_send_count)
        
        return JsonResponse(result)
        
    except Shift.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Shift with ID {shift_id} not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Auto-send error for shift {shift_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error sending offers: {str(e)}'
        }, status=500)


# ============================================================================
# ENHANCED AGENCY COORDINATION API (Task 2 - Phase 1)
# ============================================================================

@login_required
def agency_coordination_test_page(request):
    """Render the Agency Coordination test interface"""
    return render(request, 'scheduling/agency_coordination_test.html', {
        'page_title': 'Enhanced Agency Coordination Test'
    })


@login_required
@require_http_methods(["GET"])
def agency_recommendations_api(request, shift_id):
    """
    API endpoint: Get scored agency recommendations for a shift
    
    GET /api/agency-coordination/<shift_id>/
    
    Response:
    {
        "success": true,
        "shift_id": 12345,
        "recommendations": [
            {
                "agency_name": "Premier Care",
                "total_score": 85.2,
                "cost_score": 78.0,
                "response_time_score": 92.0,
                "availability_score": 85.0,
                "quality_score": 88.0,
                "relationship_score": 80.0,
                "estimated_cost": 168.50,
                "breakdown": {...}
            },
            ...
        ]
    }
    """
    from .agency_coordinator import get_agency_recommendations
    
    try:
        shift = get_object_or_404(Shift, id=shift_id)
        
        # Check permissions
        if not (request.user.is_superuser or request.user.is_operational_manager or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only managers can access agency coordination.'
            }, status=403)
        
        # Get recommendations
        limit = int(request.GET.get('limit', 5))
        recommendations = get_agency_recommendations(shift, max_recommendations=limit)
        
        return JsonResponse({
            'success': True,
            **recommendations
        })
        
    except Shift.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Shift with ID {shift_id} not found'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Agency recommendations error for shift {shift_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error generating agency recommendations: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def auto_coordinate_agencies_api(request, cover_request_id):
    """
    API endpoint: Auto-coordinate multi-agency outreach
    
    POST /api/agency-coordination/<cover_request_id>/auto-coordinate/
    
    Body:
    {
        "max_agencies": 5  // Optional, default 5
    }
    
    Response:
    {
        "success": true,
        "tier_1_agencies": [...],
        "escalation_plan": [...],
        "total_tiers": 3,
        "estimated_resolution_time": 20
    }
    """
    from .agency_coordinator import auto_coordinate_agencies
    from .models_automated_workflow import StaffingCoverRequest
    
    try:
        cover_request = get_object_or_404(StaffingCoverRequest, id=cover_request_id)
        
        # Check permissions
        if not (request.user.is_superuser or request.user.is_operational_manager or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only managers can coordinate agencies.'
            }, status=403)
        
        # Parse request body
        data = json.loads(request.body) if request.body else {}
        max_agencies = int(data.get('max_agencies', 5))
        
        # Validate
        if max_agencies < 1 or max_agencies > 10:
            return JsonResponse({
                'success': False,
                'error': 'max_agencies must be between 1 and 10'
            }, status=400)
        
        # Auto-coordinate
        result = auto_coordinate_agencies(cover_request, max_agencies=max_agencies)
        
        return JsonResponse(result)
        
    except StaffingCoverRequest.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Cover request with ID {cover_request_id} not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Auto-coordinate error for cover request {cover_request_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error coordinating agencies: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def auto_send_smart_offers_api(request, shift_id):
    """
    API endpoint: Automatically send OT offers to top-matched staff
    
    POST /api/smart-matching/<shift_id>/send-offers/
    
    Body:
    {
        "auto_send_count": 3  // Optional, default 3
    }
    
    Response:
    {
        "success": true,
        "offers_sent": 3,
        "batch_id": 456,
        "deadline": "2025-12-13T11:00:00Z",
        "top_recommendations": [...]
    }
    """
    from .staff_matching import auto_send_smart_offers
    
    try:
        shift = get_object_or_404(Shift, id=shift_id)
        
        # Check permissions (managers/admins only)
        if not (request.user.is_superuser or request.user.is_operational_manager or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only managers can send OT offers.'
            }, status=403)
        
        # Parse request body
        data = json.loads(request.body) if request.body else {}
        auto_send_count = int(data.get('auto_send_count', 3))
        
        # Validate count
        if auto_send_count < 1 or auto_send_count > 20:
            return JsonResponse({
                'success': False,
                'error': 'auto_send_count must be between 1 and 20'
            }, status=400)
        
        # Auto-send offers
        result = auto_send_smart_offers(shift, auto_send_count=auto_send_count)
        
        return JsonResponse(result)
        
    except Shift.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Shift with ID {shift_id} not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Auto-send error for shift {shift_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error sending offers: {str(e)}'
        }, status=500)


# ============================================================================
# INTELLIGENT SHIFT SWAP AUTO-APPROVAL API (Task 3 - Phase 1)
# ============================================================================

@login_required
def shift_swap_test_page(request):
    """Render Shift Swap Auto-Approval test interface"""
    return render(request, 'scheduling/shift_swap_test.html', {
        'page_title': 'Intelligent Shift Swap Auto-Approval Test'
    })


@login_required
@require_http_methods(["POST"])
def request_shift_swap_api(request):
    """
    API endpoint: Create a new shift swap request with auto-approval evaluation
    
    POST /api/shift-swaps/request/
    
    Body:
    {
        "requesting_shift_id": 123,
        "target_shift_id": 456,
        "reason": "Family emergency"
    }
    
    Response:
    {
        "success": true,
        "swap_request_id": 789,
        "status": "AUTO_APPROVED" or "MANUAL_REVIEW",
        "auto_approved": true/false,
        "approval_notes": "...",
        "qualification_score": 95.5
    }
    """
    from .models import ShiftSwapRequest, Shift
    from .swap_intelligence import auto_approve_if_eligible
    
    try:
        # Parse request body
        data = json.loads(request.body)
        requesting_shift_id = data.get('requesting_shift_id')
        target_shift_id = data.get('target_shift_id')
        reason = data.get('reason', '')
        
        if not requesting_shift_id or not target_shift_id:
            return JsonResponse({
                'success': False,
                'error': 'requesting_shift_id and target_shift_id are required'
            }, status=400)
        
        # Get shifts
        requesting_shift = get_object_or_404(Shift, id=requesting_shift_id)
        target_shift = get_object_or_404(Shift, id=target_shift_id)
        
        # Verify requesting user owns the requesting shift
        if requesting_shift.user != request.user and not request.user.is_staff:
            return JsonResponse({
                'success': False,
                'error': 'You can only swap your own shifts'
            }, status=403)
        
        # Create swap request
        swap_request = ShiftSwapRequest.objects.create(
            requesting_user=requesting_shift.user,
            target_user=target_shift.user,
            requesting_shift=requesting_shift,
            target_shift=target_shift,
            reason=reason,
            status='PENDING'
        )
        
        # Evaluate auto-approval
        result = auto_approve_if_eligible(swap_request, acting_user=request.user)
        
        # Refresh to get updated fields
        swap_request.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'swap_request_id': swap_request.id,
            'status': swap_request.status,
            'auto_approved': result['auto_approved'],
            'approval_notes': result['approval_notes'],
            'qualification_score': float(swap_request.qualification_match_score),
            'created_at': swap_request.created_at.isoformat()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Shift.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'One or both shifts not found'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Shift swap request error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error creating swap request: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_swap_recommendations_api(request, shift_id):
    """
    API endpoint: Get recommended staff for swapping with a shift
    
    GET /api/shift-swaps/<shift_id>/recommendations/
    """
    from .swap_intelligence import get_swap_recommendations
    
    try:
        shift = get_object_or_404(Shift, id=shift_id)
        
        # Check permissions
        if shift.user != request.user and not (request.user.is_staff or request.user.is_operational_manager):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        limit = int(request.GET.get('limit', 5))
        recommendations = get_swap_recommendations(shift, max_recommendations=limit)
        
        return JsonResponse({'success': True, 'shift_id': shift_id, **recommendations})
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Swap recommendations error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_swap_status_api(request, swap_id):
    """
    API endpoint: Get status of a shift swap request
    
    GET /api/shift-swaps/<swap_id>/status/
    """
    from .models import ShiftSwapRequest
    
    try:
        swap = get_object_or_404(ShiftSwapRequest, id=swap_id)
        
        # Check permissions
        if (swap.requesting_user != request.user and 
            swap.target_user != request.user and 
            not (request.user.is_staff or request.user.is_operational_manager)):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        return JsonResponse({
            'success': True,
            'swap_id': swap.id,
            'status': swap.status,
            'requesting_user': swap.requesting_user.full_name,
            'target_user': swap.target_user.full_name,
            'requesting_shift_date': swap.requesting_shift.date.isoformat(),
            'target_shift_date': swap.target_shift.date.isoformat(),
            'automated_decision': swap.automated_decision,
            'qualification_score': float(swap.qualification_match_score),
            'approval_notes': swap.approval_notes,
            'created_at': swap.created_at.isoformat(),
            'approval_date': swap.approval_date.isoformat() if swap.approval_date else None
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Swap status error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==============================================================================
# PREDICTIVE SHORTAGE ALERT SYSTEM (ML) - Task 5, Phase 2
# ==============================================================================

@login_required
def shortage_predictor_test_page(request):
    """Render Predictive Shortage Alerts test interface"""
    from .models import Unit
    
    context = {
        'page_title': 'Predictive Shortage Alert System (ML)',
        'units': Unit.objects.filter(is_active=True)
    }
    
    return render(request, 'scheduling/shortage_predictor_test.html', context)


@login_required
@require_http_methods(["POST"])
def train_shortage_model_api(request):
    """
    API endpoint: Train shortage predictor model on historical data
    
    POST /api/shortage-predictor/train/
    Body: {"months_back": 6, "save_model": true}
    
    Response: {
        "success": true,
        "metrics": {
            "accuracy": 0.87,
            "roc_auc": 0.92,
            "cv_accuracy": 0.85,
            "train_size": 450,
            "test_size": 112
        },
        "feature_importance": {
            "historical_sickness_avg": 0.35,
            "day_of_week": 0.25,
            "scheduled_leave_count": 0.20,
            ...
        }
    }
    
    Permissions: Manager/Admin only
    """
    # Check permissions
    if not (request.user.is_staff or request.user.is_operational_manager or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Manager/Admin permission required'}, status=403)
    
    try:
        import json
        from .shortage_predictor import train_shortage_predictor
        
        # Parse request body
        body = json.loads(request.body.decode('utf-8'))
        months_back = body.get('months_back', 6)
        save_model = body.get('save_model', True)
        
        # Train model
        predictor = train_shortage_predictor(months_back=months_back, save_model=save_model)
        
        # Log activity
        from .models import ActivityLog
        ActivityLog.objects.create(
            user=request.user,
            action_type='ML_MODEL_TRAINED',
            description=f'Trained shortage predictor model ({months_back} months historical data)',
            details=json.dumps(predictor.train_metrics)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Model trained successfully on {months_back} months of data',
            'metrics': predictor.train_metrics,
            'feature_importance': predictor.feature_importance,
            'model_saved': save_model
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Shortage predictor training error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_shortage_alerts_api(request):
    """
    API endpoint: Get predicted shortage alerts for upcoming days
    
    GET /api/shortage-predictor/alerts/?days_ahead=7&min_probability=0.5
    
    Response: {
        "success": true,
        "alerts": [
            {
                "date": "2025-12-28",
                "day_name": "Saturday",
                "unit": "OG Mulberry",
                "probability": 0.78,
                "confidence": 0.85,
                "predicted_gap": 2,
                "days_ahead": 3,
                "top_factors": [
                    {
                        "factor": "day_of_week",
                        "weight": 0.30,
                        "description": "Saturday (high risk)"
                    },
                    ...
                ]
            },
            ...
        ],
        "summary": "Found 3 high-risk shortage alerts in next 7 days"
    }
    
    Permissions: All staff (read-only)
    """
    try:
        from .shortage_predictor import get_shortage_alerts
        
        # Parse query parameters
        days_ahead = int(request.GET.get('days_ahead', 7))
        min_probability = float(request.GET.get('min_probability', 0.5))
        
        # Validate parameters
        if days_ahead < 1 or days_ahead > 30:
            return JsonResponse({'success': False, 'error': 'days_ahead must be 1-30'}, status=400)
        
        if min_probability < 0.0 or min_probability > 1.0:
            return JsonResponse({'success': False, 'error': 'min_probability must be 0.0-1.0'}, status=400)
        
        # Get predictions
        alerts = get_shortage_alerts(
            days_ahead=days_ahead,
            min_probability=min_probability,
            load_saved_model=True
        )
        
        # Create summary
        if len(alerts) == 0:
            summary = f"✅ No shortage alerts predicted for next {days_ahead} days"
        else:
            high_risk = len([a for a in alerts if a['probability'] >= 0.7])
            summary = f"⚠️ Found {len(alerts)} shortage alert(s) in next {days_ahead} days ({high_risk} high-risk ≥70%)"
        
        return JsonResponse({
            'success': True,
            'alerts': alerts,
            'alert_count': len(alerts),
            'high_risk_count': len([a for a in alerts if a['probability'] >= 0.7]),
            'summary': summary,
            'days_ahead': days_ahead,
            'min_probability': min_probability
        })
        
    except FileNotFoundError:
        return JsonResponse({
            'success': False, 
            'error': 'Model not trained yet. Please train the model first using /api/shortage-predictor/train/',
            'trained': False
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Shortage alerts error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_feature_importance_api(request):
    """
    API endpoint: Get ML feature importance scores
    
    GET /api/shortage-predictor/features/
    
    Response: {
        "success": true,
        "feature_importance": {
            "historical_sickness_avg": 0.35,
            "day_of_week": 0.25,
            "scheduled_leave_count": 0.20,
            "month": 0.10,
            "is_school_holiday": 0.05,
            ...
        },
        "top_3_factors": [
            {"feature": "historical_sickness_avg", "importance": 0.35},
            {"feature": "day_of_week", "importance": 0.25},
            {"feature": "scheduled_leave_count", "importance": 0.20}
        ]
    }
    """
    try:
        from .shortage_predictor import get_feature_importance
        
        # Get feature importance
        importance = get_feature_importance()
        
        # Sort by importance
        sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)
        
        top_3 = [
            {'feature': name, 'importance': round(score, 3)}
            for name, score in sorted_features[:3]
        ]
        
        return JsonResponse({
            'success': True,
            'feature_importance': importance,
            'top_3_factors': top_3,
            'total_features': len(importance)
        })
        
    except FileNotFoundError:
        return JsonResponse({
            'success': False,
            'error': 'Model not trained yet. Please train the model first.',
            'trained': False
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Feature importance error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================================
# TASK 11: AI ASSISTANT FEEDBACK & LEARNING SYSTEM
# Import API endpoints from views_compliance
# ============================================================================

from .views_compliance import (
    ai_assistant_suggestions_api,
    ai_assistant_feedback_api,
    ai_assistant_analytics_api,
    ai_assistant_insights_api
)


# ============================================================================
# TASK 19: PDF EXPORT FUNCTIONALITY (Phase 2)
# ============================================================================

# Temporarily disabled - using ReportLab instead
# from .utils.pdf_export import (
#     RotaPDFExporter,
#     LeaveReportPDFExporter,
#     ShiftAllocationPDFExporter
# )


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def export_weekly_rota_pdf(request, home_id):
    """Export weekly rota as PDF"""
    from .models_multi_home import CareHome
    from datetime import datetime, timedelta
    
    home = get_object_or_404(CareHome, pk=home_id)
    
    # Get week start date from query params or use current week
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
    else:
        # Default to start of current week (Monday)
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    
    exporter = RotaPDFExporter()
    return exporter.export_weekly_rota(home, week_start)


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def export_monthly_rota_pdf(request, home_id):
    """Export monthly rota as PDF"""
    from .models_multi_home import CareHome
    
    home = get_object_or_404(CareHome, pk=home_id)
    
    # Get month/year from query params or use current month
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))
    
    exporter = RotaPDFExporter()
    return exporter.export_monthly_rota(home, month, year)


@login_required
def export_staff_schedule_pdf(request, staff_id):
    """Export individual staff schedule as PDF"""
    from datetime import datetime, timedelta
    
    staff = get_object_or_404(User, pk=staff_id)
    
    # Staff can only export their own schedule unless they're management
    if not (request.user == staff or (request.user.role and request.user.role.is_management)):
        messages.error(request, "You don't have permission to view this schedule")
        return redirect('staff_dashboard')
    
    # Get date range from query params or default to next 4 weeks
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = date.today()
        end_date = start_date + timedelta(days=28)
    
    exporter = RotaPDFExporter()
    return exporter.export_staff_schedule(staff, start_date, end_date)


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def export_leave_summary_pdf(request):
    """Export leave request summary as PDF"""
    from datetime import datetime, timedelta
    from .models_multi_home import CareHome
    
    # Get date range and optional home filter
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    home_id = request.GET.get('home_id')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Default to current month
        today = date.today()
        start_date = date(today.year, today.month, 1)
        # Last day of month
        if today.month == 12:
            end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    home = None
    if home_id:
        home = get_object_or_404(CareHome, pk=home_id)
    
    exporter = LeaveReportPDFExporter()
    return exporter.export_leave_summary(start_date, end_date, home)


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def export_allocation_summary_pdf(request, home_id):
    """Export shift allocation summary as PDF"""
    from .models_multi_home import CareHome
    from datetime import datetime, timedelta
    
    home = get_object_or_404(CareHome, pk=home_id)
    
    # Get week start date from query params or use current week
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    
    exporter = ShiftAllocationPDFExporter()
    return exporter.export_allocation_summary(home, week_start)


# ============================================================================
# TASK 19: PDF EXPORT - Downloadable Shift Schedules and Reports
# ============================================================================

@login_required
def export_my_shifts_pdf(request):
    """Export current user's shifts as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    
    # Get date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Default: next 2 weeks
        start_date = date.today()
        end_date = start_date + timedelta(days=14)
    
    # Get user's shifts
    shifts = Shift.objects.filter(
        staff=request.user,
        date__range=[start_date, end_date]
    ).select_related('shift_type', 'unit', 'care_home').order_by('date', 'start_time')
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0066FF'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"My Shift Schedule", title_style))
    elements.append(Paragraph(f"{request.user.get_full_name()} ({request.user.sap})", styles['Normal']))
    elements.append(Paragraph(f"{start_date.strftime('%d %B %Y')} - {end_date.strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table data
    data = [['Date', 'Day', 'Shift', 'Time', 'Unit', 'Home']]
    
    for shift in shifts:
        data.append([
            shift.date.strftime('%d/%m/%Y'),
            shift.date.strftime('%A'),
            shift.shift_type.name if shift.shift_type else 'N/A',
            f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}",
            shift.unit.name if shift.unit else 'N/A',
            shift.care_home.name if shift.care_home else 'N/A'
        ])
    
    # Create table
    table = Table(data, colWidths=[3*cm, 3*cm, 3*cm, 3.5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Total shifts: {shifts.count()}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Italic']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"my_shifts_{start_date}_{end_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@user_passes_test(lambda u: u.role and u.role.is_management)
def export_weekly_rota_pdf(request, home_id):
    """Export weekly rota for a care home as PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    home = get_object_or_404(CareHome, pk=home_id)
    
    # Get week start date
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=6)
    
    # Get all shifts for the week
    shifts = Shift.objects.filter(
        care_home=home,
        date__range=[week_start, week_end]
    ).select_related('staff', 'shift_type', 'unit').order_by('date', 'start_time', 'unit__name')
    
    # Create PDF (landscape for weekly view)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0066FF'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Weekly Rota - {home.name}", title_style))
    elements.append(Paragraph(f"Week: {week_start.strftime('%d %B %Y')} - {week_end.strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*cm))
    
    # Group shifts by day and unit
    days = [(week_start + timedelta(days=i)) for i in range(7)]
    units = home.units.all().order_by('name')
    
    # Table header
    data = [['Unit'] + [day.strftime('%a\n%d/%m') for day in days]]
    
    # Table data - one row per unit
    for unit in units:
        row = [unit.name]
        for day in days:
            day_shifts = shifts.filter(unit=unit, date=day)
            if day_shifts.exists():
                shift_text = '\n'.join([
                    f"{s.staff.first_name} {s.staff.last_name[0]}. ({s.shift_type.code if s.shift_type else 'N/A'})"
                    for s in day_shifts
                ])
                row.append(shift_text)
            else:
                row.append('-')
        data.append(row)
    
    # Create table
    col_width = 3.5*cm
    table = Table(data, colWidths=[4*cm] + [col_width]*7)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"Total shifts: {shifts.count()} | Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Italic']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"weekly_rota_{home.name.replace(' ', '_')}_{week_start}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ============================================================================
# TASK 20: EXCEL EXPORT - Shift Data to Excel Spreadsheets
# ============================================================================

@login_required
def export_my_shifts_excel(request):
    """Export current user's shifts as Excel spreadsheet"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    # Get date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Default: next 4 weeks
        start_date = date.today()
        end_date = start_date + timedelta(days=28)
    
    # Get user's shifts
    shifts = Shift.objects.filter(
        staff=request.user,
        date__range=[start_date, end_date]
    ).select_related('shift_type', 'unit', 'care_home').order_by('date', 'start_time')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "My Shifts"
    
    # Header styling
    header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = f"My Shift Schedule - {request.user.get_full_name()}"
    ws['A1'].font = Font(size=16, bold=True, color="0066FF")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"{start_date.strftime('%d %B %Y')} - {end_date.strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Headers (row 4)
    headers = ['Date', 'Day', 'Shift Type', 'Start Time', 'End Time', 'Unit', 'Care Home']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    row_num = 5
    for shift in shifts:
        ws.cell(row=row_num, column=1, value=shift.date.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value=shift.date.strftime('%A'))
        ws.cell(row=row_num, column=3, value=shift.shift_type.name if shift.shift_type else 'N/A')
        ws.cell(row=row_num, column=4, value=shift.start_time.strftime('%H:%M'))
        ws.cell(row=row_num, column=5, value=shift.end_time.strftime('%H:%M'))
        ws.cell(row=row_num, column=6, value=shift.unit.name if shift.unit else 'N/A')
        ws.cell(row=row_num, column=7, value=shift.care_home.name if shift.care_home else 'N/A')
        
        # Apply borders and alignment
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        
        row_num += 1
    
    # Summary
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1, value=f"Total Shifts: {shifts.count()}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1, value=f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, size=9)
    
    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"my_shifts_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@user_passes_test(lambda u: u.role and u.role.is_management)
def export_weekly_rota_excel(request, home_id):
    """Export weekly rota for a care home as Excel spreadsheet"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    home = get_object_or_404(CareHome, pk=home_id)
    
    # Get week start date
    week_start_str = request.GET.get('week_start')
    if week_start_str:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=6)
    
    # Get all shifts for the week
    shifts = Shift.objects.filter(
        care_home=home,
        date__range=[week_start, week_end]
    ).select_related('staff', 'shift_type', 'unit').order_by('date', 'start_time')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Rota"
    
    # Styling
    header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Weekly Rota - {home.name}"
    ws['A1'].font = Font(size=16, bold=True, color="0066FF")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Week: {week_start.strftime('%d %B %Y')} - {week_end.strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Headers
    days = [(week_start + timedelta(days=i)) for i in range(7)]
    headers = ['Unit'] + [day.strftime('%a\n%d/%m') for day in days]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data - one row per unit
    units = home.units.all().order_by('name')
    row_num = 5
    
    for unit in units:
        ws.cell(row=row_num, column=1, value=unit.name)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).border = thin_border
        
        for col_num, day in enumerate(days, 2):
            day_shifts = shifts.filter(unit=unit, date=day)
            
            if day_shifts.exists():
                shift_text = '\n'.join([
                    f"{s.staff.first_name} {s.staff.last_name[0]}. ({s.shift_type.code if s.shift_type else 'N/A'})"
                    for s in day_shifts
                ])
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = shift_text
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                cell.border = thin_border
                
                # Weekend highlighting
                if day.weekday() >= 5:  # Saturday or Sunday
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                elif row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            else:
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                if day.weekday() >= 5:
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                elif row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        
        row_num += 1
    
    # Summary
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1, value=f"Total Shifts: {shifts.count()}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1, value=f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, size=9)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Adjust row heights for better readability
    for row in range(5, row_num):
        ws.row_dimensions[row].height = 40
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"weekly_rota_{home.name.replace(' ', '_')}_{week_start}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ============================================================================
# TASK 21: EMAIL NOTIFICATIONS - Automated Email Alerts
# ============================================================================

@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def send_test_email(request):
    """
    Send a test email to verify configuration (managers only)
    """
    from scheduling.email_notifications import send_shift_reminder_email
    from django.contrib import messages
    
    # Get a future shift for the current user or create test data
    test_shift = Shift.objects.filter(
        date__gte=date.today()
    ).select_related('staff', 'shift_type', 'unit', 'care_home').first()
    
    if test_shift:
        success = send_shift_reminder_email(test_shift)
        if success:
            messages.success(request, f"Test email sent to {test_shift.staff.email}")
        else:
            messages.error(request, "Failed to send test email. Check email configuration.")
    else:
        messages.warning(request, "No shifts found to send test email.")
    
    return redirect('scheduling:dashboard')


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def trigger_weekly_rotas(request):
    """
    Manually trigger weekly rota emails (managers only)
    """
    from scheduling.tasks import send_weekly_rotas
    from django.contrib import messages
    
    try:
        result = send_weekly_rotas()
        messages.success(request, f"Weekly rotas triggered: {result}")
    except Exception as e:
        messages.error(request, f"Failed to trigger weekly rotas: {str(e)}")
    
    return redirect('scheduling:dashboard')


# Integrate email notifications with leave request approval/rejection
# (These would be called from existing leave approval views)

def approve_leave_request_with_email(leave_request, approved_by, manager_notes=''):
    """
    Approve leave request and send confirmation email
    Helper function to be called from existing approval views
    """
    from scheduling.email_notifications import send_leave_approved_email
    
    leave_request.status = 'APPROVED'
    leave_request.approved_by = approved_by
    leave_request.save()
    
    # Send approval email
    send_leave_approved_email(leave_request, approved_by, manager_notes)
    
    return True


def reject_leave_request_with_email(leave_request, rejected_by, manager_notes=''):
    """
    Reject leave request and send notification email
    Helper function to be called from existing rejection views
    """
    from scheduling.email_notifications import send_leave_rejected_email
    
    leave_request.status = 'REJECTED'
    leave_request.rejected_by = rejected_by
    leave_request.save()
    
    # Send rejection email
    send_leave_rejected_email(leave_request, rejected_by, manager_notes)
    
    return True


def process_shift_swap_with_email(staff_member, original_shift, swap_with, new_shift=None):
    """
    Process shift swap and send confirmation emails to both parties
    Helper function to be called from existing swap views
    """
    from scheduling.email_notifications import send_shift_swap_email
    
    # Send email to first staff member
    send_shift_swap_email(staff_member, original_shift, swap_with, new_shift)
    
    # Send email to second staff member (reversed)
    if new_shift:
        send_shift_swap_email(swap_with, new_shift, staff_member, original_shift)
    else:
        send_shift_swap_email(swap_with, original_shift, staff_member, None)
    
    return True


# ============================================================================
# TASK 22: SMS NOTIFICATIONS - Urgent Alerts via SMS
# ============================================================================

@login_required
def sms_preferences(request):
    """
    Manage SMS notification preferences
    """
    from django.contrib import messages
    
    if request.method == 'POST':
        # Handle opt-in/opt-out
        action = request.POST.get('action')
        
        if action == 'enable':
            request.user.sms_notifications_enabled = True
            request.user.sms_emergency_only = request.POST.get('emergency_only') == 'on'
            request.user.sms_opted_in_date = timezone.now()
            request.user.save()
            messages.success(request, 'SMS notifications enabled successfully')
        
        elif action == 'disable':
            request.user.sms_notifications_enabled = False
            request.user.save()
            messages.success(request, 'SMS notifications disabled')
        
        elif action == 'update_phone':
            phone = request.POST.get('phone_number', '').strip()
            if phone:
                request.user.phone_number = phone
                request.user.save()
                messages.success(request, 'Phone number updated')
            else:
                messages.error(request, 'Please enter a valid phone number')
        
        return redirect('scheduling:sms_preferences')
    
    context = {
        'user': request.user,
        'twilio_enabled': getattr(settings, 'TWILIO_ENABLED', False),
    }
    
    return render(request, 'scheduling/sms_preferences.html', context)


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def send_test_sms(request):
    """
    Send a test SMS to verify Twilio configuration (managers only)
    """
    from scheduling.sms_notifications import send_sms_notification
    from django.contrib import messages
    
    if request.method == 'POST':
        phone = request.POST.get('phone_number')
        
        if phone:
            message = "Test SMS from Staff Rota System. Your SMS notifications are working correctly!"
            success, msg_sid = send_sms_notification(phone, message)
            
            if success:
                messages.success(request, f'Test SMS sent successfully to {phone} (ID: {msg_sid})')
            else:
                messages.error(request, 'Failed to send test SMS. Check Twilio configuration.')
        else:
            messages.error(request, 'Please enter a phone number')
    
    return redirect('scheduling:dashboard')


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def send_bulk_emergency_sms(request):
    """
    Send emergency SMS to multiple staff (managers only)
    """
    from scheduling.sms_notifications import send_bulk_sms
    from django.contrib import messages
    
    if request.method == 'POST':
        message_text = request.POST.get('message', '').strip()
        recipient_saps = request.POST.getlist('recipients')
        
        if not message_text:
            messages.error(request, 'Please enter a message')
            return redirect('scheduling:dashboard')
        
        if not recipient_saps:
            messages.error(request, 'Please select recipients')
            return redirect('scheduling:dashboard')
        
        # Get recipients
        recipients = User.objects.filter(sap__in=recipient_saps)
        
        # Send bulk SMS
        results = send_bulk_sms(recipients, message_text)
        
        messages.success(
            request,
            f"Bulk SMS sent: {results['sent']} successful, {results['failed']} failed"
        )
    
    return redirect('scheduling:dashboard')


@login_required
@user_passes_test(lambda u: u.role and u.role.is_management)
def sms_opt_in_report(request):
    """
    Report showing SMS opt-in status for all staff (managers only)
    """
    staff_list = User.objects.filter(
        is_active=True
    ).select_related('role', 'unit').order_by('last_name', 'first_name')
    
    stats = {
        'total_staff': staff_list.count(),
        'has_phone': staff_list.exclude(phone_number__isnull=True).exclude(phone_number='').count(),
        'sms_enabled': staff_list.filter(sms_notifications_enabled=True).count(),
        'emergency_only': staff_list.filter(sms_emergency_only=True).count(),
    }
    
    context = {
        'staff_list': staff_list,
        'stats': stats,
    }
    
    return render(request, 'scheduling/sms_opt_in_report.html', context)


# Helper functions for SMS integration with existing features

def send_emergency_shift_sms_alerts(shift, max_recipients=20):
    """
    Send emergency SMS to available staff for shift coverage
    Helper function to integrate with existing shift allocation
    """
    from scheduling.sms_notifications import send_emergency_coverage_alert
    
    # Find available staff (same role, not on leave, SMS enabled)
    available_staff = User.objects.filter(
        is_active=True,
        role=shift.shift_type.required_role if shift.shift_type else None,
        sms_notifications_enabled=True,
        phone_number__isnull=False
    ).exclude(phone_number='')[:max_recipients]
    
    sent_count = 0
    for staff in available_staff:
        if send_emergency_coverage_alert(staff, shift):
            sent_count += 1
    
    return sent_count


def notify_late_clockin_via_sms(staff_member, shift):
    """
    Send late clock-in SMS reminder
    Helper function to integrate with attendance tracking
    """
    from scheduling.sms_notifications import send_late_clockin_alert
    
    return send_late_clockin_alert(staff_member, shift)


def notify_manager_approval_sms(manager, approval_type, details):
    """
    Send SMS to manager for pending approvals
    Helper function to integrate with leave/swap approval workflows
    """
    from scheduling.sms_notifications import send_approval_required_alert
    
    return send_approval_required_alert(manager, approval_type, details)


# ============================================
# TASK 23: CALENDAR SYNC (iCal/Google Calendar Export)
# ============================================

@login_required
def export_my_shifts_ical(request):
    """
    Export user's shifts as iCal (.ics) file
    Allows import to Google Calendar, Apple Calendar, Outlook, etc.
    """
    from scheduling.calendar_sync import generate_shift_ical
    from datetime import date, timedelta
    
    # Get date range (default: next 12 weeks)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = date.today()
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = start_date + timedelta(weeks=12)
    
    # Get user's shifts
    shifts = Shift.objects.filter(
        staff=request.user,
        date__range=[start_date, end_date]
    ).select_related(
        'shift_type', 'unit', 'care_home'
    ).order_by('date', 'start_time')
    
    if not shifts.exists():
        messages.warning(request, "No shifts found for the selected date range.")
        return redirect('my_schedule')
    
    # Generate iCal
    calendar_name = f"{request.user.get_full_name()}'s Shifts"
    ical_content = generate_shift_ical(shifts, calendar_name)
    
    # Create HTTP response with iCal file
    response = HttpResponse(ical_content, content_type='text/calendar; charset=utf-8')
    filename = f"my_shifts_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.ics"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(ical_content)
    
    return response


@login_required
def calendar_feed(request, sap, token):
    """
    Personal calendar feed URL (subscribable in calendar apps)
    Updates automatically when rota changes
    
    URL: /calendar/feed/<sap>/<token>/
    Subscribe with: webcal://domain/calendar/feed/<sap>/<token>/
    """
    from scheduling.calendar_sync import verify_calendar_token, generate_shift_ical
    from datetime import date, timedelta
    
    # Verify token
    if not verify_calendar_token(sap, token):
        return HttpResponse("Invalid calendar feed token", status=403)
    
    # Get user
    try:
        user = User.objects.get(sap=sap)
    except User.DoesNotExist:
        return HttpResponse("User not found", status=404)
    
    # Get shifts (next 8 weeks by default, configurable)
    weeks = int(request.GET.get('weeks', 8))
    start_date = date.today()
    end_date = start_date + timedelta(weeks=weeks)
    
    shifts = Shift.objects.filter(
        staff=user,
        date__range=[start_date, end_date]
    ).select_related(
        'shift_type', 'unit', 'care_home'
    ).order_by('date', 'start_time')
    
    # Generate iCal
    calendar_name = f"{user.get_full_name()}'s Shifts (Auto-Update)"
    ical_content = generate_shift_ical(shifts, calendar_name)
    
    # Response with calendar MIME type
    response = HttpResponse(ical_content, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = f'inline; filename="shift_calendar.ics"'
    response['Cache-Control'] = 'no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    
    return response


@login_required
def export_leave_ical(request):
    """
    Export approved leave requests as iCal file
    """
    from scheduling.calendar_sync import generate_leave_ical
    from datetime import date, timedelta
    
    # Get date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = date.today()
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = start_date + timedelta(weeks=52)  # 1 year
    
    # Get approved leave requests
    leave_requests = LeaveRequest.objects.filter(
        staff=request.user,
        status='APPROVED',
        start_date__lte=end_date,
        end_date__gte=start_date
    ).order_by('start_date')
    
    if not leave_requests.exists():
        messages.warning(request, "No approved leave requests found.")
        return redirect('my_schedule')
    
    # Generate iCal
    calendar_name = f"{request.user.get_full_name()}'s Leave"
    ical_content = generate_leave_ical(leave_requests, calendar_name)
    
    response = HttpResponse(ical_content, content_type='text/calendar; charset=utf-8')
    filename = f"my_leave_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.ics"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def add_shift_to_calendar(request, shift_id):
    """
    Download single shift as iCal for "Add to Calendar" button
    """
    from scheduling.calendar_sync import create_single_shift_event
    
    # Get shift
    shift = get_object_or_404(
        Shift.objects.select_related('shift_type', 'unit', 'care_home'),
        id=shift_id,
        staff=request.user
    )
    
    # Generate single event
    ical_content = create_single_shift_event(shift)
    
    response = HttpResponse(ical_content, content_type='text/calendar; charset=utf-8')
    filename = f"shift_{shift.date.strftime('%Y%m%d')}_{shift.start_time.strftime('%H%M')}.ics"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def google_calendar_redirect(request, shift_id):
    """
    Redirect to Google Calendar "Add Event" page for a shift
    """
    from scheduling.calendar_sync import generate_google_calendar_url
    
    shift = get_object_or_404(
        Shift.objects.select_related('shift_type', 'unit', 'care_home'),
        id=shift_id,
        staff=request.user
    )
    
    google_url = generate_google_calendar_url(shift)
    return redirect(google_url)


@login_required
def outlook_calendar_redirect(request, shift_id):
    """
    Redirect to Outlook.com "Add Event" page for a shift
    """
    from scheduling.calendar_sync import generate_outlook_calendar_url
    
    shift = get_object_or_404(
        Shift.objects.select_related('shift_type', 'unit', 'care_home'),
        id=shift_id,
        staff=request.user
    )
    
    outlook_url = generate_outlook_calendar_url(shift)
    return redirect(outlook_url)


@login_required
def my_calendar_feed_info(request):
    """
    Display personal calendar feed URL and instructions
    """
    from scheduling.calendar_sync import generate_personal_calendar_token
    
    token = generate_personal_calendar_token(request.user)
    
    # Build feed URL
    feed_url = request.build_absolute_uri(
        reverse('calendar_feed', kwargs={'sap': request.user.sap, 'token': token})
    )
    
    # Webcal URL (for subscription)
    webcal_url = feed_url.replace('http://', 'webcal://').replace('https://', 'webcal://')
    
    context = {
        'feed_url': feed_url,
        'webcal_url': webcal_url,
        'token': token,
    }
    
    return render(request, 'scheduling/calendar_feed_info.html', context)


# ============================================
# TASK 24: BULK OPERATIONS
# ============================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def bulk_operations_menu(request):
    """
    Main menu for bulk operations
    """
    context = {
        'care_homes': CareHome.objects.all(),
    }
    return render(request, 'scheduling/bulk_operations_menu.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def bulk_assign_shifts(request):
    """
    Bulk assign shifts to multiple staff across date range
    """
    from scheduling.bulk_operations import (
        bulk_assign_shifts as do_bulk_assign,
        validate_bulk_operation,
        get_bulk_operation_preview,
        BulkOperationHistory,
        BulkOperationError
    )
    
    if request.method == 'POST':
        # Get form data
        care_home_id = request.POST.get('care_home')
        unit_id = request.POST.get('unit')
        shift_type_id = request.POST.get('shift_type')
        staff_ids = request.POST.getlist('staff')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Parse dates
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format")
            return redirect('bulk_assign_shifts')
        
        # Get objects
        try:
            care_home = CareHome.objects.get(id=care_home_id)
            unit = Unit.objects.get(id=unit_id)
            shift_type = ShiftType.objects.get(id=shift_type_id)
            staff_list = User.objects.filter(id__in=staff_ids)
        except (CareHome.DoesNotExist, Unit.DoesNotExist, ShiftType.DoesNotExist):
            messages.error(request, "Invalid selection")
            return redirect('bulk_assign_shifts')
        
        # Validate
        try:
            validation = validate_bulk_operation(
                'assign',
                staff_list=staff_list,
                date_range=(start_date, end_date)
            )
            
            # Show warnings
            for warning in validation.get('warnings', []):
                messages.warning(request, warning)
        
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('bulk_assign_shifts')
        
        # Execute bulk assign
        try:
            result = do_bulk_assign(
                staff_list=staff_list,
                date_range=(start_date, end_date),
                shift_type=shift_type,
                unit=unit,
                care_home=care_home,
                created_by=request.user
            )
            
            # Save to history for undo
            history = BulkOperationHistory(request.session)
            history.add_operation(
                operation_type='assign',
                affected_shifts=result['shift_ids'],
                rollback_data=result['shift_ids']
            )
            
            # Success message
            messages.success(
                request,
                f"Created {result['created']} shifts. "
                f"Skipped {result['skipped']} duplicates."
            )
            
            # Show errors if any
            for error in result['errors']:
                messages.warning(request, error)
        
        except BulkOperationError as e:
            messages.error(request, f"Bulk assign failed: {str(e)}")
        
        return redirect('bulk_operations_menu')
    
    # GET request - show form
    context = {
        'care_homes': CareHome.objects.all(),
        'shift_types': ShiftType.objects.all(),
    }
    return render(request, 'scheduling/bulk_assign_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def bulk_delete_shifts(request):
    """
    Bulk delete shifts by criteria
    """
    from scheduling.bulk_operations import (
        bulk_delete_shifts as do_bulk_delete,
        validate_bulk_operation,
        BulkOperationHistory,
        BulkOperationError
    )
    
    if request.method == 'POST':
        # Get criteria
        care_home_id = request.POST.get('care_home')
        unit_id = request.POST.get('unit')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        staff_ids = request.POST.getlist('staff')
        
        # Parse dates
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format")
            return redirect('bulk_delete_shifts')
        
        # Build queryset
        shifts = Shift.objects.filter(
            care_home_id=care_home_id,
            date__range=[start_date, end_date]
        )
        
        if unit_id:
            shifts = shifts.filter(unit_id=unit_id)
        
        if staff_ids:
            shifts = shifts.filter(staff_id__in=staff_ids)
        
        # Validate
        try:
            validation = validate_bulk_operation(
                'delete',
                shift_queryset=shifts
            )
            
            for warning in validation.get('warnings', []):
                messages.warning(request, warning)
        
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('bulk_delete_shifts')
        
        # Execute delete
        try:
            result = do_bulk_delete(
                shift_queryset=shifts,
                deleted_by=request.user
            )
            
            # Save to history
            history = BulkOperationHistory(request.session)
            history.add_operation(
                operation_type='delete',
                affected_shifts=[],
                rollback_data=result['rollback_data']
            )
            
            messages.success(request, f"Deleted {result['deleted']} shifts")
        
        except BulkOperationError as e:
            messages.error(request, f"Bulk delete failed: {str(e)}")
        
        return redirect('bulk_operations_menu')
    
    # GET request
    context = {
        'care_homes': CareHome.objects.all(),
    }
    return render(request, 'scheduling/bulk_delete_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def bulk_copy_week(request):
    """
    Copy entire week's schedule to another week
    """
    from scheduling.bulk_operations import (
        bulk_copy_week as do_bulk_copy,
        validate_bulk_operation,
        BulkOperationHistory,
        BulkOperationError
    )
    
    if request.method == 'POST':
        care_home_id = request.POST.get('care_home')
        source_week = request.POST.get('source_week')
        target_week = request.POST.get('target_week')
        unit_ids = request.POST.getlist('units')
        staff_ids = request.POST.getlist('staff')
        
        # Parse weeks (expecting YYYY-MM-DD for Monday)
        try:
            source_week_start = datetime.strptime(source_week, '%Y-%m-%d').date()
            target_week_start = datetime.strptime(target_week, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid week format")
            return redirect('bulk_copy_week')
        
        # Get objects
        try:
            care_home = CareHome.objects.get(id=care_home_id)
            units = Unit.objects.filter(id__in=unit_ids) if unit_ids else None
            staff_list = User.objects.filter(id__in=staff_ids) if staff_ids else None
        except CareHome.DoesNotExist:
            messages.error(request, "Invalid care home")
            return redirect('bulk_copy_week')
        
        # Validate
        try:
            validation = validate_bulk_operation(
                'copy',
                source_week_start=source_week_start,
                target_week_start=target_week_start
            )
        
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('bulk_copy_week')
        
        # Execute copy
        try:
            result = do_bulk_copy(
                source_week_start=source_week_start,
                target_week_start=target_week_start,
                care_home=care_home,
                units=units,
                staff_list=staff_list,
                created_by=request.user
            )
            
            # Save to history
            history = BulkOperationHistory(request.session)
            history.add_operation(
                operation_type='copy',
                affected_shifts=result['shift_ids'],
                rollback_data=result['shift_ids']
            )
            
            messages.success(
                request,
                f"Copied {result['copied']} shifts. "
                f"Skipped {result['skipped']} duplicates."
            )
        
        except BulkOperationError as e:
            messages.error(request, f"Bulk copy failed: {str(e)}")
        
        return redirect('bulk_operations_menu')
    
    # GET request
    context = {
        'care_homes': CareHome.objects.all(),
    }
    return render(request, 'scheduling/bulk_copy_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def bulk_swap_staff(request):
    """
    Swap shifts between two staff members
    """
    from scheduling.bulk_operations import (
        bulk_swap_staff as do_bulk_swap,
        validate_bulk_operation,
        BulkOperationHistory,
        BulkOperationError
    )
    
    if request.method == 'POST':
        care_home_id = request.POST.get('care_home')
        staff_a_id = request.POST.get('staff_a')
        staff_b_id = request.POST.get('staff_b')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        unit_ids = request.POST.getlist('units')
        
        # Parse dates
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format")
            return redirect('bulk_swap_staff')
        
        # Get objects
        try:
            care_home = CareHome.objects.get(id=care_home_id)
            staff_a = User.objects.get(id=staff_a_id)
            staff_b = User.objects.get(id=staff_b_id)
            units = Unit.objects.filter(id__in=unit_ids) if unit_ids else None
        except (CareHome.DoesNotExist, User.DoesNotExist):
            messages.error(request, "Invalid selection")
            return redirect('bulk_swap_staff')
        
        # Validate
        try:
            validation = validate_bulk_operation(
                'swap',
                staff_a=staff_a,
                staff_b=staff_b
            )
        
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('bulk_swap_staff')
        
        # Execute swap
        try:
            result = do_bulk_swap(
                staff_a=staff_a,
                staff_b=staff_b,
                date_range=(start_date, end_date),
                care_home=care_home,
                units=units
            )
            
            messages.success(
                request,
                f"Swapped {result['swapped']} shifts between "
                f"{staff_a.get_full_name()} and {staff_b.get_full_name()}"
            )
        
        except BulkOperationError as e:
            messages.error(request, f"Bulk swap failed: {str(e)}")
        
        return redirect('bulk_operations_menu')
    
    # GET request
    context = {
        'care_homes': CareHome.objects.all(),
    }
    return render(request, 'scheduling/bulk_swap_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def undo_last_bulk_operation(request):
    """
    Undo the last bulk operation
    """
    from scheduling.bulk_operations import (
        undo_bulk_operation,
        BulkOperationHistory,
        BulkOperationError
    )
    
    history = BulkOperationHistory(request.session)
    last_operation = history.get_last_operation()
    
    if not last_operation:
        messages.warning(request, "No operations to undo")
        return redirect('bulk_operations_menu')
    
    try:
        result = undo_bulk_operation(last_operation)
        
        messages.success(
            request,
            f"Undone {result['undone']} changes from {last_operation['type']} operation"
        )
        
        # Remove from history after successful undo
        history_list = request.session.get('bulk_operation_history', [])
        if history_list:
            history_list.pop()
            request.session['bulk_operation_history'] = history_list
            request.session.modified = True
    
    except BulkOperationError as e:
        messages.error(request, f"Undo failed: {str(e)}")
    
    return redirect('bulk_operations_menu')


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def get_units_for_home_ajax(request):
    """
    AJAX endpoint to get units for a care home
    """
    care_home_id = request.GET.get('care_home_id')
    
    if not care_home_id:
        return JsonResponse({'units': []})
    
    units = Unit.objects.filter(care_home_id=care_home_id).values('id', 'name')
    return JsonResponse({'units': list(units)})


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def get_staff_for_home_ajax(request):
    """
    AJAX endpoint to get staff for a care home
    """
    care_home_id = request.GET.get('care_home_id')
    unit_id = request.GET.get('unit_id')
    
    if not care_home_id:
        return JsonResponse({'staff': []})
    
    staff = User.objects.filter(care_home_id=care_home_id, is_active=True)
    
    if unit_id:
        staff = staff.filter(unit_id=unit_id)
    
    staff_data = staff.values('id', 'first_name', 'last_name', 'sap').order_by('last_name', 'first_name')
    
    # Format names
    staff_list = [
        {
            'id': s['id'],
            'name': f"{s['last_name']}, {s['first_name']} ({s['sap']})"
        }
        for s in staff_data
    ]
    
    return JsonResponse({'staff': staff_list})


# ============================================================================
# ANALYTICS & REPORTING VIEWS (Phase 3)
# ============================================================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def analytics_dashboard(request):
    """
    Main analytics dashboard with KPIs and charts
    """
    from .analytics import get_dashboard_summary, get_trending_data
    
    # Get filter parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    date_range = request.GET.get('date_range', 'week')
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Get dashboard data
    dashboard_data = get_dashboard_summary(care_home, unit, date_range)
    
    # Get trending data for charts
    staffing_trend = get_trending_data(care_home, unit, 'staffing', periods=12)
    costs_trend = get_trending_data(care_home, unit, 'costs', periods=12)
    occupancy_trend = get_trending_data(care_home, unit, 'occupancy', periods=12)
    compliance_trend = get_trending_data(care_home, unit, 'compliance', periods=12)
    
    context = {
        'dashboard_data': dashboard_data,
        'staffing_trend': staffing_trend,
        'costs_trend': costs_trend,
        'occupancy_trend': occupancy_trend,
        'compliance_trend': compliance_trend,
        'care_homes': CareHome.objects.all(),
        'units': Unit.objects.all() if not care_home else care_home.units.all(),
        'selected_care_home': care_home,
        'selected_unit': unit,
        'selected_date_range': date_range,
    }
    
    return render(request, 'scheduling/analytics_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def analytics_detailed_report(request):
    """
    Detailed analytics report with drill-down
    """
    from .analytics import (
        calculate_staffing_levels,
        calculate_overtime_metrics,
        calculate_cost_metrics,
        calculate_compliance_metrics,
        get_shift_distribution
    )
    
    # Get parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Parse dates
    try:
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to current month
            from .analytics import get_date_range
            start_date, end_date = get_date_range('month')
    except ValueError:
        messages.error(request, "Invalid date format. Using current month.")
        from .analytics import get_date_range
        start_date, end_date = get_date_range('month')
    
    # Get detailed metrics
    staffing = calculate_staffing_levels(care_home, unit, start_date, end_date)
    overtime = calculate_overtime_metrics(care_home, unit, start_date, end_date)
    costs = calculate_cost_metrics(care_home, unit, start_date, end_date)
    compliance = calculate_compliance_metrics(care_home, unit, start_date, end_date)
    distribution = get_shift_distribution(care_home, unit, start_date, end_date)
    
    context = {
        'staffing': staffing,
        'overtime': overtime,
        'costs': costs,
        'compliance': compliance,
        'distribution': distribution,
        'start_date': start_date,
        'end_date': end_date,
        'care_homes': CareHome.objects.all(),
        'units': Unit.objects.all() if not care_home else care_home.units.all(),
        'selected_care_home': care_home,
        'selected_unit': unit,
    }
    
    return render(request, 'scheduling/analytics_detailed_report.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def analytics_export_data(request):
    """
    Export analytics data to JSON (for charts/API)
    """
    from .analytics import get_dashboard_summary, get_trending_data
    
    # Get parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    date_range = request.GET.get('date_range', 'week')
    metric = request.GET.get('metric', 'all')
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    if metric == 'all':
        # Full dashboard data
        data = get_dashboard_summary(care_home, unit, date_range)
    else:
        # Specific trending data
        data = get_trending_data(care_home, unit, metric, periods=12)
    
    return JsonResponse(data, safe=False)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def analytics_kpi_widget(request, kpi_type):
    """
    Get single KPI widget data (AJAX endpoint)
    """
    from .analytics import (
        calculate_occupancy_rate,
        calculate_staffing_levels,
        calculate_overtime_metrics,
        calculate_cost_metrics,
        calculate_compliance_metrics
    )
    
    # Get parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    date_range = request.GET.get('date_range', 'week')
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Get date range
    from .analytics import get_date_range
    start_date, end_date = get_date_range(date_range)
    
    # Get KPI data based on type
    if kpi_type == 'occupancy':
        data = calculate_occupancy_rate(care_home, unit, start_date, end_date)
    elif kpi_type == 'staffing':
        data = calculate_staffing_levels(care_home, unit, start_date, end_date)
    elif kpi_type == 'overtime':
        data = calculate_overtime_metrics(care_home, unit, start_date, end_date)
    elif kpi_type == 'costs':
        data = calculate_cost_metrics(care_home, unit, start_date, end_date)
    elif kpi_type == 'compliance':
        data = calculate_compliance_metrics(care_home, unit, start_date, end_date)
    else:
        return JsonResponse({'error': 'Invalid KPI type'}, status=400)
    
    return JsonResponse(data)


# ============================================================================
# PREDICTIVE STAFFING VIEWS (Phase 3 - Task 26)
# ============================================================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def predictive_staffing_dashboard(request):
    """
    Predictive staffing dashboard with ML forecasts
    """
    from .predictive_staffing import (
        get_ml_model_stats,
        get_staffing_recommendations,
        detect_staffing_patterns
    )
    
    # Get filter parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    days_ahead = int(request.GET.get('days_ahead', 14))
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Get ML model stats
    model_stats = get_ml_model_stats(care_home, unit)
    
    # Get staffing recommendations
    recommendations = get_staffing_recommendations(care_home, unit, days_ahead)
    
    # Get staffing patterns
    patterns = detect_staffing_patterns(care_home, unit)
    
    context = {
        'model_stats': model_stats,
        'recommendations': recommendations,
        'patterns': patterns,
        'care_homes': CareHome.objects.all(),
        'units': Unit.objects.all() if not care_home else care_home.units.all(),
        'selected_care_home': care_home,
        'selected_unit': unit,
        'days_ahead': days_ahead,
    }
    
    return render(request, 'scheduling/predictive_staffing_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def predictive_week_forecast(request):
    """
    Weekly staffing forecast view
    """
    from .predictive_staffing import predict_week_staffing
    from datetime import datetime, timedelta
    
    # Get parameters
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    week_start_str = request.GET.get('week_start')
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
        care_home = unit.care_home if unit else None
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Parse week start (default to next Monday)
    if week_start_str:
        try:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        except ValueError:
            week_start = timezone.now().date()
            # Adjust to next Monday
            days_ahead = (7 - week_start.weekday()) % 7
            week_start = week_start + timedelta(days=days_ahead)
    else:
        week_start = timezone.now().date()
        days_ahead = (7 - week_start.weekday()) % 7
        week_start = week_start + timedelta(days=days_ahead)
    
    # Get predictions
    predictions = predict_week_staffing(week_start, care_home, unit)
    
    # Calculate summary
    total_predicted = sum(p['predicted_shifts'] for p in predictions)
    total_scheduled = sum(p['current_scheduled'] for p in predictions)
    total_gap = sum(p['gap'] for p in predictions)
    
    summary = {
        'total_predicted': total_predicted,
        'total_scheduled': total_scheduled,
        'total_gap': total_gap,
        'avg_confidence': round(sum(p['confidence'] for p in predictions) / len(predictions), 2) if predictions else 0
    }
    
    context = {
        'predictions': predictions,
        'summary': summary,
        'week_start': week_start,
        'week_end': week_start + timedelta(days=6),
        'care_homes': CareHome.objects.all(),
        'units': Unit.objects.all() if not care_home else care_home.units.all(),
        'selected_care_home': care_home,
        'selected_unit': unit,
    }
    
    return render(request, 'scheduling/predictive_week_forecast.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def predictive_single_day(request):
    """
    Single day prediction (AJAX endpoint)
    """
    from .predictive_staffing import predict_staffing_needs
    from datetime import datetime
    
    # Get parameters
    date_str = request.GET.get('date')
    care_home_id = request.GET.get('care_home')
    unit_id = request.GET.get('unit')
    leave_count = int(request.GET.get('leave_count', 0))
    
    if not date_str:
        return JsonResponse({'error': 'Date required'}, status=400)
    
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    care_home = None
    unit = None
    
    if unit_id:
        unit = Unit.objects.filter(id=unit_id).first()
    elif care_home_id:
        care_home = CareHome.objects.filter(id=care_home_id).first()
    
    # Get prediction
    prediction = predict_staffing_needs(target_date, care_home, unit, leave_count)
    
    return JsonResponse(prediction)


# ============================================================================
# TASK 27: CUSTOM REPORT BUILDER VIEWS
# ============================================================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_builder_dashboard(request):
    """
    Main report builder dashboard - shows saved templates and recent reports
    """
    from .models import SavedReport, ScheduledReport
    
    # Get user's saved reports
    user_reports = SavedReport.objects.filter(created_by=request.user)
    
    # Get public reports from other users
    public_reports = SavedReport.objects.filter(is_public=True).exclude(created_by=request.user)
    
    # Get scheduled reports
    scheduled_reports = ScheduledReport.objects.filter(
        created_by=request.user,
        is_active=True
    ).select_related('report_template')
    
    context = {
        'user_reports': user_reports,
        'public_reports': public_reports,
        'scheduled_reports': scheduled_reports,
        'report_types': SavedReport.REPORT_TYPE_CHOICES,
    }
    
    return render(request, 'scheduling/report_builder_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_builder_create(request):
    """
    Report builder interface - drag-and-drop field selection and filters
    """
    from .models import SavedReport
    from .report_builder import AVAILABLE_FIELDS
    
    # Get filter options
    care_homes = CareHome.objects.all()
    units = Unit.objects.all()
    roles = Role.objects.all()
    
    # Get report type from query params or default to CUSTOM
    report_type = request.GET.get('type', 'CUSTOM')
    
    # Get template if editing existing
    template_id = request.GET.get('template')
    template = None
    if template_id:
        template = SavedReport.objects.filter(id=template_id).first()
    
    # Get available fields for selected report type
    available_fields = AVAILABLE_FIELDS.get(report_type, {})
    
    context = {
        'report_types': SavedReport.REPORT_TYPE_CHOICES,
        'output_formats': SavedReport.OUTPUT_FORMAT_CHOICES,
        'available_fields': available_fields,
        'care_homes': care_homes,
        'units': units,
        'roles': roles,
        'template': template,
        'selected_report_type': report_type,
    }
    
    return render(request, 'scheduling/report_builder_create.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_execute(request):
    """
    Execute a report and return results in selected format
    """
    from .models import SavedReport
    from .report_builder import (
        build_report_query,
        extract_report_data,
        generate_csv_report,
        generate_excel_report,
        generate_pdf_report,
        calculate_report_summary
    )
    from django.http import HttpResponse
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    # Get report configuration from POST data
    data = json.loads(request.body)
    
    report_type = data.get('report_type', 'CUSTOM')
    selected_fields = data.get('selected_fields', [])
    filters = data.get('filters', {})
    output_format = data.get('output_format', 'PDF')
    report_name = data.get('report_name', 'Custom Report')
    
    # Validate
    if not selected_fields:
        return JsonResponse({'error': 'No fields selected'}, status=400)
    
    # Build query
    queryset = build_report_query(report_type, selected_fields, filters)
    
    if queryset is None:
        return JsonResponse({'error': 'Invalid report type'}, status=400)
    
    # Extract data
    report_data = extract_report_data(queryset, selected_fields, report_type)
    
    # Calculate summary
    summary = calculate_report_summary(report_data, report_type)
    
    # Generate output
    try:
        if output_format == 'CSV':
            output = generate_csv_report(report_data)
            content_type = 'text/csv'
            filename = f'{report_name.replace(" ", "_")}.csv'
        
        elif output_format == 'EXCEL':
            output = generate_excel_report(report_data, report_name)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'{report_name.replace(" ", "_")}.xlsx'
        
        elif output_format == 'PDF':
            output = generate_pdf_report(report_data, report_name, filters)
            content_type = 'application/pdf'
            filename = f'{report_name.replace(" ", "_")}.pdf'
        
        else:
            return JsonResponse({'error': 'Invalid output format'}, status=400)
        
        # Return file response
        response = HttpResponse(output.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    except ImportError as e:
        return JsonResponse({
            'error': f'Required library not installed: {str(e)}',
            'suggestion': 'Install openpyxl for Excel or reportlab for PDF support'
        }, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_preview(request):
    """
    Preview report data (AJAX) - returns JSON for display
    """
    from .report_builder import build_report_query, extract_report_data, calculate_report_summary
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    # Get report configuration
    data = json.loads(request.body)
    
    report_type = data.get('report_type', 'CUSTOM')
    selected_fields = data.get('selected_fields', [])
    filters = data.get('filters', {})
    
    if not selected_fields:
        return JsonResponse({'error': 'No fields selected'}, status=400)
    
    # Build query
    queryset = build_report_query(report_type, selected_fields, filters)
    
    if queryset is None:
        return JsonResponse({'error': 'Invalid report type'}, status=400)
    
    # Limit preview to 100 rows
    queryset = queryset[:100]
    
    # Extract data
    report_data = extract_report_data(queryset, selected_fields, report_type)
    
    # Calculate summary
    summary = calculate_report_summary(report_data, report_type)
    
    return JsonResponse({
        'data': report_data,
        'summary': summary,
        'preview_limit': 100,
        'total_available': queryset.count() if hasattr(queryset, 'count') else len(report_data)
    })


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_template_save(request):
    """
    Save report configuration as template (AJAX)
    """
    from .models import SavedReport
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    # Get template configuration
    data = json.loads(request.body)
    
    template_id = data.get('template_id')
    template_name = data.get('name')
    description = data.get('description', '')
    report_type = data.get('report_type', 'CUSTOM')
    selected_fields = data.get('selected_fields', [])
    filters = data.get('filters', {})
    grouping = data.get('grouping', {})
    sorting = data.get('sorting', [])
    output_format = data.get('output_format', 'PDF')
    is_public = data.get('is_public', False)
    
    # Validate
    if not template_name:
        return JsonResponse({'error': 'Template name required'}, status=400)
    
    if not selected_fields:
        return JsonResponse({'error': 'No fields selected'}, status=400)
    
    # Create or update template
    if template_id:
        # Update existing
        template = SavedReport.objects.filter(id=template_id, created_by=request.user).first()
        if not template:
            return JsonResponse({'error': 'Template not found'}, status=404)
        
        template.name = template_name
        template.description = description
        template.report_type = report_type
        template.selected_fields = selected_fields
        template.filters = filters
        template.grouping = grouping
        template.sorting = sorting
        template.output_format = output_format
        template.is_public = is_public
        template.save()
        
        message = 'Template updated successfully'
    else:
        # Create new
        template = SavedReport.objects.create(
            name=template_name,
            description=description,
            report_type=report_type,
            created_by=request.user,
            selected_fields=selected_fields,
            filters=filters,
            grouping=grouping,
            sorting=sorting,
            output_format=output_format,
            is_public=is_public
        )
        
        message = 'Template saved successfully'
    
    return JsonResponse({
        'success': True,
        'message': message,
        'template_id': template.id
    })


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_template_delete(request, template_id):
    """
    Delete a saved report template
    """
    from .models import SavedReport
    
    template = SavedReport.objects.filter(id=template_id, created_by=request.user).first()
    
    if not template:
        messages.error(request, 'Template not found or access denied.')
        return redirect('report_builder_dashboard')
    
    template_name = template.name
    template.delete()
    
    messages.success(request, f'Template "{template_name}" deleted successfully.')
    return redirect('report_builder_dashboard')


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_schedule_create(request):
    """
    Create a scheduled report
    """
    from .models import SavedReport, ScheduledReport
    from datetime import time
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    data = json.loads(request.body)
    
    template_id = data.get('template_id')
    frequency = data.get('frequency', 'WEEKLY')
    weekday = data.get('weekday')
    day_of_month = data.get('day_of_month')
    time_str = data.get('time', '09:00')
    recipients = data.get('recipients', [])
    
    # Validate
    template = SavedReport.objects.filter(id=template_id).first()
    if not template:
        return JsonResponse({'error': 'Template not found'}, status=404)
    
    # Parse time
    try:
        hour, minute = map(int, time_str.split(':'))
        time_of_day = time(hour, minute)
    except:
        return JsonResponse({'error': 'Invalid time format'}, status=400)
    
    # Create scheduled report
    scheduled = ScheduledReport.objects.create(
        report_template=template,
        created_by=request.user,
        frequency=frequency,
        weekday=weekday,
        day_of_month=day_of_month,
        time_of_day=time_of_day,
        recipients=recipients,
        is_active=True
    )
    
    # Calculate first run time
    scheduled.calculate_next_run()
    
    return JsonResponse({
        'success': True,
        'message': 'Report scheduled successfully',
        'schedule_id': scheduled.id,
        'next_run': scheduled.next_run.strftime('%Y-%m-%d %H:%M') if scheduled.next_run else None
    })


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def report_schedule_delete(request, schedule_id):
    """
    Delete a scheduled report
    """
    from .models import ScheduledReport
    
    scheduled = ScheduledReport.objects.filter(id=schedule_id, created_by=request.user).first()
    
    if not scheduled:
        messages.error(request, 'Scheduled report not found or access denied.')
        return redirect('report_builder_dashboard')
    
    template_name = scheduled.report_template.name
    scheduled.delete()
    
    messages.success(request, f'Scheduled report "{template_name}" deleted successfully.')
    return redirect('report_builder_dashboard')


# ==========================================================================================
# KPI TRACKING VIEWS (PHASE 3 - TASK 28)
# ==========================================================================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def kpi_dashboard(request):
    """
    KPI tracking dashboard with real-time performance metrics
    """
    from .models import KPIDefinition, KPIMeasurement
    from . import kpi_tracking
    
    # Get user's care home
    care_home = request.user.care_home if not request.user.is_superuser else None
    
    # Get period filter
    period = request.GET.get('period', 'month')
    
    # Get KPI summary
    summary = kpi_tracking.get_kpi_summary(care_home=care_home, period=period)
    
    # Get all active KPIs with latest measurements
    kpis = KPIDefinition.objects.filter(is_active=True)
    if care_home:
        kpis = kpis.filter(Q(care_home=care_home) | Q(care_home__isnull=True))
    
    kpi_data = []
    for kpi in kpis:
        latest = KPIMeasurement.objects.filter(kpi=kpi).order_by('-measurement_date').first()
        
        kpi_data.append({
            'definition': kpi,
            'latest_measurement': latest,
            'has_alert': latest.alert_generated if latest else False,
        })
    
    context = {
        'summary': summary,
        'kpi_data': kpi_data,
        'period': period,
        'care_home': care_home,
    }
    
    return render(request, 'scheduling/kpi_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def kpi_detail(request, kpi_id):
    """
    Detailed view of a single KPI with historical trend
    """
    from .models import KPIDefinition, KPITarget
    from . import kpi_tracking
    
    kpi = get_object_or_404(KPIDefinition, id=kpi_id, is_active=True)
    
    # Check access
    if not request.user.is_superuser and kpi.care_home and kpi.care_home != request.user.care_home:
        messages.error(request, 'Access denied to this KPI.')
        return redirect('kpi_dashboard')
    
    # Get trend data
    days_back = int(request.GET.get('days', 90))
    trend = kpi_tracking.get_kpi_trend(kpi, days_back=days_back)
    
    # Get current targets
    current_year = timezone.now().year
    targets = KPITarget.objects.filter(
        kpi=kpi,
        year=current_year
    ).order_by('month', 'quarter')
    
    # Prepare chart data
    chart_labels = [m.measurement_date.strftime('%Y-%m-%d') for m in trend]
    chart_values = [float(m.measured_value) for m in trend]
    chart_targets = [float(m.target_value) if m.target_value else None for m in trend]
    
    context = {
        'kpi': kpi,
        'trend': trend,
        'targets': targets,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
        'chart_targets': chart_targets,
        'days_back': days_back,
    }
    
    return render(request, 'scheduling/kpi_detail.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def kpi_calculate(request):
    """
    AJAX endpoint to trigger KPI calculation
    """
    from .models import KPIDefinition
    from . import kpi_tracking
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)
    
    data = json.loads(request.body)
    kpi_id = data.get('kpi_id')
    
    if not kpi_id:
        # Calculate all KPIs
        care_home = request.user.care_home if not request.user.is_superuser else None
        measurements = kpi_tracking.calculate_all_kpis(care_home=care_home)
        
        return JsonResponse({
            'success': True,
            'message': f'{len(measurements)} KPIs calculated successfully',
            'count': len(measurements)
        })
    
    # Calculate specific KPI
    kpi = get_object_or_404(KPIDefinition, id=kpi_id, is_active=True)
    
    # Check access
    if not request.user.is_superuser and kpi.care_home and kpi.care_home != request.user.care_home:
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        measurement = kpi_tracking.record_kpi_measurement(kpi)
        
        return JsonResponse({
            'success': True,
            'message': f'KPI calculated successfully',
            'value': float(measurement.measured_value),
            'status': measurement.status,
            'alert': measurement.alert_message if measurement.alert_generated else None
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def kpi_target_manage(request, kpi_id):
    """
    Manage KPI targets
    """
    from .models import KPIDefinition, KPITarget
    
    kpi = get_object_or_404(KPIDefinition, id=kpi_id, is_active=True)
    
    # Check access
    if not request.user.is_superuser and kpi.care_home and kpi.care_home != request.user.care_home:
        messages.error(request, 'Access denied to this KPI.')
        return redirect('kpi_dashboard')
    
    if request.method == 'POST':
        # Create or update target
        year = int(request.POST.get('year'))
        quarter = request.POST.get('quarter')
        month = request.POST.get('month')
        target_value = Decimal(request.POST.get('target_value'))
        stretch_target = request.POST.get('stretch_target')
        minimum_acceptable = request.POST.get('minimum_acceptable')
        notes = request.POST.get('notes', '')
        
        # Convert to integers or None
        quarter = int(quarter) if quarter else None
        month = int(month) if month else None
        stretch_target = Decimal(stretch_target) if stretch_target else None
        minimum_acceptable = Decimal(minimum_acceptable) if minimum_acceptable else None
        
        # Create or update
        target, created = KPITarget.objects.update_or_create(
            kpi=kpi,
            year=year,
            quarter=quarter,
            month=month,
            defaults={
                'target_value': target_value,
                'stretch_target': stretch_target,
                'minimum_acceptable': minimum_acceptable,
                'notes': notes,
                'set_by': request.user,
            }
        )
        
        action = 'created' if created else 'updated'
        messages.success(request, f'Target {action} successfully.')
        return redirect('kpi_target_manage', kpi_id=kpi_id)
    
    # GET request - show form
    current_year = timezone.now().year
    years = range(current_year, current_year + 3)
    
    # Get existing targets
    targets = KPITarget.objects.filter(kpi=kpi).order_by('-year', 'month', 'quarter')
    
    context = {
        'kpi': kpi,
        'targets': targets,
        'years': years,
    }
    
    return render(request, 'scheduling/kpi_target_manage.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def kpi_executive_summary(request):
    """
    Executive summary report with all KPIs
    """
    from .models import KPIDefinition, KPIMeasurement
    from . import kpi_tracking
    
    # Get user's care home
    care_home = request.user.care_home if not request.user.is_superuser else None
    
    # Get period
    period = request.GET.get('period', 'month')
    
    # Get summary
    summary = kpi_tracking.get_kpi_summary(care_home=care_home, period=period)
    
    # Get all categories
    categories = ['STAFFING', 'OCCUPANCY', 'FINANCIAL', 'COMPLIANCE', 'QUALITY', 'EFFICIENCY']
    
    category_data = {}
    for category in categories:
        kpis = KPIDefinition.objects.filter(
            is_active=True,
            category=category
        )
        
        if care_home:
            kpis = kpis.filter(Q(care_home=care_home) | Q(care_home__isnull=True))
        
        kpi_list = []
        for kpi in kpis:
            latest = KPIMeasurement.objects.filter(kpi=kpi).order_by('-measurement_date').first()
            if latest:
                kpi_list.append({
                    'definition': kpi,
                    'measurement': latest,
                })
        
        category_data[category] = kpi_list
    
    context = {
        'summary': summary,
        'category_data': category_data,
        'period': period,
        'care_home': care_home,
    }
    
    return render(request, 'scheduling/kpi_executive_summary.html', context)


# ==========================================================================================
# DATA VISUALIZATION VIEWS (PHASE 3 - TASK 29)
# ==========================================================================================

@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def dashboard_builder(request):
    """
    Dashboard builder interface - list and manage dashboards
    """
    from .models import DashboardLayout
    
    # Get user's care home
    care_home = request.user.care_home if not request.user.is_superuser else None
    
    # Get user's dashboards
    my_dashboards = DashboardLayout.objects.filter(created_by=request.user)
    
    # Get public dashboards
    public_dashboards = DashboardLayout.objects.filter(
        is_public=True
    ).exclude(created_by=request.user)
    
    if care_home:
        public_dashboards = public_dashboards.filter(
            Q(care_home=care_home) | Q(care_home__isnull=True)
        )
    
    # Get default dashboard
    default_dashboard = None
    if care_home:
        default_dashboard = DashboardLayout.objects.filter(
            care_home=care_home,
            is_default=True
        ).first()
    
    context = {
        'my_dashboards': my_dashboards,
        'public_dashboards': public_dashboards,
        'default_dashboard': default_dashboard,
        'care_home': care_home,
    }
    
    return render(request, 'scheduling/dashboard_builder.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def dashboard_view(request, dashboard_id):
    """
    View a specific dashboard with all widgets
    """
    from .models import DashboardLayout, ChartWidget
    from . import visualization_engine
    
    dashboard = get_object_or_404(DashboardLayout, id=dashboard_id)
    
    # Check access
    if not request.user.is_superuser:
        if not dashboard.is_public and dashboard.created_by != request.user:
            messages.error(request, 'Access denied to this dashboard.')
            return redirect('dashboard_builder')
    
    # Get widgets
    widgets = dashboard.widgets.all()
    
    # Prepare widget data
    widget_data = []
    for widget in widgets:
        # Fetch data based on widget configuration
        care_home = dashboard.care_home if dashboard.care_home else request.user.care_home
        
        try:
            if widget.chart_type == 'STAT_CARD':
                data = visualization_engine.generate_stat_card_data(
                    widget.data_source,
                    widget.data_config,
                    care_home=care_home
                )
                chart_config = None
            else:
                result = visualization_engine.fetch_widget_data(
                    widget.data_source,
                    widget.data_config,
                    care_home=care_home
                )
                data = result['data']
                chart_config = result['chart_config']
            
            widget_data.append({
                'widget': widget,
                'data': data,
                'chart_config': json.dumps(chart_config) if chart_config else None
            })
        except Exception as e:
            # Log error but continue with other widgets
            widget_data.append({
                'widget': widget,
                'data': {'error': str(e)},
                'chart_config': None
            })
    
    context = {
        'dashboard': dashboard,
        'widget_data': widget_data,
    }
    
    return render(request, 'scheduling/dashboard_view.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def dashboard_create(request):
    """
    Create a new dashboard
    """
    from .models import DashboardLayout, CareHome
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        care_home_id = request.POST.get('care_home')
        is_public = request.POST.get('is_public') == 'on'
        is_default = request.POST.get('is_default') == 'on'
        
        # Validate
        if not name:
            messages.error(request, 'Dashboard name is required.')
            return redirect('dashboard_create')
        
        # Get care home
        care_home = None
        if care_home_id:
            care_home = CareHome.objects.get(id=care_home_id)
        
        # Create dashboard
        dashboard = DashboardLayout.objects.create(
            name=name,
            description=description,
            care_home=care_home,
            created_by=request.user,
            is_public=is_public,
            is_default=is_default,
            layout_config={'grid': {'rows': 3, 'cols': 3}}
        )
        
        # If set as default, unset other defaults
        if is_default and care_home:
            DashboardLayout.objects.filter(
                care_home=care_home,
                is_default=True
            ).exclude(id=dashboard.id).update(is_default=False)
        
        messages.success(request, f'Dashboard "{name}" created successfully.')
        return redirect('dashboard_edit', dashboard_id=dashboard.id)
    
    # GET - show form
    care_homes = CareHome.objects.all()
    
    context = {
        'care_homes': care_homes,
    }
    
    return render(request, 'scheduling/dashboard_create.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def dashboard_edit(request, dashboard_id):
    """
    Edit dashboard - add/remove/configure widgets
    """
    from .models import DashboardLayout, ChartWidget
    
    dashboard = get_object_or_404(DashboardLayout, id=dashboard_id)
    
    # Check ownership
    if dashboard.created_by != request.user and not request.user.is_superuser:
        messages.error(request, 'You can only edit your own dashboards.')
        return redirect('dashboard_builder')
    
    # Get widgets
    widgets = dashboard.widgets.all()
    
    # Widget options
    chart_types = ChartWidget.CHART_TYPES
    data_sources = ChartWidget.DATA_SOURCES
    
    context = {
        'dashboard': dashboard,
        'widgets': widgets,
        'chart_types': chart_types,
        'data_sources': data_sources,
    }
    
    return render(request, 'scheduling/dashboard_edit.html', context)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def widget_add(request, dashboard_id):
    """
    AJAX endpoint to add a widget to dashboard
    """
    from .models import DashboardLayout, ChartWidget
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)
    
    dashboard = get_object_or_404(DashboardLayout, id=dashboard_id)
    
    # Check ownership
    if dashboard.created_by != request.user and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    data = json.loads(request.body)
    
    title = data.get('title')
    chart_type = data.get('chart_type')
    data_source = data.get('data_source')
    data_config = data.get('data_config', {})
    grid_position = data.get('grid_position', {})
    
    # Validate
    if not title or not chart_type or not data_source:
        return JsonResponse({
            'success': False,
            'error': 'Title, chart type, and data source are required'
        }, status=400)
    
    # Create widget
    widget = ChartWidget.objects.create(
        dashboard=dashboard,
        title=title,
        chart_type=chart_type,
        data_source=data_source,
        data_config=data_config,
        grid_position=grid_position
    )
    
    return JsonResponse({
        'success': True,
        'message': 'Widget added successfully',
        'widget_id': widget.id
    })


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def widget_delete(request, widget_id):
    """
    Delete a widget
    """
    from .models import ChartWidget
    
    widget = get_object_or_404(ChartWidget, id=widget_id)
    
    # Check ownership
    if widget.dashboard.created_by != request.user and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('dashboard_builder')
    
    dashboard_id = widget.dashboard.id
    widget.delete()
    
    messages.success(request, 'Widget deleted successfully.')
    return redirect('dashboard_edit', dashboard_id=dashboard_id)


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def dashboard_delete(request, dashboard_id):
    """
    Delete a dashboard
    """
    from .models import DashboardLayout
    
    dashboard = get_object_or_404(DashboardLayout, id=dashboard_id)
    
    # Check ownership
    if dashboard.created_by != request.user and not request.user.is_superuser:
        messages.error(request, 'You can only delete your own dashboards.')
        return redirect('dashboard_builder')
    
    name = dashboard.name
    dashboard.delete()
    
    messages.success(request, f'Dashboard "{name}" deleted successfully.')
    return redirect('dashboard_builder')


@login_required
@user_passes_test(lambda u: u.is_manager or u.is_head_of_service or u.is_superuser)
def widget_preview(request):
    """
    AJAX endpoint to preview widget data
    """
    from . import visualization_engine
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)
    
    data = json.loads(request.body)
    
    chart_type = data.get('chart_type')
    data_source = data.get('data_source')
    data_config = data.get('data_config', {})
    
    care_home = request.user.care_home if not request.user.is_superuser else None
    
    try:
        if chart_type == 'STAT_CARD':
            result = visualization_engine.generate_stat_card_data(
                data_source,
                data_config,
                care_home=care_home
            )
            chart_config = None
        else:
            result = visualization_engine.fetch_widget_data(
                data_source,
                data_config,
                care_home=care_home
            )
            chart_config = result['chart_config']
        
        return JsonResponse({
            'success': True,
            'data': result,
            'chart_config': chart_config
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
